#*********************************************************************************************************************
import re
from PySide6.QtCore import (Qt, QTimer, QDate, Signal,QEvent)
from PySide6 import QtCore
from PySide6.QtWidgets import (QMainWindow, QMessageBox, QPushButton,
                               QLabel, QVBoxLayout,
                               QMenu,QTableWidgetItem,QCheckBox,QApplication,QToolButton,QHeaderView,QCompleter,
                               QComboBox,QInputDialog,QProgressDialog,QDialog,QLineEdit,QWidget,QHBoxLayout,QProgressBar,QFormLayout,QTextEdit)
from PySide6.QtGui import (QDoubleValidator, QIcon, QColor, QPixmap,QBrush,
                           QAction,QMovie,QShortcut,QKeySequence,QPainterPath,QPainter)
from login import Login
from mane_python import Ui_MainWindow
from database import DataBase
import sys
import locale
from config_senha import TrocarSenha
from threading import Thread
from PySide6.QtWidgets import QSystemTrayIcon
from atualizarprodutos import AtualizarProduto
from tabelaprodutos import TabelaProdutos
from configuracoes import Configuracoes_Login
from tabelausuario import TabelaUsuario
from atualizarusuario import AtualizarUsuario
from pg_configuracoes import Pagina_Configuracoes
from estoqueprodutos import EstoqueProduto
from historicousuario import Pagina_Usuarios
from shiboken6 import isValid
from utils import DialogoAvatar
from utils import DialogoRecorteImagem
from utils import MostrarSenha,configurar_frame_valores,caminho_recurso,salvar_dialogo_memoria,abrir_dialogo_memoria
from utils import Temas
from utils import salvar_imagem_otimizada
from tutorial_overlay import TutorialOverlay
from clientes_juridicos import Clientes_Juridicos
from clientes_fisicos import Clientes_Fisicos
from dialogos import EscolherPlanilhaDialog
from dotenv import load_dotenv
from download_thread import DownloadThread
import smtplib
from email.message import EmailMessage
import json
import sqlite3
import os
import uuid
import datetime
import traceback
from datetime import datetime
from datetime import timedelta
import random
import string
import subprocess
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import requests
from packaging import version  # Para comparar versões
import shutil
import atexit
import socket
 

VERSAO_ATUAL = "1.1.11"  # O arquivo versao.json precisa ser maior que essa para chegar a atualização

TEMAS_PROGRESSO = {
            "escuro": {
                "texto": "#FFFFFF",
                "fundo": "#2E2E2E",
                "chunk": "#0078D4",
            },
            "claro": {
                "texto": "#000000",
                "fundo": "#E6E6E6",
                "chunk": "#00C853",
            },
            "classico": {
                "texto": "#000000",
                "fundo": "#D6D6D6",
                "chunk": "#0063B1",
            }
        }

ASSUNTOS_CONFIG = {
    "Sugestão de melhoria": {
        "label": "Descreva sua sugestão de melhoria (máx. 500 caracteres):",
        "email_titulo": "Nova sugestão de melhoria recebida pelo Sistema de Gerenciamento"
    },
    "Inclusão de funcionalidade": {
        "label": "Descreva qual funcionalidade você deseja sugerir (máx. 500 caracteres):",
        "email_titulo": "Nova sugestão de inclusão de funcionalidade recebida pelo Sistema de Gerenciamento"
    },
    "Relatar bug": {
        "label": "Descreva o bug encontrado (máx. 500 caracteres):",
        "email_titulo": "Novo relato de bug recebido pelo Sistema de Gerenciamento"
    },
    "Dúvida": {
        "label": "Descreva sua dúvida (máx. 500 caracteres):",
        "email_titulo": "Nova dúvida recebida pelo Sistema de Gerenciamento"
    },
    "Outro": {
        "label": "Descreva sua mensagem (máx. 500 caracteres):",
        "email_titulo": "Nova mensagem recebida pelo Sistema de Gerenciamento"
    }
}

PASTA_CONFIG = os.path.join(os.getenv("APPDATA") or ".", "SistemaGerenciamento")
os.makedirs(PASTA_CONFIG, exist_ok=True)

def log_excecao(tipo, valor, tb):
    arq = os.path.join(PASTA_CONFIG, "erros.log")
    with open(arq, "a", encoding="utf-8") as f:
        f.write("\n" + "="*60 + "\n")
        f.write(datetime.now().isoformat() + "\n")
        f.write("".join(traceback.format_exception(tipo, valor, tb)))

sys.excepthook = log_excecao

class MainWindow(QMainWindow, Ui_MainWindow):
    fechar_janela_login_signal = Signal(str)
    def __init__(self, user=None, login_window=None, tipo_usuario=None, connection=None,app=None):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        self.aplicar_icones()
        self.setWindowTitle("Sistema de Gerenciamento")
        self.historico_pausado = False
        self.historico_pausado_clientes_juridicos = False
        self.historico_pausado_clientes_fisicos = False
        self.historico_usuario_pausado = False
        self.imagem_removida_usuario = False
        self.usuario_tem_imagem_salva = False
        # Inicialize o banco de dados antes de qualquer operação que dependa dele
        self.db = DataBase('banco_de_dados.db')
        self.iniciar_verificacao_primeiro_acesso()
        self.temas = Temas()
        self.atalhos = {}  # dicionário com os atalhos definidos
        atexit.register(self.aplicar_atualizacao_automatica)
        
        # Carregar tema do sistema ou  configuração
        config = self.temas.carregar_config_arquivo()  # função de config do JSON
        self.tema_atual = config.get("tema", "claro")
        
        self.tray = QSystemTrayIcon(self)
        self.tray.setIcon(QIcon(caminho_recurso("imagens/favicon.ico")))
        self.tray.setVisible(True)  # importante: tem que ficar visível
            
        self.tipo_usuario = (tipo_usuario or "").strip().lower()
        
        
        self.limpar_pycache_pendente()
        
        self.janela_config = None

        self.atualizador_ja_iniciado = False
        

        self.table_base.verticalHeader().setFixedWidth(20)  # você pode ajustar o valor

        # Atalho F5 global
        self.atalho_f5 = QShortcut(QKeySequence("F5"), self)
        self.atalho_f5.activated.connect(self.tratar_f5_global)

        # Dentro do seu método, após criar o QDateEdit:
        calendario = self.dateEdit_3.calendarWidget()  # pega o QCalendarWidget do QDateEdit
        calendario.setFixedSize(220, 200)           # define tamanho menor do popup
        
        
        self.login_window = login_window
        # Se já vier uma conexão, usa ela. Caso contrário, cria uma nova
        if connection is None:
            self.connection = sqlite3.connect(self.db.db_path)
        else:
            self.connection = connection


        self.exibir_senha = MostrarSenha(self,self.txt_senha_cadastro)
        self.exibir_senha_usuario = MostrarSenha(self,self.txt_confirmar_senha)
        
        self.carregar_env()


        # Mapeia os campos com identificadores únicos
        self.campos_com_autocomplete = {
            "txt_nome": self.txt_nome,
            "txt_cep": self.txt_cep,
            "txt_cliente_3": self.txt_cliente_3,
            "txt_codigo_item": self.txt_codigo_item,
            "txt_complemento": self.txt_complemento,
            "txt_confirmar_senha": self.txt_confirmar_senha,
            "txt_senha_cadastro": self.txt_senha_cadastro,
            "txt_data_nascimento": self.txt_data_nascimento,
            "txt_cpf": self.txt_cpf,
            "txt_descricao_produto_3": self.txt_descricao_produto_3,
            "txt_endereco": self.txt_endereco,
            "txt_numero": self.txt_numero,
            "txt_produto": self.txt_produto,
            "txt_valor_produto_3": self.txt_valor_produto_3,
            "txt_telefone": self.txt_telefone,
            "txt_email": self.txt_email,
            "txt_rg": self.txt_rg,
            "txt_quantidade": self.txt_quantidade,
            "line_clientes": self.line_clientes,
            "txt_usuario_cadastro": self.txt_usuario_cadastro,
            "txt_cnpj": self.txt_cnpj
        }
        
        self.table_base.verticalHeader().setVisible(True)
        self.table_saida.verticalHeader().setVisible(True)
        self.table_ativos.verticalHeader().setVisible(True)
        self.table_saida.horizontalHeader().setVisible(True)
        self.table_inativos.verticalHeader().setVisible(True)
        self.table_massa_usuarios.verticalHeader().setVisible(True)
        self.table_massa_produtos.verticalHeader().setVisible(True)
        self.table_clientes_fisicos.verticalHeader().setVisible(True)
        self.table_clientes_juridicos.verticalHeader().setVisible(True)
        self.table_base.setShowGrid(True)
        self.table_saida.setShowGrid(True)
        self.table_ativos.setShowGrid(True)
        self.table_inativos.setShowGrid(True)
        self.table_massa_usuarios.setShowGrid(True)
        self.table_massa_produtos.setShowGrid(True)
        
        

        # funções que precisam do banco de dados
        self.erros_frames_produtos()
        self.erros_frames_usuarios()
        
        # Carregar informações ao iniciar
        self.carregar_informacoes_tabelas()

        # Exibir notificação de status de conexão
        self.exibir_notificacao()

        # Inicializar as configurações antes de chamar fazer_login_automatico
        self.config = Configuracoes_Login(self)

        for acao,tecla in self.config.obter_todos_atalhos().items():
            self.registrar_atalhos(acao,tecla)


        # Aplica o tamanho da fonte salvo no JSON
        self.definir_tamanho_fonte(self.config.tamanho_fonte_percentual)
        
        self._campos_select_all = set()

        # Aplica completer individual a cada campo
        for nome_campo, campo in self.campos_com_autocomplete.items():
            historico = self.config.carregar_historico_autocompletar(nome_campo)
            
            completer = QCompleter(historico,self)
            completer.setCaseSensitivity(Qt.CaseInsensitive)
            completer.setCompletionMode(QCompleter.PopupCompletion)

            campo.setCompleter(completer)
            
            #  duplo clique = selecionar tudo (via eventFilter)
            campo.installEventFilter(self)
            self._campos_select_all.add(campo)
            
            campo.editingFinished.connect(lambda nc=nome_campo, c=campo: self.auto_completar(nc, c))
        

        # Caminho para o arquivo GIF
        gif_path = caminho_recurso("imagens/oie_3193441C0mFhFhk.gif")

        # Configurar o QMovie com o GIF
        self.movie = QMovie(gif_path)
        self.label_imagem_sistema.setMovie(self.movie)
        # Iniciar a animação
        self.movie.start()


        # Configuração do produto
        self.produto_original = {}
        self.produtos_pendentes = []
        self.imagem_carregada_produto = None
        self.imagem_usuario_carregada = None
    
        # Variáveis para armazenar o estado de edição e o ID do usuário selecionado
        self.is_editing = False
        self.selected_user_id = None
        self.is_editing_produto = False
        self.selected_produto_id = None
        
        # Criação dos botões      
        self.btn_avancar = QPushButton(self)
        self.btn_avancar.setGeometry(35, 5, 30, 30)
        self.btn_avancar.setToolTip("Avançar")  # Adiciona uma dica de ferramenta
      
        
        self.btn_retroceder = QPushButton(self)
        self.btn_retroceder.setGeometry(5, 5, 30, 30) 
        self.btn_retroceder.setToolTip("Retroceder") # Adiciona uma dica de ferramenta
        
    

        # Criar o menu dentro do botão btn_opcoes
        self.menu_opcoes = QMenu(self.btn_mais_opcoes)
        
        self.pagina_configuracoes = Pagina_Configuracoes(self,self,self,self.paginas_sistemas,self.centralwidget,
                                                         self.frame_pag_estoque,self.frame_line_cadastro_produtos,
                                                         self.pg_cadastrar_usuario,self.frame_pag_cadastrar_usuario,
                                                         self.btn_mais_opcoes,self.btn_avancar,self.btn_retroceder,self.btn_home,self.btn_verificar_estoque,
                                                         self.btn_cadastrar_produto, self.btn_cadastrar_usuarios, self.btn_clientes,
                                                         self.btn_gerar_saida,self.btn_gerar_estorno,
                                                         self.label_cadastramento,self.label_cadastramento_produtos,self.frame_valor_total_produtos,
                                                         self.frame_valor_do_desconto,self.frame_valor_com_desconto1,self.frame_quantidade,self.login_window)

        


        self.pagina_clientes_juridicos = Clientes_Juridicos(self.line_clientes,self,self.btn_adicionar_cliente_juridico,self.btn_editar_clientes,
                                        self.btn_excluir_clientes,self.btn_gerar_relatorio_clientes,self.btn_marcar_como_clientes,
                                        self.btn_historico_clientes)
        
        self.pagina_clientes_fisicos = Clientes_Fisicos(self.line_clientes_fisicos,self,self.btn_adicionar_cliente_fisico,self.btn_editar_clientes_fisicos,
                                                        self.btn_excluir_clientes_fisicos,self.btn_gerar_relatorio_clientes_fisicos,self.btn_historico_clientes_fisicos,
                                                        self.btn_marcar_como_clientes_fisicos)
        
        

        # Crie o layout para o frame_imagem_cadastro e adicione o QLabel
        self.label_imagem_usuario = QLabel(self)
        layout_usuario = QVBoxLayout()  # Definindo layout_usuario aqui
        layout_usuario.addWidget(self.label_imagem_usuario)
        self.frame_imagem_cadastro.setLayout(layout_usuario)
        
        


        
        self.frame_imagem_produto_3.setLayout(QVBoxLayout())
        self.label_imagem_produto = QLabel(self.frame_imagem_produto_3)
        self.label_imagem_produto.setGeometry(0, 0, self.frame_imagem_produto_3.width(), self.frame_imagem_produto_3.height())
        self.label_imagem_produto.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label_imagem_produto.setScaledContents(False)
        self.frame_imagem_produto_3.layout().addWidget(self.label_imagem_produto)  # Adicionar QLabel ao layout do frame
        self.imagem_carregada_produto = False


        # Criar as ações do menu
        self.action_sair = QAction("Sair do Sistema", self)
        self.action_configuracoes = QAction("Configurações", self)
        self.action_contato = QAction("Contato", self)
        self.action_reiniciar = QAction("Reiniciar Sistema", self)
        self.action_planilhas_exemplo = QAction("Planilhas de exemplo", self)
        self.action_em_massa_produtos = QAction("Cadastrar produtos em massa", self)
        self.action_em_massa_usuarios = QAction("Cadastrar usuários em massa", self)
        self.action_informacoes_sistema = QAction("Informações do sistema", self)
        self.action_limpar_cache = QAction("Limpar Cache", self)
        

        # Adicionar as ações ao menu (FUNDAMENTAL!)
        self.menu_opcoes.addAction(self.action_sair)
        self.menu_opcoes.addAction(self.action_configuracoes)
        self.menu_opcoes.addAction(self.action_contato)
        self.menu_opcoes.addAction(self.action_reiniciar)
        self.menu_opcoes.addAction(self.action_planilhas_exemplo)
        self.menu_opcoes.addAction(self.action_em_massa_produtos)
        self.menu_opcoes.addAction(self.action_em_massa_usuarios)
        self.menu_opcoes.addAction(self.action_informacoes_sistema)
        self.menu_opcoes.addAction(self.action_limpar_cache)
        

        # Associar o menu ao botão
        self.btn_mais_opcoes.setMenu(self.menu_opcoes)
        self.btn_mais_opcoes.setPopupMode(QToolButton.InstantPopup)
        
        # --- Caixa de pesquisa personalizada ---
        self.widget_pesquisa = QWidget(self)
        self.widget_pesquisa.setFixedSize(260, 55)

        # Layout principal vertical
        layout_pesquisa_vertical = QVBoxLayout(self.widget_pesquisa)
        layout_pesquisa_vertical.setContentsMargins(2, 2, 2, 2)
        layout_pesquisa_vertical.setSpacing(2)

        # Layout horizontal para QLineEdit e botão fechar
        layout_linha_superior = QHBoxLayout()
        layout_linha_superior.setContentsMargins(0, 0, 0, 0)
        
        # Adiciona a contagem de resultados
        self.label_contagem = QLabel("")
        layout_linha_superior.addWidget(self.label_contagem)
        
        # Adiciona botões de navegação
        self.btn_anterior = QPushButton("↑")
        self.btn_anterior.setFixedSize(25, 25)
        self.btn_anterior.setEnabled(False)  # Começa desabilitado
        layout_linha_superior.addWidget(self.btn_anterior)
        
        self.btn_proximo = QPushButton("↓")
        self.btn_proximo.setFixedSize(25, 25)
        self.btn_proximo.setEnabled(False)  # Começa desabilitado
        layout_linha_superior.addWidget(self.btn_proximo)

        self.caixa_pesquisa = QLineEdit()
        self.caixa_pesquisa.setPlaceholderText("Procurar na página")
        layout_linha_superior.addWidget(self.caixa_pesquisa)

        self.checkbox_maiusculas = QCheckBox("Diferenciar maiúsculas")

        # Monta o layout
        layout_pesquisa_vertical.addLayout(layout_linha_superior)
        layout_pesquisa_vertical.addWidget(self.checkbox_maiusculas)
        

        btn_fechar = QPushButton("X")
        btn_fechar.setFixedSize(25, 25)
        btn_fechar.clicked.connect(self.pagina_configuracoes.fechar_pesquisa)
        layout_linha_superior.addWidget(btn_fechar)

        # Esconde no início
        self.widget_pesquisa.hide()
        
        # Atalho ESC para fechar a pesquisa
        atalho_esc = QShortcut(QKeySequence("Esc"), self.widget_pesquisa)
        atalho_esc.activated.connect(self.pagina_configuracoes.fechar_pesquisa)

        self.pagina_configuracoes.configurar_pesquisa()

        # Conectar as ações do menu aos slots correspondentes
        self.action_sair.triggered.connect(self.desconectarUsuario)
        self.action_configuracoes.triggered.connect(self.show_menu_opcoes)
        self.action_contato.triggered.connect(self.show_pg_contato)
        self.action_reiniciar.triggered.connect(self.reiniciar_sistema)
        self.action_planilhas_exemplo.triggered.connect(self.exibir_planilhas_exemplo)
        self.action_em_massa_produtos.triggered.connect(self.abrir_pg_massa_produtos)
        self.action_em_massa_usuarios.triggered.connect(self.abrir_pg_massa_usuarios)
        self.action_informacoes_sistema.triggered.connect(self.show_mensagem_sistema)
        self.action_limpar_cache.triggered.connect(self.limpar_cache_sistema)

        self.fechar_janela_login_signal.connect(self.fechar_janela_login)


        

        

        self.carregar_configuracoes()
        
        

        self.pagina_usuarios = Pagina_Usuarios(self, self.btn_cadastrar_novo_usuario,
                                               self.btn_historico_usuarios,self.btn_atualizar_ativos,
                                               self.btn_atualizar_inativos,self.btn_limpar_tabelas_usuarios,self.btn_gerar_saida_usuarios,
                                               self.btn_importar_usuarios,
                                               self.btn_abrir_planilha_massa_usuarios,self.btn_fazer_cadastro_massa_usuarios,self.progress_massa_usuarios,
                                               self.line_edit_massa_usuarios)

        self.estoque_produtos = EstoqueProduto(self,self.btn_gerar_excel,self.btn_gerar_estorno,
                                               self.btn_gerar_saida,self.btn_limpar_tabelas,
                                               self.btn_atualizar_saida,self.btn_atualizar_estoque,self.btn_historico,self.btn_fazer_cadastro_massa_produtos,
                                               self.btn_abrir_planilha_massa_produtos,self.progress_massa_produtos,self.line_edit_massa_produtos)
        
        self.configuracoes_senha = TrocarSenha(self)  

        

        # Criar instância de TabelaProdutos passando uma referência à MainWindow
        self.atualizar_produto_dialog = AtualizarProduto(self)

        # Criar instância de AtualizarProduto passando uma referência à MainWindow e à tabela de produtos
        self.tabela_produtos_dialog = TabelaProdutos(self, self.dateEdit_3)

        # Criar instância da TabelaUsuario
        self.tabela_usuario_dialogo = TabelaUsuario(self)
        
        self.produto_id = None

        self.criar_botoes_avancar_voltar()

        self.dateEdit_3.setDate(QDate.currentDate())

        self.txt_cep.textChanged.connect(self.formatar_cep)
        self.txt_cpf.textChanged.connect(self.formatar_cpf)
        self.txt_rg.textChanged.connect(self.formatar_rg)
        self.txt_cnpj.textChanged.connect(self.formatar_cnpj)
        self.txt_data_nascimento.textChanged.connect(self.formatar_data_nascimento)
        self.txt_telefone.textChanged.connect(self.formatar_telefone)
   

        self.btn_editar_cadastro.clicked.connect(self.mostrar_tabela_usuarios)
        self.btn_carregar_imagem_4.clicked.connect(self.carregar_imagem_usuario)
 

        # Lista de páginas na ordem desejada
        self.paginas = [self.home_pag, self.pag_estoque, self.pg_cadastrar_produto, 
                        self.pg_cadastrar_usuario, self.pg_clientes,
                        self.pg_contato]
        self.pagina_atual_index = 0  # Índice da página atual na lista
        self.historico_paginas = []  # Lista para armazenar o histórico de páginas

        # Configurar a página inicial
        self.paginas_sistemas.setCurrentWidget(self.paginas[self.pagina_atual_index])

        self.label_valor_total_produtos = configurar_frame_valores(self.frame_valor_total_produtos, "Valor total de produtos sem desconto")
        self.label_valor_do_desconto = configurar_frame_valores(self.frame_valor_do_desconto, "Valor do desconto")
        self.label_valor_desconto = configurar_frame_valores(self.frame_valor_com_desconto1, "Valor do produto com desconto")
        self.label_quantidade = configurar_frame_valores(self.frame_quantidade, "Quantidade total de produtos",valor_monetario=False)


        # Centralizar o texto 
        for label in [
            self.label_valor_total_produtos,
            self.label_valor_do_desconto,
            self.label_valor_desconto,
            self.label_quantidade,
        ]:
            label.setAlignment(Qt.AlignCenter)

        validator = QDoubleValidator()
        validator.setNotation(QDoubleValidator.StandardNotation)  
        validator.setDecimals(2)
        validator.setTop(1000000000)  
        self.txt_valor_produto_3.setValidator(validator)


        self.txt_valor_produto_3.editingFinished.connect(self.formatar_moeda)

        validador = QDoubleValidator()
        validador.setNotation(QDoubleValidator.StandardNotation)  
        validador.setRange(0.00, 100.00)  
        validador.setDecimals(2)
        self.txt_desconto_3.setValidator(validador)
        self.txt_desconto_3.editingFinished.connect(self.formatar_porcentagem)


        self.btn_home.clicked.connect(lambda: self.paginas_sistemas.setCurrentWidget(self.home_pag))    
        self.btn_ver_item.clicked.connect(lambda: self.paginas_sistemas.setCurrentWidget(self.pag_estoque))
        self.btn_ver_usuario.clicked.connect(lambda: self.paginas_sistemas.setCurrentWidget(self.page_verificar_usuarios))
        self.btn_cadastrar_novo_usuario.clicked.connect(lambda: self.paginas_sistemas.setCurrentWidget(self.pg_cadastrar_usuario))
        self.btn_editar_massa_produtos.clicked.connect(lambda: self.paginas_sistemas.setCurrentWidget(self.pg_cadastrar_produto))
        self.btn_editar_massa_usuario.clicked.connect(lambda: self.paginas_sistemas.setCurrentWidget(self.pg_cadastrar_usuario))
        

        
        self.btn_clientes.clicked.connect(self.mostrar_page_clientes)
        self.btn_verificar_usuarios.clicked.connect(self.mostrar_page_verificar_usuarios)
        self.btn_novo_produto.clicked.connect(self.abrir_pg_cadastrar_produto)
        self.btn_cadastrar_usuarios.clicked.connect(self.abrir_pg_cadastro_usuario)
        self.btn_cadastrar_produto.clicked.connect(self.abrir_pg_cadastrar_produto)
        self.btn_remover_imagem.clicked.connect(self.retirar_imagem_produto)
        self.btn_limpar_campos.clicked.connect(self.apagar_campos_produtos)
        self.btn_adicionar_produto.clicked.connect(self.adicionar_produto)
        self.btn_confirmar.clicked.connect(self.confirmar_produtos)
        self.btn_atualizar_cadastro.clicked.connect(self.atualizar_usuario_no_bd)
        self.btn_verificar_estoque.clicked.connect(self.mostrar_page_estoque)
        self.btn_ver_clientes_juridicos.clicked.connect(self.mostrar_page_clientes)
        self.btn_apagar_cadastro.clicked.connect(self.eliminar_campos_usuarios)
        self.btn_remover_imagem_usuario.clicked.connect(self.retirar_imagem_usuario)
        self.btn_sair_modo_edicao.clicked.connect(self.sair_modo_edicao_usuarios)
        self.btn_sair_modo_edicao_produtos.clicked.connect(self.sair_modo_edicao_produto)

        self.btn_fazer_cadastro.clicked.connect(self.subscribe_user)
        self.btn_editar.clicked.connect(self.mostrar_tabela_produtos)
        self.btn_atualizar_produto.clicked.connect(self.atualizar_produto)
        self.btn_carregar_imagem.clicked.connect(self.carregar_imagem_produto)
        
        # Conecta ao evento de redimensionamento
        self.resizeEvent = lambda event: (self.reposicionar_pesquisa(), QMainWindow.resizeEvent(self, event))

        # Sempre centralizar ao mostrar
        self.widget_pesquisa.showEvent = lambda event: self.reposicionar_pesquisa()

        self.txt_cep.editingFinished.connect(self.on_cep_editing_finished)
        
        
        self.label_avatar.setCursor(Qt.PointingHandCursor)
        self.label_avatar.setAlignment(Qt.AlignCenter)
        self.label_avatar.setFixedSize(40, 40)
                

        # Evento de clique no avatar
        def avatar_click(event):
            if event.button() == Qt.LeftButton:
                self.abrir_modal_avatar()

        self.label_avatar.mousePressEvent = avatar_click


        # Atualizações iniciais
        self.atualizar_usuario_logado()
        self.atualizar_avatar()
        
        usuario_logado = (self.config.usuario or "").strip()
        self.tipo_usuario = (self.db.obter_tipo_usuario(usuario_logado) or "").strip().lower()
        
        QTimer.singleShot(0, lambda: self.aplicar_permissoes_ui(self.tipo_usuario))
       
  
            
    def normalizar_tipo_usuario(tipo_usuario: str | None) -> str:
        if not tipo_usuario:
            return ""
        return str(tipo_usuario).strip().lower()
    
    def aplicar_permissoes_ui(self, tipo_usuario: str | None):
        role = (tipo_usuario or "").strip().lower()
        
        #Exemplo de papéis
        admin = role in ["administrador","admin"]
        user  = role in ["comum","usuário","usuario","user"]
        convidado = role in ["convidado","guest"]
        
        # --- ADMIN: vê tudo ---
        if admin:
            return
        
        if user:
            if hasattr(self, "btn_cadastrar_usuarios"):
                self.btn_cadastrar_usuarios.setVisible(False)
            if hasattr(self, "btn_verificar_usuarios"):
                self.btn_verificar_usuarios.setVisible(False)
            if hasattr(self, "btn_clientes"): 
                self.btn_clientes.setVisible(False)
            if hasattr(self,"btn_historico"):
                self.btn_historico.setVisible(False)
            if hasattr(self, "btn_ver_clientes_juridicos"):
                self.btn_ver_clientes_juridicos.setVisible(False)
            if hasattr(self, "btn_ver_item"):
                self.btn_ver_item.setVisible(False)
            self.action_em_massa_usuarios.setVisible(False)
        if convidado:
            if hasattr(self, "btn_cadastrar_usuarios"):
                self.btn_cadastrar_usuarios.setVisible(False)
            if hasattr(self, "btn_verificar_usuarios"):
                self.btn_verificar_usuarios.setVisible(False)
            if hasattr(self, "btn_clientes"): 
                self.btn_clientes.setVisible(False)
            if hasattr(self,"btn_historico"):
                self.btn_historico.setVisible(False)
            if hasattr(self, "btn_ver_clientes_juridicos"):
                self.btn_ver_clientes_juridicos.setVisible(False)
            if hasattr(self, "btn_ver_item"):
                self.btn_ver_item.setVisible(False)
            self.action_em_massa_usuarios.setVisible(False)
        
    def atualizar_usuario_logado(self):
        nome_completo = self.get_nome_completo_usuario() 

        if not nome_completo:
            self.label_nome_usuario.setText("Nenhum usuário logado")
            return

        self.label_nome_usuario.setText(nome_completo)

        
    def recortar_imagem_circular(self, pixmap, size=300):
        pixmap = pixmap.scaled(
            size, size,
            Qt.KeepAspectRatioByExpanding,
            Qt.SmoothTransformation
        )

        result = QPixmap(size, size)
        result.fill(Qt.transparent)

        painter = QPainter(result)
        painter.setRenderHint(QPainter.Antialiasing)

        path = QPainterPath()
        path.addEllipse(0, 0, size, size)
        painter.setClipPath(path)

        painter.drawPixmap(0, 0, pixmap)
        painter.end()

        return result
    
    def abrir_modal_avatar(self):
        dialog = DialogoAvatar(
            parent=self,
            caminho_imagem=self.caminho_foto_usuario
            if self.caminho_foto_usuario and os.path.exists(self.caminho_foto_usuario)
            else None
        )
        dialog.exec()


    def atualizar_avatar(self):
        usuario = self.get_usuario_logado()
        nome_completo = self.get_nome_completo_usuario() or "Usuário"

        caminho = self.db.obter_imagem_usuario(usuario)
        self.caminho_foto_usuario = caminho

        self.label_avatar.clear()
        self.label_avatar.setAlignment(Qt.AlignCenter)
        self.label_avatar.setStyleSheet("")

        if caminho and os.path.exists(caminho):
            pixmap = QPixmap(caminho)
            pixmap = self.pixmap_circular(pixmap, 40)
            self.label_avatar.setPixmap(pixmap)
        else:
            self.label_avatar.setText(self.gerar_iniciais(nome_completo))
            self.label_avatar.setStyleSheet("""
                background-color: #4f46e5;
                color: white;
                font-weight: bold;
                font-size: 14px;
                border-radius: 20px;
            """)

        self.label_avatar.repaint()


    def ver_foto_usuario(self):
        if not os.path.exists(self.caminho_foto_usuario):
            QMessageBox.information(self, "Foto", "Usuário não possui foto cadastrada.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Foto do Usuário")

        label = QLabel(dialog)
        pixmap = QPixmap(self.caminho_foto_usuario).scaled(
            300, 300, Qt.KeepAspectRatio, Qt.SmoothTransformation
        )
        label.setPixmap(pixmap)

        layout = QVBoxLayout(dialog)
        layout.addWidget(label)

        dialog.exec()
        
    def alterar_foto_usuario(self,retornar_caminho=False):
        caminho = abrir_dialogo_memoria(
            parent=self,
            chave="foto_usuario",
            titulo="Selecionar foto do usuário",
            filtro="Imagens (*.png *.jpg *.jpeg *.webp)"
        )


        if not caminho:
            return

        pixmap = QPixmap(caminho)
        if pixmap.isNull():
            return

        #  abre o preview + recorte
        dialog = DialogoRecorteImagem(pixmap, self)
        if not dialog.exec():
            return

        pixmap_recortado = dialog.obter_pixmap_recortado()

        usuario = self.get_usuario_logado()
        pasta = caminho_recurso("media/usuarios/perfil")
        os.makedirs(pasta, exist_ok=True)

        nome_arquivo = f"{usuario}_{uuid.uuid4().hex}.jpg"
        caminho_final = os.path.join(pasta, nome_arquivo)

        salvar_imagem_otimizada(
            pixmap_recortado,
            caminho_final,
            tamanho_max=512,
            qualidade=75
        )

        # remove imagem antiga
        caminho_antigo = self.db.obter_imagem_usuario(usuario)
        if caminho_antigo and os.path.exists(caminho_antigo):
            os.remove(caminho_antigo)

        self.db.salvar_imagem_usuario(usuario, caminho_final)

        # atualiza estados
        self.caminho_imagem = caminho_final
        self.atualizar_avatar()
        
        if retornar_caminho:
            return caminho_final


        
    def pixmap_circular(self, pixmap, size=40):
        pixmap = pixmap.scaled(
            size, size,
            Qt.KeepAspectRatioByExpanding,
            Qt.SmoothTransformation
        )

        result = QPixmap(size, size)
        result.fill(Qt.transparent)

        painter = QPainter(result)
        painter.setRenderHint(QPainter.Antialiasing)
        path = QPainterPath()
        path.addEllipse(0, 0, size, size)
        painter.setClipPath(path)
        painter.drawPixmap(0, 0, pixmap)
        painter.end()

        return result
#*********************************************************************************************************************
    def gerar_iniciais(self,nome):
        partes = nome.strip().split()
        if len(partes) == 1:
            return partes[0][:2].upper()
        return (partes[0][0] + partes[-1][0]).upper()
#*********************************************************************************************************************
    def get_nome_completo_usuario(self):
        return self.config.obter_nome_completo_usuario()
#*********************************************************************************************************************
    def get_usuario_logado(self):
        # Obtenha o usuário logado das configurações
        return self.config.obter_usuario_logado()
#*********************************************************************************************************************        
    def carregar_env(self):
        if getattr(sys, "frozen", False):
            # MODO EXE → mesma pasta do .exe
            base_path = os.path.dirname(sys.executable)
        else:
            # MODO PYTHON → raiz do projeto
            base_path = os.getcwd()

        env_path = os.path.join(base_path, ".env")
        load_dotenv(env_path)
#*********************************************************************************************************************            
    def caminho_historico_atualizacoes(self):
        if getattr(sys, "frozen", False):
            return os.path.join(
                self.pasta_estado(),
                "historico_atualizacoes.log"
            )
        else:
            return os.path.join(
                self.pasta_do_sistema(),
                "historico_atualizacoes.log"
            )
#*********************************************************************************************************************   
    def pasta_do_sistema(self):
        return os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.getcwd()
#*********************************************************************************************************************    
    def pasta_estado(self):
        base = os.getenv("APPDATA")
        pasta = os.path.join(base, "SistemaGerenciamento")
        os.makedirs(pasta, exist_ok=True)
        return pasta
#*********************************************************************************************************************
    def aplicar_atualizacao_automatica(self):
        try:
            caminho_estado = self.pasta_estado()
            flag_path = os.path.join(caminho_estado, "update_pending.flag")

            # Verifica se há atualização pendente
            if not os.path.exists(flag_path):
                return

            # Lê a versão remota salva
            with open(flag_path, "r", encoding="utf-8") as f:
                versao_remota = f.read().strip()

            # =========================
            # SALVA VERSÃO INSTALADA
            # =========================
            if getattr(sys, "frozen", False):
                versao_path = os.path.join(caminho_estado, "versao_instalada.txt")
            else:
                versao_path = os.path.join(self.pasta_do_sistema(), "versao_instalada.txt")

            with open(versao_path, "w", encoding="utf-8") as f:
                f.write(versao_remota)

            # =========================
            # REGISTRA HISTÓRICO
            # =========================
            try:
                log_path = self.caminho_historico_atualizacoes()
                os.makedirs(os.path.dirname(log_path), exist_ok=True)

                agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                with open(log_path, "a", encoding="utf-8") as log:
                    log.write(
                        f"[{agora}] Atualizado para a versão {versao_remota} | STATUS=SUCESSO\n"
                    )
            except Exception as e:
                print(f"Erro ao gravar histórico: {e}")

            # =========================
            # REMOVE FLAG
            # =========================
            os.remove(flag_path)

        except Exception as e:
            print(f"Erro ao aplicar atualização automática: {e}")

            try:
                log_path = self.caminho_historico_atualizacoes()
                agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                with open(log_path, "a", encoding="utf-8") as log:
                    log.write(
                        f"[{agora}] Falha ao aplicar atualização | STATUS=ERRO | {str(e)}\n"
                    )
            except:
                pass

#*********************************************************************************************************************
    def verificar_atualizacao_automatica(self):
        try:
            url = "https://github.com/KevenLucas025/Sistema-Atualizador/raw/refs/heads/main/versao.json"
            response = requests.get(url, timeout=5)
            dados = response.json()

            versao_remota = dados.get("versao")
            link_download = dados.get("url_download") or dados.get("link_download")
            versao_local = self.versao_instalada_local()

            if not versao_remota or not link_download:
                return

            if version.parse(versao_remota) > version.parse(versao_local):

                if not self.config.atualizacoes_automaticas:
                    if not self.config.nao_mostrar_aviso_atualizacoes:
                        QMessageBox.information(
                            self,
                            "Atualização disponível",
                            f"Uma nova versão ({versao_remota}) está disponível!\n\n"
                            "Vá em Mais opções > Configurações > Atualizações > Verificar se há atualizações para baixar."
                        )
                    return


                # Download automático = agora via thread
                self.baixar_atualizacao(versao_remota, link_download, automatico=True)

        except Exception as e:
            print(f"Erro ao verificar atualização automática: {e}")

    def aplicar_tema_progress(self,progress: QProgressDialog,tema:str):
        cores = TEMAS_PROGRESSO.get(tema,TEMAS_PROGRESSO["claro"])
        
        barra = progress.findChild(QProgressBar)
        label = progress.findChild(QLabel)
        
        if barra:
                barra.setStyleSheet(f"""
                    QProgressBar {{
                        background-color: {cores["fundo"]};
                        border-radius: 13px;
                        height: 20px;
                        text-align: center;
                        padding: 4px;
                    }}

                    QProgressBar::chunk {{
                        background-color: {cores["chunk"]};
                        border-radius: 4px;
                        margin: 0px;
                        min-width: 7px;
                    }}
                """)
        if label:
            label.setStyleSheet(f"""
                QLabel {{
                background: transparent;
                color: {cores["texto"]};
                }}
            
            """)
            
    
    def baixar_atualizacao(self, versao_remota, link_download, automatico=False):
        caminho = self.pasta_do_sistema()
        destino_temp = os.path.join(caminho, "SistemadeGerenciamento_tmp.exe")

        self.download_thread = DownloadThread(link_download, destino_temp)

        # =======================
        # MODO MANUAL (COM BARRA)
        # =======================
        if not automatico:
            self.progress = QProgressDialog("Baixando atualização...", "Cancelar", 0, 100, self.janela_config)
            self.progress.setWindowTitle("Atualização")
            self.progress.setWindowModality(Qt.WindowModal)
            self.progress.show()
            self.aplicar_tema_progress(self.progress, self.tema_atual)

            # Conectar progresso → dialog
            self.download_thread.progresso.connect(self.progress.setValue)

            # Cancelar download
            def cancelar():
                self.download_thread.cancelar()


            self.progress.canceled.connect(cancelar)

        # =======================
        # SINAIS DO THREAD
        # =======================
        def finalizado(destino):
            caminho_estado = self.pasta_estado()
            flag = os.path.join(caminho_estado, "update_pending.flag")
            with open(flag, "w") as f:
                f.write(versao_remota)

            if automatico:
                QMessageBox.information(self, "Atualização baixada",
                    "A atualização foi baixada automaticamente.\nReinicie o sistema.")
            else:
                QMessageBox.information(self, "Concluído",
                    "Download finalizado! Reinicie o sistema para aplicar.")

            if not automatico:
                self.progress.setValue(100)

        def erro(msg):
            QMessageBox.warning(self, "Erro ao baixar", msg)
            if not automatico:
                self.progress.cancel()

        self.download_thread.concluido.connect(finalizado)
        self.download_thread.erro.connect(erro)

        QTimer.singleShot(150, self.download_thread.start)


        
    def versao_instalada_local(self):
        # ==========================
        # MODO EXE → SOMENTE APPDATA
        # ==========================
        if getattr(sys, "frozen", False):
            caminho_appdata = os.path.join(
                self.pasta_estado(),
                "versao_instalada.txt"
            )

            if os.path.exists(caminho_appdata):
                try:
                    versao = open(caminho_appdata, "r", encoding="utf-8").read().strip()
                    if versao:
                        return versao
                except:
                    pass

            return VERSAO_ATUAL

        # ==========================
        # MODO PYTHON → SOMENTE PROJETO
        # ==========================
        caminho_local = os.path.join(
            self.pasta_do_sistema(),
            "versao_instalada.txt"
        )

        if os.path.exists(caminho_local):
            try:
                versao = open(caminho_local, "r", encoding="utf-8").read().strip()
                if versao:
                    return versao
            except:
                pass

        return VERSAO_ATUAL 

    def obter_atualizador(self):
        caminho_sistema = self.pasta_do_sistema()

        if getattr(sys, "frozen", False):
            atualizador = os.path.join(caminho_sistema, "Atualizador.exe")
            cwd = caminho_sistema
        else:
            atualizador = os.path.join(caminho_sistema, "dist", "Atualizador.exe")
            cwd = os.path.join(caminho_sistema, "dist")

        return atualizador, cwd

    def iniciar_atualizador_se_necessario(self):
        if self.atualizador_ja_iniciado:
            return

        caminho_estado = self.pasta_estado()
        caminho_sistema = self.pasta_do_sistema()

        flag = os.path.join(caminho_estado, "update_pending.flag")
        tmp_exe = os.path.join(caminho_sistema, "SistemadeGerenciamento_tmp.exe")

        # ==========================
        # NÃO EXISTE UPDATE REAL
        # ==========================
        if not os.path.exists(tmp_exe):
            if os.path.exists(flag):
                os.remove(flag)
            return

        # ==========================
        # OBTÉM ATUALIZADOR (PY / EXE)
        # ==========================
        atualizador, cwd = self.obter_atualizador()

        # ==========================
        # CHAMA ATUALIZADOR
        # ==========================
        if os.path.exists(flag) and os.path.exists(atualizador):
            try:
                subprocess.Popen(
                    [atualizador],
                    cwd=cwd,
                    creationflags=subprocess.CREATE_NO_WINDOW
                )
                self.atualizador_ja_iniciado = True
            except Exception as e:
                print(f"Erro ao iniciar atualizador: {e}")


    def carregar_config_padrao(self):
        return {
            "tema": "classico",
            "usuario": "",
            "senha": "",
            "mantem_conectado": False,
            "nao_mostrar_mensagem_boas_vindas": False,
            "nao_mostrar_aviso_irreversivel": False,
            "nao_mostrar_mensagem_arquivo_excel": False,
            "nao_mostrar_mensagem_arquivo_excel_fisicos": False,
            "historico_autocompletes": {},
            "nome_usuario": "",
            "email": ""
        }

    def caminho_configuracao_json(self):
        # Usa o MESMO caminho do Configuracoes_Login
        return self.config.caminho_config_json()

    def carregar_config_arquivo(self, caminho=None):
        # Se não passar caminho, usa o padrão correto (AppData no EXE / cwd no Python)
        if caminho is None:
            caminho = self.caminho_configuracao_json()

        if not os.path.exists(caminho):
            print(f"Arquivo {caminho} não encontrado. Usando configuração padrão.")
            return self.carregar_config_padrao()

        try:
            with open(caminho, "r", encoding="utf-8") as f:
                txt = f.read().strip()
                if not txt:
                    print(f"Arquivo {caminho} vazio. Usando configuração padrão.")
                    return self.carregar_config_padrao()
                return json.loads(txt)
        except json.JSONDecodeError as e:
            print(f"Erro ao decodificar {caminho}: {e}. Usando configuração padrão.")
            return self.carregar_config_padrao()
        
    def boas_vindas(self):
        if self.config.nao_mostrar_mensagem_boas_vindas:
            return
        
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Boas Vindas")
        msg.setText(f"{self.config.obter_usuario_logado()}, seja bem-vindo(a) ao sistema de gerenciamento.\n"
                    "É um prazer tê-lo(a) conosco em nosso sistema.\n"
                    "Esperamos que as informações aqui contidas sejam de grande ajuda.\n"
                    "Pedimos que sempre que possível, você possa avaliar e sugerir melhorias.")
        msg.setStandardButtons(QMessageBox.Ok)

        checkbox_nao_mostrar_mensagem = QCheckBox("Não mostrar essa mensagem novamente")
        msg.setCheckBox(checkbox_nao_mostrar_mensagem)
        msg.exec()

        if checkbox_nao_mostrar_mensagem.isChecked():
            self.config.nao_mostrar_mensagem_boas_vindas = True
            self.config.salvar_configuracoes(self.config.usuario, self.config.senha, self.config.mantem_conectado)
    

    def verificar_conexao(self):
        """Verifica se há conexão com a internet."""
        try:
            # Tenta conectar ao Google para verificar a internet
            socket.create_connection(("8.8.8.8", 53), timeout=5)
            return True
        except OSError:
            return False
 
    def exibir_notificacao(self):
        conectado = self.verificar_conexao()

        if conectado:
            return  # Não faz nada se estiver online

        mensagem = "Sem conexão com a internet. Envio de sugestões não irá funcionar corretamente."

        self.tray.showMessage(
            "Status da Conexão",
            mensagem,
            QSystemTrayIcon.Warning, 
            5000
        )
            
    def criar_item(self, texto):
        item = QTableWidgetItem(texto)
        item.setTextAlignment(Qt.AlignCenter)  # Centraliza o texto
        item.setForeground(QBrush(QColor(255, 255, 255)))  # Define a cor do texto como branco
        return item
    

    def carregar_informacoes_tabelas(self):
        # Limpa as tabelas antes de carregar novas informações
        self.table_base.setRowCount(0)
        self.table_saida.setRowCount(0)
        self.table_ativos.setRowCount(0)
        self.table_inativos.setRowCount(0)

        # Carregar dados da `table_base`
        produtos_base = self.db.obter_produtos_base()
        for produto in produtos_base:
            row_position = self.table_base.rowCount()
            self.table_base.insertRow(row_position)
            for col, data in enumerate(produto):
                self.table_base.setItem(row_position, col, self.criar_item(str(data)))

        # Ajusta colunas e linhas automaticamente após preencher
        self.table_base.resizeColumnsToContents()
        self.table_base.resizeRowsToContents()

        # Ajuste refinado nas colunas da table_base
        header = self.table_base.horizontalHeader()
        for i in range(self.table_base.columnCount() - 1):
            header.setSectionResizeMode(i, QHeaderView.ResizeToContents)

        # Última coluna ("Usuário") com tamanho fixo
        header.setSectionResizeMode(self.table_base.columnCount() - 1, QHeaderView.Interactive)
        self.table_base.setColumnWidth(self.table_base.columnCount() - 1, 120)
            

        # Carregar dados da `table_saida`
        produtos_saida = self.db.obter_produtos_saida()
        for produto in produtos_saida:
            row_position = self.table_saida.rowCount()
            self.table_saida.insertRow(row_position)
            for col, data in enumerate(produto):
                self.table_saida.setItem(row_position, col, self.criar_item(str(data)))
        
        self.table_saida.resizeColumnsToContents()
        self.table_saida.resizeRowsToContents()

        # Carregar dados da `table_ativos`   
        usuarios_ativos = self.db.obter_usuarios_ativos()
        for usuario in usuarios_ativos:
            row_position = self.table_ativos.rowCount()
            self.table_ativos.insertRow(row_position)
            for col, data in enumerate(usuario):
                self.table_ativos.setItem(row_position, col,self.criar_item(str(data)))

        self.table_ativos.resizeColumnsToContents()
        self.table_ativos.resizeRowsToContents()

        # Carregar dados da `table_inativos` 
        usuarios_inativos = self.db.obter_usuarios_inativos()
        for usuario in usuarios_inativos:
            row_position = self.table_inativos.rowCount()
            self.table_inativos.insertRow(row_position)
            for col, data in enumerate(usuario):
                self.table_inativos.setItem(row_position, col,self.criar_item(str(data)))

        # Carregar dados da table_clientes_juridicos
        clientes = self.db.obter_clientes_juridicos()
        for cliente in clientes:
            row_position = self.table_clientes_juridicos.rowCount()
            self.table_clientes_juridicos.insertRow(row_position)
            for col, data in enumerate(cliente):
                self.table_clientes_juridicos.setItem(row_position, col,self.criar_item(str(data)))

        # Carregar dados da table_clientes_fisicos
        clientes_fisicos = self.db.obter_clientes_fisicos()
        for cliente in clientes_fisicos:
            row_position = self.table_clientes_fisicos.rowCount()
            self.table_clientes_fisicos.insertRow(row_position)
            for col, data in enumerate(cliente):
                self.table_clientes_fisicos.setItem(row_position,col,self.criar_item(str(data)))

        self.table_inativos.resizeColumnsToContents()
        self.table_inativos.resizeRowsToContents()
        self.table_massa_produtos.resizeColumnsToContents()
        self.table_massa_produtos.resizeRowsToContents()
        self.table_clientes_juridicos.resizeColumnsToContents()
        self.table_clientes_juridicos.resizeRowsToContents()
        self.table_clientes_fisicos.resizeColumnsToContents()
        self.table_clientes_fisicos.resizeRowsToContents()
        
    def enviar_sugestao_email(self, nome, email_usuario, assunto, mensagem):
        EMAIL_SISTEMA = os.getenv("EMAIL_SISTEMA")
        SENHA_APP = os.getenv("EMAIL_APP_SENHA")

        if not EMAIL_SISTEMA or not SENHA_APP:
            raise Exception("Credenciais de e-mail não configuradas.")

        # Configuração baseada no assunto
        config = ASSUNTOS_CONFIG.get(assunto, {})

        titulo_email = config.get(
            "email_titulo",
            "Nova mensagem recebida pelo Sistema de Gerenciamento"
        )

        msg = EmailMessage()
        msg["From"] = EMAIL_SISTEMA
        msg["To"] = EMAIL_SISTEMA
        msg["Reply-To"] = email_usuario
        msg["Subject"] = f"[Sistema] {assunto}"

        # =========================
        # TEXTO SIMPLES (fallback)
        # =========================
        msg.set_content(
            f"Nova mensagem recebida.\n\n"
            f"Nome: {nome}\n"
            f"E-mail: {email_usuario}\n"
            f"Assunto: {assunto}\n\n"
            f"Mensagem:\n{mensagem}"
        )

        # =========================
        # HTML BONITO
        # =========================
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
        <meta charset="UTF-8">
        </head>
        <body style="font-family: Arial, sans-serif; background-color: #f4f6f8; padding: 20px;">

        <div style="
            max-width: 600px;
            margin: auto;
            background: #ffffff;
            border-radius: 8px;
            padding: 24px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        ">

            <h2 style="color: #2c3e50; margin-top: 0;">
            {titulo_email}
            </h2>

            <hr style="border: none; border-top: 1px solid #ddd; margin: 16px 0;">

            <p><strong>👤 Nome:</strong><br>{nome}</p>

            <p><strong>📧 E-mail para contato:</strong><br>{email_usuario}</p>

            <p><strong>📌 Assunto:</strong><br>{assunto}</p>

            <p><strong>📝 Mensagem:</strong></p>

            <div style="
                border: 1px solid #ddd;
                border-radius: 4px;
                padding: 12px;
                background: #fafafa;
                white-space: pre-wrap;
            ">
            {mensagem}
            </div>

            <hr style="border: none; border-top: 1px solid #ddd; margin: 24px 0;">

            <p style="font-size: 12px; color: #777;">
            Mensagem enviada automaticamente pelo Sistema de Gerenciamento.
            </p>

        </div>

        </body>
        </html>
        """

        msg.add_alternative(html, subtype="html")

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL_SISTEMA, SENHA_APP)
            smtp.send_message(msg)

        
    def mostrar_sugestao(self):
        janela = QMainWindow(self)
        janela.setWindowTitle("Feedback")
        janela.resize(500, 350)

        central_widget = QWidget()
        layout_principal = QVBoxLayout(central_widget)
        
        if self.tema_atual == "escuro":
            lineedit_bg = "#303030"
            combobox_style = """
            
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView {
                    background-color: #2b2b2b;
                    color: #f0f0f0;
                    selection-background-color: #696969;
                    border: 1px solid #555555;
                }

                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                } 
            """ 
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """
        elif self.tema_atual == "claro":
            lineedit_bg = "white"
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """
        else:
            lineedit_bg = "white"
            combobox_style = """
            QComboBox {
                background-color: white;
                border: 3px solid rgb(50,150,250);
                border-radius: 5px;
                color: black;
                padding: 5px;
            }

            QComboBox QAbstractItemView {
                background-color: white;
                color: black;
                border: 3px solid white;
                selection-background-color: rgb(120,120,120);
                selection-color: black;
            }

            /* Scrollbar vertical */
            QComboBox QScrollBar:vertical {
                background-color: rgb(240,240,240);  /* Trilha visível */
                width: 10px;
                margin: 1px;
                border: 1px solid white;
                border-radius: 5px;
            }

            /* Alça (handle) da scrollbar */
            QComboBox QScrollBar::handle:vertical {
                background: rgb(120,120,120);  /* Cor da barra frontal,sobe e desce*/
                min-height: 30px;
                border-radius: 5px;
            }

            /* Trilha atrás do handle — essa parte faz toda a diferença */
            QComboBox QScrollBar::add-page:vertical,
            QComboBox QScrollBar::sub-page:vertical {
                background: transparent;  /* barra atrás do scroll */
            }

            /* Esconde setas */
            QComboBox QScrollBar::add-line:vertical,
            QComboBox QScrollBar::sub-line:vertical {
                height: 0px;
                background: none;
                border: none;
            }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """
            

        # =========================
        # FORMULÁRIO (Nome / Assunto)
        # =========================
        form = QFormLayout()

        line_nome = QLineEdit()
        line_nome.setPlaceholderText("Digite seu nome")
        
        line_email = QLineEdit()
        line_email.setPlaceholderText("Digite seu e-mail")
        
        line_nome.setStyleSheet(lineedit_style)
        line_email.setStyleSheet(lineedit_style)

        assunto = QComboBox()
        assunto.addItems([
            "Selecione um assunto",
            "Sugestão de melhoria",
            "Inclusão de funcionalidade",
            "Relatar bug",
            "Dúvida",
            "Outro"
        ])
        assunto.setCurrentIndex(0)
        assunto.setObjectName("combobox_assunto")
        assunto.setStyleSheet(combobox_style)
        

        form.addRow("Nome:", line_nome)
        form.addRow("E-mail",line_email)
        form.addRow("Assunto:", assunto)
        

        layout_principal.addLayout(form)
        
        label_msg = QLabel("Descreva sua sugestão (máx. 500 caracteres):")

        # =========================
        # CAMPO DE TEXTO (500 chars)
        # =========================
        def atualizar_label_assunto():
            assunto_sel = assunto.currentText()

            config = ASSUNTOS_CONFIG.get(assunto_sel)
            if config:
                label_msg.setText(config["label"])
            else:
                label_msg.setText("Descreva sua mensagem (máx. 500 caracteres):")

        assunto.currentIndexChanged.connect(atualizar_label_assunto)


        texto = QTextEdit()
        texto.setPlaceholderText("Digite aqui sua sugestão...")
        texto.setMaximumHeight(120)

        # "Quadrado" (borda)
        texto.setStyleSheet("""
            QTextEdit {
                border: 1px solid #999;
                border-radius: 4px;
                padding: 6px;
            }
        """)

        contador = QLabel("0 / 500 caracteres")

        def atualizar_contador():
            conteudo = texto.toPlainText()
            if len(conteudo) > 500:
                texto.blockSignals(True)
                texto.setPlainText(conteudo[:500])
                texto.blockSignals(False)
                cursor = texto.textCursor()
                cursor.movePosition(cursor.End)
                texto.setTextCursor(cursor)

            contador.setText(f"{len(texto.toPlainText())} / 500 caracteres")

        texto.textChanged.connect(atualizar_contador)

        layout_principal.addWidget(label_msg)
        layout_principal.addWidget(texto)
        layout_principal.addWidget(contador)

        # =========================
        # BOTÃO ENVIAR
        # =========================
        botao_enviar = QPushButton("Enviar Sugestão")

        def enviar():
            nome = line_nome.text().strip()
            email = line_email.text().strip()
            assunto_sel = assunto.currentText()
            mensagem = texto.toPlainText().strip()
            
            if not self.email_valido(email):
                QMessageBox.warning(janela, "Atenção", "E-mail inválido.")
                return


            if not nome or not email:
                QMessageBox.warning(janela, "Atenção", "Preencha nome e e-mail.")
                return

            if assunto.currentIndex() == 0:
                QMessageBox.warning(janela, "Atenção", "Selecione um assunto.")
                return

            if not mensagem:
                QMessageBox.warning(janela, "Atenção", "Digite sua sugestão.")
                return

            try:
                self.enviar_sugestao_email(nome, email, assunto_sel, mensagem)
                QMessageBox.information(
                    janela,
                    "Obrigado!",
                    "Sua sugestão foi enviada com sucesso 😊"
                )
                janela.close()

            except Exception as e:
                QMessageBox.critical(
                    janela,
                    "Erro",
                    f"Erro ao enviar sugestão:\n{str(e)}"
                )


        botao_enviar.clicked.connect(enviar)

        layout_principal.addStretch()
        layout_principal.addWidget(botao_enviar, alignment=Qt.AlignRight)

        janela.setCentralWidget(central_widget)
        janela.show()

    def versao_do_sistema(self):
        return self.versao_instalada_local()
    

    def show_mensagem_sistema(self):
        versao_atual = self.versao_do_sistema()
        
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Informações do Sistema")
        msg.setText(
            f"Bem-vindo ao Sistema de Gerenciamento!\n\n"
            f"Versão atual: {versao_atual}\n\n"
            "Sempre que possível, realize sugestões de melhorias para que possamos "
            "evoluir continuamente o sistema."
        )
        msg.exec()
        

    # EXIBE A PÁGINA DE CONFIGURAÇÕES AO CLICAR NA OPÇÃO CONFIGURAÇÕES DENTRO DO MENU DO BOTÃO MAIS OPÇÕES
    def show_menu_opcoes(self):
        # chama a função do pg_configuracoes
        self.pagina_configuracoes.configurar_menu_opcoes(self.btn_mais_opcoes)

    def show_combobox(self):
        self.combobox.show()
        
    # EXIBE A PÁGINA DE CONTATO AO CLICAR NA OPÇÃO CONTATO DENTRO DO MENU DO BOTÃO MAIS OPÇÕES
    def show_pg_contato(self):
        selected_action = self.sender()
        if selected_action == self.action_contato:
            self.paginas_sistemas.setCurrentWidget(self.pg_contato)
#*********************************************************************************************************************
    def atualizar_usuario_no_bd(self):
        if not self.db.verificar_conexao():
            QMessageBox.critical(self, "Erro", "Conexão com o banco de dados não estabelecida!")
            return
        self.connection = self.db.connection 

        if not self.is_editing or not self.selected_user_id:
            QMessageBox.warning(self, "Erro", "Nenhum usuário selecionado para atualizar")
            return

        tem_imagem = (
            hasattr(self, 'label_imagem_usuario') and 
            self.label_imagem_usuario is not None and 
            self.label_imagem_usuario.pixmap() is not None
        )

        if getattr(self, 'usuario_tem_imagem_salva', False) and tem_imagem and not self.imagem_removida_usuario:
            QMessageBox.warning(self, "Remoção de Imagem",
                                "Remova a imagem do usuário e inclua novamente para seguir com a atualização")
            return

        try:        
            cursor = self.connection.cursor()
            cursor.execute("SELECT COUNT(1) FROM users WHERE id = ?", (self.selected_user_id,))
            user_exists = cursor.fetchone()[0]

            if not user_exists:
                QMessageBox.warning(self, "Aviso", "Usuário não encontrado!")
                return
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao verificar usuário: {str(e)}")
            return

        # Obter os dados da interface
        usuario_nome = self.txt_nome.text()
        usuario_usuario = self.txt_usuario_cadastro.text()
        usuario_telefone = self.txt_telefone.text()
        usuario_endereco = self.txt_endereco.text()
        usuario_numero = self.txt_numero.text()
        usuario_complemento = self.txt_complemento.text()
        usuario_cidade = self.txt_cidade.text()
        usuario_bairro = self.txt_bairro.text()
        usuario_email = self.txt_email.text()
        usuario_data_nascimento = self.txt_data_nascimento.text()
        usuario_rg = self.txt_rg.text()
        usuario_cpf = self.txt_cpf.text()
        usuario_cnpj = self.txt_cnpj.text()
        usuario_cep = self.txt_cep.text()
        usuario_estado = self.perfil_estado.currentText()
        usuario_senha = self.txt_senha_cadastro.text()
        usuario_confirmar_senha = self.txt_confirmar_senha.text()
        usuario_acesso = self.perfil_usuarios.currentText()
        usuario_logado = self.get_usuario_logado()
       


        # Verificar campos obrigatórios
        campos_nao_preenchidos_usuarios = []
        
        if not usuario_nome: campos_nao_preenchidos_usuarios.append("nome")
        if not usuario_usuario: campos_nao_preenchidos_usuarios.append("usuario")
        if not usuario_telefone: campos_nao_preenchidos_usuarios.append("telefone")
        if not usuario_endereco: campos_nao_preenchidos_usuarios.append("endereco")
        if not usuario_numero: campos_nao_preenchidos_usuarios.append("numero")
        if not usuario_complemento: campos_nao_preenchidos_usuarios.append("complemento")
        if not usuario_cidade: campos_nao_preenchidos_usuarios.append("cidade")
        if not usuario_bairro: campos_nao_preenchidos_usuarios.append("bairro")
        if not usuario_email: campos_nao_preenchidos_usuarios.append("email")
        if not usuario_data_nascimento: campos_nao_preenchidos_usuarios.append("data_nascimento")
        if not usuario_rg: campos_nao_preenchidos_usuarios.append("rg")
        if not usuario_cpf: campos_nao_preenchidos_usuarios.append("cpf")
        if not usuario_cep: campos_nao_preenchidos_usuarios.append("cep")
        if not usuario_estado: campos_nao_preenchidos_usuarios.append("estado")
        if not usuario_senha: campos_nao_preenchidos_usuarios.append("senha")
        if not usuario_confirmar_senha: campos_nao_preenchidos_usuarios.append("confirmar_senha")
        if not usuario_acesso: campos_nao_preenchidos_usuarios.append("perfil_usuarios")

        if campos_nao_preenchidos_usuarios:
            self.exibir_asteriscos_usuarios(campos_nao_preenchidos_usuarios)
            QMessageBox.warning(self, "Aviso", "Todos os campos obrigatórios devem ser preenchidos!")
            return

        if usuario_senha != usuario_confirmar_senha:
            self.exibir_asteriscos_usuarios(["senha", "confirmar_senha"])
            QMessageBox.warning(self, "Aviso", "As senhas não coincidem!")
            return

        if not self.validar_cpf(usuario_cpf):
            self.exibir_asteriscos_usuarios(["cpf"])
            QMessageBox.warning(self, "Aviso", "CPF inválido!")
            return

        if not self.validar_email(usuario_email):
            self.exibir_asteriscos_usuarios(["email"])
            QMessageBox.warning(self, "Aviso", "E-mail inválido!")
            return

        if not self.validar_telefone(usuario_telefone):
            self.exibir_asteriscos_usuarios(["telefone"])
            QMessageBox.warning(self, "Aviso", "Número de telefone inválido!")
            return
        
        if usuario_cnpj and not self.validar_cnpj(usuario_cnpj):
            self.exibir_asteriscos_usuarios(["cnpj"])
            QMessageBox.warning(self, "Aviso", "CNPJ informado está incorreto!")
            return
        
        if self.perfil_usuarios.currentIndex() == 0:
            self.exibir_asteriscos_usuarios(["perfil"])
            QMessageBox.warning(self, "Aviso", "Por favor informe um tipo de acesso válido")
            return


        # Obter imagem, se houver
        usuario_imagem_original = None
        if tem_imagem:
            pixmap = self.label_imagem_usuario.pixmap()
            if pixmap:
                usuario = self.get_usuario_logado()
                pasta = caminho_recurso("media/usuarios/original")
                os.makedirs(pasta, exist_ok=True)  # garante que a pasta exista
                
                # Nome único para a imagem
                nome_arquivo = f"{usuario}_{uuid.uuid4().hex}.jpg"
                caminho_final = os.path.join(pasta, nome_arquivo)
                
                # Salvar imagem otimizada
                salvar_imagem_otimizada(pixmap, caminho_final, tamanho_max=512, qualidade=75)
                
                usuario_imagem_original = caminho_final

        # SQL e valores com imagem
        if usuario_imagem_original:
            sql = """
                UPDATE users 
                SET Nome=?, Telefone=?, Endereço=?, Número=?, Cidade=?, Bairro=?, Complemento=?, 
                    Email=?, "Data de Nascimento"=?, RG=?, CPF=?, CNPJ=?, CEP=?, Estado=?, 
                    Senha=?, "Imagem Original" =?, "Confirmar Senha"=?,Acesso=?,"Usuário Logado"=?
                WHERE id=?
            """
            valores = (
                usuario_nome, usuario_telefone, usuario_endereco, usuario_numero, usuario_cidade, usuario_bairro, usuario_complemento,
                usuario_email, usuario_data_nascimento, usuario_rg, usuario_cpf, usuario_cnpj, usuario_cep, usuario_estado,
                usuario_senha, usuario_imagem_original, usuario_confirmar_senha,usuario_acesso,usuario_logado, self.selected_user_id
            )
        else:
            sql = """
                UPDATE users 
                SET Nome=?, Telefone=?, Endereço=?, Número=?, Cidade=?, Bairro=?, Complemento=?, 
                    Email=?, "Data de Nascimento"=?, RG=?, CPF=?, CNPJ=?, CEP=?, Estado=?, 
                    Senha=?, "Confirmar Senha"=?,Acesso=?,"Usuário Logado"=?
                WHERE id=?
            """
            valores = (
                usuario_nome, usuario_telefone, usuario_endereco, usuario_numero, usuario_cidade, usuario_bairro, usuario_complemento,
                usuario_email, usuario_data_nascimento, usuario_rg, usuario_cpf, usuario_cnpj, usuario_cep, usuario_estado,
                usuario_senha, usuario_confirmar_senha, usuario_acesso,usuario_logado,self.selected_user_id
            )

        try:
            with self.connection:
                cursor = self.connection.cursor()
                cursor.execute(sql, valores)
            QMessageBox.information(self, "Sucesso", "Usuário atualizado com sucesso!")
            self.limpar_campos_após_atualizar_usuario()
            self.is_editing = False
            self.selected_user_id = None
            descricao = f"O usuário {usuario_usuario} foi atualizado com sucesso"
            self.registrar_historico_usuarios("Atualização de Usuário", descricao)
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao atualizar o usuário: {str(e)}")
      
#*********************************************************************************************************************
    def mostrar_tabela_usuarios(self):
        # Criar uma instância da classe AtualizarUsuario
        self.tabela_usuario_dialogo = AtualizarUsuario(self)
        # Exibir a janela da tabela de usuários
        self.tabela_usuario_dialogo.show()
        self.tabela_usuario_dialogo.listar_usuarios()

    def mostrar_tabela_produtos(self):
        # Criar uma instância da classe AtualizarProduto
        self.tabela_produtos_dialogo = AtualizarProduto(self)
        # Exibir a janela da tabela de produtos
        self.tabela_produtos_dialogo.show()
        self.tabela_produtos_dialogo.listar_produtos()

#*********************************************************************************************************************
    def carregar_dados_usuario(self, user_id):
        try:
            cursor = self.connection.cursor()
            cursor.execute("SELECT * FROM users WHERE id=?", (user_id,))
            usuario = cursor.fetchone()
            
            if usuario:
                # Supondo que o resultado esteja na ordem dos campos
                self.txt_nome.setText(usuario[1])
                self.txt_usuario_cadastro.setText(usuario[2])
                self.txt_telefone.setText(usuario[3])
                self.txt_endereco.setText(usuario[4])
                self.txt_numero.setText(usuario[5])
                self.txt_complemento.setText(usuario[6])
                self.txt_email.setText(usuario[7])
                self.txt_data_nascimento.setText(usuario[8])
                self.txt_rg.setText(usuario[9])
                self.txt_cpf.setText(usuario[10])
                self.txt_cep.setText(usuario[11])
                self.txt_estado.setText(usuario[12])
                self.txt_senha_cadastro.setText(usuario[13])
                self.txt_confirmar_senha.setText(usuario[14])
                
                # Se tiver uma imagem, carregar a imagem
                if usuario[15]:
                    self.exibir_imagem_em_label_usuario(usuario[15])

                
                self.is_editing = True
                self.selected_user_id = user_id
            else:
                QMessageBox.warning(self, "Aviso", "Usuário não encontrado!")
        
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar os dados do usuário: {str(e)}")

#*********************************************************************************************************************
    def validar_cpf(self, text, widget=None):
            if widget is None:
                widget = self.sender()
                if widget is None:
                    return False

            if text == "Não Cadastrado":
                widget.setText(text)
                return True

            # Remover caracteres não numéricos
            cpf = re.sub('[^0-9]', '', text)

            # Verificar se o CPF tem 11 dígitos
            if len(cpf) != 11:
                return False

            return True
#*********************************************************************************************************************
    def validar_email(self, text, widget=None):
            if widget is None:
                widget = self.sender()
                if widget is None:
                    return False

            if text == "Não Cadastrado":
                widget.setText(text)
                return True

            # Expressão regular para e-mail
            regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            return bool(re.match(regex, text))



#*********************************************************************************************************************
    def validar_cnpj(self, text, widget=None):
        if widget is None:
            widget = self.sender()
            if widget is None:
                return False

        if text == "Não Cadastrado":
            widget.setText(text)
            return True

        # Remover caracteres não numéricos
        cnpj = re.sub('[^0-9]', '', text)

        # Verificar se o CNPJ tem 14 dígitos
        if len(cnpj) != 14:
            return False

        return True
#********************************************************************************************************************* 
    def iniciar_verificacao_primeiro_acesso(self):
        self.timer_primeiro_acesso = QTimer(self)
        self.timer_primeiro_acesso.setInterval(60_000)  # 1 minuto (teste)

        self._ultimo_aviso_primeiro_acesso = None

        def tarefa():
            # TESTE: 1 minuto (produção: 1 dia)
            intervalo_regra = timedelta(minutes=1)
            periodo_aviso = timedelta(minutes=1)  # produção: timedelta(days=1)

            # 1) tenta corrigir automaticamente (silencioso)
            self.db.atualizar_primeiro_acesso(intervalo=intervalo_regra)

            # 2) verifica se ainda sobrou pendência (inclui datas inválidas)
            pendentes = self.db.contar_primeiro_acesso_pendente(intervalo=intervalo_regra)

            agora = datetime.now()
            pode_avisar = (
                self._ultimo_aviso_primeiro_acesso is None
                or (agora - self._ultimo_aviso_primeiro_acesso) >= periodo_aviso
            )

            if pendentes > 0 and pode_avisar:
                self._ultimo_aviso_primeiro_acesso = agora
                QMessageBox.warning(
                    self,
                    "Atenção",
                    f"Existem {pendentes} usuário(s) com 'Usuário Logado' em estado inválido ('Primeiro Acesso').\n\n"
                    "Atualize esse(s) registro(s) manualmente para 'Não Logado' e/ou corrija a 'Data da Inclusão do Usuário'."
                )

        self.timer_primeiro_acesso.timeout.connect(tarefa)
        self.timer_primeiro_acesso.start()
        tarefa()
#*********************************************************************************************************************
    def validar_telefone(self, text, widget=None):
        if widget is None:
            widget = self.sender()
            if widget is None:
                return False

        if text == "Não Cadastrado":
            widget.setText(text)
            return True

        # Remover caracteres não numéricos
        telefone = re.sub('[^0-9]', '', text)

        # Verificar se o telefone tem 11 dígitos
        if len(telefone) != 11:
            return False

        return True
#*********************************************************************************************************************
    def criar_botoes_avancar_voltar(self):
        self.btn_avancar.clicked.connect(self.avancar_pagina)
        self.btn_retroceder.clicked.connect(self.retroceder_pagina)

        if self.tema_atual == "escuro":
            css = """
            QPushButton {
                border-radius: 8px;
                background: qlineargradient(
                    x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgb(50, 50, 50),
                    stop:1 rgb(80, 80, 80)
                );
                border: 4px solid transparent;
            }

            QPushButton:hover {
                background: qlineargradient(
                    x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgb(60, 60, 60),
                    stop:1 rgb(100, 100, 100)
                );
            }
            """
        elif self.tema_atual == "claro":
            css = """
            QPushButton {
                border-radius: 8px;
                background: qlineargradient(
                    x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgb(220, 220, 220),
                    stop:1 rgb(245, 245, 245)
                );
                border: 4px solid transparent;
            }
            """
        else:
            css = """
            QPushButton {
                border-radius: 8px;
                background: qlineargradient(
                    x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgb(50,150,250),
                    stop:1 rgb(100,200,255)
                );
                border: 4px solid transparent;
            }

            QPushButton:hover {
                background: qlineargradient(
                    x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgb(100,180,255),
                    stop:1 rgb(150,220,255)
                );
                border: 4px solid transparent;
            }
            """

        for btn in (self.btn_avancar, self.btn_retroceder):
            btn.setStyleSheet(css)
            btn.setCursor(Qt.PointingHandCursor)
#*********************************************************************************************************************
    def exibir_botao_mostrar_usuarios(self):
        self.tabela_usuario_dialogo.btn_mostrar_usuarios.setVisible(True)
#*********************************************************************************************************************
    def carregar_imagem_usuario(self):
        # Abrir uma caixa de diálogo de seleção de arquivo
        caminho = abrir_dialogo_memoria(
            parent=self,
            chave="carregar_imagem_usuario",
            titulo="Selecionar Imagem Usuário Original",
            filtro="Imagens (*.png *.jpg *.jpeg *.webp)"
        )

 
        if not caminho:
            return
        
        pixmap = QPixmap(caminho)
        if pixmap.isNull():
            QMessageBox.warning(self, "Aviso", "Não foi possível carregar a imagem.")
            return
        
        # ===============================
        # 📁 Pasta padrão do sistema
        # ===============================
        pasta = caminho_recurso("media/usuarios/original")
        os.makedirs(pasta,exist_ok=True)
        
        nome_arquivo = f"{uuid.uuid4().hex}.jpg"
        caminho_final = os.path.join(pasta,nome_arquivo)
        
        salvar_imagem_otimizada(
            pixmap,
            caminho_final,
            tamanho_max=300,
            qualidade=80
        )
        
         # 🔥 ESTE É O ESTADO QUE O CADASTRO PRECISA
        self.caminho_imagem = caminho_final

        # ===============================
        # 🖼️ Atualizar QLabel
        # ===============================
        self.limpar_imagem_usuario()
        
        self.label_imagem_usuario = QLabel(self.frame_imagem_cadastro)
        self.label_imagem_usuario.setObjectName("label_imagem_usuario")

        frame_size = self.frame_imagem_cadastro.size()
        self.label_imagem_usuario.setFixedSize(frame_size.width(), frame_size.height())

        pixmap = pixmap.scaled(
            self.label_imagem_usuario.width(),
            self.label_imagem_usuario.height(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )

        self.label_imagem_usuario.setPixmap(pixmap)
        self.label_imagem_usuario.setAlignment(Qt.AlignCenter)
        self.label_imagem_usuario.show()

        self.imagem_usuario_carregada = True
                
        
#*********************************************************************************************************************
    def avancar_pagina(self):
        # Armazenar a página atual no histórico de páginas
        self.historico_paginas.append(self.pagina_atual_index)
        
        # Avançar para a próxima página na lista
        self.pagina_atual_index += 1
        if self.pagina_atual_index >= len(self.paginas):
            self.pagina_atual_index = 0  # Voltar para a primeira página se atingir o final da lista
        self.paginas_sistemas.setCurrentWidget(self.paginas[self.pagina_atual_index])
#*********************************************************************************************************************
    def retroceder_pagina(self):
        if self.historico_paginas:
            # Remover a página atual do histórico de páginas
            self.historico_paginas.pop()
            
            # Verificar se há páginas no histórico
            if self.historico_paginas:
                # Retroceder para a página anterior no histórico
                self.pagina_atual_index = self.historico_paginas[-1]
            else:
                # Se não houver mais páginas no histórico, retroceder para a página anterior na lista
                self.pagina_atual_index -= 1
                if self.pagina_atual_index < 0:
                    self.pagina_atual_index = len(self.paginas) - 1
            self.paginas_sistemas.setCurrentWidget(self.paginas[self.pagina_atual_index])
#*********************************************************************************************************************
# Configuração do local
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
#*********************************************************************************************************************
    
#*********************************************************************************************************************
    def desconectarUsuario(self):
        msgBox = QMessageBox(self)
        msgBox.setIcon(QMessageBox.Question)
        msgBox.setText("Tem certeza que deseja sair?")
        msgBox.setWindowTitle("Aviso")

        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)

        botao_sim = msgBox.button(QMessageBox.Yes)
        botao_sim.setText("Sim")

        botao_nao = msgBox.button(QMessageBox.No)
        botao_nao.setText("Não")

        resposta = msgBox.exec()

        if resposta != QMessageBox.Yes:
            return

        #  Limpa as configurações direto pelo self.config (não depende do login_window)
        self.config.salvar_configuracoes(
            nome_usuario="",
            usuario="",
            senha="",
            email="",
            mantem_conectado=False
        )

        # Fecha a janela principal
        self.close()

        # ✅ Abre a tela de login novamente (cria uma nova se não existir)
        if self.login_window is None:
            from login import Login
            self.login_window = Login(login_window=None)

        # Garante que não fica nada preenchido
        if hasattr(self.login_window, "limpar_campos"):
            self.login_window.limpar_campos()
        if hasattr(self.login_window, "btn_manter_conectado"):
            self.login_window.btn_manter_conectado.setChecked(False)

        self.login_window.show()

#*********************************************************************************************************************
    def fechar_janela_login_delay(self):
        if self.login_window.isVisible():
            self.login_window.close()
        else:
            print("Janela de login não está visível.")

    def fechar_janela_login(self):
        QTimer.singleShot(100, self.fechar_janela_login_delay)
#********************************************************************************************************************* 
    def obter_senha_salva(self):
        try:
            query = "SELECT Senha FROM users WHERE Usuário = ?"
            cursor = self.connection.cursor()
            cursor.execute(query, (self.config.usuario,))
            result = cursor.fetchone()
            if result:
                return result[0]
            else:
                return ""
        except Exception as e:
            print("Erro ao obter senha:", e)
            return ""
#*********************************************************************************************************************            
    def carregar_configuracoes(self):
        self.txt_usuario_cadastro.setText(self.config.usuario or "")

        # Se a tela de login existir (login manual), atualiza o checkbox nela
        if self.login_window is not None:
            self.login_window.btn_manter_conectado.setChecked(
                bool(self.config.mantem_conectado)
            )

#*********************************************************************************************************************
    def formatar_porcentagem(self):
        valor = self.txt_desconto_3.text()
        if valor:
            try:
                valor_int = int(float(valor))  # Converte para inteiro, mesmo se digitar "12.0"
            except ValueError:
                QMessageBox.warning(self, "Erro", "Valor inválido")
                return
            if valor_int < 5:
                self.mostrar_erro_desconto()
                return
            valor_formatado = f"{valor_int}%"
            self.txt_desconto_3.setText(valor_formatado)
#*********************************************************************************************************************
    def mostrar_erro_desconto(self):
        detalhes_msg = QMessageBox(self)
        detalhes_msg.setIcon(QMessageBox.Warning)
        detalhes_msg.setWindowTitle("Erro")
        detalhes_msg.setText("Por favor adicione um valor de desconto correto.")

        ok_btn = QPushButton("OK")
        detalhes_msg.addButton(ok_btn, QMessageBox.RejectRole)

        detalhes_btn = QPushButton("Detalhes")
        detalhes_msg.addButton(detalhes_btn, QMessageBox.ActionRole)

        detalhes_msg.setDefaultButton(ok_btn)
        detalhes_msg.exec()

        if detalhes_msg.clickedButton() == detalhes_btn:
            detalhes_msg_detalhes = QMessageBox()
            detalhes_msg_detalhes.setIcon(QMessageBox.Information)
            detalhes_msg_detalhes.setWindowTitle("Detalhes do Erro")
            detalhes_msg_detalhes.setText("O seu desconto não pode ser menor que 5%. \n Somente descontos maiores serão válidos para esta ação")
            detalhes_msg_detalhes.exec()
#*********************************************************************************************************************
    def subscribe_user(self):
        # Verificar se está no modo de edição
        if self.is_editing:
            QMessageBox.warning(self, "Modo de Edição Ativo", 
                                "Você está editando um usuário.\nAtualize o cadastro em vez de criar um novo.")
            return

        try:
            # Coleta de dados dos campos
            nome = self.txt_nome.text().strip()
            usuario = self.gerar_codigo_usuarios()
            senha = self.txt_senha_cadastro.text()
            confirmar_senha = self.txt_confirmar_senha.text()
            cep = self.txt_cep.text().strip()
            endereco = self.txt_endereco.text().strip()
            numero = self.txt_numero.text().strip()
            cidade = self.txt_cidade.text().strip()
            bairro = self.txt_bairro.text().strip()
            estado = self.perfil_estado.currentText()
            complemento = self.txt_complemento.text().strip()
            telefone = self.txt_telefone.text().strip()
            email = self.txt_email.text().strip()
            data_nascimento = self.txt_data_nascimento.text().strip()
            rg = self.txt_rg.text().strip()
            cpf = self.txt_cpf.text().strip()
            cnpj = self.txt_cnpj.text().strip()
            imagem = self.converter_imagem_usuario()
            segredo = "Não Cadastrado"
            acesso = self.perfil_usuarios.currentText()

            usuario_logado = self.config.obter_usuario_logado()
            self.db.salvar_usuario_logado(usuario_logado)

            # Validação de campos obrigatórios
            campos_vazios = []
            for campo, widget in self.campos_obrigatorios_usuarios.items():
                valor = widget.currentText() if isinstance(widget, QComboBox) else widget.text()
                if valor.strip() == "":
                    campos_vazios.append(campo)

            if campos_vazios:
                self.exibir_asteriscos_usuarios(campos_vazios)
                QMessageBox.warning(self, "Erro", f"O campo {campos_vazios[0]} é obrigatório.")
                return

            # Validação de senha
            if senha != confirmar_senha:
                self.exibir_asteriscos_usuarios(["senha", "confirmar senha"])
                QMessageBox.warning(self, "Erro", "As senhas não coincidem.")
                return
            
            # Verifica se a senha é válida
            if not self.configuracoes_senha.validar_senha(senha,confirmar_senha):
                self.exibir_asteriscos_usuarios(["senha", "confirmar senha"])
                QMessageBox.warning(self, "Erro", "A senha deve conter pelo menos 8 caracteres, incluindo letras e números.")
                return
                
            # Validação de e-mail
            if not self.email_valido(email):
                self.exibir_asteriscos_usuarios(["email"])
                QMessageBox.warning(self, "Erro", "E-mail inválido.")
                return
            
            if cnpj and not self.validar_cnpj(cnpj):
                self.exibir_asteriscos_usuarios(["cnpj"])
                QMessageBox.warning(self, "Aviso", "CNPJ informado está incorreto!")
                return
            
            if self.perfil_usuarios.currentIndex() == 0:
                self.exibir_asteriscos_usuarios(["acesso"])
                QMessageBox.warning(self, "Aviso", "Selecione um tipo de acesso válido!")
                return


            # Verificar se usuário já existe
            campo_duplicado = self.db.user_exists(usuario,telefone,email,rg,cpf,cnpj)
            if campo_duplicado:
                mensagens = {
                    'usuario': f"Já existe um usuário cadastrado com o login {usuario}.",
                    'telefone': f"Já existe um usuário cadastrado com o telefone {telefone}.",
                    'email': f"Já existe um usuário cadastrado com o e-mail {email}.",
                    'rg': f"Já existe um usuário cadastrado com o RG {rg}.",
                    'cpf': f"Já existe um usuário cadastrado com o CPF {cpf}.",
                    'cnpj': f"Já existe um usuário cadastrado com o CNPJ {cnpj}.",
                }
                # Exibe o asterisco no campo duplicado
                self.exibir_asteriscos_usuarios([campo_duplicado])
                # Exibe a mensagem de erro
                QMessageBox.warning(self, "Erro de Cadastro", mensagens[campo_duplicado])
                return
                
            # Inserir novo usuário
            self.db.insert_user(
                nome=nome,
                usuario=usuario,
                senha=senha,
                confirmar_senha=confirmar_senha,
                cep=cep,
                endereco=endereco,
                numero=numero,
                cidade=cidade,
                bairro=bairro,
                estado=estado,
                complemento=complemento,
                telefone=telefone,
                email=email,
                data_nascimento=data_nascimento,
                rg=rg,
                cpf=cpf,
                cnpj=cnpj,
                segredo=segredo,
                imagem=imagem,
                usuario_logado=usuario_logado,
                acesso=acesso,
            )

            self.registrar_historico_usuarios("Cadastro de Usuários", f"Usuário {usuario} cadastrado com sucesso.")
            QMessageBox.information(self, "Cadastro de Usuário", "Cadastro realizado com sucesso.")


            # Limpar campos após cadastro
            self.eliminar_campos_usuarios()

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao cadastrar usuário:\n{e}")



    def buscar_cep(self,cep:str):
        cep = cep.replace("-", "").strip()
        if len(cep) != 8 or not cep.isdigit():
            QMessageBox.warning(self, "ERRO", "CEP inválido")
            return
        try:
            url = f"https://viacep.com.br/ws/{cep}/json/"
            resposta = requests.get(url)
            if resposta.status_code == 200:
                dados = resposta.json()
                if "erro"  in dados:
                    QMessageBox.warning(self, "ERRO", "CEP não encontrado")
                    return None
                return dados
            else:
                QMessageBox.warning(self,"Erro de conexão","Não foi possível consultar o CEP")
                return None
        except Exception as e:
            QMessageBox.warning(self,"Erro ao consultar CEP",f"Erro: {str(e)}")
            return None

    def preencher_campos_cep(self, dados):
        if dados is None:
            return  # Não faz nada se não tem dados
        
        self.txt_endereco.setText(dados.get("logradouro", ""))
        self.txt_bairro.setText(dados.get("bairro", ""))
        self.txt_cidade.setText(dados.get("localidade", ""))


        complemento = dados.get("complemento", "")
        if any(char.isdigit() for char in complemento):
            self.txt_numero.setText(complemento)
        #self.txt_complemento.setText(complemento)
        # Só ajustar o estado, já que é o único campo extra que você tem
        estado = dados.get("uf", "")
        index_estado = self.perfil_estado.findText(estado)
        if index_estado != -1:
            self.perfil_estado.setCurrentIndex(index_estado)


    def on_cep_editing_finished(self):
        cep_digitado = self.txt_cep.text()
        dados_cep = self.buscar_cep(cep_digitado)
        if dados_cep:
            self.preencher_campos_cep(dados_cep)

    def eliminar_campos_usuarios(self):
        # Limpar campos de texto
        self.txt_nome.clear()
        self.txt_usuario_cadastro.clear()
        self.txt_telefone.clear()
        self.txt_endereco.clear()
        self.txt_cidade.clear()
        self.txt_bairro.clear()
        self.txt_numero.clear()
        self.txt_complemento.clear()
        self.txt_email.clear()
        self.txt_data_nascimento.clear()
        self.txt_rg.clear()
        self.txt_cpf.clear()
        self.txt_cep.clear()
        self.txt_senha_cadastro.clear()
        self.txt_cnpj.clear()
        self.txt_confirmar_senha.clear()
        # Limpar campos de combo box
        self.perfil_estado.setCurrentIndex(0)  # Se for um combo box, o índice padrão é 0 (ou o valor inicial)
        self.perfil_usuarios.setCurrentIndex(0)
        frame = self.frame_imagem_cadastro
        if frame:
            if frame is not None:
                for widget in frame.children():
                    if isinstance(widget,QLabel) and widget.pixmap():
                        widget.clear()
                        widget.setPixmap(QPixmap())
                        widget.hide()
                        print("Imagem do usuário removida com sucesso!!")
                       
            else:
                QMessageBox.warning(self, "Erro","Não foi possível remover a imagem do usuário\n"
                                                "Tente remover pelo botão remover imagem")
#*********************************************************************************************************************
    def converter_imagem_usuario(self):
        """
        Retorna o caminho da imagem do usuário (TEXT),
        compatível com o novo modelo.
        """
        if hasattr(self, "caminho_imagem") and self.caminho_imagem:
            if os.path.exists(self.caminho_imagem):
                return self.caminho_imagem

        return "Não Cadastrado"
#*********************************************************************************************************************    
    def exibir_imagem_em_label_usuario(self, caminho_imagem):
        if not caminho_imagem:
            print("Caminho da imagem vazio.")
            return

        if not os.path.exists(caminho_imagem):
            print("Imagem não encontrada:", caminho_imagem)
            return

        pixmap = QPixmap(caminho_imagem)
        if pixmap.isNull():
            print("Falha ao carregar imagem:", caminho_imagem)
            return
        
        tamanho = self.label_imagem_usuario.size()
        pixmap = pixmap.scaled(
            tamanho,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )

        self.label_imagem_usuario.setPixmap(pixmap)
        self.label_imagem_usuario.repaint()

        print("Imagem carregada com sucesso no QLabel.")

#*********************************************************************************************************************
    def erros_frames_produtos(self):
        # Definir os campos obrigatórios e seus respectivos frames de erro
        self.campos_obrigatorios = {
            'produto': self.txt_produto,
            'quantidade': self.txt_quantidade,
            'valor_produto': self.txt_valor_produto_3,
            'data_cadastro': self.dateEdit_3,
            'cliente': self.txt_cliente_3,
            'descricao': self.txt_descricao_produto_3
        }

        self.frames_erros = {
            'produto': self.frame_erro_produto,
            'quantidade': self.frame_erro_quantidade,
            'valor_produto': self.frame_erro_valor_produto,
            'data_cadastro': self.frame_erro_data_cadastro,
            'cliente': self.frame_erro_cliente,
            'descricao': self.frame_erro_descricao
        }

        # Esconder todos os frames de erro inicialmente
        for frame in self.frames_erros.values():
            frame.hide()

        # Conectar o sinal focusIn ao método esconder_asteriscos
        for widget in self.campos_obrigatorios.values():
            widget.installEventFilter(self)

    def registrar_atalhos(self, acao, tecla_str):
        seq = QKeySequence(tecla_str)

        # Se o atalho já existir, desconecta o anterior antes de recriar
        if hasattr(self, "atalhos") and acao in self.atalhos:
            antigo_shortcut = self.atalhos[acao]
            try:
                antigo_shortcut.disconnect()  # desconecta o sinal
            except Exception:
                pass
            antigo_shortcut.setParent(None)  # remove da hierarquia de widgets
            del self.atalhos[acao]  # remove do dicionário

        # Cria novo atalho
        shortcut = QShortcut(seq, self)

        # Garante que o dicionário existe
        if not hasattr(self, "atalhos"):
            self.atalhos = {}

        self.atalhos[acao] = shortcut

        # Conecta de acordo com a ação
        if acao == "Pesquisar":
            shortcut.activated.connect(self.abrir_pesquisa)
        elif acao == "Abrir Mais Opções":
            shortcut.activated.connect(self.abrir_mais_opcoes)
        elif acao == "Abrir Configurações":
            shortcut.activated.connect(self.abrir_configuracoes)
        elif acao == "Abrir Página Inicial":
            shortcut.activated.connect(self.ir_pagina_inicial)

     # Garante que sempre fique centralizado no topo
    def reposicionar_pesquisa(self):
        largura_janela = self.width()
        largura_widget = self.widget_pesquisa.width()
        x = (largura_janela - largura_widget) // 2
        self.widget_pesquisa.move(x, 20)  # 20 px do topo


    def abrir_pesquisa(self):
        self.widget_pesquisa.show()
        self.reposicionar_pesquisa()  # garante que apareça centralizado
        self.caixa_pesquisa.setFocus()

    def abrir_mais_opcoes(self):
        if hasattr(self, "btn_mais_opcoes") and self.btn_mais_opcoes.menu():
            menu = self.btn_mais_opcoes.menu()
            # Abre o menu abaixo do botão
            menu.exec(self.btn_mais_opcoes.mapToGlobal(
                self.btn_mais_opcoes.rect().bottomLeft()
            ))

    def abrir_configuracoes(self):
        self.pagina_configuracoes.configurar_menu_opcoes(self)
        self.pagina_configuracoes.janela_config.show()
        self.pagina_configuracoes.janela_config.raise_()
        self.pagina_configuracoes.janela_config.activateWindow()

    def ir_pagina_inicial(self):
        if hasattr(self,"home_pag"):
            self.paginas_sistemas.setCurrentWidget(self.home_pag)



#*********************************************************************************************************************
    def exibir_asteriscos_produtos(self, campos_nao_preenchidos):
        for campo in campos_nao_preenchidos:
            frame = self.frames_erros.get(campo)
            if frame is None:
                print(f"[ERRO] Campo '{campo}' não tem frame correspondente! ")
                continue
            name_label_asterisco_produtos = f'label_asterisco_produtos_{campo}'

            if not hasattr(self,name_label_asterisco_produtos):
                #Define o QLabel de asterisco correspondente ao campo
                label = QLabel(frame)
                #Carregar a imagem do frame do asterisco, redimensionando para 12x12, mantendo a proporção original
                asterisco_produtos_pixmap = QPixmap(caminho_recurso("imagens/Imagem1.png").scaled(12,12,Qt.KeepAspectRatio, Qt.SmoothTransformation))
                # Definir o tamanho do QLabel para ser o mesmo que o QFrame
                label.setPixmap(asterisco_produtos_pixmap)
                # Alinhar o QLabel ao centro do frame
                label.setAlignment(Qt.AlignCenter)
                # Aplicar um layout ao frame
                layout = QVBoxLayout(frame)
                layout.setContentsMargins(0,0,0,0) # Remover margens
                # Adicionar o QLabel ao layout
                layout.addWidget(label)
                setattr(self, name_label_asterisco_produtos,label) # Armazena a referência do QLabel
                # Adiciona o print para verificar se o asterisco foi realmente adicionado
                print(f"Asterisco adicionado ao campo de: {campo}")
            else:
                label = getattr(self, name_label_asterisco_produtos)
                label.show()  # Exibir o QLabel do asterisco

            # Exibir o frame de erro
            frame.show()

    def esconder_asteriscos_produtos(self):
        for campo, frame in self.frames_erros.items():
            frame.hide()
            label_name = f'label_asterisco_produtos_{campo}'
            if hasattr(self, label_name):
                getattr(self, label_name).hide()  # Esconder o QLabel do asterisco
#*********************************************************************************************************************
    def erros_frames_usuarios(self):
        self.campos_obrigatorios_usuarios = {
        'nome': self.txt_nome,
        'usuario': self.txt_usuario_cadastro,
        'telefone': self.txt_telefone,
        'endereco': self.txt_endereco,            # <--- corrigido
        'numero': self.txt_numero,          # <--- corrigido
        'complemento': self.txt_complemento,
        'email': self.txt_email,               # <--- corrigido
        'data_nascimento': self.txt_data_nascimento,  # <--- corrigido
        'rg': self.txt_rg,
        'cpf': self.txt_cpf,
        'cep': self.txt_cep,
        'estado': self.perfil_estado,
        'senha': self.txt_senha_cadastro,
        'confirmar senha': self.txt_confirmar_senha,
        'perfil': self.perfil_usuarios
        }
        
        self.campos_evento_usuarios = {
            **self.campos_obrigatorios_usuarios,
            'cnpj': self.txt_cnpj
        }

        self.frames_erros_usuarios = {
            'nome': self.frame_erro_nome,
            'usuario': self.frame_erro_usuario,
            'telefone': self.frame_erro_telefone,
            'endereco': self.frame_erro_endereco,
            'cidade': self.frame_erro_cidade,
            'bairro': self.frame_erro_bairro,
            'numero': self.frame_erro_numero,
            'complemento': self.frame_erro_complemento,
            'email': self.frame_erro_email,
            'data_nascimento': self.frame_data_nascimento,
            'rg': self.frame_erro_rg,
            'cpf': self.frame_erro_cpf,
            'cnpj': self.frame_erro_cnpj,
            'cep': self.frame_erro_cep,
            'estado': self.frame_erro_estado,
            'senha': self.frame_senha,
            'confirmar senha': self.frame_confirmar_senha,
            'perfil': self.frame_erro_perfil
        }

        # Esconder todos os frames de erro inicialmente
        for frame in self.frames_erros_usuarios.values():
            frame.hide()

        # Conectar o sinal focusIn ao método esconder_asteriscos
        for widget in self.campos_evento_usuarios.values():
            widget.installEventFilter(self)
#*********************************************************************************************************************
    def exibir_asteriscos_usuarios(self, campos_nao_preenchidos_usuarios):
        for campo in campos_nao_preenchidos_usuarios:
            frame = self.frames_erros_usuarios.get(campo)
            if frame is None:
                print(f"[ERRO] Campo '{campo}' não tem frame correspondente!")
                continue

            name_label_asterisco = f'label_asterisco_usuarios_{campo}'

            if not hasattr(self, name_label_asterisco):
                # Define o QLabel para o asterisco
                label = QLabel(frame)
                # Carregar a imagem do asterisco, redimensionando para 12x12, mantendo a proporção original
                asterisco_pixmap = QPixmap(
                    caminho_recurso("imagens/Imagem1.png")
                ).scaled(
                    12, 12,
                    Qt.KeepAspectRatio,
                    Qt.SmoothTransformation
                )
                # Definir o tamanho do QLabel para ser o mesmo que o QFrame
                label.setPixmap(asterisco_pixmap)
                # Alinhar o QLabel ao centro do frame
                label.setAlignment(Qt.AlignCenter)
                # Aplicar um layout ao frame
                layout = QVBoxLayout(frame)
                layout.setContentsMargins(0, 0, 0, 0)  # Remove margens
                # Adicionar o QLabel ao layout
                layout.addWidget(label)
                setattr(self, name_label_asterisco, label)  # Armazena a referência do QLabel
                # Adiciona o print para verificar se o asterisco foi realmente adicionado
                print(f"Asterisco adicionado ao frame de: {campo}")
            else:
                label = getattr(self, name_label_asterisco)
                label.show()  # Exibir o QLabel do asterisco
            
            # Exibir o frame de erro
            frame.show()

#*********************************************************************************************************************
    def esconder_asteriscos_usuarios(self):
        for campo, frame in self.frames_erros_usuarios.items():
            frame.hide()
            label_name = f'label_asterisco_usuarios_{campo}'
            if hasattr(self, label_name):
                getattr(self, label_name).hide()
#*********************************************************************************************************************
    def eventFilter(self, obj, event):
        # ✅ 1) Duplo clique: selecionar tudo (sem autocompletar)
        if event.type() == QEvent.MouseButtonDblClick:
            # Se quiser restringir só aos campos que têm completer:
            if obj in self.campos_com_autocomplete.values():
                # QLineEdit tem selectAll()
                try:
                    obj.selectAll()
                    return True  # evento tratado
                except Exception:
                    pass

        # ✅ 2) Seu comportamento atual: FocusIn esconde erros
        if event.type() == QEvent.FocusIn:
            # Esconder somente o erro do campo de USUÁRIO que recebeu foco
            for campo, widget in self.campos_evento_usuarios.items():
                if obj is widget:
                    frame = self.frames_erros_usuarios.get(campo)
                    if frame:
                        frame.hide()
                    label_name = f'label_asterisco_usuarios_{campo}'
                    if hasattr(self, label_name):
                        getattr(self, label_name).hide()
                    break

            # Esconder somente o erro do campo de PRODUTO que recebeu foco
            for campo, widget in self.campos_obrigatorios.items():
                if obj is widget:
                    frame = self.frames_erros.get(campo)
                    if frame:
                        frame.hide()
                    label_name = f'label_asterisco_produtos{campo}'
                    if hasattr(self, label_name):
                        getattr(self, label_name).hide()
                    break

        return super().eventFilter(obj, event)
#*********************************************************************************************************************  
    def atualizar_valores_frames_na_hora_do_cadastro(self, quantidade, valor_do_desconto, valor_com_desconto):
        # Formatar os valores corretamente
        valor_com_desconto_formatado = locale.currency(valor_com_desconto, grouping=True)

        valor_do_desconto_formatado = (
            "Sem desconto" if valor_do_desconto == 0 else locale.currency(valor_do_desconto, grouping=True)
        )

        valor_total_formatado = locale.currency(valor_com_desconto + valor_do_desconto, grouping=True)

        # Atualizar os textos das LABELS (não dos frames!)
        self.label_valor_desconto.setText(valor_com_desconto_formatado)
        self.label_valor_do_desconto.setText(valor_do_desconto_formatado)
        self.label_quantidade.setText("{:.0f}".format(quantidade))
        self.label_valor_total_produtos.setText(valor_total_formatado)

        # Garantir que os textos fiquem centralizados nos labels
        self.label_valor_total_produtos.setAlignment(Qt.AlignCenter)
        self.label_valor_do_desconto.setAlignment(Qt.AlignCenter)
        self.label_valor_desconto.setAlignment(Qt.AlignCenter)
        self.label_quantidade.setAlignment(Qt.AlignCenter)
#*********************************************************************************************************************
    def confirmar_produtos(self):
        if self.is_editing_produto:
            QMessageBox.warning(self, "Modo de Edição Ativo",
                                "Você está editando um produto.\nAtualize o produto em vez de criar um novo.")
            return
        # Verificar se há produtos pendentes para confirmar
        if (not hasattr(self, 'produtos_pendentes_por_cliente') or
            not any(self.produtos_pendentes_por_cliente.values())):
            
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro")  
            msg.setText("Não há produtos preenchidos para confirmar.")
            msg.exec()
            return None

        # Verificar se todos os campos obrigatórios estão preenchidos
        if not self.campos_obrigatorios_preenchidos():
            return
            
        # Verificar se a imagem está carregada
        if not self.imagem_carregada_produto:
            # Perguntar ao usuário se ele deseja seguir sem uma imagem
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Question)
            msgBox.setText("O produto não contém imagem. Deseja confirmar mesmo assim?")
            msgBox.setWindowTitle("Aviso")

            # Adicionando os botões de forma apropriada
            button_sim = msgBox.addButton("Sim", QMessageBox.YesRole)
            button_nao = msgBox.addButton("Não", QMessageBox.NoRole)

            resposta = msgBox.exec()

            if msgBox.clickedButton() == button_nao:
                return
                
        # Perguntar ao usuário se ele tem certeza de que deseja cadastrar o produto
        confirmar_msg = QMessageBox()
        confirmar_msg.setIcon(QMessageBox.Question)
        confirmar_msg.setWindowTitle("Confirmação")
        confirmar_msg.setText("Tem certeza de que deseja cadastrar o produto?")

        button_sim = confirmar_msg.addButton("Sim", QMessageBox.YesRole)
        button_nao = confirmar_msg.addButton("Não", QMessageBox.NoRole)
        
        resposta = confirmar_msg.exec()

        if confirmar_msg.clickedButton() == button_nao:
            return
                
        # Para salvar, por exemplo:
        for lista_produtos in self.produtos_pendentes_por_cliente.values():
            for produto in lista_produtos:
                self.inserir_produto_no_bd(produto)

        # Depois limpar:
        self.produtos_pendentes_por_cliente.clear()

        # Limpar os campos de entrada
        self.limpar_campos_produtos()
        
        # Limpar a imagem
        # Limpar a imagem (sem estourar erro se o QLabel já tiver sido destruído)
        if hasattr(self, "label_imagem_produto") and isValid(self.label_imagem_produto):
            self.label_imagem_produto.clear()

        self.imagem_carregada_produto = None

        self.dateEdit_3.setDate(QDate.currentDate())  # Define a data atual

        # Exibir mensagem de sucesso apenas se todos os campos estiverem preenchidos
        self.mostrar_mensagem_sucesso()

    def definir_tamanho_fonte(self, percentual):
        # pega a fonte padrão do QApplication
        fonte_base = QApplication.font()
        tamanho_base = fonte_base.pointSizeF()  # tamanho padrão do sistema, ex: 9pt
        novo_tamanho = tamanho_base * percentual / 100.0

        # aplica CSS apenas para a fonte
        style = f"""
        QWidget {{
            font-size: {novo_tamanho}pt;
        }}
        QMenuBar, QMenu, QToolButton, QLabel, QPushButton, QLineEdit, QTextEdit, QTableWidget {{
            font-size: {novo_tamanho}pt;
        }}
        QHeaderView::section {{
            font-size: {novo_tamanho}pt;
        }}
        """
        QApplication.instance().setStyleSheet(style)

        # salva no JSON
        self.config.salvar_percentual_fonte(percentual)
#*********************************************************************************************************************
    def inserir_produto_no_bd(self, produto_info, registrar_historico=True):
        try:
            cursor = self.db.connection.cursor()

            # ===============================
            # 1️⃣ Verificar se o cliente existe
            # ===============================
            if produto_info.get("cnpj"):
                cursor.execute(
                    "SELECT CPF, CNPJ FROM clientes_juridicos WHERE CNPJ = ?",
                    (produto_info["cnpj"],)
                )
            elif produto_info.get("cpf"):
                cursor.execute(
                    "SELECT CPF FROM clientes_fisicos WHERE CPF = ?",
                    (produto_info["cpf"],)
                )
            else:
                QMessageBox.critical(self, "Erro", "Cliente sem CPF ou CNPJ.")
                return

            row = cursor.fetchone()
            if not row:
                QMessageBox.critical(self, "Erro", "Cliente não encontrado no banco.")
                return

            # ===============================
            # 2️⃣ Definir CPF / CNPJ finais
            # ===============================
            cpf_final = None
            cnpj_final = None

            if produto_info.get("cnpj"):
                cpf_final, cnpj_final = row
            else:
                cpf_final = row[0]

            # ===============================
            # 3️⃣ Preparar dados do produto
            # ===============================
            valor_real_formatado = produto_info["valor_produto"]
            desconto_formatado = (
                "Sem desconto"
                if produto_info["desconto"] == 0
                else f"{int(produto_info['desconto'])}%"
            )

            # ===============================
            # 4️⃣ Converter imagem
            # ===============================
            caminho_imagem = None
            if self.imagem_carregada_produto:
                pasta = caminho_recurso("media/produtos")
                os.makedirs(pasta, exist_ok=True)
                
                nome_arquivo = f"produto_{uuid.uuid4().hex}.jpg"
                caminho_imagem = os.path.join(pasta, nome_arquivo)
                
                salvar_imagem_otimizada(
                    self.label_imagem_produto.pixmap(),
                    caminho_imagem,
                    tamanho_max=600,
                    qualidade=80
                )

            usuario_logado = self.get_usuario_logado()

            # ===============================
            # 5️⃣ Inserir produto
            # ===============================
            self.db.insert_product(
                produto_info["produto"],
                produto_info["quantidade"],
                valor_real_formatado,
                desconto_formatado,
                produto_info["total_sem_desconto"],
                produto_info["valor_total"],
                produto_info["data_cadastro"],
                produto_info["codigo_item"],
                produto_info["cliente"],
                produto_info["descricao_produto"],
                usuario_logado,
                cnpj_final,
                cpf_final,
                caminho_imagem  
            )

            # ===============================
            # 6️⃣ Somar valor total do cliente
            # ===============================
            if cnpj_final:
                cursor.execute(
                    'SELECT "Total Com Desconto" FROM products WHERE CNPJ = ?',
                    (cnpj_final,)
                )
            else:
                cursor.execute(
                    'SELECT "Total Com Desconto" FROM products WHERE CPF = ?',
                    (cpf_final,)
                )

            linhas = cursor.fetchall()
            valor_total_cliente = 0.0

            for linha in linhas:
                valor_str = (
                    str(linha[0])
                    .replace("R$", "")
                    .replace(".", "")
                    .replace(",", ".")
                )
                try:
                    valor_total_cliente += float(valor_str)
                except Exception as e:
                    print(f"Erro ao converter valor: '{valor_str}' -> {e}")

            valor_total_formatado = (
                f"R$ {valor_total_cliente:,.2f}"
                .replace(",", "X")
                .replace(".", ",")
                .replace("X", ".")
            )

            # ===============================
            # 7️⃣ Verificar modo do cliente
            # ===============================
            if cnpj_final:
                cursor.execute(
                    'SELECT "Modo Valor Gasto" FROM clientes_juridicos WHERE CNPJ = ?',
                    (cnpj_final,)
                )
            else:
                cursor.execute(
                    'SELECT "Modo Valor Gasto" FROM clientes_fisicos WHERE CPF = ?',
                    (cpf_final,)
                )

            resultado = cursor.fetchone()
            if not resultado:
                print("[ERRO] Cliente sem 'Modo Valor Gasto'")
                return

            modo = resultado[0].strip().lower()

            # ===============================
            # 8️⃣ Se for manual, parar aqui
            # ===============================
            if "manual" in modo:
                self.db.connection.commit()
                return

            # ===============================
            # 9️⃣ Atualizar cliente
            # ===============================
            if cnpj_final:
                cursor.execute("""
                    UPDATE clientes_juridicos
                    SET "Valor Gasto Total" = ?, "Última Compra" = ?
                    WHERE CNPJ = ?
                """, (valor_total_formatado, produto_info["data_cadastro"], cnpj_final))
            else:
                cursor.execute("""
                    UPDATE clientes_fisicos
                    SET "Valor Gasto Total" = ?, "Última Compra" = ?
                    WHERE CPF = ?
                """, (valor_total_formatado, produto_info["data_cadastro"], cpf_final))

            self.db.connection.commit()

            # ===============================
            # 🔁 Atualizar telas
            # ===============================
            if cnpj_final and hasattr(self, "pagina_clientes_juridicos"):
                self.pagina_clientes_juridicos.carregar_clientes_juridicos()

            if cpf_final and hasattr(self, "pagina_clientes_fisicos"):
                self.pagina_clientes_fisicos.carregar_clientes_fisicos()

            # ===============================
            # 🕘 Histórico
            # ===============================
            if registrar_historico:
                descricao = (
                    f"Produto {produto_info['produto']} foi cadastrado "
                    f"com quantidade {produto_info['quantidade']} "
                    f"e valor {valor_real_formatado}."
                )
                self.registrar_historico("Cadastro de Produto", descricao)

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao cadastrar produto: {str(e)}")

#*********************************************************************************************************************
    def adicionar_produto(self):
        # Verificar se todos os campos estão preenchidos
        campos_preenchidos = all([
            self.txt_produto.text().strip(),
            self.txt_quantidade.text().strip(),
            self.txt_valor_produto_3.text().strip(),
            self.dateEdit_3.date().isValid(),
            self.txt_cliente_3.text().strip(),
            self.txt_descricao_produto_3.text().strip()
        ])

        # Gerar código automaticamente se estiver vazio
        if campos_preenchidos and (not self.txt_codigo_item.text().strip() or self.txt_codigo_item.text() == "0"):
            self.txt_codigo_item.setText(self.gerar_codigo_aleatorio())

        produto_original = self.produto_selecionado if hasattr(self, 'produto_selecionado') else None
        produto_info = self.subscribe_produto()

        if produto_info is not None:
            quantidade, valor_do_desconto, valor_com_desconto = produto_info

            # Verificar se houve alteração
            if produto_original and not self.verificar_alteracoes_produto(produto_original):
                msg = QMessageBox(self)
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("Informação")
                msg.setText("Nenhuma alteração foi feita no produto.")
                msg.exec()
                return
            
            self.atualizar_valores_frames_na_hora_do_cadastro(quantidade, valor_do_desconto, valor_com_desconto)

            # Limpar produto selecionado após a adição
            self.produto_selecionado = None
#*********************************************************************************************************************
    def subscribe_produto(self): 
        # Verificar se todos os campos obrigatórios estão preenchidos
        campos_nao_preenchidos = [
            campo for campo, widget in self.campos_obrigatorios.items()
            if not widget.text()
        ]

        if campos_nao_preenchidos:
            self.exibir_asteriscos_produtos(campos_nao_preenchidos)  # Mostrar os asteriscos nos campos obrigatórios
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro")
            msg.setText("Todos os campos obrigatórios precisam ser preenchidos.")
            msg.exec()
            return
        
        # Verificar se o nome do produto começa com um número ou caractere especial
        if re.match(r'^[\d\W]', self.txt_produto.text()):
            self.exibir_asteriscos_produtos(["produto"])  # Mostrar o asterisco ao lado do campo Produto

            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Erro")
            msg.setText("O nome do produto não pode começar com um número ou caractere especial.")
            msg.exec()
            return
        else:
            self.esconder_asteriscos_produtos()  # Esconder os asteriscos se não houver erro

        # Verificar se não estamos no modo de edição
        if not self.is_editing_produto:
            codigo_item = self.gerar_codigo_aleatorio()
            self.txt_codigo_item.setText(codigo_item)

        # Obter os valores dos campos
        valor_produto_str = self.txt_valor_produto_3.text().replace('R$', '').replace('.', '').replace(',', '.').strip()
        if not valor_produto_str:
            self.txt_valor_produto_3.setText("Não Cadastrado")
            valor_produto = 0.0
        else:
            valor_produto = float(valor_produto_str)

        quantidade_str = self.txt_quantidade.text().strip()
        if not quantidade_str:
            self.txt_quantidade.setText("Não Cadastrado")
            quantidade = 0
        else:
            quantidade = int(quantidade_str)

        desconto_str = self.txt_desconto_3.text().replace('%', '').strip().replace(',', '.')  # Removendo o símbolo de porcentagem
        desconto = float(desconto_str) if desconto_str and desconto_str != 'Sem desconto' else 0.0

        # Atualizar o valor do desconto na QLineEdit
        self.txt_desconto_3.setText("Sem desconto" if desconto == 0 else f"{int(desconto)}%")

        # Atualizar o valor do desconto na QLineEdit se não tiver desconto
        if desconto == 0:
            self.txt_desconto_3.setText("Sem desconto")
            desconto = 0.0  # Definir desconto como zero para cálculo e banco de dados

        # Calcular o valor total do produto antes do desconto
        valor_total = quantidade * valor_produto
        valor_desconto = valor_total * (desconto / 100) # Cálculo usando porcentagem real
        valor_com_desconto = valor_total - valor_desconto
        
        # Formatação monetária padrão BR
        valor_unitario_formatado = (
            f"R$ {valor_produto:,.2f}"
            .replace(",", "X")
            .replace(".", ",")
            .replace("X", ".")
        )

        total_sem_desconto_formatado = (
            f"R$ {valor_total:,.2f}"
            .replace(",", "X")
            .replace(".", ",")
            .replace("X", ".")
        )

        valor_total_formatado = (
            f"R$ {valor_com_desconto:,.2f}"
            .replace(",", "X")
            .replace(".", ",")
            .replace("X", ".")
        )

        # Adicionar os dados do produto aos produtos pendentes
        produto_info = {
            "produto": self.txt_produto.text(),
            "quantidade": quantidade,
            "valor_produto": valor_unitario_formatado,   
            "desconto": desconto,
            "total_sem_desconto": total_sem_desconto_formatado,
            "valor_total": valor_total_formatado,
            "data_cadastro": self.dateEdit_3.date().toString("dd/MM/yyyy"),
            "codigo_item": self.txt_codigo_item.text(),
            "cliente": self.txt_cliente_3.text(),
            "descricao_produto": self.txt_descricao_produto_3.text()
        }
         # Verificar se o cliente existe antes de adicionar
        cursor = self.db.connection.cursor()

        cliente_nome = produto_info["cliente"]
        
        # Buscar em clientes_fisicos
        cursor.execute("""
            SELECT rowid, "Nome do Cliente", CPF 
            FROM clientes_fisicos 
            WHERE "Nome do Cliente" = ?
        """, (cliente_nome,))
        clientes_fisicos = [(rowid, nome, "Físico", cpf) for rowid, nome, cpf in cursor.fetchall()]

        # Buscar em clientes_juridicos
        cursor.execute("""
            SELECT rowid, "Nome do Cliente", CNPJ 
            FROM clientes_juridicos 
            WHERE "Nome do Cliente" = ?
        """, (cliente_nome,))
        clientes_juridicos = [(rowid, nome, "Jurídico", cnpj) for rowid, nome, cnpj in cursor.fetchall()]


        # Junta todos os resultados
        clientes_encontrados = clientes_fisicos + clientes_juridicos

        if not clientes_encontrados:
            QMessageBox.warning(
                self,
                "Cliente não Encontrado",
                f"O cliente {produto_info['cliente']} precisa estar cadastrado antes de adicionar um produto"
            )
            return

        
        # Se houver mais de um cliente com o mesmo nome, perguntar por qual seguir
        if len(clientes_encontrados) > 1:
            nomes_formatados = "\n".join(
                [f"{i+1}. {tipo} - {nome} - {'CNPJ' if tipo == 'Jurídico' else 'CPF'}: {doc}"
                for i, (_, nome, tipo, doc) in enumerate(clientes_encontrados)]
            )
            escolha, ok = QInputDialog.getInt(
                self,
                "Cliente duplicado encontrado",
                f"Foram encontrados {len(clientes_encontrados)} clientes com o mesmo nome '{cliente_nome}':\n\n{nomes_formatados}\n\nDigite o número do cliente que deseja usar:",
                1,  # valor inicial
                1,  # mínimo
                len(clientes_encontrados)  # máximo
            )
            
            if not ok:
                return  # usuário cancelou

            cliente_escolhido = clientes_encontrados[escolha - 1]
        else:
            cliente_escolhido = clientes_encontrados[0]

        rowid_cliente, nome_cliente, tipo_cliente,documento_cliente = cliente_escolhido

        # Armazena no produto_info
        produto_info["rowid_cliente"] = rowid_cliente
        if tipo_cliente == "Jurídico":
            produto_info["cnpj"] = documento_cliente
            produto_info["cpf"] = None
        else:
            produto_info["cpf"] = documento_cliente
            produto_info["cnpj"] = None


        # 6) Inicializar dicionário se não existir para o cliente
        if not hasattr(self, 'produtos_pendentes_por_cliente'):
            self.produtos_pendentes_por_cliente = {}

        if rowid_cliente not in self.produtos_pendentes_por_cliente:
            self.produtos_pendentes_por_cliente[rowid_cliente] = []


        if not self.is_editing_produto:
            if rowid_cliente not in self.produtos_pendentes_por_cliente:
                self.produtos_pendentes_por_cliente[rowid_cliente] = []
            self.produtos_pendentes_por_cliente[rowid_cliente].append(produto_info)
        else:
            # lógica para editar produto, se houver
            pass
        # Retornar os valores calculados para exibição
        return quantidade, valor_desconto, valor_com_desconto

#*********************************************************************************************************************
    def verificar_alteracoes_produto(self, produto_original):
        try:
            valor_produto = float(self.txt_valor_produto_3.text().replace('R$', '').replace('.', '').replace(',', '.').strip())
        except ValueError:
            valor_produto = 0.0

        try:
            desconto_str = self.txt_desconto_3.text().replace('%', '').replace(',', '.').strip()
            desconto = float(desconto_str) if desconto_str and desconto_str.lower() != 'sem desconto' else 0.0
        except ValueError:
            desconto = 0.0

        produto_atual = {
            "produto": self.txt_produto.text().strip(),
            "quantidade": int(self.txt_quantidade.text()) if self.txt_quantidade.text().isdigit() else 0,
            "valor_produto": round(valor_produto, 2),
            "desconto": round(desconto, 2),
            "data_cadastro": self.dateEdit_3.date().toString("dd/MM/yyyy"),
            "codigo_item": self.txt_codigo_item.text().strip(),
            "cliente": self.txt_cliente_3.text().strip(),
            "descricao_produto": self.txt_descricao_produto_3.text().strip()
        }

        print("Produto Original:", produto_original)
        print("Produto Atual   :", produto_atual)

        for chave in produto_atual:
            valor_original = produto_original.get(chave)
            valor_atual = produto_atual[chave]
            if valor_original != valor_atual:
                print(f"[ALTERADO] Campo '{chave}': Original = {valor_original}, Atual = {valor_atual}")
                return True

        return False

#*********************************************************************************************************************
    def exibir_tabela_produtos(self):
        dialog_atualizacao = AtualizarProduto(self)
        dialog_atualizacao.exec()
       
    def gerar_codigo_aleatorio(self, length=12):
        caracteres = string.ascii_uppercase + string.digits
        codigo_aleatorio = ''.join(random.choice(caracteres) for _ in range(length))
        return codigo_aleatorio
#*********************************************************************************************************************
    def gerar_codigo_usuarios(self, length=7,prefixo="SIG"):
        caracteres = string.ascii_uppercase + string.digits
        tentativa_maxima = 1000  # Para evitar loop infinito
        for _ in range(tentativa_maxima):
            codigo_aleatorio = ''.join(random.choice(caracteres) for _ in range(length - len(prefixo)))
            return codigo_aleatorio
#*********************************************************************************************************************
    def selecionar_produto_tabela(self):
        produto = self.obter_produto_selecionado()

        if produto and produto.imagem_caminho:
            pixmap = QPixmap(produto.imagem_caminho)
            if not pixmap.isNull():
                self.limpar_imagem_produto()  # Limpar imagem anterior
                self.label_imagem_produto.setPixmap(pixmap)
                self.label_imagem_produto.show()
                self.imagem_carregada_produto = True
            else:
                self.limpar_imagem_produto()
                self.imagem_carregada_produto = False
        else:
            self.limpar_imagem_produto()
#*********************************************************************************************************************
    def retirar_imagem_produto(self):
        frame = self.frame_imagem_produto_3
        if frame is not None:
            for widget in frame.children():
                if isinstance(widget, QLabel) and widget.pixmap() is not None and not widget.pixmap().isNull():
                    widget.clear()  # Limpar o QLabel
                    widget.setPixmap(QPixmap())  # Definir um pixmap vazio ou padrão
                    widget.hide()  # Esconder o QLabel para garantir que não fique visível
                    print("Imagem removida com sucesso")
                    QMessageBox.information(self, "Sucesso", "Imagem removida com sucesso")
                    return
        print("Não há imagem do produto para remover.")
        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowTitle("Aviso")
        msg_box.setText("Não há imagem do produto para remover!")
        msg_box.exec()
#*********************************************************************************************************************
    def retirar_imagem_usuario(self):
        frame = self.frame_imagem_cadastro
        if frame is not None:
            for widget in frame.children():
                if isinstance(widget,QLabel) and widget.pixmap() is not None and not widget.pixmap().isNull():
                    widget.clear()
                    widget.setPixmap(QPixmap())
                    widget.hide()
                    self.imagem_removida_usuario = True
                    print("Imagem removida do usuário com sucesso")
                    QMessageBox.information(self, "Sucesso", "Imagem removida do usuário com sucesso")
                    return
        print("Não há imagem do usuário para remover.")
        msg_box = QMessageBox(self, "Erro", "Não há imagem do usuário para remover")
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowTitle("Aviso")
        msg_box.setText("Não há imagem do usuário para remover")
        msg_box.exec()
#*********************************************************************************************************************
    def carregar_imagem_produto(self):
        caminho = abrir_dialogo_memoria(
            parent=self,
            chave="imagem_produto",
            titulo="Selecionar Imagem do Produto",
            filtro="Imagens (*.png *.jpg *.jpeg *.webp)"
        )

        
        if caminho:
            pixmap = QPixmap(caminho)
            if not pixmap.isNull():
                self.limpar_imagem_produto()  # Limpar qualquer QLabel existente no frame

                # Verificar se já existe um QLabel para a imagem
                label_imagem = None
                for widget in self.frame_imagem_produto_3.children():
                    if isinstance(widget, QLabel):
                        label_imagem = widget
                        break

                # Se não houver QLabel, criar um novo
                if label_imagem is None:
                    label_imagem = QLabel(self.frame_imagem_produto_3)
                    label_imagem.setObjectName("label_imagem_produto")

                # Definir tamanho do QLabel para ser o mesmo que o QFrame
                frame_size = self.frame_imagem_produto_3.size()
                label_imagem.setFixedSize(frame_size.width(), frame_size.height())

                # Redimensionar o pixmap para se ajustar ao QLabel
                pixmap = pixmap.scaled(label_imagem.width(), label_imagem.height(), Qt.KeepAspectRatio)

                # Definir o pixmap no QLabel
                label_imagem.setPixmap(pixmap)

                # Ajustar o alinhamento da imagem no QLabel
                label_imagem.setAlignment(Qt.AlignCenter)

                # Mostrar o QLabel
                label_imagem.show()
                self.imagem_carregada_produto = True

                # Salvar a imagem carregada para o atributo nova_imagem
                self.nova_imagem = caminho
            else:
                QMessageBox.warning(self, "Aviso", "Não foi possível carregar a imagem.")
        else:
            pass
#*********************************************************************************************************************
    def limpar_imagem_produto(self):
        for widget in self.frame_imagem_produto_3.children():
            if isinstance(widget, QLabel):
                widget.clear()
                widget.setPixmap(QPixmap())
                widget.hide()  # Esconder o QLabel para garantir que não fique visível
        self.imagem_carregada_produto = False
#*********************************************************************************************************************  
    def limpar_imagem_usuario(self):
        for widget in self.frame_imagem_cadastro.children():
            if isinstance(widget, QLabel):
                widget.clear()
                widget.setPixmap(QPixmap())
                widget.hide()  # Esconder o QLabel para garantir que não fique visível
        self.imagem_usuario_carregada = False
#*********************************************************************************************************************
    def carregar_nova_imagem_produto(self, nova_imagem):
        # Este método simula o carregamento de uma nova imagem no QLabel dentro do frame_imagem_produto
        for widget in self.frame_imagem_produto_3.children():
            if isinstance(widget, QLabel):
                widget.setPixmap(QPixmap(nova_imagem))
                widget.show()  # Mostrar o QLabel novamente após carregar a nova imagem
                self.imagem_carregada_produto = True
                return
#*********************************************************************************************************************
    def limpar_imagem_produto_após_atualizar(self):
        frame = self.frame_imagem_produto_3
        if frame is not None:
            for widget in frame.children():
                if isinstance(widget, QLabel):
                    widget.deleteLater()

            # Adiciona um QLabel vazio de volta ao frame
            novo_label = QLabel(frame)
            novo_label.setPixmap(QPixmap())
            layout = frame.layout()
            if layout is None:
                layout = QVBoxLayout(frame)
                frame.setLayout(layout)
            layout.addWidget(novo_label)
            self.label_imagem_produto = novo_label  # Atualiza a referência ao novo QLabel
#*********************************************************************************************************************
    def atualizar_produto(self):
        if not self.is_editing_produto or not self.produto_id:
            QMessageBox.warning(self, "Erro", "Nenhum produto selecionado para atualizar")
            return

        if not hasattr(self, 'produto_id') or not self.produto_id:
            msgBox = QMessageBox(self)
            msgBox.setIcon(QMessageBox.Critical)
            msgBox.setText("Não há produto selecionado para seguir.")
            msgBox.setWindowTitle("Aviso")
            detalhes_button = msgBox.addButton("Detalhes", QMessageBox.ActionRole)
            msgBox.addButton(QMessageBox.Ok)
            clicked_button = msgBox.exec()

            if clicked_button == QMessageBox.Ok:
                return
            elif msgBox.clickedButton() == detalhes_button:
                detalhesMsgBox = QMessageBox(self)
                detalhesMsgBox.setIcon(QMessageBox.Information)
                detalhesMsgBox.setText("Para usar essa opção deverá ser necessário editar um produto!")
                detalhesMsgBox.setWindowTitle("Detalhes")
                detalhesMsgBox.exec()
                return

        # Verificar se a imagem foi alterada
        imagem_alterada = False
        produto_imagem = None
        if hasattr(self, 'nova_imagem') and self.nova_imagem:
           if os.path.exists(self.nova_imagem):
                imagem_alterada = True
                pasta = caminho_recurso("media/produtos")
                os.makedirs(pasta, exist_ok=True)
                
                nome_arquivo = f"produto_{uuid.uuid4().hex}.jpg"
                caminho_final = os.path.join(pasta, nome_arquivo)
                
                pixmap = QPixmap(self.nova_imagem)
                if pixmap.isNull():
                    QMessageBox.warning(self,"Aviso","Imagem do produto inválida")
                    return
                
                salvar_imagem_otimizada(
                    pixmap,
                    caminho_final,
                    tamanho_max=512,
                    qualidade=80
                )
                
                produto_imagem = caminho_final  # SALVA O CAMINHO

        # Verificar se os campos foram alterados
        alteracao_campo = False
        if hasattr(self, 'produto_original'):
            alteracao_campo = self.verificar_alteracoes_produto(self.produto_original)

        # Se nada foi alterado, exibir mensagem e retornar
        if not alteracao_campo and not imagem_alterada:
            QMessageBox.information(self, "Aviso", "Nenhuma alteração foi feita no produto.")
            return

        # Obter os dados dos campos
        produto_nome = self.txt_produto.text()
        produto_quantidade = self.txt_quantidade.text()
        produto_valor_real = self.txt_valor_produto_3.text()
        produto_desconto = self.txt_desconto_3.text().strip()

        valor_real = float(produto_valor_real.replace('R$', '').replace('.', '').replace(',', '.').strip())
        
        quantidade = int(produto_quantidade)

        total_sem_desconto = valor_real * quantidade

        desconto_float = 0.0
        
        if produto_desconto:
            desconto_limpo = (
                produto_desconto
                .replace("%","")
                .replace(",",".")
                .strip()
            )
            
            if desconto_limpo.replace(".","",1).isdigit():
                desconto_float = float(desconto_limpo)
        
        produto_valor_total = total_sem_desconto * (1 - desconto_float / 100)

        produto_total_sem_desconto = f"R$ {total_sem_desconto:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        produto_valor_total = f"R$ {produto_valor_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


        produto_data_cadastro = self.dateEdit_3.date().toString("dd/MM/yyyy")
        produto_codigo_item = self.txt_codigo_item.text()
        produto_cliente = self.txt_cliente_3.text()
        produto_descricao = self.txt_descricao_produto_3.text()
        produto_id = self.produto_id


        try:
            self.db.atualizar_produto(
                produto_id, produto_nome, produto_quantidade,produto_valor_real,
                produto_desconto,produto_total_sem_desconto,produto_valor_total,produto_data_cadastro,
                produto_codigo_item, produto_cliente, produto_descricao, produto_imagem
            )
            
            self.db.atualizar_valor_gasto_cliente(produto_cliente)
            self.db.atualizar_valor_gasto_cliente_fisico(produto_cliente)
            QMessageBox.information(self, "Sucesso", "Produto atualizado com sucesso!")
            self.limpar_imagem_produto_após_atualizar()
            self.limpar_campos_produtos()
            self.is_editing_produto = False
            self.selected_produto_id = None
            if hasattr(self, 'produto_original'):
                del self.produto_original
            if hasattr(self,'pagina_clientes_juridicos'):
                self.pagina_clientes_juridicos.carregar_clientes_juridicos()
            if hasattr(self, 'paginas_clientes_fisicos'):
                self.pagina_clientes_fisicos.carregar_clientes_fisicos()
                
                
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao atualizar o produto: {str(e)}")

#*******************************************************************************************************
    def registrar_historico(self, acao, descricao):
        # Verifica se o histórico está pausado
        if self.historico_pausado:
            print("Histórico pausado. Registro não será feito.")
            return  # Se o histórico estiver pausado, não faz nada

        usuario = self.get_usuario_logado()  # Obtenha o usuário logado
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        cursor = self.db.connection.cursor()
        cursor.execute("""
            INSERT INTO historico ('Data e Hora', Usuário, Ação, Descrição)
            VALUES (?, ?, ?, ?)
        """, (data_hora, usuario, acao, descricao))
        self.db.connection.commit()
#*******************************************************************************************************
    def registrar_historico_usuarios(self,acao,descricao):
        #Verifica se o histórico está pausado
        if self.historico_usuario_pausado:
            print("Histórico pausado. Registro não será feito")
            return # Se o histórico estiver pausado, não faz 
        usuario = self.get_usuario_logado()
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        cursor = self.db.connection.cursor()
        cursor.execute("""
            INSERT INTO historico_usuarios('Data e Hora', Usuário, Ação, Descrição)
            VALUES (?,?,?,?)
        """,(data_hora,usuario,acao,descricao))
        self.db.connection.commit()
#*******************************************************************************************************        
    def registrar_historico_clientes_juridicos(self,acao,descricao):
        # Verifica se o histórico está pausado
        if self.historico_pausado_clientes_juridicos:
            print("Histórico pausado. Registro não será feito")
            return # Se o histórico estiver pausado, não faz nada
        usuario = self.get_usuario_logado()  # Obtenha o usuário logado
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        cursor = self.db.connection.cursor()
        cursor.execute("""
            INSERT INTO historico_clientes_juridicos ('Data e Hora', Usuário, Ação, Descrição)
            VALUES (?,?,?,?)
        """,(data_hora,usuario,acao,descricao))
        self.db.connection.commit()

    def registrar_historico_clientes_fisicos(self,acao,descricao):
        # Verifica se o histórico está pausado
        if self.historico_pausado_clientes_fisicos:
            print("Histórico pausado. O registro não será feito")
            return # Se o histórico estiver pausado, não faz nada
        usuario = self.get_usuario_logado()
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M")

        cursor = self.db.connection.cursor()
        cursor.execute("""
            INSERT INTO historico_clientes_fisicos ('Data e Hora',Usuário,Ação,Descrição)
            VALUES (?,?,?,?)
        """,(data_hora,usuario,acao,descricao))
        self.db.connection.commit()

#*******************************************************************************************************
    def campos_obrigatorios_preenchidos(self):
        # Verificar se todos os campos obrigatórios estão preenchidos
        campos = [
            self.txt_produto.text(),
            self.txt_quantidade.text(),
            self.txt_valor_produto_3.text(),
            self.dateEdit_3.text(),
            self.txt_codigo_item.text(),
            self.txt_cliente_3.text(),
            self.txt_descricao_produto_3.text()
        ]
        for campo in campos:
            if not campo:
                msg = QMessageBox(self)
                msg.setIcon(QMessageBox.Warning)
                msg.setWindowTitle("Erro")
                msg.setText("Todos os campos precisam ser preenchidos.")
                msg.exec()
                return False
        return True
#*********************************************************************************************************************
    def mostrar_mensagem_sucesso(self):
        success_msg = QMessageBox(self)
        success_msg.setIcon(QMessageBox.Information)
        success_msg.setWindowTitle("Sucesso")
        success_msg.setText("Produtos confirmados e cadastrados com sucesso.")
        success_msg.exec()
#*********************************************************************************************************************
    def formatar_cep(self, text, widget=None):
        if widget is None:
            widget = self.sender()  # pega o QLineEdit que disparou o sinal
            if widget is None:
                return

        numero_limpo = ''.join(filter(str.isdigit, text))[:8]

        if len(numero_limpo) >= 5:
            cep_formatado = f"{numero_limpo[:5]}-{numero_limpo[5:]}"
        else:
            cep_formatado = numero_limpo

        widget.blockSignals(True)
        widget.setText(cep_formatado)
        widget.blockSignals(False)

#*********************************************************************************************************************
    def formatar_cpf(self, text, widget=None):
        if widget is None:
            widget = self.sender()
            if widget is None:
                return
            
        if text == "Não Cadastrado":
            widget.setText(text)
            return

        numero_cpf = ''.join(filter(str.isdigit, text))[:11]  # Limita a 11 dígitos

        if len(numero_cpf) >= 9:
            cpf_formatado = "{}.{}.{}-{}".format(
                numero_cpf[:3], numero_cpf[3:6], numero_cpf[6:9], numero_cpf[9:]
            )
        else:
            cpf_formatado = numero_cpf

        widget.blockSignals(True)
        widget.setText(cpf_formatado)
        widget.blockSignals(False)
#*********************************************************************************************************************
    def formatar_cnpj(self, text, widget=None):
        if widget is None:
            widget = self.sender()
            if widget is None:
                return

        if text == "Não Cadastrado":
            widget.setText(text)
            return

        # Remover caracteres não numéricos
        numero_cnpj = ''.join(filter(str.isdigit, text))

        # Verifica se há pelo menos 14 dígitos
        if len(numero_cnpj) >= 14:
            numero_cnpj = numero_cnpj[:14]  # Limita a 14 dígitos

            # Formata para o padrão nacional: 11.111.111/1111-11
            cnpj_formatado = "{}.{}.{}/{}-{}".format(
                numero_cnpj[:2],
                numero_cnpj[2:5],
                numero_cnpj[5:8],
                numero_cnpj[8:12],
                numero_cnpj[12:]
            )

            # Atualiza o texto do widget com formatação
            widget.blockSignals(True)
            widget.setText(cnpj_formatado)
            widget.blockSignals(False)
        else:
            widget.blockSignals(True)
            widget.setText(numero_cnpj)
            widget.blockSignals(False)
#*********************************************************************************************************************
    def formatar_moeda(self, widget=None):
        if widget is None:
            widget = self.sender()
            if widget is None:
                return
                
        valor = widget.text().strip()
        
        # Permitir "Não Cadastrado"
        if valor.lower() == "não cadastrado":
            widget.blockSignals(True)
            widget.setText("Não Cadastrado")
            widget.blockSignals(False)
            return

        if not valor:
            return

        try:
            # Remove R$, espaços e separadores
            valor_limpo = (
                valor.replace("R$", "")
                    .replace(" ", "")
                    .replace(".", "")
                    .replace(",", ".")
            )
            valor_float = float(valor_limpo)
        except ValueError:
            QMessageBox.warning(self, "Erro", "Valor inválido.")
            return

        # Formatar como moeda
        valor_formatado = locale.currency(valor_float, grouping=True)

        widget.blockSignals(True)
        widget.setText(valor_formatado)
        widget.blockSignals(False)

#*********************************************************************************************************************
    def formatar_rg(self, text,widget=None):
        if widget is None:
            widget = self.sender()
            if widget is None:
                return

        if text == "Não Cadastrado":
            widget.setText(text)
            return
    
        # Remover caracteres não numéricos
        numero_rg = ''.join(filter(str.isdigit, text))
        
        # Verificar se o RG tem pelo menos 9 dígitos
        if len(numero_rg) >= 9:
            # Formatar o RG com pontos e hífen
            rg_formatado = "{}.{}.{}-{}".format(numero_rg[:2], numero_rg[2:5], numero_rg[5:8], numero_rg[8:9])

            # Atualiza o texto do widget com formatação
            widget.blockSignals(True)
            widget.setText(rg_formatado)
            widget.blockSignals(False)
        else:
            # Atualiza o texto do widget com formatação
            widget.blockSignals(True)
            widget.setText(numero_rg)
            widget.blockSignals(False)

    def formatar_cnh(self, text, widget=None):
        if widget is None:
            widget = self.sender()
            if widget is None:
                return

        if text == "Não Cadastrado":
            widget.setText(text)
            return
        
        # Remove tudo que não for dígito
        apenas_numeros = ''.join(filter(str.isdigit, text))
        
        # Limita a 11 dígitos
        cnh_formatada = apenas_numeros[:11]
        
        # Atualiza o campo
        widget.setText(cnh_formatada)

#*********************************************************************************************************************
    def formatar_data_nascimento(self, text, widget=None):
        if widget is None:
            widget = self.sender()
            if widget is None:
                return

        if text == "Não Cadastrado":
            widget.setText(text)
            return
        
        # Remover todos os caracteres que não são dígitos
        numeros = ''.join(filter(str.isdigit, text))
        data_formatada = ""

        if len(numeros) >= 1:
            data_formatada = numeros[:2] + "/"
            if len(numeros) >= 3:
                data_formatada += numeros[2:4] + "/"
                if len(numeros) >= 5:
                    data_formatada += numeros[4:8]

        widget.blockSignals(True)
        widget.setText(data_formatada)
        widget.blockSignals(False)

    def validar_data_quando_finalizar(self, text, widget=None):
        if widget is None:
            widget = self.sender()
            if widget is None:
                return
            
        if len(text) == 10:  # formato completo dd/mm/yyyy
            try:
                datetime.strptime(text, "%d/%m/%Y")
            except ValueError:
                QMessageBox.warning(self, "Data inválida", f"A data '{text}' é inválida.")
                widget.setText("")
                widget.setFocus()
#*********************************************************************************************************************
    def formatar_telefone(self, text, widget=None):   
        if widget is None:
            widget = self.sender()
            if widget is None:
                return
        
        numero_limpo = ''.join(filter(str.isdigit, text))[:14]

        if len(numero_limpo) >= 2:
            numero_formatado = f"({numero_limpo[:2]}) "

            if len(numero_limpo) >= 8:
                numero_formatado += f"{numero_limpo[2:7]}-{numero_limpo[7:11]}"
            else:
                numero_formatado += numero_limpo[2:]

            widget.blockSignals(True)
            widget.setText(numero_formatado)
            widget.blockSignals(False)

#*******************************************************************************************************
    def conectar_botao_adicionar_produto(self):
        self.btn_adicionar_produto.clicked.connect(self.adicionar_produto)
#*********************************************************************************************************************
    def selecionar_imagem_usuario(self, row):
    # Obter o ID do usuário selecionado
        id_usuario = self.tabela_usuario_dialogo.item(row, 0).text()
        self.produto_id = id_usuario
        
        # Atualizar os campos e a imagem com base no produto selecionado
        self.atualizar_imagem()
#*******************************************************************************************************
    def mostrar_erro_desconto(self):
        detalhes_msg_detalhes = QMessageBox(self)
        detalhes_msg_detalhes.setIcon(QMessageBox.Information)
        detalhes_msg_detalhes.setWindowTitle("Detalhes do Erro")
        detalhes_msg_detalhes.setText("Os campos obrigatórios precisam estar preenchidos.")
        detalhes_msg_detalhes.exec()
#*******************************************************************************************************
    def limpar_campos_produtos(self):
        self.txt_produto.clear()
        self.txt_quantidade.clear()
        self.txt_valor_produto_3.clear()
        self.txt_desconto_3.clear()
        self.dateEdit_3.clear()
        self.txt_codigo_item.clear()
        self.txt_cliente_3.clear()
        self.txt_descricao_produto_3.clear()
        
        # Resetar os QLabel dos frames para valores padrão
        self.label_valor_total_produtos.setText("R$ 0,00")
        self.label_valor_desconto.setText("R$ 0,00")
        self.label_quantidade.setText("0")  # Quantidade é um número simples
        self.label_valor_do_desconto.setText("R$ 0,00")

        
        # Limpar o campo dateEdit e configurar para a data atual
        self.dateEdit_3.setDate(QDate.currentDate())
#*******************************************************************************************************
    def apagar_campos_produtos(self):
        # Verificar se algum campo já está preenchido
        if any([
            self.txt_produto.text(),
            self.txt_quantidade.text(),
            self.txt_valor_produto_3.text(),
            self.txt_desconto_3.text(),
            self.txt_codigo_item.text(),
            self.txt_cliente_3.text(),
            self.txt_descricao_produto_3.text()
        ]):
            # Limpar todos os campos das QLineEdit
            self.limpar_campos_produtos()

            # Limpar a imagem
            self.apagar_imagem_produto_btn_apagar_campos()

            # Mensagem de sucesso
            successMsgBox = QMessageBox(self)
            successMsgBox.setIcon(QMessageBox.Information)
            successMsgBox.setText("Campos limpos com sucesso!")
            successMsgBox.setWindowTitle("Sucesso")
            successMsgBox.exec()
        else:
            # Mostrar mensagem de aviso se nenhum campo estiver preenchido
            QMessageBox.information(self,"Aviso","Não há campos preenchidos para limpar.")
            
#*******************************************************************************************************
    def apagar_imagem_produto_btn_apagar_campos(self):
        # Verificar se o QFrame contém um QLabel para imagem
        if self.frame_imagem_produto_3 is not None:
            for widget in self.frame_imagem_produto_3.children():
                if isinstance(widget, QLabel):
                    widget.clear()  # Limpar o QLabel
                    widget.setPixmap(QPixmap())  # Definir um pixmap vazio ou padrão
#*******************************************************************************************************
    def email_valido(self, email):
        email = email.strip()
        email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
        return bool(email_pattern.match(email))
# **********************************************************************************************************************
    def configurar_geometria_frames(self):
        altura = 101
        largura = 335
        self.frame_valor_total_produtos.setGeometry(335, 101, largura, altura)
#**********************************************************************************************************************
     # Método para limpar o frame após o cadastro do usuário
    def limpar_frame_cadastro_usuario(self): 
        # Remover imagem do QLabel
        self.frame_imagem_cadastro.clear()
#*******************************************************************************************************
    def atualizar_imagem(self):
        if hasattr(self, 'produto_imagem') and self.produto_imagem:
            caminho_imagem = self.produto_imagem
            if caminho_imagem:
                # Carregar a imagem e exibi-la em um QLabel
                pixmap = QPixmap(caminho_imagem)
                if not pixmap.isNull():
                    # Redimensionar a imagem para se ajustar ao QLabel (se necessário)
                    pixmap = pixmap.scaled(200, 200, Qt.KeepAspectRatio)
                    self.label_imagem_produto.setPixmap(pixmap)
                else:
                    QMessageBox.warning(self, "Aviso", "Não foi possível carregar a imagem.")
            else:
                QMessageBox.warning(self, "Aviso", "Não foi possível encontrar o caminho da imagem.")
        else:
            QMessageBox.warning(self, "Aviso", "Nenhuma imagem definida para o produto.")
#**************************************************************************************************************            
    def limpar_campos_após_atualizar_usuario(self):
        self.txt_nome.clear()
        self.txt_usuario_cadastro.clear()
        self.txt_telefone.clear()
        self.txt_endereco.clear()
        self.txt_numero.clear()
        self.txt_cidade.clear()
        self.txt_bairro.clear()
        self.txt_complemento.clear()
        self.txt_email.clear()
        self.txt_data_nascimento.clear()
        self.txt_rg.clear()
        self.txt_cep.clear()
        self.txt_cpf.clear()
        self.txt_cnpj.clear()
        self.txt_senha_cadastro.clear()
        self.txt_confirmar_senha.clear()
        self.perfil_estado.setCurrentIndex(0)
        self.perfil_usuarios.setCurrentIndex(0)
        self.label_imagem_usuario.clear()


    def reiniciar_sistema(self):
        caminho_estado = self.pasta_estado()
        caminho_sistema = self.pasta_do_sistema()

        flag = os.path.join(caminho_estado, "update_pending.flag")
        tmp_exe = os.path.join(caminho_sistema, "SistemadeGerenciamento_tmp.exe")

        # =====================================
        # MODO PYTHON
        # =====================================
        if not getattr(sys, "frozen", False):
            subprocess.Popen(
                [sys.executable, os.path.abspath(sys.argv[0])],
                cwd=caminho_sistema
            )
            QTimer.singleShot(100, QApplication.quit)
            return

        # =====================================
        # MODO EXE COM UPDATE
        # =====================================
        if os.path.exists(flag) and os.path.exists(tmp_exe):
            atualizador, cwd = self.obter_atualizador()
            subprocess.Popen([atualizador], cwd=cwd, creationflags=subprocess.CREATE_NO_WINDOW)
            QTimer.singleShot(100, QApplication.quit)
            return  # 🚫 NÃO reinicia aqui


        # =====================================
        # MODO EXE SEM UPDATE
        # =====================================
        subprocess.Popen(
            [sys.argv[0]],
            cwd=caminho_sistema,
            creationflags=subprocess.DETACHED_PROCESS
        )

        QTimer.singleShot(100, QApplication.quit)


    def auto_completar(self, nome_campo, campo):
        texto = campo.text().strip()
        if not texto:
            return

        self.config.adicionar_ao_historico(nome_campo, texto)

        historico = self.config.carregar_historico_autocompletar(nome_campo)
        novo_completer = QCompleter(historico, self)
        novo_completer.setCaseSensitivity(Qt.CaseInsensitive)
        novo_completer.setCompletionMode(QCompleter.PopupCompletion)
        campo.setCompleter(novo_completer)

    # Método para tratar o duplo clique e preencher o campo
    def completar_por_duplo_clique(self, event, nome_campo, campo):
        if event.type() == QEvent.MouseButtonDblClick:
            # Recupera o completador e o modelo de dados
            completer = campo.completer()
            model = completer.model()

            # Pegando o primeiro item de sugestão do modelo
            if model.rowCount() > 0:
                index = model.index(0, 0)  # Pega o primeiro item
                texto_completo = model.data(index)  # Pega o texto do item
                campo.setText(texto_completo)  # Preenche o campo com a sugestão

                # Atualiza o histórico de autocompletar
                self.config.adicionar_ao_historico(nome_campo, texto_completo)
                return True
        return False


    def exibir_planilhas_exemplo(self):
        # Abre o dialog customizado
        dialog = EscolherPlanilhaDialog(self,tipo_usuario=self.tipo_usuario)
        if dialog.exec() != QDialog.Accepted:
            print("Operação cancelada pelo usuário.")
            return

        escolha = dialog.escolha()
        if not escolha:
            print("Operação cancelada pelo usuário.")
            return

        # Agora, a escolha é diretamente mapeada
        if escolha == "Planilha de Exemplo 1 - Produtos":
            nome_sugestao = "Produtos Exemplo.xlsx"
            sheet_name = 'Produtos'
            dados = {
                "Produto": ["Exemplo 1", "Exemplo 2", "Exemplo 3", "Exemplo 4"],
                "Quantidade": [10, 50, 5, 20],
                "Valor Unitário": [4500.00, 150.00, 1200.00, 900.00],
                "Desconto": [5, "Sem desconto", 10, "Sem desconto"],
                "Data do Cadastro": ["10/05/2024", "10/04/2024", "10/03/2024", "10/02/2024"],
                "Código do Item": ["AUTO12345", "AUTO67890", "AUTO11223", "AUTO33445"],
                "Cliente": ["João da Silva", "Maria Oliveira", "Pedro Santos", "Carla Lima"],
                "Descrição": [
                    "Notebook com 16GB RAM, SSD",
                    "Mouse sem fio, modelo M170",
                    "Impressora multifuncional",
                    "Monitor 24”, Full HD"
                ]
            }
        else:
            nome_sugestao = "Usuários Exemplos.xlsx"
            sheet_name = 'Usuários'
            dados = {
                "Nome": ["Keven Lucas da Silva Jesus", "Maria Oliveira da Silva", "Pedro Santos Reis", "Carla Lima Medeiros"],
                "Usuário": ["Gerado automaticamente", "Gerado automaticamente", "Gerado automaticamente", "Gerado automaticamente"],
                "Senha": ["senha123", "senha456", "senha789", "senha101"],
                "Confirmar Senha": ["senha123", "senha456", "senha789", "senha101"],
                "CEP": ["00000-0001", "00000-0002", "00000-0003", "00000-0004"],
                "Endereço": ["Rua A, 123", "Rua B, 456", "Rua C, 789", "Rua D, 101"],
                "Número": [123, 456, 789, 101],
                "Cidade": ["Vila Madalena", "Sumaré", "Campinas", "Monte Mor"],
                "Bairro": ["Centro", "Zona Sul", "Zona Norte", "Zona Leste"],
                "Estado": ["SP", "RJ", "MG", "PR"],
                "Complemento": ["Apto 1", "Opcional", "Opcional", "465"],
                "Telefone": ["(11) 00000-0001", "(21) 00000-0002", "(31) 00000-0003", "(41) 00000-0004"],
                "E-mail": ["keven.lucas00@dhdfge.com", "mariolieira1000@gmail.com","pedrosantos00123@gmail.com","carlalima14520@gmail.com"],
                "Data de Nascimento": ["01/01/2000", "02/02/1995", "03/03/1990", "04/04/1985"],
                "RG": ["00.000.000-1", "00.000.000-2", "00.000.000-3", "00.000.000-4"],
                "CPF": ["000.000.000-01", "000.000.000-02", "000.000.000-03", "000.000.000-04"],
                "CNPJ": ["00.000.000/0001-01", "00.000.000/0001-02", "00.000.000/0001-03", "00.000.000/0001-04"],   
                "Acesso": ["Convidado", "Usuário", "Usuário", "Convidado"]
            }

        # Gerar a planilha com os dados corretos
        df = pd.DataFrame(dados)

        caminho = salvar_dialogo_memoria(
            parent=self,
            chave="excel_planilhas_exemplos",
            titulo="Salvar Excel",
            nome_padrao=nome_sugestao,
            filtro="Excel (*.xlsx)"
        )


        if not caminho:
            print("Operação cancelada pelo usuário.")
            return

        # Salvar usando openpyxl para aplicar estilos
        df.to_excel(caminho, index=False, engine='openpyxl', sheet_name=sheet_name)

        # Abrir a planilha para formatação
        wb = load_workbook(caminho)
        ws = wb[sheet_name]  # Usar o nome da aba dinamicamente

        # Formatar as colunas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
                ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(caminho)

        QMessageBox.information(
            self,
            "Sucesso",
            f"Planilha '{escolha}' salva com sucesso em {caminho}!",
            QMessageBox.Ok
        )

     
    def abrir_pg_massa_produtos(self):
        self.paginas_sistemas.setCurrentWidget(self.page_cadastrar_massa_produtos)

        self.mostrar_tutorial_pagina(
            chave="tutorial_pg_massa_produtos",
            pagina_widget=self.page_cadastrar_massa_produtos,
            titulo="Cadastro em Massa Produtos",
            texto="Nesta página você pode cadastrar vários produtos de uma vez usando uma planilha.\n\n"
                "1) Importe a planilha\n"
                "2) Clique em Fazer o cadastro em massa.\n"
                "3) Editar produto leva a página de cadastro de produtos\n\n"
                "Observação 1: revise os dados antes de confirmar para evitar erros no estoque.\n"
                "Observação 2: gere a planilha de exemplos para ter uma ideia de como importar e preencher a planilha"
        )
        
    def abrir_pg_massa_usuarios(self):
        self.paginas_sistemas.setCurrentWidget(self.page_cadastrar_massa_usuarios)

        self.mostrar_tutorial_pagina(
            chave="tutorial_pg_massa_usuarios",
            pagina_widget=self.page_cadastrar_massa_usuarios,
            titulo="Cadastro em Massa Usuários",
            texto="Nesta página você pode cadastrar vários usuários de uma vez usando uma planilha.\n\n"
                "1) Importe a planilha\n"
                "2) Clique em Fazer o cadastro em massa.\n"
                "3) Editar usuário leva a página de cadastro de usuários\n\n"
                "Observação 1: revise os dados antes de confirmar para evitar erros.\n"
                "Observação 2: gere a planilha de exemplos para ter uma ideia de como importar e preencher a planilha"
        )

    # Sair do modo edição na página de cadastrar usuários
    def sair_modo_edicao_usuarios(self):
        if not self.is_editing:
            QMessageBox.warning(self, "Aviso", "Você não está no modo de edição.")
            return

        self.is_editing = False
        self.selected_user = None
         # Limpa todos os campos (pode reaproveitar o trecho do subscribe_user)
        self.txt_nome.clear()
        self.txt_usuario_cadastro.clear()
        self.txt_senha_cadastro.clear()
        self.txt_confirmar_senha.clear()
        self.txt_cpf.clear()
        self.txt_cnpj.clear()
        self.txt_email.clear()
        self.txt_numero.clear()
        self.txt_endereco.clear()
        self.txt_bairro.clear()
        self.txt_cidade.clear()
        self.txt_cep.clear()
        self.txt_complemento.clear()
        self.txt_rg.clear()
        self.txt_telefone.clear()
        self.txt_data_nascimento.clear()
        self.perfil_estado.setCurrentIndex(0)
        self.perfil_usuarios.setCurrentIndex(0)
        self.label_imagem_usuario.clear()

        # Se quiser mostrar visualmente que saiu do modo edição (opcional)
        QMessageBox.information(self, "Edição cancelada", "Você saiu do modo de edição.")

    def sair_modo_edicao_produto(self):
        if not self.is_editing_produto:
            QMessageBox.warning(self, "Aviso", "Você não está no modo de edição de produtos.")
            return

        self.is_editing_produto = False
        self.selected_produto_id = None

        # Se quiser mostrar visualmente que saiu do modo edição (opcional)
        QMessageBox.information(self, "Edição cancelada", "Você saiu do modo de edição.")
    
    
    def formatar_tamanho(self,bytes_tamanho):
        for unidade in ['B', 'KB', 'MB', 'GB']:
            if bytes_tamanho < 1024:
                return f"{bytes_tamanho:.2f} {unidade}"
            bytes_tamanho /= 1024
        return f"{bytes_tamanho:.2f} TB"

    
    def limpar_cache_sistema(self):
        pasta_cache = "imagens_temporarias"

        tamanho_total = 0
        erros = []

        progresso = QProgressDialog(
            "Limpando cache do sistema...",
            None,
            0,
            0,
            self
        )
        progresso.setWindowTitle("Limpeza de Cache")
        progresso.setWindowModality(Qt.ApplicationModal)
        progresso.setMinimumDuration(0)
        progresso.show()
        progresso.repaint()
        QApplication.processEvents()

        if os.path.exists(pasta_cache):
            for item in os.listdir(pasta_cache):
                caminho = os.path.join(pasta_cache, item)
                try:
                    if os.path.isfile(caminho) or os.path.islink(caminho):
                        tamanho_total += os.path.getsize(caminho)
                        os.remove(caminho)

                    elif os.path.isdir(caminho):
                        for root, _, files in os.walk(caminho):
                            for file in files:
                                try:
                                    tamanho_total += os.path.getsize(
                                        os.path.join(root, file)
                                    )
                                except OSError:
                                    pass
                        shutil.rmtree(caminho)

                except Exception as e:
                    erros.append(f"{caminho}: {e}")

                QApplication.processEvents()

        progresso.close()

        #  Marca limpeza de __pycache__ no próximo start
        self.marcar_limpeza_pycache()

        mensagem = (
            f"Cache do sistema limpo com sucesso.\n"
            f"Espaço liberado: {self.formatar_tamanho(tamanho_total)}"
        )

        if erros:
            mensagem += "\n\nAlguns itens não puderam ser removidos:\n"
            mensagem += "\n".join(erros[:5])
            if len(erros) > 5:
                mensagem += "\n..."

        QMessageBox.information(self, "Limpeza concluída", mensagem)
        
    def limpar_pycache_pendente(self):
        caminho_estado = self.pasta_estado()
        flag = os.path.join(caminho_estado, "limpar_pycache.flag")

        if not os.path.exists(flag):
            return

        #  __pycache__ só existe em modo Python
        if getattr(sys, "frozen", False):
            os.remove(flag)
            return

        base = self.pasta_do_sistema()  # 👈 caminho correto do projeto

        for root, dirs, _ in os.walk(base):
            if "__pycache__" in dirs:
                try:
                    shutil.rmtree(os.path.join(root, "__pycache__"))
                except Exception:
                    pass

        os.remove(flag)



    def marcar_limpeza_pycache(self):
        caminho_estado = self.pasta_estado()
        os.makedirs(caminho_estado, exist_ok=True)

        flag = os.path.join(caminho_estado, "limpar_pycache.flag")
        with open(flag, "w", encoding="utf-8") as f:
            f.write("1")



    def tratar_f5_global(self):
        pagina_atual = self.paginas_sistemas.currentWidget()
        print(f"[DEBUG] Página atual: {type(pagina_atual)}")

        if pagina_atual == self.pg_clientes:
            if self.pagina_clientes_juridicos.table_clientes_juridicos.isVisible():
                QMessageBox.information(self,"Aviso","Dados atualizados com sucesso!")
                self.pagina_clientes_juridicos.carregar_clientes_juridicos()
            elif self.pagina_clientes_fisicos.table_clientes_fisicos.isVisible():
                QMessageBox.information(self,"Aviso","Dados atualizados com sucesso!")
                self.pagina_clientes_fisicos.carregar_clientes_fisicos()
            else:
                print("Nenhuma tabela de cliente visível")
        else:
            print("Página atual não tem método de atualização com F5.")

    def aplicar_icones(self):
        self.btn_editar.setIcon(
            QIcon(caminho_recurso("imagens/editar.png"))
        )

        self.btn_adicionar_produto.setIcon(
            QIcon(caminho_recurso("imagens/pngwing.com.png"))
        )

        self.btn_sair_modo_edicao_produtos.setIcon(
            QIcon(caminho_recurso("imagens/sair.png"))
        )

        self.btn_atualizar_produto.setIcon(
            QIcon(caminho_recurso("imagens/toppng.com-update-512x512.png"))
        )

        self.btn_limpar_campos.setIcon(
            QIcon(caminho_recurso("imagens/Delete-Button-PNG-Download-Image.png"))
        )
        
        self.btn_remover_imagem.setIcon(
            QIcon(caminho_recurso("imagens/icons8-remover-imagem-16.png"))
        )
        self.btn_carregar_imagem.setIcon(
            QIcon(caminho_recurso("imagens/014upload2_99941.png"))
        )
        self.btn_confirmar.setIcon(
            QIcon(caminho_recurso("imagens/confirmacao.png"))
        )
        self.btn_ver_clientes_juridicos.setIcon(
            QIcon(caminho_recurso("imagens/74472.png"))
        )
        self.btn_ver_item.setIcon(
            QIcon(caminho_recurso("imagens/pasta.png"))
        )
        self.btn_editar_cadastro.setIcon(
            QIcon(caminho_recurso("imagens/editar.png"))
        )
        self.btn_carregar_imagem_4.setIcon(
            QIcon(caminho_recurso("imagens/014upload2_99941.png"))
        )
        self.btn_apagar_cadastro.setIcon(
            QIcon(caminho_recurso("imagens/Delete-Button-PNG-Download-Image.png"))
        )
        self.btn_atualizar_cadastro.setIcon(
            QIcon(caminho_recurso("imagens/toppng.com-update-512x512.png"))
        )
        self.btn_sair_modo_edicao.setIcon(
            QIcon(caminho_recurso("imagens/sair.png"))
        )
        self.btn_remover_imagem_usuario.setIcon(
            QIcon(caminho_recurso("imagens/icons8-remover-imagem-16.png"))
        )
        self.btn_fazer_cadastro.setIcon( 
            QIcon(caminho_recurso("imagens/confirmacao.png"))
        )
        self.btn_ver_usuario.setIcon(
            QIcon(caminho_recurso("imagens/74472.png"))
        )
        self.label_icone_wpp.setPixmap(QPixmap(caminho_recurso("imagens/icone_whatsapp.png")))
        self.label_icone_email.setPixmap(QPixmap(caminho_recurso("imagens/icone_email.png")))
    
    def mostrar_tutorial_pagina(self, chave: str, pagina_widget,titulo: str, texto: str):
        if self.config.tutorial_ja_visto(chave):
            return

        overlay_attr = f"_tutorial_overlay_{chave}"
        overlay = getattr(self, overlay_attr, None)
        if overlay is None:
            overlay = TutorialOverlay(pagina_widget)
            setattr(self, overlay_attr, overlay)

        def on_close(nao_mostrar: bool):
            if nao_mostrar:
                self.config.marcar_tutorial_visto(chave)

        QTimer.singleShot(0, lambda: overlay.show_tutorial(titulo,texto, on_close=on_close))
        
    def abrir_pg_cadastrar_produto(self):
        self.paginas_sistemas.setCurrentWidget(self.pg_cadastrar_produto)

        self.mostrar_tutorial_pagina(
            chave="tutorial_pg_cadastrar_produto",
            pagina_widget=self.pg_cadastrar_produto,
            titulo="Cadastro de Produtos",
            texto="Nesta página você cadastra produtos, edita, aplica desconto e vincula ao cliente.\n"
                "Preencha os campos, adicione o produto e confirme para salvar no sistema."
        )
    def abrir_pg_cadastro_usuario(self):
        self.paginas_sistemas.setCurrentWidget(self.pg_cadastrar_usuario)

        self.mostrar_tutorial_pagina(
            chave="tutorial_pg_cadastrar_usuario",
            pagina_widget=self.pg_cadastrar_usuario,
            titulo="Cadastro de Usuários",
            texto="Nesta página você pode cadastrar, editar e atualizar usuários do sistema.\n"
                "Preencha os campos obrigatórios, defina o nível de acesso e clique em confirmar para salvar."
        )
    
    def mostrar_page_estoque(self):
        # Navegar para a página de estoque
        self.paginas_sistemas.setCurrentWidget(self.pag_estoque)

        self.mostrar_tutorial_pagina(
            chave="tutorial_pg_estoque",
            pagina_widget=self.pag_estoque,
            titulo="Estoque",
            texto="Nesta página você acompanha o estoque e o histórico, registra saídas/estornos e gera relatórios.\n"
                 "Use 'Atualizar' para recarregar os dados e 'Gerar Excel' para exportar."
        )
        
        # Atualizar a tabela na página de estoque
        self.estoque_produtos.carregar_tabela_estoque()
        
    def mostrar_page_verificar_usuarios(self):
        self.paginas_sistemas.setCurrentWidget(self.page_verificar_usuarios)

        self.mostrar_tutorial_pagina(
            chave="tutorial_pg_verificar_usuarios",
            pagina_widget=self.page_verificar_usuarios,
            titulo="Verificar Usuários",
            texto="Nesta página você pode visualizar os usuários cadastrados, "
                "gerar a saída de usuários, importar registros e acompanhar o histórico de movimentações.\n\n"
                "⚠️ Atenção: usuários que tiverem saída gerada não poderão ser estornados. "
                "Essa ação é irreversível."
        )
    
    def mostrar_page_clientes(self):
        self.paginas_sistemas.setCurrentWidget(self.pg_clientes)

        self.mostrar_tutorial_pagina(
            chave="tutorial_pg_clientes",
            pagina_widget=self.pg_clientes,
            titulo="Clientes",
            texto="Nesta seção você gerencia os clientes do sistema, tanto físicos quanto jurídicos.\n\n"
                "Aqui é possível cadastrar, editar e excluir clientes, gerar relatórios "
                "e acompanhar o histórico de movimentações.\n\n"
                "As informações são organizadas por tipo de cliente, mas seguem o mesmo padrão de controle."
        )
        
    
            
    def closeEvent(self, event):
        event.accept()


# Função principal
if __name__ == '__main__':
    app = QApplication(sys.argv)

    app.setStyle("WindowsVista")
    app.setWindowIcon(QIcon(caminho_recurso("imagens/ícone_sistema_provisório.png")))

    temas = Temas()
    temas.aplicar_tema_global(app)

    # Carrega configurações ANTES de criar janelas
    config = Configuracoes_Login(None)
    config.carregar()

    # Tenta login automático sem criar/mostrar a tela de login
    tipo_usuario = None
    if config.mantem_conectado and config.usuario:
        senha_salva = config.obter_senha_salva()
        if senha_salva:
            db = DataBase("banco_de_dados.db")
            db.connecta()
            tipo_usuario = db.check_user(config.usuario, senha_salva)

    if tipo_usuario:
        #  Abre direto a MainWindow (login nem aparece)
        main_window = MainWindow(
            user=tipo_usuario.lower(),
            tipo_usuario=tipo_usuario,
            login_window=None,
            app=app
        )
        main_window.show()
    else:
        #  Cai no fluxo normal: mostra login
        login_window = Login(login_window=None)
        login_window.show()
        
    

    sys.exit(app.exec())






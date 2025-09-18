from PySide6.QtWidgets import (QDialog, QPushButton, QVBoxLayout, QTableWidget, 
                               QTableWidgetItem, QMessageBox, QCheckBox, QLabel, QLineEdit, 
                               QLabel, QFrame,QMainWindow,QWidget,QComboBox,QHeaderView,
                               QAbstractItemView,QHBoxLayout,QFileDialog)
from PySide6 import QtWidgets
from PySide6.QtGui import QPixmap, Qt, QImage,QBrush, QColor
from PySide6.QtCore import  Qt,QEvent,QTimer
from PySide6 import QtCore
from database import DataBase, sqlite3
import base64
import re
import os
from configuracoes import Configuracoes_Login
from utils import Temas
from dialogos import FiltroUsuarioDialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import json



class TabelaUsuario(QMainWindow):
    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main_window = main_window
        self.frame_imagem_cadastro = QFrame()
        self.db = DataBase()
        

        self.setWindowTitle("Tabela de Usuários")
        self.resize(800, 600)

        self.temas = Temas()


        self.config = Configuracoes_Login(self)
        self.config.carregar()

        self.usuarios = self.db.get_users()
        self.usuario_selecionado = False
        
        self.checkboxes = []  # Lista para armazenar os checkboxes
        
        self.coluna_checkboxes_adicionada = False  # Variável para controlar se a coluna de checkbox foi adicionada

       

        self.limpar_campos_de_texto()

        config = self.temas.carregar_config_arquivo()
        self.tema = config.get("tema", "claro")

        # Criar tabela
        self.table_widget = QTableWidget(self)
        self.table_widget.setObjectName("tabelaUsuarios")
        self.setStyleSheet(self.aplicar_tema(self.tema))
        self.table_widget.setColumnCount(24)
        self.table_widget.setFocusPolicy(Qt.StrongFocus)
        self.table_widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.table_widget.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)


         # Widget central e layout principal vertical
        widget_central = QWidget()
        self.layout_tabela = QVBoxLayout(widget_central)

        # Criar QLabel para imagem
        self.label_imagem_usuario = QLabel()
        self.label_imagem_usuario.setScaledContents(True)
        layout_usuario = QVBoxLayout()
        layout_usuario.addWidget(self.label_imagem_usuario)
        self.main_window.frame_imagem_cadastro.setLayout(layout_usuario)

        # Botões (verticais para ficar empilhados e largos)
        self.btn_apagar_usuario = QPushButton("Apagar Usuários")
        self.btn_editar_usuario = QPushButton("Atualizar Usuário")
        self.btn_filtrar_usuario = QPushButton("Filtrar Usuários")
        self.checkbox_selecionar = QCheckBox("Selecionar")
        self.btn_ordenar_usuario = QPushButton("Ordenar Usuários")
        self.btn_visualizar_imagem = QPushButton("Visualizar Imagem")
        self.btn_atualizar_tabela = QPushButton("Atualizar Tabela")
        self.btn_gerar_excel = QPushButton("Gerar arquivo Excel")

        # Colocar todos os botões numa lista
        botoes = [
            self.btn_apagar_usuario,
            self.btn_editar_usuario,
            self.btn_filtrar_usuario,
            self.btn_ordenar_usuario,
            self.btn_visualizar_imagem,
            self.btn_atualizar_tabela,
            self.btn_gerar_excel
        ]
        for btn in botoes:
            btn.setCursor(Qt.PointingHandCursor)


        layout_botoes = QVBoxLayout()
        layout_botoes.addWidget(self.btn_apagar_usuario)
        layout_botoes.addWidget(self.btn_editar_usuario)
        layout_botoes.addWidget(self.btn_filtrar_usuario)
        layout_botoes.addWidget(self.btn_ordenar_usuario)
        layout_botoes.addWidget(self.btn_visualizar_imagem)
        layout_botoes.addWidget(self.btn_atualizar_tabela)
        layout_botoes.addWidget(self.btn_gerar_excel)
        layout_botoes.addWidget(self.checkbox_selecionar)
        

        # Faz os botões ficarem do tamanho da largura da janela
        for btn in [self.btn_apagar_usuario, self.btn_editar_usuario, self.btn_filtrar_usuario,
                    self.checkbox_selecionar, self.btn_ordenar_usuario, self.btn_visualizar_imagem,
                    self.btn_atualizar_tabela,self.btn_gerar_excel]:
            btn.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)

        # Adicionar layout dos botões em cima
        self.layout_tabela.addLayout(layout_botoes)

        # Depois adicionar a tabela (ocupando o restante)
        self.layout_tabela.addWidget(self.table_widget)

        # Adicionar o frame de imagem (se quiser que fique embaixo da tabela)
        self.layout_tabela.addWidget(self.frame_imagem_cadastro)

        # Definir widget central
        self.setCentralWidget(widget_central)

        # Inicializar a tabela e limpar campos
        self.limpar_campos_de_texto()
        self.preencher_tabela_usuario()


        # Conectar sinais dos botões
        self.btn_apagar_usuario.clicked.connect(self.confirmar_apagar_usuario)
        self.btn_editar_usuario.clicked.connect(self.editar_usuario)
        self.btn_filtrar_usuario.clicked.connect(self.filtrar_usuario)
        self.btn_ordenar_usuario.clicked.connect(self.ordenar_usuario)
        self.btn_visualizar_imagem.clicked.connect(self.visualizar_imagem_usuario)
        self.checkbox_selecionar.stateChanged.connect(self.ao_clicar_em_selecionar)
        self.btn_atualizar_tabela.clicked.connect(self.atualizar_tabela_usuario)
        self.btn_gerar_excel.clicked.connect(self.gerar_arquivo_excel_usuarios)

    
    

    
    def aplicar_tema(self, tema: str) -> str:
        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    border-radius: 8px;
                    background: qlineargradient(
                        x1:0, y1:0, x2:0, y2:1,
                        stop:0 rgb(60, 60, 60),   /* topo */
                        stop:1 rgb(100, 100, 100) /* base */
                    );
                    font-size: 12px;
                    padding: 3px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
                
            """
            
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            table_view_style = """
                QTableView {
                    background-color: white;
                    color: black;
                    gridline-color: #ccc;
                    selection-background-color: #e5f3ff;
                    selection-color: black;
                }

                QHeaderView::section {
                    background-color: #f0f0f0;
                    color: black;
                    border: 1px solid #ccc;
                    padding: 1px;
                }

                QTabWidget::pane {
                    border: 1px solid #ccc;
                    background-color: white;
                }

                /* Scrollbar vertical */
                QTableView QScrollBar:horizontal {
                    border: none;
                    background-color: #f0f0f0;
                    width: 12px;
                    margin: 0px;
                    border-radius: 5px;
                }
                
                /* Scrollbar vertical */
                QTableView QScrollBar:vertical {
                    border: none;
                    background-color: #f0f0f0;
                    width: 12px;
                    margin: 0px;
                    border-radius: 5px;
                }

                /* Parte que você arrasta - Handle */
                QTableView QScrollBar::handle:vertical {
                    background-color: #b0b0b0;  /* cinza claro */
                    min-height: 22px;
                    border-radius: 5px;
                }

                QTableView QScrollBar::handle:horizontal {
                    background-color: #b0b0b0;
                    min-width: 22px;
                    border-radius: 5px;
                }

                /* Groove horizontal */
                QTableView QScrollBar::groove:horizontal {
                    background-color: #e0e0e0;
                    border-radius: 5px;
                    height: 15px;
                    margin: 0px 10px 0px 10px;
                }

                /* Groove vertical */
                QTableView QScrollBar::groove:vertical {
                    background-color: #e0e0e0;
                    border-radius: 5px;
                    width: 15px;
                    margin: 10px 0px 10px 0px;
                }

                QTableWidget::item:selected {
                    background-color: #cce7ff;
                    color: black;
                }

                QTableCornerButton::section {
                    background-color: #f0f0f0;
                    border: 1px solid #ccc;
                }
                /* Forçar cor do texto do QCheckBox */
                QCheckBox {
                    color: white;
            }
            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """
        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                border-radius: 8px;
                background: qlineargradient(
                    x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgb(220, 220, 220),  /* topo */
                    stop:1 rgb(245, 245, 245)   /* base */
                );
                font-size: 14px;
                color: #000000; /* texto escuro */
                }

                QPushButton:hover {
                    background-color: #e0e0e0;
                }

                QPushButton:pressed {
                    background-color: #d0d0d0;
                    border: 2px solid #aaaaaa;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            scroll_style = """
                /* Scrollbar vertical */
                QScrollBar:vertical {
                    background: #f0f0f0;  /* trilho claro */
                    width: 12px;
                    margin: 0px;
                    border-radius: 6px;
                }

                QScrollBar::handle:vertical {
                    background: #b0b0b0;  /* cor do handle */
                    border-radius: 6px;
                    min-height: 20px;
                }

                QScrollBar::handle:vertical:hover {
                    background: #a0a0a0;  /* hover no handle */
                }

                QScrollBar::add-line:vertical,
                QScrollBar::sub-line:vertical {
                    background: none;
                    height: 0px;
                }

                QScrollBar::add-page:vertical,
                QScrollBar::sub-page:vertical {
                    background: none;
                }
                """
            table_view_style = """
                QTableView {
                    background-color: white;
                    color: black;
                    gridline-color: #ccc;
                    selection-background-color: #e5f3ff;
                    selection-color: black;
                }
                QHeaderView:vertical {
                    background-color: white; 
                    border: none;              
                }

                QHeaderView::section {
                    background-color: #f0f0f0;
                    color: black;
                    border: 1px solid #ccc;
                    padding: 1px;
                }

                QTabWidget::pane {
                    border: 1px solid #ccc;
                    background-color: white;
                }

                /* Scrollbar horizontal */
                QTableView QScrollBar:horizontal {
                    border: none;
                    background-color: #f0f0f0;
                    width: 12px;
                    margin: 0px;
                    border-radius: 5px;
                }
                
                /* Scrollbar vertical */
                QTableView QScrollBar:vertical {
                    border: none;
                    background-color: #f0f0f0;
                    width: 12px;
                    margin: 0px;
                    border-radius: 5px;
                }

                /* Parte que você arrasta - Handle */
                QTableView QScrollBar::handle:vertical {
                    background-color: #b0b0b0;  /* cinza claro */
                    min-height: 22px;
                    border-radius: 5px;
                }

                QTableView QScrollBar::handle:horizontal {
                    background-color: #b0b0b0;
                    min-width: 22px;
                    border-radius: 5px;
                }

                /* Groove horizontal */
                QTableView QScrollBar::groove:horizontal {
                    background-color: #e0e0e0;
                    border-radius: 5px;
                    height: 15px;
                    margin: 0px 10px 0px 10px;
                }

                /* Groove vertical */
                QTableView QScrollBar::groove:vertical {
                    background-color: #e0e0e0;
                    border-radius: 5px;
                    width: 15px;
                    margin: 10px 0px 10px 0px;
                }

                QTableWidget::item:selected {
                    background-color: #cce7ff;
                    color: black;
                }

                QTableCornerButton::section {
                    background-color: #f0f0f0;
                    border: 1px solid #ccc;
                }
                /* Forçar cor do texto do QCheckBox */
                QCheckBox {
                    color: black;
                }
                
            """

            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """
        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    color: rgb(255, 255, 255);
                    border-radius: 8px;
                    font-size: 12px;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgb(50, 150, 250), stop:1 rgb(100, 200, 255)); /* Gradiente de azul claro para azul mais claro */
                    border: 4px solid transparent;
                }

                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgb(100, 180, 255), stop:1 rgb(150, 220, 255)); /* Gradiente de azul mais claro para azul ainda mais claro */
                    color: black;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 3px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """

            #  Scroll geral (scroll_style)
            scroll_style = """
                QScrollBar:vertical {
                    border: none;
                    background-color: rgb(255, 255, 255); /* branco */
                    width: 30px;
                    margin: 0px 10px 0px 10px;
                }
                QScrollBar::handle:vertical {
                    background-color: rgb(180, 180,180);  /* cinza */
                    min-height: 30px;
                    border-radius: 5px;
                }
            """

            #  Estilo específico para QTableView (table_view_style)
            table_view_style = """
                QTableView {
                    background-color: rgb(0,80,121);
                    color: white;
                    gridline-color: black;
                    border: 1px solid white;
                    selection-background-color: #007acc;
                    selection-color: white;
                }

                QHeaderView::section {
                    background-color: white;
                    color: black;
                    border: 1px solid #eeeeee;  /* Borda branco-acinzentada */
                    padding: 1px;
                }

                QTabWidget::pane {
                    border: 1px solid #004466;
                    background-color: #003355;
                }

                /* Scrollbars da QTableView - vertical */
                QTableView QScrollBar:vertical{
                    border: none;
                    background-color: rgb(255, 255, 255); /* fundo do track */
                    border-radius: 5px;
                    width: 10px; /* largura da barra vertical */
                    margin: 0px;
                
                }
                /* Scrollbars da QTableView - horizontal */
                QTableView QScrollBar:horizontal {
                    height: 11px;
                    background-color: rgb(255, 255, 255);
                    margin: 0px;
                }

                /* Handle dos scrolls (a parte que você arrasta) */
                QTableView QScrollBar::handle:vertical {
                    background-color: rgb(180, 180, 150);  /* cor do handle */
                    min-height: 10px;
                    min-width: 10px;
                    border-radius: 5px;
                
                }
                /* Handle dos scrolls (a parte que você arrasta) */
                QTableView QScrollBar::handle:horizontal {
                    background-color: rgb(180, 180, 150);
                    min-width: 20px;
                    border-radius: 5px;
                }

                /* Groove vertical */
                QTableView QScrollBar::groove:vertical {
                    background-color: rgb(100, 240, 240);  /* faixa visível no vertical */
                    border-radius: 2px;
                    width: 10px;
                    margin: 0px 10px 0px 10px;
                }

                /* Groove horizontal (faixa por onde o handle desliza) */
                QTableView QScrollBar::groove:horizontal {
                    background-color: rgb(220, 220, 220);
                    border-radius: 5px;
                    height: 10px;
                }

                QTableWidget::item:selected {
                    background-color: rgb(0, 120, 215);
                    color: white;
                }

                QTableCornerButton::section {
                    background-color: white;
                }
            """

            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        estilo_completo = f"""
        QMainWindow {{
            background-color: {bg_cor};
        }}
        {button_style}
        {lineedit_style}
        {combobox_style}
        {table_view_style}
        {scroll_style}
        """
        return estilo_completo

    
    def preencher_tabela_usuario(self):
        self.table_widget.setRowCount(0)

        column_titles = [
            "ID","Nome", "Usuário", "Senha", "Confirmar Senha", "CEP", "Endereço",
            "Número", "Cidade", "Bairro", "Estado", "Complemento", "Telefone", "Email",
            "Data de Nascimento", "RG", "CPF", "CNPJ",
            "Última Troca de Senha", "Data da Senha Cadastrada",
            "Data da Inclusão do Usuário", "Segredo", "Usuário Logado", "Acesso"
        ]

        for col, title in enumerate(column_titles):
            self.table_widget.setHorizontalHeaderItem(col, QTableWidgetItem(title))

        try:
            self.db.connecta()
            usuarios = self.db.obter_usuarios_sem_imagem()

            for usuario in usuarios:
                row_position = self.table_widget.rowCount()
                self.table_widget.insertRow(row_position)


                for i, dado in enumerate(usuario):
                    item = QTableWidgetItem(str(dado))
                    self.table_widget.setItem(row_position, i, item)
         

            self.table_widget.resizeColumnsToContents()
            self.table_widget.resizeRowsToContents()

            if self.table_widget.rowCount() == 0:
                self.exibir_mensagem_sem_usuarios()
            else:
                self.ocultar_mensagem_sem_usuarios()

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao acessar o banco de dados: {str(e)}")
#*******************************************************************************************************
    def exibir_mensagem_sem_usuarios(self):
        # Verificar se a QLabel já existe
        if not hasattr(self, 'label_sem_usuario'):
            self.label_sem_usuario = QLabel("Usuários cadastrados serão exibidos aqui...")
            self.label_sem_usuario.setAlignment(Qt.AlignCenter)
            self.label_sem_usuario.setStyleSheet("font-size: 16px; color: black;")
            
            # Verificar se o widget tem um layout
            if not self.table_widget.layout():
                self.main_layout = QVBoxLayout(self.table_widget)
                self.table_widget.setLayout(self.main_layout)
            else:
                self.main_layout = self.table_widget.layout()

            self.main_layout.addWidget(self.label_sem_usuario)

        self.label_sem_usuario.show()

    def ocultar_mensagem_sem_usuarios(self):
        if hasattr(self,"label_sem_usuario"):
            self.label_sem_usuario.hide()
#*******************************************************************************************************
    def confirmar_apagar_usuario(self):
        # Verificar se uma linha está selecionada
        if self.table_widget.currentRow() >= 0:
            # Exibir uma mensagem de confirmação
            msg_box = QMessageBox()
            msg_box.setWindowTitle("Confirmar")
            msg_box.setText("Você tem certeza que deseja apagar este usuário?")

            
            sim_button = QPushButton("Sim")
            sim_button.clicked.connect(self.apagar_usuario_confirmado)
            msg_box.addButton(sim_button, QMessageBox.YesRole)
            
            nao_button = QPushButton("Não")
            nao_button.clicked.connect(msg_box.reject)
            msg_box.addButton(nao_button, QMessageBox.NoRole)

            # Exibir a caixa de mensagem
            msg_box.exec()
        else:
            QMessageBox.warning(None, "Aviso", "Nenhum usuário selecionado para apagar.")
#*******************************************************************************************************
    def apagar_usuario_confirmado(self):
        # Obter o índice da linha selecionada
        index = self.table_widget.currentRow()
        
        # Obter o item da célula
        item = self.table_widget.item(index, 0)
        
        # Verificar se o item não é None e se o texto não está vazio
        if item is not None and item.text():
            # Obter o ID do usuário da coluna 0 (ID)
            id_usuario = int(item.text())
            
            #Obter o nome do usuário 
            nome_usuario = self.table_widget.item(index, 1).text()
            
            # Remover a linha da tabela
            self.table_widget.removeRow(index)
            
            # Remover o usuário do banco de dados
            db = DataBase()
            try:
                db.connecta()
                db.remover_usuario(id_usuario)
                QMessageBox.information(self, "Sucesso", "Usuário removido com sucesso!")
                
                usuario = f"O Usuário {nome_usuario} foi removido."
                self.main_window.registrar_historico_usuarios("Removação de Usuário", usuario)
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao remover o usuário: {str(e)}")
            finally:
                pass
        else:
            QMessageBox.warning(self, "Aviso", "Nenhuma célula selecionada ou célula vazia.")
#*******************************************************************************************************
    def recuperar_imagem_do_banco_usuarios(self, id_usuario):
        imagem_blob = None
        
        try:
            connection = self.db.connecta()
            if connection:
                cursor = connection.cursor()
                cursor.execute("SELECT Imagem FROM users WHERE id = ?", (id_usuario,))
                result = cursor.fetchone()

                if result:
                    imagem_blob = result[0]
                else:
                    print(f"Imagem não encontrada para o usuário: {id_usuario}")
                    return None

        except Exception as e:
            print(f"Erro ao recuperar imagem do banco de dados: {str(e)}")
            return None

        if imagem_blob and isinstance(imagem_blob,str) and imagem_blob.strip().lower() != "não cadastrado":
            try:
                imagem_data = base64.b64decode(imagem_blob)
                return imagem_data  # Retorna bytes decodificados
                
            except Exception as e:
                print(f"Erro ao decodificar imagem: {str(e)}")
                return None
        else:
            print(f"Imagem não cadastrada ou inválida para o usuário: {id_usuario}")
            return None
#*******************************************************************************************************
    def editar_usuario(self):
        print("Função editar_usuario chamada")

        row_index = -1
        id_item = None # Inicialmente vazio

        #Verifica quantos checkboxes estão marcados (se estiverem visíveis)
        if self.coluna_checkboxes_adicionada:
            selecionados = [cb for cb in self.checkboxes if cb.isChecked()]

            if len(selecionados) == 0:
                QMessageBox.warning(None, "Aviso", "Nenhum usuário selecionado para editar.")
                return
            elif len(selecionados) > 1:
                QMessageBox.warning(None, "Aviso", "Você só pode editar um usuário por vez.")
                return
            
            # Encontrar o índice da linha do checkbox marcado
            for i in range(self.table_widget.rowCount()):
                widget = self.table_widget.cellWidget(i, 0)
                if widget:
                    checkbox = widget.findChild(QCheckBox)
                    if checkbox and checkbox.isChecked():
                        row_index = i
                        id_item = self.table_widget.item(row_index, 1)  # ID na coluna 1 quando tem checkbox
                        break
        else:        
            row_index = self.table_widget.currentRow()
            if row_index < 0:
                QMessageBox.warning(None, "Aviso", "Selecione um usuário na tabela.")
                return
            id_item = self.table_widget.item(row_index, 0) # ID na coluna 0 quando não tem checkbox

        # Validar ID
        if not id_item or not id_item.text().isdigit():
            print("id_item está vazio!")
            QMessageBox.warning(None, "Aviso", "ID de usuário inválido ou não encontrado.")
            return

        
        id_usuario = int(id_item.text())
        print(f"ID selecionado: {id_usuario}")

        # Obter a imagem do banco de dados
        imagem_data = self.recuperar_imagem_do_banco_usuarios(id_usuario)

        if not hasattr(self, 'label_imagem_usuario') or self.label_imagem_usuario is None:
            self.label_imagem_usuario = QLabel()
            self.label_imagem_usuario.setScaledContents(True)

        self.main_window.usuario_tem_imagem_salva = bool(imagem_data)

        if imagem_data:
            try:
                image = QImage.fromData(imagem_data)
                if not image.isNull():
                    pixmap = QPixmap.fromImage(image)

                    # Redimensionar a imagem para o tamanho do QLabel
                    pixmap = pixmap.scaled(self.label_imagem_usuario.size(), Qt.KeepAspectRatio)

                    # Definir a imagem no QLabel
                    self.label_imagem_usuario.setPixmap(pixmap)
                    self.label_imagem_usuario.repaint()
                    print("Imagem definida no QLabel")
                else:
                    print("Erro: Imagem está vazia")
            except Exception as e:
                print(f"Erro ao processar imagem: {str(e)}")
        else:
            print("Imagem não encontrada no banco de dados.")

        # Recuperar dados do usuário diretamente do banco
        dados_usuario = self.db.recuperar_usuario_por_id(id_usuario)
        if dados_usuario:
            (usuario_nome, usuario_usuario, usuario_senha, usuario_confirmar_senha, usuario_cep, usuario_endereco,
            usuario_numero, usuario_cidade, usuario_bairro, usuario_estado, usuario_complemento, usuario_telefone,
            usuario_email, usuario_data_nascimento, usuario_rg, usuario_cpf, usuario_cnpj, usuario_acesso) = dados_usuario

            # Preencher os campos da MainWindow com os dados do usuário selecionado
            self.main_window.txt_nome.setText(usuario_nome or "")
            self.main_window.txt_usuario_cadastro.setText(usuario_usuario or "")
            self.main_window.txt_senha_cadastro.setText(usuario_senha or "")
            self.main_window.txt_confirmar_senha.setText(usuario_confirmar_senha or "")
            self.main_window.txt_cep.setText(usuario_cep or "")
            self.main_window.txt_endereco.setText(usuario_endereco or "")
            self.main_window.txt_numero.setText(usuario_numero or "")
            self.main_window.txt_cidade.setText(usuario_cidade or "")
            self.main_window.txt_bairro.setText(usuario_bairro or "")
            self.main_window.perfil_estado.setCurrentText(usuario_estado or "")
            self.main_window.txt_complemento.setText(usuario_complemento or "")
            self.main_window.txt_telefone.setText(usuario_telefone or "")
            self.main_window.txt_email.setText(usuario_email or "")
            self.main_window.txt_data_nascimento.setText(usuario_data_nascimento or "")
            self.main_window.txt_rg.setText(usuario_rg or "")
            self.main_window.txt_cpf.setText(usuario_cpf or "")
            self.main_window.txt_cnpj.setText(usuario_cnpj or "")
            self.main_window.perfil_usuarios.setCurrentText(usuario_acesso or "")


            # Atualizar o layout do frame_imagem_cadastro com o QLabel
            self.main_window.frame_imagem_cadastro.setVisible(True)  # Garante que o frame esteja visível
            self.main_window.frame_imagem_cadastro.layout().addWidget(self.label_imagem_usuario)

            self.main_window.is_editing = True  # Indica que um usuário está em edição
            self.main_window.selected_user_id = id_usuario  # Guarda o ID do usuário selecionado
            self.main_window.imagem_removida_usuario = False  # Indica que a imagem não foi removida

            self.usuario_selecionado = True  # Indica que um usuário foi selecionado

            self.close()
        else:
            print("Usuário não encontrado no banco de dados.")

    def usuario_foi_selecionado(self):
        return self.usuario_selecionado     

#*******************************************************************************************************
    def atualizar_tabela_usuario_filtrada(self, usuarios):
        self.table_widget.setRowCount(0)

        # Cabeçalhos: sem a coluna "Imagem"
        column_titles = [
            "ID","Nome", "Usuário", "Senha", "Confirmar Senha", "CEP", "Endereço",
            "Número", "Cidade", "Bairro", "Estado", "Complemento", "Telefone", "Email",
            "Data de Nascimento", "RG", "CPF", "CNPJ",
            "Última Troca de Senha", "Data da Senha Cadastrada",
            "Data da Inclusão do Usuário", "Segredo", "Usuário Logado", "Acesso"
        ]

        self.table_widget.setColumnCount(len(column_titles))
        for col, title in enumerate(column_titles):
            self.table_widget.setHorizontalHeaderItem(col, QTableWidgetItem(title))

        usuario_logado = self.config.obter_usuario_logado()

        for usuario in usuarios:
            # Remover a coluna "Imagem" (índice 18) de uma cópia da tupla
            dados = list(usuario)
            if len(dados) >= 25:
                del dados[18]  # Remove "Imagem"
                dados[23] = usuario_logado


            row_position = self.table_widget.rowCount()
            self.table_widget.insertRow(row_position)

            for col, data in enumerate(dados):
                item = self.formatar_texto(str(data))
                self.table_widget.setItem(row_position, col, item)

        self.table_widget.resizeColumnsToContents()
        self.table_widget.resizeRowsToContents()
#*******************************************************************************************************
    def obter_usuarios_por_filtro(self, campo, valor):
        query = f"""
            SELECT "ID","Nome", "Usuário", "Senha", "Confirmar Senha", "CEP", "Endereço",
            "Número", "Cidade", "Bairro", "Estado", "Complemento", "Telefone", "Email",
            "Data de Nascimento", "RG", "CPF", "CNPJ",
            "Última Troca de Senha", "Data da Senha Cadastrada",
            "Data da Inclusão do Usuário", "Segredo", "Usuário Logado", "Acesso"
            FROM users
            WHERE "{campo}" LIKE ?
        """
        parameters = (f"%{valor}%",)
        
        try:
            db = DataBase()
            connection = db.connecta()
            cursor = connection.cursor()
            cursor.execute(query, parameters)
            return cursor.fetchall()
        except Exception as e:
            print(f"Erro ao filtrar usuários por {campo}:", e)
            return []
        finally:
            if cursor:
                cursor.close()
#*******************************************************************************************************
    def filtrar_usuario(self):
        if getattr(self, "checkbox_header_users", None) and self.checkbox_header_users.isChecked():
            QMessageBox.warning(
                None,
                "Aviso",
                "Desmarque o checkbox antes de filtrar os usuários."
            )
            return

        dialog = FiltroUsuarioDialog(self)
        if dialog.exec():
            criterio, valor = dialog.get_valores()

            mapeamento_campo = {
                "Filtrar por Nome": "Nome",
                "Filtrar Por Usuário": "Usuário",
                "Filtrar Por Acesso": "Acesso",
                "Filtrar Por Telefone": "Telefone",
                "Filtrar Por Email": "Email",
                "Filtrar Por RG": "RG",
                "Filtrar Por CPF": "CPF",
                "Filtrar Por CNPJ": "CNPJ"
            }

            campo_bd = mapeamento_campo.get(criterio)
            if campo_bd:
                usuarios = self.obter_usuarios_por_filtro(campo_bd, valor)

                if not usuarios:
                    QMessageBox.warning(
                        dialog,
                        "Nenhum resultado encontrado",
                        f"Nenhum usuário com {campo_bd} '{valor}' foi encontrado no sistema."
                    )
                else:
                    self.atualizar_tabela_usuario_filtrada(usuarios)


    def ao_clicar_em_selecionar(self, estado):
        if estado == Qt.Checked:
            self.adicionar_coluna_checkboxes()  # Adiciona os checkboxes na tabela
        else:
            # Desmarcar o checkbox do cabeçalho, se estiver marcado
            if hasattr(self, "checkbox_header_users"):
                self.checkbox_header_users.blockSignals(True)
                self.checkbox_header_users.setChecked(False)
                self.checkbox_header_users.blockSignals(False)

            self.selecionar_users_individual()  # Remove a coluna de checkboxes

#*******************************************************************************************************
    def selecionar_todos_users(self):
        if not self.coluna_checkboxes_adicionada:
            # Se a coluna não está ativa, chama para garantir que tudo fique desativado
            self.selecionar_users_individual()
            return

        estado = self.checkbox_header_users.checkState()

        # Se desmarcou o checkbox do cabeçalho E todos os checkboxes individuais estão desmarcados,
        # aí remove a coluna inteira (checkboxes)
        if estado == Qt.Unchecked and all(not cb.isChecked() for cb in self.checkboxes):
            self.selecionar_users_individual()
            return

        # Marca ou desmarca todos os checkboxes individuais de acordo com o estado do cabeçalho
        estado_checked = estado == Qt.Checked
        for checkbox in self.checkboxes:
            checkbox.blockSignals(True)
            checkbox.setChecked(estado_checked)
            checkbox.blockSignals(False)


    def selecionar_users_individual(self):
        if self.table_widget.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Não há usuários na tabela para selecionar.")
            return

        if self.coluna_checkboxes_adicionada:
            self.table_widget.removeColumn(0)
            self.table_widget.verticalHeader().setVisible(True)
            self.coluna_checkboxes_adicionada = False

            if hasattr(self, "checkbox_header_users"):
                self.checkbox_header_users.blockSignals(True)
                self.checkbox_header_users.setChecked(False)  # Garante que não fique marcado
                self.checkbox_header_users.blockSignals(False)
                
                self.checkbox_header_users.deleteLater()
                del self.checkbox_header_users

            self.checkboxes.clear()
            return

        self.table_widget.insertColumn(0)
        self.table_widget.setHorizontalHeaderItem(0, QTableWidgetItem(""))
        self.table_widget.setColumnWidth(0, 30)
        self.table_widget.horizontalHeader().setMinimumSectionSize(30)
        self.table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)

        # Checkbox do cabeçalho
        header_usuarios = self.table_widget.horizontalHeader()
        self.checkbox_header_users = QCheckBox(header_usuarios.viewport())
        self.checkbox_header_users.setToolTip("Selecionar todos")
        self.checkbox_header_users.setStyleSheet("""QCheckBox{background: transparent;}""")
        self.checkbox_header_users.setChecked(False)
        self.checkbox_header_users.setFixedSize(20, 20)
        self.checkbox_header_users.stateChanged.connect(self.selecionar_todos_users)
        self.checkbox_header_users.show()

        QTimer.singleShot(0, self.atualizar_posicao_checkbox_header_users)

        

        self.checkboxes.clear()

        for row in range(self.table_widget.rowCount()):
            checkbox = QCheckBox()
            checkbox.stateChanged.connect(self.atualizar_selecao_todos_usuarios)

            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(checkbox)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)

            self.table_widget.setCellWidget(row, 0, container)
            self.checkboxes.append(checkbox)

        self.table_widget.verticalHeader().setVisible(False)
        self.table_widget.horizontalScrollBar().valueChanged.connect(self.atualizar_posicao_checkbox_header_users)
        header_usuarios.sectionResized.connect(self.atualizar_posicao_checkbox_header_users)
        header_usuarios.geometriesChanged.connect(self.atualizar_posicao_checkbox_header_users)
        self.coluna_checkboxes_adicionada = True

        
    def atualizar_selecao_todos_usuarios(self):
        self.checkbox_header_users.blockSignals(True)

        # Atualizar o estado do "Selecionar Todos"
        all_checked = all(checkbox.isChecked() for checkbox in self.checkboxes if checkbox)
        any_checked = any(checkbox.isChecked() for checkbox in self.checkboxes if checkbox)

        if all_checked:
            self.checkbox_header_users.setCheckState(Qt.Checked)
        elif any_checked:
            self.checkbox_header_users.setCheckState(Qt.PartiallyChecked)
        else:
            self.checkbox_header_users.setCheckState(Qt.Unchecked)

        self.checkbox_header_users.blockSignals(False)

    def atualizar_posicao_checkbox_header_users(self):
        if hasattr(self, "checkbox_header_users") and self.coluna_checkboxes_adicionada:
            header = self.table_widget.horizontalHeader()

            x = header.sectionViewportPosition(0) + (header.sectionSize(0) - self.checkbox_header_users.width()) // 2 + 4
            y = (header.height() - self.checkbox_header_users.height()) // 2
            self.checkbox_header_users.move(x, y)

#*******************************************************************************************************
    def ordenar_usuario(self):
        if hasattr(self, "checkbox_header_users"):
            QMessageBox.warning(
                None,
                "Aviso",
                "Desmarque o checkbox antes de ordenar o histórico."
            )
            return
        # Implementação básica para ordenar produtos
        self.table_widget.sortItems(1, Qt.AscendingOrder)  # Ordenar pela coluna 1 em ordem ascendente
#*******************************************************************************************************
    def visualizar_imagem_usuario(self):
        # Se a coluna de checkboxes está ativa, usar os checkboxes para saber os selecionados
        if self.coluna_checkboxes_adicionada:
            selecionados = [i for i, cb in enumerate(self.checkboxes) if cb.isChecked()]

            if not selecionados:
                QMessageBox.warning(None, "Aviso", "Selecione um usuário para visualizar a imagem.")
                return

            if len(selecionados) > 1:
                QMessageBox.warning(None, "Aviso", "Só é possível visualizar a imagem de um usuário por vez.")
                return

            row_index = selecionados[0]
        else:
            # Se a seleção for pela linha da tabela
            row_index = self.table_widget.currentRow()
            if row_index < 0:
                QMessageBox.warning(None, "Aviso", "Selecione um usuário para visualizar a imagem.")
                return

        # Pega o ID da coluna correta (0 se não tiver coluna de checkbox, 1 se tiver)
        coluna_id = 1 if self.coluna_checkboxes_adicionada else 0
        item = self.table_widget.item(row_index, coluna_id)
        
        if item is None:
            QMessageBox.warning(None, "Erro", "Não foi possível identificar o ID do usuário.")
            return

        try:
            id_usuario = int(item.text())
        except ValueError:
            QMessageBox.warning(None, "Erro", "ID de usuário inválido.")
            return

        imagem_data = self.recuperar_imagem_do_banco_usuarios(id_usuario)

        if imagem_data:
            try:
                print("Dados da imagem recuperados com sucesso para visualização.")

                pixmap = QPixmap()
                pixmap.loadFromData(imagem_data)

                if pixmap.isNull():
                    print("Aviso: pixmap é nulo")
                    QMessageBox.warning(None, "Aviso", "Não foi possível carregar a imagem.")
                    return

                # Tentar abrir o arquivo com um visualizador de imagens padrão
                os.startfile("imagem_temporaria.png")

            except Exception as e:
                print(f"Erro ao processar imagem: {str(e)}")
        else:
            QMessageBox.warning(None, "Aviso", "Imagem não encontrada.")

    def limpar_campos_de_texto(self):
        self.main_window.txt_usuario_cadastro.clear()

    def atualizar_tabela_usuario(self):
        try:
            # Limpar a tabela antes de atualizar
            self.table_widget.setRowCount(0)
            self.table_widget.setColumnCount(24)  # Definir o número de colunas
            self.table_widget.setHorizontalHeaderLabels([
                "ID", "Nome", "Usuário", "Senha", "Confirmar Senha", "CEP", "Endereço",
                "Número", "Cidade", "Bairro", "Estado", "Complemento", "Telefone", "Email",
                "Data de Nascimento", "RG", "CPF", "CNPJ",
                "Última Troca de Senha", "Data da Senha Cadastrada",
                "Data da Inclusão do Usuário", "Segredo", "Usuário Logado", "Acesso"
            ])
            

            conexao = self.db.connecta()
            cursor = conexao.cursor()
            
            cursor.execute("""
                SELECT ID,Nome, Usuário, Senha, "Confirmar Senha", CEP, Endereço,Número,Cidade,
                Bairro, Estado, Complemento, Telefone, Email, "Data de Nascimento", RG, CPF,
                CNPJ, "Última Troca de Senha", "Data da Senha Cadastrada",
                "Data da Inclusão do Usuário", Segredo, "Usuário Logado", Acesso
                FROM users
            """)
            usuarios = cursor.fetchall()

            if not usuarios:
                QMessageBox.information(None, "Aviso", "Nenhum usuário encontrado no banco de dados.")


            # Preencher a tabela com os dados atualizados
            for usuario in usuarios:
                row_position = self.table_widget.rowCount()
                self.table_widget.insertRow(row_position)
                for col, dado in enumerate(usuario):
                    item = self.formatar_texto(str(dado))
                    self.table_widget.setItem(row_position, col, item)         

            # Ajustar o tamanho das colunas
            self.table_widget.resizeColumnsToContents()
            self.table_widget.resizeRowsToContents()

            QMessageBox.information(None,"Sucesso","Dados carregados com sucesso! ")

        except Exception as e:
            print(f"[Erro ao atualizar tabela]: {e}")
            QMessageBox.warning(None,"Aviso",f"Erro ao atualizar a tabela de usuários: {e}")

    def gerar_arquivo_excel_usuarios(self):
        # Obter o número de linhas e colunas na tabela
        num_linhas = self.table_widget.rowCount()
        num_colunas = self.table_widget.columnCount()

        # Verifica se a tabela está vazia ou não
        if self.table_widget.rowCount() == 0:
            QMessageBox.information(self,"Aviso","Nenhum usuário cadastrado para gerar o arquivo")
            return # Se a tabela estiver vazia, encerra a função sem prosseguir
        
         # Extrai os dados da tabela para uma lista de listas
        dados_usuarios = []
        for linha in range(num_linhas):
            linhas_dados_usuarios = []
            for coluna in range(num_colunas):
                item = self.table_widget.item(linha,coluna)
                linhas_dados_usuarios.append(item.text() if item else "")
            dados_usuarios.append(linhas_dados_usuarios)

        # Obtém os cabeçalhos da tabela
        cabecalhos = []
        for coluna in range(num_colunas):
            cabecalho_usuario = self.table_widget.horizontalHeaderItem(coluna)
            cabecalhos.append(cabecalho_usuario.text() if cabecalho_usuario else "")

        # Nome da aba dentro do excel
        nome_pagina = "Tabela de Usuários"
        
        # Cria um DataFrame do pandas com os dados da tabela
        df = pd.DataFrame(dados_usuarios, columns=cabecalhos)

        # Abre um diálogo para salvar o arquivo
        caminho_arquivo, _ = QFileDialog.getSaveFileName(self,"Salvar Arquivo Excel","Tabela de Usuários","Arquivo Excel(*.xlsx)")

        if caminho_arquivo:
            # Salvao o DataFrame como um arquivo Excel
            df.to_excel(caminho_arquivo,index=False,engine="openpyxl",sheet_name=nome_pagina)

        if not caminho_arquivo:
            print("Operação cancelada pelo usuário")
            return

        # Abrir a planilha para formatação
        openn = load_workbook(caminho_arquivo)
        sheet = openn[nome_pagina]
  
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                #Centraliza o texto
                cell.alignment = Alignment(horizontal='center',vertical='center')
                try:
                    if cell.value:
                        max_length = max(max_length,len(str(cell.value)))

                except:
                    pass
                #Ajustar largura da coluna
                sheet.column_dimensions[col_letter].width = max_length + 2
        # Salvar as alterações
        openn.save(caminho_arquivo)
        QMessageBox.information(self,"Aviso","Arquivo excel gerado com sucesso")


    # Função auxiliar para criar um QTableWidgetItem com texto centralizado e branco
    def formatar_texto(self, text):
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)

        # Define a cor com base no tema atual
        if self.tema == "claro":
            cor = QColor("black")
        else:  # Para "escuro" e "clássico"
            cor = QColor("white")

        item.setForeground(QBrush(cor))
        return item         
    
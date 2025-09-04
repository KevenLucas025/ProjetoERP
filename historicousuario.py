from PySide6.QtWidgets import (QWidget,QTableWidget,QMessageBox,
                               QTableWidgetItem,QLineEdit,QCheckBox,
                               QFileDialog,QMainWindow,QVBoxLayout,QPushButton,QHBoxLayout,
                               QLabel,QRadioButton,QGroupBox,QDialog,QHeaderView)
from PySide6.QtGui import QBrush,QColor,QGuiApplication
from PySide6.QtCore import Qt,QEvent,QTimer
from database import DataBase
import sqlite3
import pandas as pd
import csv
from datetime import datetime
from configuracoes import Configuracoes_Login
from dialogos import ComboDialog
from reportlab.lib.pagesizes import letter,landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import json


class Pagina_Usuarios(QWidget):
    def __init__(self,main_window,btn_abrir_planilha_usuarios,btn_cadastrar_novo_usuario,
                  btn_historico_usuarios,btn_atualizar_ativos,btn_atualizar_inativos,btn_limpar_tabelas_usuarios,
                  btn_gerar_saida_usuarios,line_excel_usuarios,progress_excel_usuarios,
                  btn_importar_usuarios,btn_abrir_planilha_massa_usuarios,btn_fazer_cadastro_massa_usuarios,
                  progress_massa_usuarios,line_edit_massa_usuarios,parent=None):
        super().__init__(parent)

        self.db = DataBase("banco_de_dados.db")

        self.config = Configuracoes_Login(self)


        self.checkboxes = []  # Lista para armazenar os checkboxes
        self.coluna_checkboxes_usuarios_adicionada = False
        self.checkbox_header_usuarios = None

      
        

        self.main_window = main_window
        self.table_ativos = self.main_window.table_ativos  # Referência para a tabela no main window
        self.table_inativos = self.main_window.table_inativos
        self.table_massa_usuarios = self.main_window.table_massa_usuarios
        self.btn_abrir_planilha_usuarios = btn_abrir_planilha_usuarios
        self.btn_cadastrar_novo_usuario = btn_cadastrar_novo_usuario
        self.btn_historico_usuarios = btn_historico_usuarios
        self.btn_atualizar_ativos = btn_atualizar_ativos
        self.btn_atualizar_inativos = btn_atualizar_inativos
        self.btn_limpar_tabelas_usuarios = btn_limpar_tabelas_usuarios
        self.btn_gerar_saida_usuarios = btn_gerar_saida_usuarios
        self.line_excel_usuarios = line_excel_usuarios
        self.progress_excel_usuarios = progress_excel_usuarios
        self.btn_importar_usuarios = btn_importar_usuarios
        self.btn_abrir_planilha_massa_usuarios = btn_abrir_planilha_massa_usuarios
        self.btn_fazer_cadastro_massa_usuarios = btn_fazer_cadastro_massa_usuarios
        self.progress_massa_usuarios = progress_massa_usuarios
        self.line_edit_massa_usuarios = line_edit_massa_usuarios
        
        

        self.btn_gerar_saida_usuarios.clicked.connect(self.confirmar_saida_usuarios)
        self.btn_limpar_tabelas_usuarios.clicked.connect(self.limpar_tabelas)
        self.btn_atualizar_ativos.clicked.connect(self.atualizar_ativos)
        self.btn_atualizar_inativos.clicked.connect(self.atualizar_inativos)
        self.btn_historico_usuarios.clicked.connect(self.exibir_tabela_historico_usuario)
        self.btn_abrir_planilha_usuarios.clicked.connect(self.abrir_planilha_usuarios)
        self.btn_importar_usuarios.clicked.connect(self.importar_usuario)
        self.btn_abrir_planilha_massa_usuarios.clicked.connect(self.abrir_panilha_usuarios_em_massa)
        self.btn_fazer_cadastro_massa_usuarios.clicked.connect(self.cadastrar_usuarios_em_massa)
        self.main_window.table_ativos.viewport().installEventFilter(self)
        self.main_window.table_inativos.viewport().installEventFilter(self)


    def aplicar_tema(self, tema: str) -> str:
        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    border-radius: 8px;
                    background: qlineargradient(
                        x1:0, y1:0, x2:0, y2:1,
                        stop:0 rgb(60, 60, 60),   /* topo */
                        stop:1 rgb(100, 100, 100) /* base */
                    );
                    font-size: 12px;
                    padding: 3px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
                
            """
            
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            table_view_style = """
            /* QTableView com seleção diferenciada */
            QTableView {
                background-color: #202124;
                color: white;
                gridline-color: #555555;
                selection-background-color: #7a7a7a;
                selection-color: white;
            }
            /* Coluna dos cabeçalhos */
            QHeaderView::section {
                background-color: #ffffff;
                color: black;
                border: 1px solid #aaaaaa;
                padding: 1px;
            }

            /* QTabWidget headers brancos */
            QTabWidget::pane {
                border: 1px solid #444444;
                background-color: #202124;
            }
            /* Estiliza a barra de rolagem horizontal */
            QTableView QScrollBar:horizontal {
                border: none;
                background-color: #ffffff;
                height: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* Estiliza a barra de rolagem vertical */
            QTableView QScrollBar:vertical {
                border: none;
                background-color: #ffffff;  
                width: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* QTabWidget headers brancos */
            QTabWidget::pane {
                border: 1px solid #444444;
                background-color: #202124;
            }
            /* Estiliza a barra de rolagem horizontal */
            QTableView QScrollBar:horizontal {
                border: none;
                background-color: none;
                height: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* Estiliza a barra de rolagem vertical */
            QTableView QScrollBar:vertical {
                border: none;
                background-color: none; 
                width: 12px;
                margin: 0px;
                border-radius: 5px;
            }
            

            /* Parte que você arrasta */
            QTableView QScrollBar::handle:vertical {
                background-color: #777777;  /* cinza médio */
                min-height: 22px;
                border-radius: 5px;
            }

            QTableView QScrollBar::handle:horizontal {
                background-color: #777777;
                min-width: 22px;
                border-radius: 5px;
            }

            /* Groove horizontal */
            QTableView QScrollBar::groove:horizontal {
                background-color: transparent;
                border-radius: 5px;
                height: 15px;
                margin: 0px 10px 0px 10px;
            }

            /* Groove vertical */
            QTableView QScrollBar::groove:vertical {
                background-color: transparent;
                border-radius: 5px;
                width: 25px;
                margin: 10px 0px 10px 10px;
            }

            /* Estilo para item selecionado */
            QTableWidget::item:selected {
                background-color: #555555;  /* cinza de seleção */
                color: white;
            }
            /* CornerButton (canto superior esquerdo) */
            QTableCornerButton::section {
                background-color: #ffffff;
                border: 1px solid #aaaaaa;
                padding: 2px;
            }
            /* Forçar cor do texto do QCheckBox */
            QCheckBox {
                color: white;
            }

            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """
        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """
        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(0,120,180),
                                                stop:1 rgb(0,150,220));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(0,100,160);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #007acc;
                }
                QPushButton:pressed {
                    background-color: #006bb3;
                    border: 2px solid #005c99;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 3px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            scroll_style = """
                QScrollBar:vertical {
                border: none;
                background-color: rgb(255, 255, 255); /* branco */
                width: 30px;
                margin: 0px 10px 0px 10px;
            }
                QScrollBar::handle:vertical {
                background-color: rgb(180, 180,180);  /* cinza */
                min-height: 30px;
                border-radius: 5px;
            }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
        """
        estilo_completo = f"""
        QMainWindow {{
            background-color: {bg_cor};
        }}
        {button_style}
        {lineedit_style}
        {combobox_style}
        {table_view_style}
        {scroll_style}
        """
        return estilo_completo


    def carregar_config(self):
            with open("config.json", "r", encoding="utf-8") as f:
                return json.load(f)
    

    # Função auxiliar para criar um QTableWidgetItem com texto centralizado
    def formatar_texto(self, text):
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)  # Centraliza o texto
        item.setForeground(QBrush(QColor("white"))) 
        return item

    def tabela_ativos(self):
        cn = sqlite3.connect("banco_de_dados.db")

        # Carregar dados da tabela users usando pandas
        query = """
        SELECT Nome,Usuário,Senha,"Confirmar Senha",Acesso,Endereço,CEP,CPF,Número,Estado,Email,RG,Complemento,
        Telefone,"Data de Nascimento","Última Troca de Senha","Data da Senha Cadastrada","Data da Inclusão do Usuário",Segredo
        FROM users
        """
        df = pd.read_sql_query(query, cn)

        # Verificando se há dados na consulta
        if df.empty:
            print("Nenhum dado encontrado no banco de dados 'users'")
            return
        
        # Configurando cabeçalhos das colunas
        numero_colunas = len(df.columns)
        self.table_ativos.setRowCount(len(df))
        self.table_ativos.setColumnCount(numero_colunas)

    
       # Limpando o QTableWidget antes de popular com novos dados
        self.table_ativos.clearContents()
        self.table_ativos.setRowCount(0)  # Certificando  que as linhas estão limpas 

        for row_index, row_data in df.iterrows():
            self.table_ativos.insertRow(row_index)
            for col_index, data in enumerate(row_data):
                item = self.formatar_texto(str(data))
                self.table_ativos.setItem(row_index,col_index,item)

    
    def gerar_saida_usuarios(self, usuarios_selecionados):
        saida_usuarios = []
        historico_logs = []
        data_atual = datetime.now().strftime("%d/%m/%Y %H:%M")
        linhas_para_remover = []

        total_usuarios_ativos = self.main_window.table_ativos.rowCount()
        quantidade_selecionada = len(usuarios_selecionados)
        
        if total_usuarios_ativos - quantidade_selecionada <= 1:
            QMessageBox.critical(None,"Erro","Não é possível gerar saída. O sistema deve ter pelo menos um usuário ativo")
            return


        with self.db.connecta() as conexao:
            cursor = conexao.cursor()

            for row in usuarios_selecionados:
                usuario_nome = self.main_window.table_ativos.item(row, 0).text() or ""
                usuario_usuario = self.main_window.table_ativos.item(row, 1).text() or ""
                usuario_senha = self.main_window.table_ativos.item(row, 2).text() or ""
                usuario_confirmar_senha = self.main_window.table_ativos.item(row, 3).text() or ""
                usuario_cep = self.main_window.table_ativos.item(row, 4).text() or ""
                usuario_endereco = self.main_window.table_ativos.item(row, 5).text() or ""
                usuario_numero = self.main_window.table_ativos.item(row, 6).text() or ""
                usuario_cidade = self.main_window.table_ativos.item(row, 7).text() or ""
                usuario_bairro = self.main_window.table_ativos.item(row, 8).text() or ""
                usuario_estado = self.main_window.table_ativos.item(row, 9).text() or ""
                usuario_complemento = self.main_window.table_ativos.item(row, 10).text() or ""
                usuario_telefone = self.main_window.table_ativos.item(row, 11).text() or ""
                usuario_email = self.main_window.table_ativos.item(row, 12).text() or ""
                usuario_data_nascimento = self.main_window.table_ativos.item(row, 13).text() or ""
                usuario_rg = self.main_window.table_ativos.item(row, 14).text() or ""
                usuario_cpf = self.main_window.table_ativos.item(row, 15).text() or ""
                usuario_cnpj = self.main_window.table_ativos.item(row, 16).text() or ""
                usuario_ultima_troca_senha = self.main_window.table_ativos.item(row, 17).text() or ""
                usuario_data_senha_cadastrada = self.main_window.table_ativos.item(row, 18).text() or ""
                usuario_data_inclusao_usuario = self.main_window.table_ativos.item(row, 19).text() or ""
                usuario_segredo = self.main_window.table_ativos.item(row, 20).text() or ""
                usuario_logado = self.main_window.table_ativos.item(row, 21).text() or ""
                usuario_acesso = self.main_window.table_ativos.item(row, 22).text() or ""

                

                # Recupera a imagem somente para o banco de dados
                imagem_usuario = self.recuperar_imagem_usuario_bd_users(usuario_usuario)

                # Dados completos com imagem - para o banco
                dados_para_banco = (
                    usuario_nome, usuario_usuario, usuario_senha, usuario_confirmar_senha,
                    usuario_cep, usuario_endereco, usuario_numero, usuario_cidade, usuario_bairro,
                    usuario_estado, usuario_complemento, usuario_telefone, usuario_email,
                    usuario_data_nascimento, usuario_rg, usuario_cpf, usuario_cnpj, imagem_usuario,
                    usuario_ultima_troca_senha, usuario_data_senha_cadastrada, usuario_data_inclusao_usuario,
                    data_atual,  # Data da Inatividade
                    usuario_segredo, usuario_logado, usuario_acesso
                )

                # Dados sem imagem - para exibir na tabela
                dados_para_interface = (
                    usuario_nome, usuario_usuario, usuario_senha, usuario_confirmar_senha,
                    usuario_cep, usuario_endereco, usuario_numero, usuario_cidade, usuario_bairro,
                    usuario_estado, usuario_complemento, usuario_telefone, usuario_email,
                    usuario_data_nascimento, usuario_rg, usuario_cpf, usuario_cnpj,
                    usuario_ultima_troca_senha, usuario_data_senha_cadastrada, usuario_data_inclusao_usuario,
                    data_atual, usuario_segredo, usuario_logado, usuario_acesso
                )

                # Adiciona na lista para exibir
                saida_usuarios.append(dados_para_interface)

                # Adiciona ao banco de dados
                cursor.execute("""
                    INSERT OR REPLACE INTO users_inativos (
                        Nome, Usuário, Senha, "Confirmar Senha", CEP, Endereço, Número, Cidade, Bairro, Estado,
                        Complemento, Telefone, Email, "Data de Nascimento", RG, CPF, CNPJ, Imagem,
                        "Última Troca de Senha", "Data da Senha Cadastrada", "Data da Inclusão do Usuário",
                        "Data da Inatividade do Usuário", Segredo, "Usuário Logado", Acesso
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, dados_para_banco)

                historico_logs.append(f"Usuário '{usuario_nome}' ({usuario_usuario}) foi transferido para inativos.")

                cursor.execute("""DELETE FROM users WHERE "Usuário" = ?""",(usuario_usuario,))
                linhas_para_remover.append(row)

            conexao.commit()
            self.tabela_inativos_preencher(saida_usuarios,limpar=False)

            # Remover da tabela de ativos os usuários transferidos
            for row in sorted(linhas_para_remover):
                self.main_window.table_ativos.removeRow(row)

        for texto in historico_logs:
            self.main_window.registrar_historico_usuarios("Gerado Saída de Usuário", texto)

        if saida_usuarios:
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Information)
            msg_box.setWindowTitle("Aviso")
            msg_box.setText("Saída do(s) usuário(s) gerada com sucesso!")
            msg_box.exec()

   

    def tabela_inativos_preencher(self, dados, limpar=True):
        if limpar:
            self.main_window.table_inativos.setRowCount(0)
            linha_inicial = 0
        else:
            linha_inicial = self.main_window.table_inativos.rowCount()

        for linha in dados:
            self.main_window.table_inativos.insertRow(linha_inicial)
            for coluna, valor in enumerate(linha):
                self.main_window.table_inativos.setItem(linha_inicial, coluna, self.formatar_texto(str(valor)))
            linha_inicial += 1

        self.main_window.table_inativos.resizeColumnsToContents()
        self.main_window.table_inativos.resizeRowsToContents()

        
     # Função para recuperar imagem de um produto com base no código do produto
    def recuperar_imagem_usuario_bd_users(self, id_usuario):
        conexao = sqlite3.connect('banco_de_dados.db')
        cursor = conexao.cursor()
        cursor.execute("SELECT Imagem FROM users WHERE Usuário = ?", (id_usuario,))
        
        resultado = cursor.fetchone()  # Tenta buscar uma linha
        
        if resultado is not None:
            imagem_blob = resultado[0]  # Recupera a imagem se o resultado não for None
        else:
            imagem_blob = None  # Define como None se a imagem não for encontrada
        
        conexao.close()
        return imagem_blob


    def confirmar_saida_usuarios(self):
        configuracoes = Configuracoes_Login(self.main_window)
        if configuracoes.nao_mostrar_aviso_irreversivel:
            pass
        else:
            # Cria a mensagem de aviso com o checkbox
            msg_aviso_irreversivel = QMessageBox(self)
            msg_aviso_irreversivel.setIcon(QMessageBox.Warning)  # Ícone de aviso
            msg_aviso_irreversivel.setWindowTitle("Aviso de Saída Irreversível")  # Título da janela
            msg_aviso_irreversivel.setText("A função de saída de usuário é irreversível.\n\nUsuários removidos não poderão ser restaurados.")  # Texto do aviso
            msg_aviso_irreversivel.setStandardButtons(QMessageBox.Ok)  # Botão Ok
            msg_aviso_irreversivel.setStyleSheet("""
                QMessageBox {
                    background-color: qlineargradient(
                        x1:0, y1:0, x2:0, y2:1,
                        stop:0 #FFD54F,   /* amarelo */
                        stop:1 #8D6E63    /* marrom */
                    );
                    color: black;  /* cor do texto */
                    font: bold 11pt "Segoe UI";
                }

                QMessageBox QLabel {
                    background: transparent;                             
                    color: black;
                }

                QMessageBox QPushButton {
                    background-color: #6D4C41;
                    color: white;
                    border-radius: 6px;
                    padding: 4px 12px;
                }

                QMessageBox QPushButton:hover {
                    background-color: #8D6E63;
                }

                QMessageBox QPushButton:pressed {
                    background-color: #5D4037;
                }
                QMessageBox QCheckBox{
                    background: transparent;
                    color: black;
                }
            """)


            # Adiciona o checkbox à caixa de mensagem
            checkbox = QCheckBox("Não mostrar esta mensagem novamente")
            msg_aviso_irreversivel.setCheckBox(checkbox)


            # Exibe a mensagem de aviso
            msg_aviso_irreversivel.exec()

            # Verifica se o usuário marcou a opção para não mostrar novamente
            if checkbox.isChecked():
                configuracoes.nao_mostrar_aviso_irreversivel = True  # Define que o usuário não quer mais ver este aviso
                configuracoes.salvar(configuracoes.usuario, configuracoes.senha,configuracoes.mantem_conectado)

        # Verifica a linha diretamente selecionada pelo usuário
        selecionar_linhas = self.main_window.table_ativos.selectionModel().selectedRows()
        usuarios_selecionados = [row.row() for row in selecionar_linhas]

        # Verifica se existe uma linha selecionada diretamente pelo clique
        if usuarios_selecionados:
            mensagem = "Tem certeza de que deseja gerar a saída do usuário selecionado?"

            caixa_dialogo = QMessageBox()
            caixa_dialogo.setWindowTitle("Confirmar Saída")
            caixa_dialogo.setText(mensagem)
            caixa_dialogo.setIcon(QMessageBox.Question)
            caixa_dialogo.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            caixa_dialogo.setDefaultButton(QMessageBox.Yes)

            # Personalizando o texto dos botões
            botao_nao = caixa_dialogo.button(QMessageBox.No)
            botao_nao.setText("Não")
            botao_sim = caixa_dialogo.button(QMessageBox.Yes)
            botao_sim.setText("Sim")

            resposta = caixa_dialogo.exec()
            if resposta == QMessageBox.Yes:
                self.gerar_saida_usuarios(usuarios_selecionados)

        else:
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Information)
            msg_box.setWindowTitle("Aviso")
            msg_box.setText("Nenhum usuário selecionado para gerar saída")
            msg_box.exec()


    def atualizar_ativos(self):
        try:
            # Limpar a tabela antes de atualizar
            self.table_ativos.setRowCount(0)  # Remove todas as linhas da tabela

            # Consultar todos os usuarios
            query = """
            SELECT Nome,Usuário,Senha,"Confirmar Senha",CEP,Endereço,Número,Cidade,Bairro,Estado,Complemento,Telefone,Email,"Data de Nascimento",
            RG,CPF,CNPJ,"Última Troca de Senha","Data da Senha Cadastrada","Data da Inclusão do Usuário",Segredo,"Usuário Logado",Acesso
            FROM users
            """
            usuarios = self.db.executar_query(query)

            for linha_index, linha_data in enumerate(usuarios):
                self.table_ativos.insertRow(linha_index)
                for col, value in enumerate(linha_data):
                    item = self.formatar_texto(str(value))
                    self.table_ativos.setItem(linha_index, col, item)
                    
            self.table_ativos.resizeColumnsToContents()  # Ajusta as colunas automaticamente
            self.table_ativos.resizeRowsToContents()  # Ajusta as linhas automaticamente

            # Exibir uma mensagem de sucesso
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Information)
            msg_box.setWindowTitle("Informação")
            msg_box.setText("Tabela ativos atualizada com sucesso!")       
            msg_box.exec()

        except Exception as e:
            print(f"Erro ao atualizar a tabela de ativos: {e}")
            
    def atualizar_inativos(self):
        try:
            # Limpa a tabela antes de carregar os novos dados
            self.table_inativos.setRowCount(0)
            
            # Consulta os dados da tabela de saída no banco de dados (somente usuarios que já tiveram saída gerada)
            query = """
            SELECT Nome,Usuário,Senha,"Confirmar Senha",CEP,Endereço,Número,Cidade,Bairro,Estado,Complemento,Telefone,Email,
            "Data de Nascimento",RG,CPF,CNPJ,"Última Troca de Senha","Data da Senha Cadastrada","Data da Inclusão do Usuário",
            "Data da Inatividade do Usuário",Segredo,"Usuário Logado",Acesso
            FROM users_inativos
            """
            saidas = self.db.executar_query(query)  # Método que executa a consulta e retorna os resultados

            # Preenche a tabela com os dados obtidos
            if saidas:
                for saida in saidas:
                    row_position = self.table_inativos.rowCount()
                    self.table_inativos.insertRow(row_position)
                    for column, value in enumerate(saida):
                        item = self.formatar_texto(str(value))
                        self.table_inativos.setItem(row_position, column, item)
                        
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Information)
            msg_box.setWindowTitle("Informação")
            msg_box.setText("Tabela inativos atualizada com sucesso!")       
            msg_box.exec()

        except Exception as e:
            print(f"Erro ao atualizar a tabela de inativos: {e}")


    # Função auxiliar para criar um QTableWidgetItem com texto centralizado e branco
    def formatar_texto(self, text):
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)
        item.setForeground(QBrush(QColor("white")))
        return item



    def limpar_tabelas(self):
        # Verifica se as tabelas estão vazias
        if self.table_inativos.rowCount() == 0 and self.table_ativos.rowCount() == 0:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("Aviso")
            msg.setText("As tabelas já estão vazias.")
            msg.exec()
            return

        # Limpa a tabela de ativos
        self.table_ativos.setRowCount(0)

        # Limpa a tabela de inativos
        self.table_inativos.setRowCount(0)

        # Mensagem de confirmação após a limpeza
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle("Limpeza Concluída")
        msg.setText("As tabelas foram limpas com sucesso.")
        msg.exec()


    # Limpa a coluna selecionada clicando em qualquer lugar da tabela
    def eventFilter(self, source, event):
        if event.type() == QEvent.MouseButtonPress:
            if source == self.main_window.table_ativos.viewport():
                index = self.main_window.table_ativos.indexAt(event.pos())
                if not index.isValid():
                    self.main_window.table_ativos.clearSelection()
                    self.main_window.table_ativos.clearFocus()

            elif source == self.main_window.table_inativos.viewport():
                index = self.main_window.table_inativos.indexAt(event.pos())
                if not index.isValid():
                    self.main_window.table_inativos.clearSelection()
                    self.main_window.table_inativos.clearFocus()

        return super().eventFilter(source, event)

    def exibir_tabela_historico_usuario(self):
        if self.coluna_checkboxes_usuarios_adicionada:
            self.remover_coluna_checkboxes()
        self.janela_historico = QMainWindow()
        self.janela_historico.setWindowTitle("Histórico de Ações")
        self.janela_historico.resize(800,650)

        screen = QGuiApplication.primaryScreen()
        screen_geometry = screen.availableGeometry()
        window_geometry = self.janela_historico.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.janela_historico.move(window_geometry.topLeft())
        
        central_widget = QWidget()
        layout = QVBoxLayout(central_widget)

        # Carregar tema
        config = self.carregar_config()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    border-radius: 8px;
                    background: qlineargradient(
                        x1:0, y1:0, x2:0, y2:1,
                        stop:0 rgb(60, 60, 60),   /* topo */
                        stop:1 rgb(100, 100, 100) /* base */
                    );
                    font-size: 12px;
                    padding: 3px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
                
            """
            
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            table_view_style = """
            /* QTableView com seleção diferenciada */
            QTableView {
                background-color: #202124;
                color: white;
                gridline-color: #555555;
                selection-background-color: #7a7a7a;
                selection-color: white;
            }
            /* Coluna dos cabeçalhos */
            QHeaderView::section {
                background-color: #ffffff;
                color: black;
                border: 1px solid #aaaaaa;
                padding: 1px;
            }

            /* QTabWidget headers brancos */
            QTabWidget::pane {
                border: 1px solid #444444;
                background-color: #202124;
            }
            /* Estiliza a barra de rolagem horizontal */
            QTableView QScrollBar:horizontal {
                border: none;
                background-color: #ffffff;
                height: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* Estiliza a barra de rolagem vertical */
            QTableView QScrollBar:vertical {
                border: none;
                background-color: #ffffff;  
                width: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* QTabWidget headers brancos */
            QTabWidget::pane {
                border: 1px solid #444444;
                background-color: #202124;
            }
            /* Estiliza a barra de rolagem horizontal */
            QTableView QScrollBar:horizontal {
                border: none;
                background-color: none;
                height: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* Estiliza a barra de rolagem vertical */
            QTableView QScrollBar:vertical {
                border: none;
                background-color: none; 
                width: 12px;
                margin: 0px;
                border-radius: 5px;
            }
            

            /* Parte que você arrasta */
            QTableView QScrollBar::handle:vertical {
                background-color: #777777;  /* cinza médio */
                min-height: 22px;
                border-radius: 5px;
            }

            QTableView QScrollBar::handle:horizontal {
                background-color: #777777;
                min-width: 22px;
                border-radius: 5px;
            }

            /* Groove horizontal */
            QTableView QScrollBar::groove:horizontal {
                background-color: transparent;
                border-radius: 5px;
                height: 15px;
                margin: 0px 10px 0px 10px;
            }

            /* Groove vertical */
            QTableView QScrollBar::groove:vertical {
                background-color: transparent;
                border-radius: 5px;
                width: 25px;
                margin: 10px 0px 10px 10px;
            }

            /* Estilo para item selecionado */
            QTableWidget::item:selected {
                background-color: #555555;  /* cinza de seleção */
                color: white;
            }
            /* CornerButton (canto superior esquerdo) */
            QTableCornerButton::section {
                background-color: #ffffff;
                border: 1px solid #aaaaaa;
                padding: 2px;
            }
            /* Forçar cor do texto do QCheckBox */
            QCheckBox {
                color: white;
            }

            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """
        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """
        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(0,120,180),
                                                stop:1 rgb(0,150,220));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(0,100,160);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #007acc;
                }
                QPushButton:pressed {
                    background-color: #006bb3;
                    border: 2px solid #005c99;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 3px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            scroll_style = """
                QScrollBar:vertical {
                border: none;
                background-color: rgb(255, 255, 255); /* branco */
                width: 30px;
                margin: 0px 10px 0px 10px;
            }
                QScrollBar::handle:vertical {
                background-color: rgb(180, 180,180);  /* cinza */
                min-height: 30px;
                border-radius: 5px;
            }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
        """

        # Tabela do histórico
        self.tabela_historico_usuarios = QTableWidget()
        self.tabela_historico_usuarios.setColumnCount(4)
        self.tabela_historico_usuarios.setHorizontalHeaderLabels(["Data/Hora", "Usuário", "Ação", "Descrição"])

        #Botão Atualizar
        botao_atualizar = QPushButton("Atualizar Histórico")
        botao_atualizar.clicked.connect(self.atualizar_historico_usuario)
        botao_atualizar.setStyleSheet(button_style)

        # Botão Apagar
        botao_apagar = QPushButton("Apagar Histórico")
        botao_apagar.clicked.connect(self.apagar_historico_usuario)
        botao_apagar.setStyleSheet(button_style)

        botao_exportar_csv = QPushButton("Exportar para CSV")
        botao_exportar_csv.clicked.connect(self.exportar_csv_usuarios)
        botao_exportar_csv.setStyleSheet(button_style)

        botao_exportar_excel = QPushButton("Exportar para Excel")
        botao_exportar_excel.clicked.connect(self.exportar_excel_usuarios)
        botao_exportar_excel.setStyleSheet(button_style)

        botao_exportar_pdf = QPushButton("Exportar PDF")
        botao_exportar_pdf.clicked.connect(self.exportar_pdf_usuarios)
        botao_exportar_pdf.setStyleSheet(button_style)

        botao_pausar_historico = QPushButton("Pausar Histórico")
        botao_pausar_historico.clicked.connect(self.pausar_historico_usuario)
        botao_pausar_historico.setStyleSheet(button_style)

        botao_filtrar_historico = QPushButton("Filtrar Histórico")
        botao_filtrar_historico.clicked.connect(self.filtrar_historico_usuarios)
        botao_filtrar_historico.setStyleSheet(button_style)

        botao_ordenar_historico = QPushButton("Ordenar Histórico")
        botao_ordenar_historico.clicked.connect(self.ordenar_historico_usuario)
        botao_ordenar_historico.setStyleSheet(button_style)

        # Criar checkbox "Selecionar Individualmente" toda vez que a janela for aberta
        self.checkbox_selecionar = QCheckBox("Selecionar")
        self.checkbox_selecionar.stateChanged.connect(self.selecionar_usuarios_individual)
        self.checkbox_selecionar.setStyleSheet(f"color: {text_cor};")

        botoes = [
            botao_atualizar,
            botao_apagar,
            botao_exportar_csv,
            botao_exportar_excel,
            botao_exportar_pdf,
            botao_pausar_historico,
            botao_filtrar_historico,
            botao_ordenar_historico
        ]
        for btn in botoes:
            btn.setCursor(Qt.PointingHandCursor)

        # Adicionar outros botões ao layout
        layout.addWidget(botao_atualizar)
        layout.addWidget(botao_apagar)
        layout.addWidget(botao_exportar_csv)
        layout.addWidget(botao_exportar_excel)
        layout.addWidget(botao_exportar_pdf)
        layout.addWidget(botao_pausar_historico)
        layout.addWidget(botao_ordenar_historico)
        layout.addWidget(botao_filtrar_historico)
        layout.addWidget(self.checkbox_selecionar)
        layout.addWidget(self.tabela_historico_usuarios)

        # Configurar o widget central e exibir a janela
        self.janela_historico.setCentralWidget(central_widget)
        self.janela_historico.setStyleSheet(f"background-color: {bg_cor}; color:{text_cor}")
        self.janela_historico.show()

        # Preencher tabela pela primeira vez
        self.carregar_historico_usuario()
        self.tabela_historico_usuarios.resizeColumnsToContents()
        self.tabela_historico_usuarios.setStyleSheet(table_view_style)


    def carregar_historico_usuario(self):
        with sqlite3.connect('banco_de_dados.db') as cn:
            cursor = cn.cursor()
            cursor.execute('SELECT * FROM historico_usuarios ORDER BY "Data e Hora" DESC')
            registros = cursor.fetchall()

        self.tabela_historico_usuarios.clearContents()
        self.tabela_historico_usuarios.setRowCount(len(registros))
        
        deslocamento_usuarios = 1 if self.coluna_checkboxes_usuarios_adicionada else 0
        self.checkboxes = []  # Zerar e recriar lista de checkboxes
        

        for i, (data, usuario, acao, descricao) in enumerate(registros):
            if self.coluna_checkboxes_usuarios_adicionada:
                checkbox = QCheckBox()
                checkbox.setStyleSheet("margin-left: 9px; margin-right: 9px;")
                self.tabela_historico_usuarios.setCellWidget(i, 0, checkbox)
                self.checkboxes.append(checkbox)
            self.tabela_historico_usuarios.setItem(i, 0 + deslocamento_usuarios, QTableWidgetItem(data))
            self.tabela_historico_usuarios.setItem(i, 1 + deslocamento_usuarios, QTableWidgetItem(usuario))
            self.tabela_historico_usuarios.setItem(i, 2 + deslocamento_usuarios, QTableWidgetItem(acao))
            self.tabela_historico_usuarios.setItem(i, 3 + deslocamento_usuarios, QTableWidgetItem(descricao))



    def atualizar_historico_usuario(self):
        QMessageBox.information(self, "Sucesso", "Dados carregados com sucesso!")
        self.carregar_historico_usuario()

    def apagar_historico_usuario(self):
        # Caso checkboxes estejam ativados
        if self.coluna_checkboxes_usuarios_adicionada and self.checkboxes:
            linhas_para_remover = []
            ids_para_remover = []

            # Identificar as linhas com checkboxes selecionados
            for row, checkbox in enumerate(self.checkboxes):
                if checkbox and checkbox.isChecked():
                    linhas_para_remover.append(row)
                    coluna_data_hora = 1 if self.coluna_checkboxes_usuarios_adicionada else 2
                    item_data_widget = self.tabela_historico_usuarios.item(row, coluna_data_hora)  # Coluna de Data/Hora
                    if item_data_widget:
                        item_data_text = item_data_widget.text().strip()
                        # Excluir com base na data e hora
                        if item_data_text:
                            ids_para_remover.append(item_data_text)
                        else:
                            print(f"Erro ao capturar ID para a data/hora: '{item_data_text}'")
                    else:
                        print(f"Erro ao capturar Data/Hora na linha {row}")

            if not ids_para_remover:
                QMessageBox.warning(self, "Erro", "Nenhum item válido foi selecionado para apagar!")
                return

            # Confirmar exclusão
            mensagem = (
                f"Você tem certeza que deseja apagar os {len(ids_para_remover)} itens selecionados?"
                if len(ids_para_remover) > 1
                else "Você tem certeza que deseja apagar o item selecionado?"
            )

            if not self.confirmar_historico_usuarios_apagado(mensagem):
                return

            # Excluir do banco de dados
            with sqlite3.connect('banco_de_dados.db') as cn:
                cursor = cn.cursor()
                try:
                    for item_id in ids_para_remover:
                        cursor.execute("DELETE FROM historico_usuarios WHERE 'Data e Hora' = ?", (item_id,))
                    cn.commit()
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao excluir do banco de dados: {e}")
                    return

            # Remover as linhas na interface
            for row in sorted(linhas_para_remover, reverse=True):
                self.tabela_historico_usuarios.removeRow(row)

            QMessageBox.information(self, "Sucesso", "Itens removidos com sucesso!")

        # Caso sem checkboxes (seleção manual)
        else:
            linha_selecionada = self.tabela_historico_usuarios.currentRow()

            if linha_selecionada < 0:
                QMessageBox.warning(self, "Erro", "Nenhum item foi selecionado para apagar!")
                return

            # Capturar a Data/Hora da célula correspondente (coluna 0)
            coluna_data_hora = 0 if self.coluna_checkboxes_usuarios_adicionada else 1
            item_data_widget = self.tabela_historico_usuarios.item(linha_selecionada, coluna_data_hora)  # Coluna de Data/Hora
            if not item_data_widget:
                QMessageBox.warning(self, "Erro", "Não foi possível identificar a Data/Hora do item a ser apagado!")
                return

            item_data_text = item_data_widget.text().strip()

            # Conectar ao banco de dados para buscar o ID relacionado à Data/Hora
            with sqlite3.connect('banco_de_dados.db') as cn:
                cursor = cn.cursor()
                try:
                    # Buscar o ID com base na Data/Hora, removendo espaços ou caracteres extras
                    cursor.execute('SELECT id FROM historico_usuarios WHERE "Data e Hora" = ?', (item_data_text,))
                    resultado = cursor.fetchone()

                    if resultado:
                        item_id = resultado[0]  # Pegamos o ID encontrado
                    else:
                        QMessageBox.warning(self, "Erro", f"Não foi encontrado um item para a Data/Hora: {item_data_text}")
                        return

                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao buscar ID: {e}")
                    return

            # Confirmar exclusão
            mensagem = "Você tem certeza que deseja apagar o item selecionado?"

            if not self.confirmar_historico_usuarios_apagado(mensagem):
                return

            # Excluir do banco de dados
            with sqlite3.connect('banco_de_dados.db') as cn:
                cursor = cn.cursor()
                try:
                    cursor.execute("DELETE FROM historico_usuarios WHERE 'Data e Hora' = ?", (item_id,))
                    print(f"Item removido do banco de dados: ID {item_id}")
                    cn.commit()
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao excluir do banco de dados: {e}")
                    return

            # Remover a linha da interface
            self.tabela_historico_usuarios.removeRow(linha_selecionada)

            QMessageBox.information(self, "Sucesso", "Item removido com sucesso!")

     # Função para desmarcar todos os checkboxes
    def desmarcar_checkboxes(self):
        for checkbox in self.checkboxes:
            if checkbox:
                checkbox.setChecked(False)

    def confirmar_historico_usuarios_apagado(self, mensagem):
        """
        Exibe uma caixa de diálogo para confirmar a exclusão.
        """
        msgbox = QMessageBox(self)
        msgbox.setWindowTitle("Confirmação")
        msgbox.setText(mensagem)

        btn_sim = QPushButton("Sim")
        btn_nao = QPushButton("Não")
        msgbox.addButton(btn_sim, QMessageBox.ButtonRole.YesRole)
        msgbox.addButton(btn_nao, QMessageBox.ButtonRole.NoRole)

        msgbox.setDefaultButton(btn_nao)
        resposta = msgbox.exec()

        return msgbox.clickedButton() == btn_sim
                
    def selecionar_todos_usuarios(self):
        if not self.coluna_checkboxes_usuarios_adicionada:
            QMessageBox.warning(self, "Aviso", "Ative a opção 'Selecionar Individualmente' antes.")
            self.checkbox_header_usuarios.setChecked(False)
            return

        estado = self.checkbox_header_usuarios.checkState() == Qt.Checked
        self.checkboxes.clear()
        
        for row in range(self.tabela_historico_usuarios.rowCount()):
            widget = self.tabela_historico_usuarios.cellWidget(row,0)
            if widget is not None:
                checkbox = widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.blockSignals(True)
                    checkbox.setChecked(estado)
                    checkbox.blockSignals(False)
                
    # Função para adicionar checkboxes selecionar_individual na tabela de histórico
    def selecionar_usuarios_individual(self):
        if self.tabela_historico_usuarios.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico para selecionar.")
            if hasattr(self, "checkbox_selecionar") and isinstance(self.checkbox_selecionar, QCheckBox):
                QTimer.singleShot(0, lambda: self.checkbox_selecionar.setChecked(False))
            return

        # ---- Se a coluna já existe, remove ----
        if self.coluna_checkboxes_usuarios_adicionada:
            self.remover_coluna_checkboxes()
            return


        # ----- Caso contrário, adiciona a coluna -----
        self.tabela_historico_usuarios.insertColumn(0)
        self.tabela_historico_usuarios.setHorizontalHeaderItem(0, QTableWidgetItem(""))
        self.tabela_historico_usuarios.setColumnWidth(0, 30)
        self.tabela_historico_usuarios.horizontalHeader().setMinimumSectionSize(30)
        self.tabela_historico_usuarios.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)

        # Checkbox do cabeçalho
        header_usuarios_hist = self.tabela_historico_usuarios.horizontalHeader()
        self.checkbox_header_usuarios = QCheckBox(header_usuarios_hist.viewport())
        self.checkbox_header_usuarios.setToolTip("Selecionar todos")
        self.checkbox_header_usuarios.setChecked(False)
        self.checkbox_header_usuarios.setStyleSheet("QCheckBox{background: transparent;}")
        self.checkbox_header_usuarios.stateChanged.connect(self.selecionar_todos_usuarios)
        self.checkbox_header_usuarios.setFixedSize(20, 20)
        self.checkbox_header_usuarios.show()

        # Centralizar após layout
        QTimer.singleShot(0, self.atualizar_posicao_checkbox_header_usuarios)

        self.checkboxes.clear()
        for row in range(self.tabela_historico_usuarios.rowCount()):
            checkbox = QCheckBox()
            checkbox.stateChanged.connect(self.atualizar_selecao_todos_usuarios)

            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(checkbox)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)

            self.tabela_historico_usuarios.setCellWidget(row, 0, container)
            self.checkboxes.append(checkbox)

        self.tabela_historico_usuarios.verticalHeader().setVisible(False)
        header_usuarios_hist.sectionResized.connect(self.atualizar_posicao_checkbox_header_usuarios)
        self.tabela_historico_usuarios.horizontalScrollBar().valueChanged.connect(self.atualizar_posicao_checkbox_header_usuarios)
        header_usuarios_hist.geometriesChanged.connect(self.atualizar_posicao_checkbox_header_usuarios)

        self.coluna_checkboxes_usuarios_adicionada = True

    def remover_coluna_checkboxes(self):
        """Remove a coluna de checkboxes de forma segura"""
        if self.checkbox_header_usuarios is not None:
            try:
                # Desconecta sinais
                self.checkbox_header_usuarios.stateChanged.disconnect()
            except (RuntimeError, TypeError):
                pass

            try:
                # Remove do parent e agenda deleção
                self.checkbox_header_usuarios.setParent(None)
                self.checkbox_header_usuarios.deleteLater()
            except RuntimeError:
                pass

            self.checkbox_header_usuarios = None

        # Remove a coluna de fato
        try:
            self.tabela_historico_usuarios.removeColumn(0)
        except Exception:
            pass

        self.tabela_historico_usuarios.verticalHeader().setVisible(True)
        self.coluna_checkboxes_usuarios_adicionada = False
        self.checkboxes.clear()



    def closeEvent(self, event):
        if getattr(self, "coluna_checkboxes_usuarios_adicionada", False):
            self.remover_coluna_checkboxes()
        # Desmarcar o checkbox principal se existir
        if hasattr(self, "checkbox_selecionar") and isinstance(self.checkbox_selecionar, QCheckBox):
            self.checkbox_selecionar.setChecked(False)
        super().closeEvent(event)

    
    def atualizar_selecao_todos_usuarios(self):
        self.checkbox_header_usuarios.blockSignals(True)

        # Atualizar o estado do "Selecionar Todos"
        all_checked = all(checkbox.isChecked() for checkbox in self.checkboxes if checkbox)
        any_checked = any(checkbox.isChecked() for checkbox in self.checkboxes if checkbox)

        if all_checked:
            self.checkbox_header_usuarios.setCheckState(Qt.Checked)
        elif any_checked:
            self.checkbox_header_usuarios.setCheckState(Qt.PartiallyChecked)
        else:
            self.checkbox_header_usuarios.setCheckState(Qt.Unchecked)
        
        self.checkbox_header_usuarios.blockSignals(False)

        
    def atualizar_posicao_checkbox_header_usuarios(self):
        if (
            getattr(self, "checkbox_header_usuarios", None) is not None
            and self.coluna_checkboxes_usuarios_adicionada
        ):
            try:
                header = self.tabela_historico_usuarios.horizontalHeader()

                # largura da seção da coluna 0
                section_width = header.sectionSize(0)
                section_pos = header.sectionViewportPosition(0)

                # centralizar horizontalmente
                x = section_pos + (section_width - self.checkbox_header_usuarios.width()) // 2 + 4

                # centralizar verticalmente
                y = (header.height() - self.checkbox_header_usuarios.height()) // 2

                self.checkbox_header_usuarios.move(x, y)
            except RuntimeError:
                # objeto já foi deletado pelo Qt
                self.checkbox_header_usuarios = None


    def ordenar_historico_usuario(self):
        if getattr(self, "checkbox_selecionar", None) and self.checkbox_selecionar.isChecked():
            QMessageBox.warning(
                self,
                "Aviso",
                "Desmarque o checkbox antes de ordernar o histórico de usuários."
            )
            return
        # Obter a coluna pela qual o usuário deseja ordenar
        coluna = self.obter_coluna_usuario_para_ordenar()  # Função fictícia para capturar escolha
        if coluna is None:
            return  # Cancela o processo todo
        
        # Determinar a direção de ordenação (ascendente ou descendente)
        direcao = self.obter_direcao_ordenacao_usuario()  # Função fictícia para capturar escolha
        if direcao is None:
            return  # Cancela o processo todo
        
        # Mapeamento de colunas para índices (ajustar conforme sua tabela)
        colunas_para_indices = {
            "Data/Hora": 0,
            "Usuário": 1,
            "Ação": 2,
            "Descrição": 3
        }
        
        # Verificar se a coluna escolhida é válida
        if coluna not in colunas_para_indices:
            QMessageBox.warning(self, "Erro", "Coluna inválida para ordenação!")
            return
        
        # Obter o índice da coluna escolhida
        indice_coluna = colunas_para_indices[coluna]
        
        # Obter os dados atuais da tabela
        dados = []
        for row in range(self.tabela_historico_usuarios.rowCount()):
            linha = [
                self.tabela_historico_usuarios.item(row, col).text() if self.tabela_historico_usuarios.item(row, col) else ""
                for col in range(self.tabela_historico_usuarios.columnCount())
            ]
            dados.append(linha)
        
        # Ordenar os dados com base na coluna escolhida e direção
        dados.sort(key=lambda x: x[indice_coluna], reverse=(direcao == "Decrescente"))
        
        # Atualizar a tabela com os dados ordenados
        self.tabela_historico_usuarios.setRowCount(0)  # Limpar tabela
        for row_data in dados:
            row = self.tabela_historico_usuarios.rowCount()
            self.tabela_historico_usuarios.insertRow(row)
            for col, value in enumerate(row_data):
                self.tabela_historico_usuarios.setItem(row, col, QTableWidgetItem(value))

    def obter_coluna_usuario_para_ordenar(self):
        colunas = ["Data/Hora", "Usuário", "Ação", "Descrição"]
        dialog = ComboDialog("Ordenar por", "Escolha a coluna:", colunas, self)
        if dialog.exec() == QDialog.Accepted:
            return dialog.escolha()
        return None

    def obter_direcao_ordenacao_usuario(self):
        direcoes = ["Crescente", "Decrescente"]
        dialog = ComboDialog("Direção da Ordenação", "Escolha a direção:", direcoes, self)
        if dialog.exec() == QDialog.Accepted:
            return dialog.escolha()
        return None
    
    def filtrar_historico_usuarios(self):
        if getattr(self, "checkbox_selecionar") and self.checkbox_selecionar.isChecked():
            QMessageBox.warning(
                self,
                "Aviso",
                "Desmarque o checkbox antes de filtrar o histórico de usuários."
            )
            return
        # Criar a janela de filtro
        janela_filtro = QDialog(self)
        janela_filtro.setWindowTitle("Filtrar Histórico")
        layout = QVBoxLayout(janela_filtro)

        # Campo para inserir a data
        campo_data = QLineEdit()
        campo_data.setPlaceholderText("DD/MM/AAAA")
        
        # Conectar ao método de formatação, passando o texto
        campo_data.textChanged.connect(lambda: self.formatar_data_usuarios(campo_data))


        # Campo para selecionar se quer o mais recente ou mais antigo (filtro por hora)
        grupo_hora = QGroupBox("Filtrar por Hora")
        layout_hora = QVBoxLayout(grupo_hora)

        radio_mais_novo = QRadioButton("Mais Recente")
        radio_mais_velho = QRadioButton("Mais Antigo")

        layout_hora.addWidget(radio_mais_novo)
        layout_hora.addWidget(radio_mais_velho)
        grupo_hora.setLayout(layout_hora)

        # Botão para aplicar o filtro
        botao_filtrar = QPushButton("Aplicar Filtro")
        def aplicar_e_fechar():
            self.aplicar_filtro_usuarios(
                campo_data.text(),
                radio_mais_novo.isChecked(),
                radio_mais_velho.isChecked()
            )
            janela_filtro.accept()
        botao_filtrar.clicked.connect(aplicar_e_fechar)
        # Adicionar widgets ao layout
        layout.addWidget(QLabel("Filtros Disponíveis"))
        layout.addWidget(campo_data)
        layout.addWidget(grupo_hora)
        layout.addWidget(botao_filtrar)

        # Exibir a janela de filtro
        janela_filtro.setLayout(layout)
        janela_filtro.exec()

    def formatar_data_usuarios(self, campo_data):
        # Obter o texto do campo de data
        texto_data = campo_data.text().replace("/", "")  # Remover as barras existentes
        texto_data = ''.join(filter(str.isdigit, texto_data))  # Permite apenas números

        # Verificar se há caracteres alfabéticos (letras)
        if any(char.isalpha() for char in texto_data):
            # Mostrar mensagem de erro caso haja letras
            QMessageBox.warning(self, "Erro", "Somente números são permitidos.")
            campo_data.clear()
            return  # Não aplica a formatação se houver letras

        # Limitar a entrada para no máximo 8 dígitos
        if len(texto_data) > 8:
            texto_data = texto_data[:8]

         # Formatar a data no formato DD/MM/AAAA
        if len(texto_data) >= 8:
            data_formatada = "{}/{}/{}".format(texto_data[:2], texto_data[2:4], texto_data[4:])  # DD/MM/AAAA
        elif len(texto_data) > 6:
            data_formatada = "{}/{}".format(texto_data[:2], texto_data[2:8])  # DD/MM
        else:
            data_formatada = texto_data[:8]  # Apenas o DD

        # Atualizar o texto do campo de data se houver mudança
        if campo_data.text() != data_formatada:
            campo_data.setText(data_formatada)  # Atualiza o texto do campo de data
            campo_data.setCursorPosition(len(data_formatada))  # Move o cursor para o final do texto

    def aplicar_filtro_usuarios(self, data_filtrada, ordenar_por_mais_recente, ordenar_por_mais_antigo):
        data_formatada = data_filtrada.strip()

        with self.db.connecta() as conexao:
            cursor = conexao.cursor()

            query_base = "SELECT * FROM historico_usuarios"
            parametros = []

            # Filtro por data
            if data_formatada:
                query_base += ' WHERE "Data e Hora" LIKE ?'
                parametros.append(f"%{data_formatada}%")

            if ordenar_por_mais_recente:
                query_base += ' ORDER BY "Data e Hora" ASC'
            elif ordenar_por_mais_antigo:
                query_base += ' ORDER BY "Data e Hora" DESC'


            cursor.execute(query_base, parametros)
            resultados = cursor.fetchall()

        # Limpar e preencher a tabela
        self.tabela_historico_usuarios.setRowCount(0)
        for i, row in enumerate(resultados):
            self.tabela_historico_usuarios.insertRow(i)
            self.tabela_historico_usuarios.setItem(i, 0, QTableWidgetItem(row[0]))  # Data/Hora
            self.tabela_historico_usuarios.setItem(i, 1, QTableWidgetItem(row[1]))  # Usuário
            self.tabela_historico_usuarios.setItem(i, 2, QTableWidgetItem(row[2]))  # Ação
            self.tabela_historico_usuarios.setItem(i, 3, QTableWidgetItem(row[3]))  # Descrição

        QMessageBox.information(self, "Filtro Aplicado", f"{len(resultados)} registro(s) encontrado(s)!")



    def exportar_csv_usuarios(self):
        num_linhas = self.tabela_historico_usuarios.rowCount()
        num_colunas = self.tabela_historico_usuarios.columnCount()

        # Verificar se a tabela está vazia
        if self.tabela_historico_usuarios.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico encontrado para gerar arquivo CSV.")
            return  # Se a tabela estiver vazia, encerra a função sem prosseguir

        nome_arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Arquivo CSV",
            "historico_usuarios.csv",
            "Arquivos CSV (*.csv)"

        )

        if not nome_arquivo:
            return
        
        #Criar o arquivo CSV
        try:
            with open(nome_arquivo, mode="w",newline="",encoding="utf-8-sig") as arquivo_csv:
                escritor = csv.writer(arquivo_csv, delimiter=";")

                 # Adicionar cabeçalhos ao CSV
                cabecalhos = [self.tabela_historico_usuarios.horizontalHeaderItem(col).text() for col in range (num_colunas)]
                escritor.writerow(cabecalhos)

                # Adicionar os dados da tabela ao CSV
                for linha in range(num_linhas):
                    dados_linhas = [
                        self.tabela_historico_usuarios.item(linha, col).text() if self.tabela_historico_usuarios.item(linha, col) else ""
                        for col in range(num_colunas)

                    ]
                    escritor.writerow(dados_linhas)

                    QMessageBox.information(self, "Sucesso", f"Arquivo CSV salvo com sucesso em:\n{nome_arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao salvar o arquivo CSV:\n{str(e)}")


    def exportar_excel_usuarios(self):
        num_linhas = self.tabela_historico_usuarios.rowCount()
        num_colunas = self.tabela_historico_usuarios.columnCount()

        # Verificar se a tabela está vazia
        if self.tabela_historico_usuarios.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico encontrado para gerar arquivo Excel.")
            return  # Se a tabela estiver vazia, encerra a função sem prosseguir
        
        nome_arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Arquivo Excel",
            "historico_usuarios.xlsx",
            "Arquivos Excel (*.xlsx)"

        )

        if not nome_arquivo:
            return
        
        # Garantir que o arquivo tenha extensão .xlsx
        if not nome_arquivo.endswith(".xlsx"):
            nome_arquivo += ".xlsx"

        # Criar uma lista para armazenar os dados da tabela
        dados = []


        for linha in range(num_linhas):
            linha_dados = []
            for coluna in range(num_colunas):
                item = self.tabela_historico_usuarios.item(linha, coluna)
                linha_dados.append(item.text() if item else "") # Adicionar o texto ou vazio se o item for None
            dados.append(linha_dados)

        # Obter os cabeçalhos da tabela        
        cabecalhos = [self.tabela_historico_usuarios.horizontalHeaderItem(coluna).text() for coluna in range (num_colunas)]
        
        try:
            # Criar um DataFrame do pandas com os dados e cabeçalhos
            df = pd.DataFrame(dados, columns=cabecalhos)

            # Exportar para Excel
            df.to_excel(nome_arquivo, index=False,engine="openpyxl")
            QMessageBox.information(self, "Sucesso",f"Arquivo Excel gerado com sucesso em: \n{nome_arquivo}")
        except Exception as e:
            QMessageBox.critical(self, "Erro",f"Erro ao salvar arquivo Excel: {str(e)}")

    def exportar_pdf_usuarios(self):
        num_linhas = self.tabela_historico.rowCount()
        num_colunas = self.tabela_historico.columnCount()

        # Verificar se a tabela está vazia
        if self.tabela_historico.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico encontrado para gerar arquivo PDF.")
            return  # Se a tabela estiver vazia, encerra a função sem prosseguir

        nome_arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Arquivo PDF",
            "historico.pdf",
            "Arquivos PDF (*.pdf)"
        )

        if not nome_arquivo:
            return

        # Garantir que o arquivo tenha extensão .pdf
        if not nome_arquivo.endswith(".pdf"):
            nome_arquivo += ".pdf"

        # Criar uma lista para armazenar os dados da tabela
        dados = []

        # Obter os cabeçalhos da tabela
        cabecalhos = [self.tabela_historico.horizontalHeaderItem(coluna).text() for coluna in range(num_colunas)]
        dados.append(cabecalhos)  # Adicionar os cabeçalhos como a primeira linha do PDF

        # Adicionar os dados da tabela
        for linha in range(num_linhas):
            linha_dados = []
            for coluna in range(num_colunas):
                item = self.tabela_historico.item(linha, coluna)
                linha_dados.append(item.text() if item else "")  # Adicionar o texto ou vazio se o item for None
            dados.append(linha_dados)

        try:
            # Criar o PDF
            pdf = SimpleDocTemplate(nome_arquivo, pagesize=landscape(letter))
            tabela = Table(dados)

            # Adicionar estilo à tabela
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Cabeçalho com fundo cinza
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Texto do cabeçalho branco
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centralizar texto
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fonte do cabeçalho
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Espaçamento inferior no cabeçalho
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Fundo das linhas de dados
                ('GRID', (0, 0), (-1, -1), 1, colors.black)  # Bordas da tabela
            ])
            tabela.setStyle(estilo)

            # Gerar o PDF
            pdf.build([tabela])
            QMessageBox.information(self, "Sucesso", f"Arquivo PDF gerado com sucesso em: \n{nome_arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar arquivo PDF: {str(e)}")

    def pausar_historico_usuario(self):
        # Criação da nova janela de histórico como QMainWindow
        self.janela_escolha = QMainWindow()
        self.janela_escolha.setWindowTitle("Pausar Histórico")
        self.janela_escolha.resize(255, 150)

        # Carregar tema
        config = self.carregar_config()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    border-radius: 8px;
                    background: qlineargradient(
                        x1:0, y1:0, x2:0, y2:1,
                        stop:0 rgb(60, 60, 60),   /* topo */
                        stop:1 rgb(100, 100, 100) /* base */
                    );
                    font-size: 12px;
                    padding: 3px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
                
            """
            
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            table_view_style = """
            /* QTableView com seleção diferenciada */
            QTableView {
                background-color: #202124;
                color: white;
                gridline-color: #555555;
                selection-background-color: #7a7a7a;
                selection-color: white;
            }
            /* Coluna dos cabeçalhos */
            QHeaderView::section {
                background-color: #ffffff;
                color: black;
                border: 1px solid #aaaaaa;
                padding: 1px;
            }

            /* QTabWidget headers brancos */
            QTabWidget::pane {
                border: 1px solid #444444;
                background-color: #202124;
            }
            /* Estiliza a barra de rolagem horizontal */
            QTableView QScrollBar:horizontal {
                border: none;
                background-color: #ffffff;
                height: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* Estiliza a barra de rolagem vertical */
            QTableView QScrollBar:vertical {
                border: none;
                background-color: #ffffff;  
                width: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* QTabWidget headers brancos */
            QTabWidget::pane {
                border: 1px solid #444444;
                background-color: #202124;
            }
            /* Estiliza a barra de rolagem horizontal */
            QTableView QScrollBar:horizontal {
                border: none;
                background-color: none;
                height: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* Estiliza a barra de rolagem vertical */
            QTableView QScrollBar:vertical {
                border: none;
                background-color: none; 
                width: 12px;
                margin: 0px;
                border-radius: 5px;
            }
            

            /* Parte que você arrasta */
            QTableView QScrollBar::handle:vertical {
                background-color: #777777;  /* cinza médio */
                min-height: 22px;
                border-radius: 5px;
            }

            QTableView QScrollBar::handle:horizontal {
                background-color: #777777;
                min-width: 22px;
                border-radius: 5px;
            }

            /* Groove horizontal */
            QTableView QScrollBar::groove:horizontal {
                background-color: transparent;
                border-radius: 5px;
                height: 15px;
                margin: 0px 10px 0px 10px;
            }

            /* Groove vertical */
            QTableView QScrollBar::groove:vertical {
                background-color: transparent;
                border-radius: 5px;
                width: 25px;
                margin: 10px 0px 10px 10px;
            }

            /* Estilo para item selecionado */
            QTableWidget::item:selected {
                background-color: #555555;  /* cinza de seleção */
                color: white;
            }
            /* CornerButton (canto superior esquerdo) */
            QTableCornerButton::section {
                background-color: #ffffff;
                border: 1px solid #aaaaaa;
                padding: 2px;
            }
            /* Forçar cor do texto do QCheckBox */
            QCheckBox {
                color: white;
            }

            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """
        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """
        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(0,120,180),
                                                stop:1 rgb(0,150,220));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(0,100,160);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #007acc;
                }
                QPushButton:pressed {
                    background-color: #006bb3;
                    border: 2px solid #005c99;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 3px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            scroll_style = """
                QScrollBar:vertical {
                border: none;
                background-color: rgb(255, 255, 255); /* branco */
                width: 30px;
                margin: 0px 10px 0px 10px;
            }
                QScrollBar::handle:vertical {
                background-color: rgb(180, 180,180);  /* cinza */
                min-height: 30px;
                border-radius: 5px;
            }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
        """


        # Botão "Sim"
        botao_sim = QPushButton("Sim")
        botao_sim.clicked.connect(self.historico_ativo_usuario)
        botao_sim.setStyleSheet(button_style)

        # Botão "Não"
        botao_nao = QPushButton("Não")
        botao_nao.clicked.connect(self.historico_inativo_usuario)
        botao_nao.setStyleSheet(button_style)


        # Criação do layout e tabela para exibir o histórico
        central_widget = QWidget()
        layout = QVBoxLayout(central_widget)

        # Texto centralizado
        label = QLabel("Deseja pausar o histórico?")
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)  # Alinha o texto ao centro

        layout.addWidget(label)  # Adiciona o texto centralizado
        layout.addWidget(botao_sim)
        layout.addWidget(botao_nao)

        self.janela_escolha.setCentralWidget(central_widget)
        self.janela_escolha.show()


    def historico_ativo_usuario(self):
        # Atualiza o estado do histórico para ativo
        self.main_window.historico_usuario_pausado = True  # Atualiza a variável no MainWindow
        QMessageBox.information(self, "Histórico", "O registro do histórico foi pausado.")
        self.janela_escolha.close()

    def historico_inativo_usuario(self):
        # Atualiza o estado do histórico para inativo (continua registrando)
        self.main_window.historico_usuario_pausado = False  # Atualiza a variável no MainWindow
        QMessageBox.information(self, "Histórico", "O registro do histórico continua ativo.")
        self.janela_escolha.close()


    def abrir_planilha_usuarios(self):
        # Abrir o diálogo para selecionar o arquivo Excel
        nome_arquivo, _ = QFileDialog.getOpenFileName(self, "Abrir Arquivo Excel", "", "Arquivos Excel (*.xlsx)")

        if not nome_arquivo:
            return  # Se o usuário cancelar a seleção do arquivo

        
        # Alterar o texto da line_excel para "Carregando arquivo Excel..."
        self.line_excel_usuarios.setText("Carregando arquivo Excel...")
        self.nome_arquivo_excel = nome_arquivo  # Salva para usar depois

        # Inicializar a barra de progresso
        self.progress_excel_usuarios.setValue(0)
        self.progresso = 0
        

        # Começar o timer para simular carregamento visual
        self.timer_excel = QTimer()
        self.timer_excel.timeout.connect(self.atualizar_progresso_excel_usuarios)
        self.timer_excel.start(20)


    def atualizar_progresso_excel_usuarios(self):
        if self.progresso < 100:
            self.progresso += 1
            self.progress_excel_usuarios.setValue(self.progresso)
        else:
            self.timer_excel.stop()

            try:
                df = pd.read_excel(self.nome_arquivo_excel, engine="openpyxl", header=0)
                df = df.fillna("Não informado")

                coluna_table_ativos = [
                    "Nome", "Usuário", "Senha", "Confirmar Senha", "Acesso",
                    "Endereço", "CEP", "CPF", "Número", "Estado", "E-mail", "RG", "Complemento", "Telefone",
                    "Data de Nascimento", "Última Troca de Senha", "Data da Senha Cadastrada",
                    "Data da Inclusão do Usuário", "Segredo", "Usuário Logado"
                ]

                if df.shape[1] != len(coluna_table_ativos):
                    QMessageBox.warning(self, "Erro", "O número de colunas no arquivo Excel não corresponde ao número esperado.")
                    self.line_excel_usuarios.clear()
                    self.resetar_progresso()
                    return

                if df.empty:
                    QMessageBox.warning(self, "Erro", "O arquivo Excel está vazio.")
                    self.line_excel_usuarios.clear()
                    self.resetar_progresso()
                    return

                self.table_ativos.setRowCount(0)

                for row in df.itertuples(index=False):
                    row_position = self.table_ativos.rowCount()
                    self.table_ativos.insertRow(row_position)
                    for column, value in enumerate(row):
                        item = self.formatar_texto(str(value))
                        self.table_ativos.setItem(row_position, column, item)

                QMessageBox.information(self, "Sucesso", "Arquivo Excel importado com sucesso!")

            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao importar o arquivo Excel: {e}")

            self.line_excel_usuarios.setText(self.nome_arquivo_excel)
            self.resetar_progresso()
            # Ajusta colunas e linhas automaticamente após preencher
            self.main_window.table_ativos.resizeColumnsToContents()
            self.main_window.table_ativos.resizeRowsToContents()

    def resetar_progresso(self):
        self.progress_excel_usuarios.setValue(0)
        self.progresso = 0
    
    def importar_usuario(self):
        if self.table_ativos.rowCount() == 0 and self.table_inativos.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum dado encontrado para gerar arquivo Excel.")
            return

        nome_arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Arquivo Excel",
            "relatório de usuários.xlsx",
            "Arquivos Excel (*.xlsx)"
        )

        if not nome_arquivo:
            return

        if not nome_arquivo.endswith(".xlsx"):
            nome_arquivo += ".xlsx"

        try:
            with pd.ExcelWriter(nome_arquivo, engine="openpyxl") as writer:
                def tabela_para_dataframe(tabela):
                    dados = []
                    cabecalhos = [tabela.horizontalHeaderItem(col).text() for col in range(tabela.columnCount())]
                    for linha in range(tabela.rowCount()):
                        linha_dados = []
                        for coluna in range(tabela.columnCount()):
                            item = tabela.item(linha, coluna)
                            linha_dados.append(item.text() if item else "")
                        dados.append(linha_dados)
                    return pd.DataFrame(dados, columns=cabecalhos)
                
                if self.table_ativos.rowCount() > 0:
                    df_ativos = tabela_para_dataframe(self.table_ativos)
                    df_ativos.to_excel(writer, sheet_name="Ativos", index=False)

                if self.table_inativos.rowCount() > 0:
                    df_inativos = tabela_para_dataframe(self.table_inativos)
                    df_inativos.to_excel(writer, sheet_name="Inativos", index=False)

            # Abrir o arquivo com openpyxl para ajustar estilos
            from openpyxl import load_workbook
            wb = load_workbook(nome_arquivo)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column  # número da coluna (1, 2, ...)
                    column_letter = get_column_letter(column)

                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    ws.column_dimensions[column_letter].width = max_length + 2  # ajustar largura

                # Negrito no cabeçalho
                for cell in ws[1]:
                    cell.font = Font(bold=True)

            wb.save(nome_arquivo)

            QMessageBox.information(self, "Sucesso", f"Arquivo Excel gerado com sucesso em:\n{nome_arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar arquivo Excel: {str(e)}")

    # Limpa a coluna selecionada clicando em qualquer lugar da tabela
    def eventFilter(self, source, event):
        if event.type() == QEvent.MouseButtonPress:
            if source == self.main_window.table_ativos.viewport():
                index = self.main_window.table_ativos.indexAt(event.pos())
                if not index.isValid():
                    self.main_window.table_ativos.clearSelection()

            elif source == self.main_window.table_inativos.viewport():
                index = self.main_window.table_inativos.indexAt(event.pos())
                if not index.isValid():
                    self.main_window.table_inativos.clearSelection()

        return super().eventFilter(source, event)
    
    def abrir_panilha_usuarios_em_massa(self):
        # Abrir o diálogo para selecionar o arquivo Excel
        nome_arquivo, _ = QFileDialog.getOpenFileName(self, "Abrir Arquivo Excel", "", "Arquivos Excel (*.xlsx)")

        # Se o usuário cancelar a seleção do arquivo
        if not nome_arquivo:
            return
        # Alterar o texto da line_excel para "Carregando arquivo Excel..."
        self.line_edit_massa_usuarios.setText("Carregando arquivo Excel...")
        self.nome_arquivo_excel_massa = nome_arquivo

        # Inicializar a barra de progresso
        self.progress_massa_usuarios.setValue(0)
        self.progresso_massa = 0

        # Começar o timer para simular carregamento visual
        self.timer_excel_massa_usuarios = QTimer()
        self.timer_excel_massa_usuarios.timeout.connect(self.atualizar_progress_line_usuarios_em_massa)
        self.timer_excel_massa_usuarios.start(20)

    def atualizar_progress_line_usuarios_em_massa(self):
        if self.progresso_massa < 100:
            self.progresso_massa += 1
            self.progress_massa_usuarios.setValue(self.progresso_massa)
        else:
            self.timer_excel_massa_usuarios.stop()

            try:
                df = pd.read_excel(self.nome_arquivo_excel_massa, engine="openpyxl", header=0)
                df = df.fillna("Não informado")

                colunas_table_massa_usuarios = ["Nome", "Usuário", "Senha", "Confirmar Senha","CEP", 
                    "Endereço", "Número","Cidade","Bairro","Estado","Complemento","Telefone","E-mail", "Data de Nascimento",
                    "RG","CPF","CNPJ", "Acesso"]

                # Verificar se o DataFrame está vazio
                if df.shape[1] != len(colunas_table_massa_usuarios):
                    QMessageBox.warning(self, "Erro", "O número de colunas em massa no arquivo Excel não corresponde ao número esperado.")
                    self.line_edit_massa_usuarios.clear()
                    # Zerando a barra de progresso
                    self.progress_massa_usuarios.setValue(0)
                    self.progresso_massa = 0
                    return 
                    
                    
                if df.shape[0] == 0:
                    QMessageBox.warning(self, "Erro", "O arquivo Excel está vazio.")
                    self.line_edit_massa_usuarios.clear()
                    # Zerando a barra de progresso
                    self.progress_massa_usuarios.setValue(0)
                    self.progresso_massa = 0
                    return  
                # Adicionar os dados à tabela
                self.table_massa_usuarios.setRowCount(0)
                
        
                for _, row in df.iterrows():
                    row_position = self.table_massa_usuarios.rowCount()
                    self.table_massa_usuarios.insertRow(row_position)
                    for column, value in enumerate(row):
                        item = self.formatar_texto_usuarios_em_massa(str(value))
                        self.table_massa_usuarios.setItem(row_position, column, item)
                QMessageBox.information(self, "Sucesso", "Arquivo Excel importado com sucesso!")

            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao importar o arquivo Excel: {e}")
             # Quando o arquivo for carregado, atualizar o texto da line_excel com o caminho do arquivo
            self.line_edit_massa_usuarios.setText(self.nome_arquivo_excel_massa)
            
            self.progress_massa_usuarios.setValue(0)
            self.progresso_massa = 0
            self.table_massa_usuarios.resizeColumnsToContents()
            self.table_massa_usuarios.resizeRowsToContents()
            
    def formatar_texto_usuarios_em_massa(self, texto):
        item = QTableWidgetItem(texto)
        item.setTextAlignment(Qt.AlignCenter)  # Centraliza o texto
        item.setForeground(QBrush(QColor("white"))) 
        return item

    def cadastrar_usuarios_em_massa(self):
        try:
            total_linhas = self.table_massa_usuarios.rowCount()
            if total_linhas == 0:
                QMessageBox.warning(self, "Erro", "Nenhum usuário encontrado para cadastrar.")
                return
            with self.db.connecta() as conexao:
                cursor = conexao.cursor()
            
            
            usuario_logado = self.config.obter_usuario_logado()
           

            for linha in range(total_linhas):
                nome = self.table_massa_usuarios.item(linha, 0).text().strip() if self.table_massa_usuarios.item(linha,0) else ""
                usuario = self.main_window.gerar_codigo_usuarios()
                senha = self.table_massa_usuarios.item(linha, 2).text().strip() if self.table_massa_usuarios.item(linha, 2) else ""
                confirmar_senha = self.table_massa_usuarios.item(linha, 3).text().strip() if self.table_massa_usuarios.item(linha, 3) else ""
                cep = self.table_massa_usuarios.item(linha, 4).text().strip() if self.table_massa_usuarios.item(linha, 4) else ""
                endereco = self.table_massa_usuarios.item(linha, 5).text().strip() if self.table_massa_usuarios.item(linha, 5) else ""
                numero = self.table_massa_usuarios.item(linha, 6).text().strip() if self.table_massa_usuarios.item(linha, 6) else ""
                cidade = self.table_massa_usuarios.item(linha, 7).text().strip() if self.table_massa_usuarios.item(linha, 7) else ""
                bairro = self.table_massa_usuarios.item(linha, 8).text().strip() if self.table_massa_usuarios.item(linha, 8) else ""
                estado = self.table_massa_usuarios.item(linha, 9).text().strip() if self.table_massa_usuarios.item(linha, 9) else ""
                complemento = self.table_massa_usuarios.item(linha, 10).text().strip() if self.table_massa_usuarios.item(linha, 10) else ""
                telefone = self.table_massa_usuarios.item(linha, 11).text().strip() if self.table_massa_usuarios.item(linha, 11) else ""
                email = self.table_massa_usuarios.item(linha, 12).text().strip() if self.table_massa_usuarios.item(linha, 12) else ""
                data_nascimento = self.table_massa_usuarios.item(linha, 13).text().strip() if self.table_massa_usuarios.item(linha, 13) else ""
                rg = self.table_massa_usuarios.item(linha, 14).text().strip() if self.table_massa_usuarios.item(linha, 14) else ""
                cpf = self.table_massa_usuarios.item(linha, 15).text().strip() if self.table_massa_usuarios.item(linha, 15) else ""
                cnpj = self.table_massa_usuarios.item(linha, 16).text().strip() if self.table_massa_usuarios.item(linha, 16) else ""
                acesso = self.table_massa_usuarios.item(linha,17).text().strip() if self.table_massa_usuarios.item(linha, 17) else ""
           
                segredo = "" # vazio por padrão no cadastro em massa
                     
                self.db.insert_user(
                    nome,usuario,senha,confirmar_senha,cep,endereco,numero,cidade,bairro,estado,complemento,
                    telefone,email,data_nascimento,rg,cpf,cnpj,segredo,usuario_logado,acesso)
                
                data_atual = datetime.now().strftime("%d/%m/%Y %H:%M")

                # Registrar no histórico
                descricao = f"Usuário {usuario} foi cadastrado no sistema!"
                cursor.execute("""
                    INSERT INTO historico_usuarios ("Data e Hora",Usuário,Ação,Descrição)
                    VALUES (?,?,?,?)
                """,(data_atual,usuario_logado, "Cadastro em Massa",descricao))

            conexao.commit()


            QMessageBox.information(self, "Sucesso", "Usuários cadastrados em massa com sucesso!")
            self.line_edit_massa_usuarios.clear()
            # Limpar a tabela após a inserção
            self.table_massa_usuarios.setRowCount(0)

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao cadastrar usuários em massa:\n{e}")

                
                
        
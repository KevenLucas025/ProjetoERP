from PySide6.QtWidgets import (QLineEdit,QTableWidgetItem,
                               QMessageBox,QMainWindow,QVBoxLayout,QWidget,QLabel,QCheckBox,
                               QPushButton,QScrollArea,QComboBox,QGridLayout,QHeaderView,QHBoxLayout,
                               QGraphicsOpacityEffect,QTableWidget,QDialog,
                               QRadioButton,QGroupBox,QFileDialog,QFormLayout,QDateEdit,QMenu,QApplication)
from PySide6.QtGui import QColor,QBrush,QGuiApplication
from PySide6.QtCore import Qt,QTimer,QPropertyAnimation,QEvent,QDate,QPoint
from database import DataBase
import sqlite3
import pandas as pd
from configuracoes import Configuracoes_Login
from utils import Temas
from datetime import datetime
from dialogos import ComboDialog,DialogoSenha
import csv
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.pagesizes import letter,landscape,A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,Paragraph,Spacer
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from fpdf import FPDF
import json



class Clientes_Juridicos(QWidget):
    def __init__(self, line_clientes: QLineEdit,main_window,btn_adicionar_cliente_juridico,
                 btn_editar_clientes,btn_excluir_clientes,btn_gerar_relatorio_clientes,btn_marcar_como_clientes,
                 btn_historico_clientes):
        super().__init__()
        self.line_clientes = line_clientes
        self.db = DataBase("banco_de_dados.db")
        self.temas = Temas()
        self.config = Configuracoes_Login(self)
        self.coluna_checkboxes_clientes_adicionada = False
        self.checkboxes_clientes = []
        
        self.timer_buscar = QTimer()
        self.timer_buscar.setSingleShot(True)
        self.timer_buscar.timeout.connect(self._executar_busca_dinamica_juridicos)
        self.line_clientes.textChanged.connect(self._iniciar_timer_busca_juridicos)
        
        
        self.main_window = main_window
        self.table_clientes_juridicos = self.main_window.table_clientes_juridicos  # Referência para a tabela no main window
        self.btn_adicionar_cliente_juridico = btn_adicionar_cliente_juridico
        self.btn_editar_clientes = btn_editar_clientes
        self.btn_excluir_clientes = btn_excluir_clientes
        self.btn_gerar_relatorio_clientes = btn_gerar_relatorio_clientes
        self.btn_marcar_como_clientes = btn_marcar_como_clientes
        self.btn_historico_clientes = btn_historico_clientes
        
        
        
        self.btn_excluir_clientes.clicked.connect(self.excluir_clientes_juridicos)
        self.btn_marcar_como_clientes.clicked.connect(self.marcar_como_clientes)
        self.btn_adicionar_cliente_juridico.clicked.connect(self.exibir_janela_cadastro_cliente)
        self.btn_editar_clientes.clicked.connect(self.editar_cliente_juridico)
        self.btn_historico_clientes.clicked.connect(self.historico_clientes_juridicos)
        self.btn_gerar_relatorio_clientes.clicked.connect(self.abrir_janela_relatorio_clientes_juridicos)
        self.imagem_line()

        
        
        # ENTER → busca manual
        self.line_clientes.returnPressed.connect(
            lambda: self.buscar_cliente_juridico_dinamico(manual=True)
        )

        # Botão lupa → busca manual
        self.botao_lupa_juridicos.clicked.connect(
            lambda: self.buscar_cliente_juridico_dinamico(manual=True)
        )

        # Digitação → busca dinâmica
        self.line_clientes.textChanged.connect(
            lambda: self.buscar_cliente_juridico_dinamico(manual=False)
        )



        self.configurar_menu_contexto()

        
        #self.configurar_line_clientes()


        
        self.installEventFilter(self)
        self.table_clientes_juridicos.viewport().installEventFilter(self)

    def _iniciar_timer_busca_juridicos(self, texto):
        self.ultimo_texto = texto
        self.timer_buscar.start(300)  # espera 300ms após digitar

    def _executar_busca_dinamica_juridicos(self):
        self.buscar_cliente_juridico_dinamico(self.ultimo_texto)

    # Função auxiliar para criar um QTableWidgetItem com texto centralizado
    def formatar_texto_juridico(self, text):
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)  # Centraliza o texto
        item.setForeground(QBrush(QColor("white")))
        return item
                
    def carregar_clientes_juridicos(self):
        try:
            with self.db.connecta() as conexao:
                cursor = conexao.cursor()

                cursor.execute("""
                    SELECT 
                        "Nome do Cliente",
                        "Razão Social",
                        "Data da Inclusão",
                        CNPJ,
                        RG,
                        CPF,
                        Email,
                        CNH,
                        "Categoria da CNH",
                        "Data de Emissão da CNH",
                        "Data de Vencimento da CNH",
                        Telefone, 
                        CEP, 
                        Endereço, 
                        Número,
                        Complemento,
                        Cidade, 
                        Bairro,
                        Estado,
                        "Status do Cliente", 
                        "Categoria do Cliente",
                        "Última Atualização",
                        "Valor Gasto Total",
                        "Última Compra"
                    FROM clientes_juridicos
                    ORDER BY "Data da Inclusão" DESC
                """)

                dados = cursor.fetchall()

                # Limpa a tabela antes de recarregar
                self.table_clientes_juridicos.clearContents()
                self.table_clientes_juridicos.setRowCount(0)

                deslocamento = 1 if self.coluna_checkboxes_clientes_adicionada else 0
                self.checkboxes_clientes = []

                for linha_idx, linha_dados in enumerate(dados):
                    self.table_clientes_juridicos.insertRow(linha_idx)

                    # Adiciona checkbox se estiver ativado
                    if self.coluna_checkboxes_clientes_adicionada:
                        checkbox = QCheckBox()
                        checkbox.setStyleSheet("margin-left: 9px; margin-right: 9px;")
                        self.table_clientes_juridicos.setCellWidget(linha_idx, 0, checkbox)
                        self.checkboxes_clientes.append(checkbox)

                    # Preenche os dados nas colunas (com deslocamento)
                    for coluna_idx, dado in enumerate(linha_dados):
                        item = self.formatar_texto_juridico(str(dado))
                        self.table_clientes_juridicos.setItem(linha_idx, coluna_idx + deslocamento, item)

                self.table_clientes_juridicos.resizeColumnsToContents()
                self.table_clientes_juridicos.resizeRowsToContents()

        except Exception as e:
            QMessageBox.critical(self.main_window, "Erro", f"Erro ao carregar clientes:\n{e}")

    def imagem_line(self):
        # Criar botão da lupa
        self.botao_lupa_juridicos = QPushButton(self.line_clientes)
        self.botao_lupa_juridicos.setCursor(Qt.PointingHandCursor)  # Muda o cursor ao passar o mouse
        self.botao_lupa_juridicos.setContentsMargins(0,0,0,0)
        self.botao_lupa_juridicos.setObjectName("botao_lupa_juridicos")
        
        # Definir tamanho do botão
        altura = self.line_clientes.height() - 4  # Ajustar altura conforme a LineEdit
        self.botao_lupa_juridicos.setFixedSize(altura, altura)

        # Posicionar o botão no canto direito da LineEdit
        self.botao_lupa_juridicos.move(self.line_clientes.width() - altura + 1, 2)

        

        # Conectar clique do botão a uma função
        self.botao_lupa_juridicos.clicked.connect(self.buscar_cliente_juridico_dinamico)

    def _buscar_clientes_juridicos(self, texto):
        """Executa a busca no banco e retorna lista de resultados."""
        if not isinstance(texto, str):
            return []

        texto = texto.strip()
        if not texto:
            return self._listar_todos_clientes()

        with sqlite3.connect('banco_de_dados.db') as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT  
                    "Nome do Cliente",
                    "Razão Social",
                    "Data da Inclusão",
                    CNPJ,
                    RG,       
                    CPF,
                    Email,
                    CNH,
                    "Categoria da CNH",
                    "Data de Emissão da CNH",
                    "Data de Vencimento da CNH",
                    Telefone,
                    CEP,
                    Endereço,
                    Número,
                    Complemento,
                    Cidade,
                    Bairro,
                    Estado,
                    "Status do Cliente",
                    "Categoria do Cliente",
                    "Última Atualização",
                    "Origem do Cliente",
                    "Valor Gasto Total",
                    "Última Compra"
                FROM clientes_juridicos
                WHERE 
                    LOWER("Nome do Cliente") LIKE LOWER(?) OR
                    CPF LIKE ? OR
                    RG LIKE ? OR
                    Telefone LIKE ?
            """, (f'%{texto}%', f'%{texto}%', f'%{texto}%', f'%{texto}%'))
            return cursor.fetchall()

    def _listar_todos_clientes(self):
        with sqlite3.connect('banco_de_dados.db') as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT  
                    "Nome do Cliente",
                    "Razão Social",
                    "Data da Inclusão",
                    CNPJ,
                    RG,       
                    CPF,
                    Email,
                    CNH,
                    "Categoria da CNH",
                    "Data de Emissão da CNH",
                    "Data de Vencimento da CNH",
                    Telefone,
                    CEP,
                    Endereço,
                    Número,
                    Complemento,
                    Cidade,
                    Bairro,
                    Estado,
                    "Status do Cliente",
                    "Categoria do Cliente",
                    "Última Atualização",
                    "Origem do Cliente",
                    "Valor Gasto Total",
                    "Última Compra"
                FROM clientes_juridicos
            """)
            return cursor.fetchall()

    def buscar_cliente_juridico_dinamico(self, texto=None, manual=False):
        # Captura o texto do campo se não foi passado
        if not isinstance(texto, str):
            texto = self.line_clientes.text()

        texto = texto.strip()

        if not texto:
            if manual:
                QMessageBox.warning(self, "Atenção", "Por favor, digite um cliente para pesquisar.")
            else:
                # Busca dinâmica sem texto → mostra todos
                self.preencher_resultado_busca(self._listar_todos_clientes())
            return

        resultados = self._buscar_clientes_juridicos(texto)

        if not resultados and manual:
            QMessageBox.information(self, "Resultado", "Nenhum cliente encontrado.")

        self.preencher_resultado_busca(resultados)

    def buscar_cliente_juridico_manual(self):
        texto = self.line_clientes_fisicos.text().strip()
        if not texto:
            self.preencher_resultado_busca(self._listar_todos_clientes())
            return

        resultados = self._buscar_clientes_fisicos(texto)
        if resultados:
            self.preencher_resultado_busca(resultados)
        else:
            QMessageBox.information(self, "Resultado", "Nenhum cliente encontrado.")
            self.preencher_resultado_busca(self._listar_todos_clientes())


    def preencher_resultado_busca(self, resultados):
        self.table_clientes_juridicos.setRowCount(0)

        for row_data in resultados:
            row_index = self.table_clientes_juridicos.rowCount()
            self.table_clientes_juridicos.insertRow(row_index)
            for col_index, data in enumerate(row_data):
                item = self.formatar_texto_juridico(str(data))
                self.table_clientes_juridicos.setItem(row_index, col_index, item)
        self.table_clientes_juridicos.resizeColumnsToContents()
        self.table_clientes_juridicos.resizeRowsToContents()

    def formatar_e_buscar_cep(self, widget):
        texto_cep = widget.text()
        self.main_window.formatar_cep(texto_cep, widget)

        dados = self.main_window.buscar_cep(texto_cep)
        if dados:
            self.campos_cliente_juridico["Endereço"].setText(dados.get("logradouro", ""))
            self.campos_cliente_juridico["Complemento"].setText(dados.get("complemento", ""))
            self.campos_cliente_juridico["Bairro"].setText(dados.get("bairro", ""))
            self.campos_cliente_juridico["Cidade"].setText(dados.get("localidade", ""))
            
            estado = dados.get("uf", "")
            index_estado = self.campos_cliente_juridico["Estado"].findText(estado)
            if index_estado >= 0:
                self.campos_cliente_juridico["Estado"].setCurrentIndex(index_estado)

                
    def exibir_edicao_clientes(self, dados_cliente: dict):
        self.dados_originais_cliente = dados_cliente.copy()
        self.alteracoes_realizadas = False

        # Inicializa o dicionário para armazenar os widgets
        self.campos_cliente_juridico = {}

        # Criar a janela
        self.janela_editar_cliente = QMainWindow()
        self.janela_editar_cliente.setWindowTitle("Editar Cliente Jurídico")
        self.janela_editar_cliente.resize(683, 600)
        self.janela_editar_cliente.setStyleSheet("background-color: rgb(0, 80, 121);")

        # Centralizar a janela
        screen = QGuiApplication.primaryScreen()
        screen_geometry = screen.availableGeometry()
        window_geometry = self.janela_editar_cliente.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.janela_editar_cliente.move(window_geometry.topLeft())

        central_widget = QWidget()
        layout = QGridLayout(central_widget)
        layout.setSpacing(10)
        layout.setContentsMargins(30, 30, 30, 30)

         # Carregar tema
        config = self.temas.carregar_config_arquivo()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"
            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(60,60,60),
                                                stop:1 rgb(100,100,100));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid #666666;
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
            """
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """

        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"
            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            scroll_style = """
                QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"
            lineedit_bg = "white"
            button_style = """
            QPushButton {
                color: rgb(255, 255, 255);
                border-radius: 8px;
                font-size: 16px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgb(50, 150, 250), stop:1 rgb(100, 200, 255)); /* Gradiente de azul claro para azul mais claro */
                border: 4px solid transparent;
            }

            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgb(100, 180, 255), stop:1 rgb(150, 220, 255)); /* Gradiente de azul mais claro para azul ainda mais claro */
                color: black;
            }
            QPushButton:pressed {
                background-color: #006bb3;
                border: 2px solid #005c99;
            }
            """
            combobox_style = """
            QComboBox {
                background-color: white;
                border: 3px solid rgb(50,150,250);
                border-radius: 5px;
                color: black;
                padding: 5px;
            }

            QComboBox QAbstractItemView {
                background-color: white;
                color: black;
                border: 3px solid white;
                selection-background-color: rgb(120,120,120);
                selection-color: black;
            }

            /* Scrollbar vertical */
            QComboBox QScrollBar:vertical {
                background-color: rgb(240,240,240);  /* Trilha visível */
                width: 10px;
                margin: 1px;
                border: 1px solid white;
                border-radius: 5px;
            }

            /* Alça (handle) da scrollbar */
            QComboBox QScrollBar::handle:vertical {
                background: rgb(120,120,120);  /* Cor da barra frontal,sobe e desce*/
                min-height: 30px;
                border-radius: 5px;
            }

            /* Trilha atrás do handle — essa parte faz toda a diferença */
            QComboBox QScrollBar::add-page:vertical,
            QComboBox QScrollBar::sub-page:vertical {
                background: transparent;  /* barra atrás do scroll */
            }

            /* Esconde setas */
            QComboBox QScrollBar::add-line:vertical,
            QComboBox QScrollBar::sub-line:vertical {
                height: 0px;
                background: none;
                border: none;
            }
            """
            scroll_style = """
                QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: #b4b4b4;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """
        
        # Widget central e layout
        conteudo = QWidget()
        conteudo.setStyleSheet(f"background-color: {bg_cor}; color: {text_cor};")
        layout = QVBoxLayout(conteudo)


         # Função auxiliar para adicionar campos
        def add_linha(titulo, widget=None):
            label = QLabel(titulo)
            label.setStyleSheet(f"color: {text_cor};")
            layout.addWidget(label)
            if widget is None:
                widget = QLineEdit()
                widget.setStyleSheet(lineedit_style)
                # Preencher valor existente
                chave = titulo.rstrip(":")
                widget.setText(dados_cliente.get(chave, ""))
            layout.addWidget(widget)
            chave_sem_ponto = titulo.rstrip(":")
            self.campos_cliente_juridico[chave_sem_ponto] = widget
            return label, widget

        # Campos básicos
        add_linha("Nome do Cliente")
        add_linha("Razão Social")
        add_linha("Data da Inclusão")
        add_linha("CNPJ")
        add_linha("RG")
        add_linha("CPF")
        add_linha("Email")
        add_linha("CNH")

        # Categoria CNH
        combobox_categoria_cnh = QComboBox()
        combobox_categoria_cnh.addItems(["Selecionar","AB","A","B","C","D","E","Nenhuma"])
        valor_categoria = dados_cliente.get("Categoria da CNH", "Selecionar")
        index_categoria = combobox_categoria_cnh.findText(valor_categoria)
        combobox_categoria_cnh.setCurrentIndex(index_categoria if index_categoria >=0 else 0)
        combobox_categoria_cnh.setStyleSheet(combobox_style)
        label_categoria, widget_categoria = add_linha("Categoria da CNH", combobox_categoria_cnh)

         # Datas CNH
        label_emissao_cnh, widget_emissao_cnh = add_linha("Data de Emissão da CNH")
        widget_emissao_cnh.setText(dados_cliente.get("Data de Emissão da CNH",""))
        label_vencimento_cnh, widget_vencimento_cnh = add_linha("Data de Vencimento da CNH")
        widget_vencimento_cnh.setText(dados_cliente.get("Data de Vencimento da CNH",""))

        # Mostrar/esconder datas conforme categoria
        def on_categoria_cnh_change(text):
            if text not in ("Selecionar","Nenhuma"):
                label_emissao_cnh.show()
                widget_emissao_cnh.show()
                label_vencimento_cnh.show()
                widget_vencimento_cnh.show()
            else:
                label_emissao_cnh.hide()
                widget_emissao_cnh.hide()
                label_vencimento_cnh.hide()
                widget_vencimento_cnh.hide()
                widget_emissao_cnh.clear()
                widget_vencimento_cnh.clear()

        combobox_categoria_cnh.currentTextChanged.connect(on_categoria_cnh_change)
        on_categoria_cnh_change(valor_categoria)

        add_linha("Telefone")
        add_linha("CEP")
        add_linha("Endereço")
        add_linha("Número")
        add_linha("Complemento")
        add_linha("Cidade")
        add_linha("Bairro")

        # Estado
        combobox_estado_cliente = QComboBox()
        combobox_estado_cliente.addItems([
            "Selecionar","AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG",
            "PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"
        ])
        valor_estado = dados_cliente.get("Estado", "Selecionar")
        index_estado = combobox_estado_cliente.findText(valor_estado)
        combobox_estado_cliente.setCurrentIndex(index_estado if index_estado >=0 else 0)
        combobox_estado_cliente.setStyleSheet(combobox_style)
        add_linha("Estado", combobox_estado_cliente)

        # Categoria do cliente
        add_linha("Categoria do Cliente")

        # Status do cliente
        combobox_status_cliente = QComboBox()
        combobox_status_cliente.addItems(["Selecionar","Ativo","Inativo","Pendente","Bloqueado"])
        valor_status = dados_cliente.get("Status do Cliente","Selecionar")
        index_status = combobox_status_cliente.findText(valor_status)
        combobox_status_cliente.setCurrentIndex(index_status if index_status >=0 else 0)
        combobox_status_cliente.setStyleSheet(combobox_style)
        add_linha("Status do Cliente", combobox_status_cliente)

        add_linha("Última Atualização")


        # Valor gasto total
        add_linha("Valor Gasto Total")
        self.campos_cliente_juridico["Valor Gasto Total"].setText(dados_cliente.get("Valor Gasto Total",""))

        # Última Compra
        add_linha("Última Compra")
        self.campos_cliente_juridico["Última Compra"].setText(dados_cliente.get("Última Compra",""))

                
        # Conectar formatações
        self.campos_cliente_juridico["CNPJ"].textChanged.connect(
            lambda text: self.main_window.formatar_cnpj(text, self.campos_cliente_juridico["CNPJ"])
        )
        self.campos_cliente_juridico["RG"].textChanged.connect(
            lambda text: self.main_window.formatar_rg(text, self.campos_cliente_juridico["RG"])
        )
        self.campos_cliente_juridico["CPF"].textChanged.connect(
            lambda text: self.main_window.formatar_cpf(text, self.campos_cliente_juridico["CPF"])
        )
        self.campos_cliente_juridico["Email"].textChanged.connect(
            lambda text: self.main_window.validar_email(text, self.campos_cliente_juridico["Email"])
        )
        self.campos_cliente_juridico["CNH"].textChanged.connect(
            lambda text: self.main_window.formatar_cnh(text, self.campos_cliente_juridico["CNH"])
        )
        widget_emissao_cnh.textChanged.connect(
            lambda text: self.main_window.formatar_data_nascimento(text, widget_emissao_cnh)
        )
        widget_vencimento_cnh.textChanged.connect(
            lambda text: self.main_window.formatar_data_nascimento(text, widget_vencimento_cnh)
        )
        self.campos_cliente_juridico["Telefone"].textChanged.connect(
            lambda text: self.main_window.formatar_telefone(text, self.campos_cliente_juridico["Telefone"])
        )
        self.campos_cliente_juridico["CEP"].textChanged.connect(
            lambda text: self.main_window.formatar_cep(text, self.campos_cliente_juridico["CEP"])
        )
        self.campos_cliente_juridico["CEP"].editingFinished.connect(
            lambda: self.formatar_e_buscar_cep(self.campos_cliente_juridico["CEP"])
        )
        self.campos_cliente_juridico["Valor Gasto Total"].editingFinished.connect(
            lambda: self.main_window.formatar_moeda(self.campos_cliente_juridico["Valor Gasto Total"])
        )
        
         # Campos de data conectados corretamente
        campos_data = ["Data de Emissão da CNH", "Data de Vencimento da CNH",
                   "Última Compra", "Data da Inclusão", "Última Atualização"]
        
        for campo in campos_data:
            if campo in self.campos_cliente_juridico:
                widget = self.campos_cliente_juridico[campo]
                widget.textChanged.connect(
                    lambda texto, w=widget: self.main_window.formatar_data_nascimento(texto, w)
                )
                widget.editingFinished.connect(
                    lambda w=widget: self.main_window.validar_data_quando_finalizar(w.text(), w)
                )
        for campo, widget in self.campos_cliente_juridico.items():
            if isinstance(widget, QLineEdit):
                widget.textChanged.connect(self.marcar_alteracao)
            elif isinstance(widget, QComboBox):
                widget.currentTextChanged.connect(self.marcar_alteracao)


        # Botão atualizar
        botao_atualizar = QPushButton("Atualizar")
        botao_atualizar.setStyleSheet(button_style)
        botao_atualizar.clicked.connect(self.atualizar_dados_clientes)
        layout.addWidget(botao_atualizar)

        # Scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(conteudo)
        scroll_area.setStyleSheet(scroll_style)

        self.janela_editar_cliente.setCentralWidget(scroll_area)
        self.janela_editar_cliente.show()
    
                
    def editar_cliente_juridico(self):
        linha_selecionada = self.table_clientes_juridicos.currentRow()
        if linha_selecionada < 0:
            QMessageBox.warning(self, "Aviso", "Nenhum cliente selecionado para edição.")
            return
        
        coluna_offset = 1 if self.coluna_checkboxes_clientes_adicionada else 0

        colunas = [
            "Nome do Cliente", "Razão Social", "Data da Inclusão", "CNPJ","RG","CPF","Email","CNH","Categoria da CNH","Data de Emissão da CNH","Data de Vencimento da CNH","Telefone","CEP", "Endereço", 
            "Número", "Complemento", "Cidade", "Bairro","Estado", "Status do Cliente","Categoria do Cliente",
            "Última Atualização","Valor Gasto Total", "Última Compra"
        ]
        

        dados_cliente = {}
        for col, nome_coluna in enumerate(colunas):
            item = self.table_clientes_juridicos.item(linha_selecionada, col + coluna_offset)
            dados_cliente[nome_coluna] = item.text() if item else ""

        self.exibir_edicao_clientes(dados_cliente)

    
    def exibir_janela_cadastro_cliente(self):
        self.campos_cliente_juridico = {}
        self.informacoes_obrigatorias_cadastro_clientes()

        # Criar a janela
        self.janela_cadastro = QMainWindow()
        self.janela_cadastro.resize(700, 550)
        self.janela_cadastro.setWindowTitle("Cadastro de Cliente Jurídico")
        self.janela_cadastro.setObjectName("janela_cadastro")

        # Centralizar a janela
        screen = QGuiApplication.primaryScreen()
        screen_geometry = screen.availableGeometry()
        window_geometry = self.janela_cadastro.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.janela_cadastro.move(window_geometry.topLeft())

        # Carregar tema
        config = self.temas.carregar_config_arquivo()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(60,60,60),
                                                stop:1 rgb(100,100,100));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid #666666;
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
            """
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """

        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"
            lineedit_bg = "white"

            button_style = """
            QPushButton {
                color: rgb(255, 255, 255);
                border-radius: 8px;
                font-size: 16px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgb(50, 150, 250), stop:1 rgb(100, 200, 255)); /* Gradiente de azul claro para azul mais claro */
                border: 4px solid transparent;
            }

            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgb(100, 180, 255), stop:1 rgb(150, 220, 255)); /* Gradiente de azul mais claro para azul ainda mais claro */
                color: black;
            }
            QPushButton:pressed {
                background-color: #006bb3;
                border: 2px solid #005c99;
            }
            """
            combobox_style = """
            QComboBox {
                background-color: white;
                border: 3px solid rgb(50,150,250);
                border-radius: 5px;
                color: black;
                padding: 5px;
            }

            QComboBox QAbstractItemView {
                background-color: white;
                color: black;
                border: 3px solid white;
                selection-background-color: rgb(120,120,120);
                selection-color: black;
            }

            /* Scrollbar vertical */
            QComboBox QScrollBar:vertical {
                background-color: rgb(240,240,240);  /* Trilha visível */
                width: 10px;
                margin: 1px;
                border: 1px solid white;
                border-radius: 5px;
            }

            /* Alça (handle) da scrollbar */
            QComboBox QScrollBar::handle:vertical {
                background: rgb(120,120,120);  /* Cor da barra frontal,sobe e desce*/
                min-height: 30px;
                border-radius: 5px;
            }

            /* Trilha atrás do handle — essa parte faz toda a diferença */
            QComboBox QScrollBar::add-page:vertical,
            QComboBox QScrollBar::sub-page:vertical {
                background: transparent;  /* barra atrás do scroll */
            }

            /* Esconde setas */
            QComboBox QScrollBar::add-line:vertical,
            QComboBox QScrollBar::sub-line:vertical {
                height: 0px;
                background: none;
                border: none;
            }
            """

            scroll_style = """
                QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: #b4b4b4;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        # --------------------------
        # Widget central e layout
        conteudo = QWidget()
        conteudo.setStyleSheet(f"background-color: {bg_cor}; color: {text_cor};")
        layout = QVBoxLayout(conteudo)

        # Função auxiliar para adicionar campos
        def add_linha(titulo, widget=None):
            label = QLabel(titulo)
            label.setStyleSheet(f"color: {text_cor};")
            layout.addWidget(label)
            if widget is None:
                widget = QLineEdit()
                widget.setStyleSheet(lineedit_style)
            layout.addWidget(widget)
            chave_sem_ponto = titulo.rstrip(":")
            self.campos_cliente_juridico[chave_sem_ponto] = widget
            return label, widget

        
        # ComboBox Categoria CNH criado antecipadamente para uso no formulário 
        combobox_categoria_cnh = QComboBox() 
        combobox_categoria_cnh.addItems(["Selecionar", "AB", "A", "B", "C", "D", "E", "Nenhuma"])
        combobox_categoria_cnh.setStyleSheet(combobox_style)

        # Exemplo de uso
        add_linha("Nome do Cliente")
        add_linha("Razão Social")
        add_linha("CNPJ")
        cnpj_widget = self.campos_cliente_juridico["CNPJ"] 
        cnpj_widget.textChanged.connect(lambda text: self.main_window.formatar_cnpj(text, cnpj_widget))
        add_linha("RG") 
        self.campos_cliente_juridico["RG"].setPlaceholderText("Opcional") 
        rg_widget = self.campos_cliente_juridico["RG"] 
        rg_widget.textChanged.connect(lambda text: self.main_window.formatar_rg(text, rg_widget))
        add_linha("CPF") 
        cpf_widget = self.campos_cliente_juridico["CPF"] 
        cpf_widget.textChanged.connect(lambda text: self.main_window.formatar_cpf(text, cpf_widget)) 
        add_linha("Email") 
        email_widget = self.campos_cliente_juridico["Email"] 
        email_widget.textChanged.connect(lambda text: self.main_window.validar_email(text,email_widget))
        add_linha("CNH") 
        self.campos_cliente_juridico["CNH"].setPlaceholderText("Opcional") 
        label_categoria_cnh, widget_categoria_cnh = add_linha("Categoria da CNH", combobox_categoria_cnh)
        # Campos Data Emissão e Vencimento CNH, inicialmente escondidos 
        label_emissao_cnh, widget_emissao_cnh = add_linha("Data de Emissão da CNH") 
        label_vencimento_cnh, widget_vencimento_cnh = add_linha("Data de Vencimento da CNH")


        # Aplicar formatação de data (mesma usada em Data de Nascimento) 
        widget_emissao_cnh.textChanged.connect(lambda text: self.main_window.formatar_data_nascimento(text, widget_emissao_cnh)) 
        widget_vencimento_cnh.textChanged.connect(lambda text: self.main_window.formatar_data_nascimento(text, widget_vencimento_cnh)) 
        label_emissao_cnh.hide() 
        widget_emissao_cnh.hide() 
        label_vencimento_cnh.hide() 
        widget_vencimento_cnh.hide()

        cnh_widget = self.campos_cliente_juridico["CNH"] 
        cnh_widget.textChanged.connect(lambda text: self.main_window.formatar_cnh(text, cnh_widget))

        # Função para mostrar/esconder datas da CNH conforme categoria selecionada 
        def on_categoria_cnh_change(text): 
            if text not in ("Selecionar", "Nenhuma"): 
                label_emissao_cnh.show() 
                widget_emissao_cnh.show() 
                label_vencimento_cnh.show() 
                widget_vencimento_cnh.show() 
            else: 
                label_emissao_cnh.hide() 
                widget_emissao_cnh.hide() 
                label_vencimento_cnh.hide() 
                widget_vencimento_cnh.hide() 
                widget_emissao_cnh.clear() 
                widget_vencimento_cnh.clear()

        combobox_categoria_cnh.currentTextChanged.connect(on_categoria_cnh_change)

        add_linha("Telefone") 
        telefone_widget = self.campos_cliente_juridico["Telefone"] 
        telefone_widget.textChanged.connect(lambda text: self.main_window.formatar_telefone(text, telefone_widget))

        add_linha("CEP") 
        cep_widget = self.campos_cliente_juridico["CEP"] 
        cep_widget.textChanged.connect(lambda text: self.main_window.formatar_cep(text, cep_widget)) 
        cep_widget.editingFinished.connect(lambda: self.on_cep_editing_finished_cadastro(cep_widget)) 
        add_linha("Endereço") 
        add_linha("Número") 
        add_linha("Complemento")

        self.campos_cliente_juridico["Complemento"].setPlaceholderText("Opcional") 
        add_linha("Cidade") 
        add_linha("Bairro")
        

        # ComboBox Estado 
        combobox_estado_cliente = QComboBox() 
        combobox_estado_cliente.addItems([ "Selecionar","AC","AL","AP","AM","BA","CE","DF","ES","GO",
                                          "MA","MT", "MS","MG","PA","PB","PR","PE","PI","RJ","RN",
                                          "RS","RO","RR","SC","SP","SE","TO" ])
        
        combobox_estado_cliente.setCurrentIndex(0) 
        combobox_estado_cliente.setStyleSheet(combobox_style)
            
        add_linha("Estado", combobox_estado_cliente) 
        add_linha("Categoria do Cliente")

        combobox_status_cliente = QComboBox()

        combobox_status_cliente.addItems(["Selecionar","Ativo","Inativo","Pendente","Bloqueado"]) 
        combobox_status_cliente.setCurrentIndex(0) 
        combobox_status_cliente.setStyleSheet(combobox_style) 
        add_linha("Status do Cliente:", combobox_status_cliente)                                  

        # Botão de cadastro
        btn_fazer_cadastro = QPushButton("Fazer o Cadastro")
        btn_fazer_cadastro.setStyleSheet(button_style)
        btn_fazer_cadastro.clicked.connect(self.cadastrar_clientes_juridicos)
        layout.addWidget(btn_fazer_cadastro)


        # Scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(conteudo)
        scroll_area.setStyleSheet(scroll_style)

        self.janela_cadastro.setCentralWidget(scroll_area)
        self.janela_cadastro.show()


    def informacoes_obrigatorias_cadastro_clientes(self):
        """Valida campos obrigatórios apenas no cadastro"""
        self.campos_obrigatorios_clientes = {
            "Nome do Cliente": "O campo Nome do Cliente é obrigatório.",
            "Razão Social": "O campo de Razão Social é obrigatório.",
            "CNPJ": "O campo CNPJ é obrigatório.",
            "RG": "O campo de RG é obrigatório.",
            "CPF": "O campo de CPF é obrigatório.",
            "Email": "O campo de E-mail é obrigatório.",
            "Telefone": "O campo Telefone é obrigatório.",
            "CEP": "O campo CEP é obrigatório.",
            "Endereço": "O campo Endereço é obrigatório.",
            "Número": "O campo Número é obrigatório.",
            "Cidade": "O campo Cidade é obrigatório.",
            "Bairro": "O campo Bairro é obrigatório.",
            "Estado": "O campo de Estado é obrigatório.",
            "Categoria do Cliente": "O campo Categoria do Cliente é obrigatório.",
            "Status do Cliente": "Você deve selecionar um status válido para o cliente."
        }

        # Se CNH foi informada, exige os outros campos relacionados
        cnh_widget = self.campos_cliente_juridico.get("CNH")
        cnh_valor = ""
        if isinstance(cnh_widget, QLineEdit):
            cnh_valor = cnh_widget.text().strip()
        elif isinstance(cnh_widget, QComboBox):
            cnh_valor = cnh_widget.currentText().strip()
                
        if cnh_valor and cnh_valor.lower() not in ["não cadastrado", "selecione"]:
            self.campos_obrigatorios_clientes.update({
                "Categoria da CNH": "Informe a Categoria da CNH.",
                "Data de Emissão da CNH": "Informe a Data de Emissão da CNH.",
                "Data de Vencimento da CNH": "Informe a Data de Vencimento da CNH."
            })

        # Validação genérica dos campos
        for campo, mensagem_erro in self.campos_obrigatorios_clientes.items():
            widget = self.campos_cliente_juridico.get(campo)
            if widget is None:
                continue
            
            valor = ""
            if isinstance(widget, QLineEdit):
                valor = widget.text().strip()
            elif isinstance(widget, QComboBox):
                valor = widget.currentText().strip()
                if valor.lower() in ["selecione", "selecionar"]:
                    valor = ""
            else:
                valor = str(widget).strip()

            print(f"[Cadastro] Validando campo '{campo}': valor '{valor}'")  # debug

        return True


    def informacoes_obrigatorias_edicao_clientes(self):
        """Valida também campos automáticos na edição"""
        # Primeiro roda a validação padrão de cadastro
        self.informacoes_obrigatorias_cadastro_clientes()

        # Adiciona os extras que só são exigidos na edição
        self.campos_obrigatorios_clientes.update({
            "Última Atualização": "O campo Última Atualização é obrigatório.",
            "Data da Inclusão": "O campo Data da Inclusão é obrigatório.",
            "Valor Gasto Total": "O campo Valor Gasto Total é obrigatório.",
            "Última Compra": "O campo Última Compra é obrigatório."
        })

        # Validação genérica novamente
        for campo, mensagem_erro in self.campos_obrigatorios_clientes.items():
            widget = self.campos_cliente_juridico.get(campo)
            if widget is None:
                continue

            valor = ""
            if isinstance(widget, QLineEdit):
                valor = widget.text().strip()
            elif isinstance(widget, QComboBox):
                valor = widget.currentText().strip()
                if valor.lower() in ["selecione", "selecionar"]:
                    valor = ""
            else:
                valor = str(widget).strip()

            print(f"[Edição] Validando campo '{campo}': valor '{valor}'")  # debug

        return True


    def atualizar_dados_clientes(self):
        if not self.informacoes_obrigatorias_edicao_clientes():
            return
        
        # Carregar tema
        config = self.temas.carregar_config_arquivo()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """
        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        else: # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"
            lineedit_bg = "white"

            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """


        # --- VALIDAÇÃO DOS CAMPOS OBRIGATÓRIOS ---
        for campo, mensagem in self.campos_obrigatorios_clientes.items():
            widget = self.campos_cliente_juridico[campo]
            if isinstance(widget, QLineEdit):
                valor = widget.text().strip()
            elif isinstance(widget, QComboBox):
                valor = widget.currentText().strip()
            else:
                valor = str(widget).strip()

            if not valor or (isinstance(widget, QComboBox) and valor.lower() == "selecionar"):
                QMessageBox.warning(self, "Atenção", mensagem)
                return


        if not self.alteracoes_realizadas:
            QMessageBox.information(self, "Sem alterações", "Nenhuma modificação foi feita.")
            return
        
        # --- COLETA DOS DADOS ---
        dados_atualizados = {}
        for campo, widget in self.campos_cliente_juridico.items():
            if isinstance(widget, QLineEdit):
                dados_atualizados[campo] = widget.text()
            elif isinstance(widget, QComboBox):
                dados_atualizados[campo] = widget.currentText()
            else:
                dados_atualizados[campo] = str(widget)

        # --- VERIFICAÇÃO DE CAMPOS SENSÍVEIS ---
        campos_sensiveis = ["Data da Inclusão","Última Atualização", "Valor Gasto Total", "Última Compra"]
        alterou_campo_sensivel = any(
            dados_atualizados[c] != self.dados_originais_cliente[c]
            for c in campos_sensiveis
        )

        if alterou_campo_sensivel:
            try:
                with open("config.json", "r", encoding="utf-8") as f:
                    config = json.load(f)
                    senha_correta = config.get("senha", "")
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Não foi possível carregar a senha do sistema\n {e}")
                return

            tentativas = 0
            while tentativas < 3:
                dialogo = DialogoSenha(self)
                if dialogo.exec() == QDialog.Accepted:
                    senha_digitada = dialogo.get_senha().strip()
                    if senha_digitada == senha_correta.strip():
                        break  # Senha correta
                    else:
                        tentativas += 1
                        if tentativas < 3:
                            QMessageBox.critical(self, "Acesso Negado",
                                                f"Senha incorreta. Tentativas restantes: {3 - tentativas}")
                        else:
                            QMessageBox.critical(self, "Acesso Negado",
                                                "Você excedeu o número máximo de tentativas.\nO sistema será encerrado e á atualização não será feita.")
                            QApplication.quit()
                            return
                else:
                    return  # Cancelou

        try:
            cursor = self.db.connection.cursor()
            cnpj = dados_atualizados["CNPJ"]  # Identificador único

            # --- VALORES PADRÃO PARA CAMPOS NÃO PREENCHIDOS ---
            if not dados_atualizados["CNH"]:
                dados_atualizados["CNH"] = "Não Cadastrado"
                dados_atualizados["Categoria da CNH"] = "Não Cadastrado"
                dados_atualizados["Data de Emissão da CNH"] = "Não Cadastrado"
                dados_atualizados["Data de Vencimento da CNH"] = "Não Cadastrado"

            if not dados_atualizados["Complemento"]:
                dados_atualizados["Complemento"] = "Não se aplica"

            query = """
                UPDATE clientes_juridicos SET
                    `Nome do Cliente` = ?, `Razão Social` = ?, `Data da Inclusão` = ?, CNPJ = ?,
                    RG = ?, CPF = ?, CNH = ?, `Categoria da CNH` = ?, `Data de Emissão da CNH` = ?, 
                    `Data de Vencimento da CNH` = ?, Telefone = ?, CEP = ?, Endereço = ?, 
                    Número = ?, Complemento = ?, Cidade = ?, Bairro = ?, Estado = ?, `Status do Cliente` = ?, `Categoria do Cliente` = ?,
                    `Última Atualização` = ?, `Valor Gasto Total` = ?, `Última Compra` = ?
                WHERE CNPJ = ?
            """
            valores = (
                dados_atualizados["Nome do Cliente"], dados_atualizados["Razão Social"], dados_atualizados["Data da Inclusão"], dados_atualizados["CNPJ"],
                dados_atualizados["RG"], dados_atualizados["CPF"], dados_atualizados["CNH"], dados_atualizados["Categoria da CNH"], dados_atualizados["Data de Emissão da CNH"], 
                dados_atualizados["Data de Vencimento da CNH"], dados_atualizados["Telefone"], dados_atualizados["CEP"], dados_atualizados["Endereço"], dados_atualizados["Número"], dados_atualizados["Complemento"],
                dados_atualizados["Cidade"], dados_atualizados["Bairro"], dados_atualizados["Estado"], dados_atualizados["Status do Cliente"], dados_atualizados["Categoria do Cliente"],
                dados_atualizados["Última Atualização"], dados_atualizados["Valor Gasto Total"], dados_atualizados["Última Compra"],
                cnpj
            )

            cursor.execute(query, valores)
            self.db.connection.commit()

            self.main_window.registrar_historico_clientes_juridicos(
                "Edição de Clientes",
                f"Cliente {dados_atualizados['Nome do Cliente']} editado com sucesso"
            )

            QMessageBox.information(self, "Sucesso", "Cliente atualizado com sucesso.")
            self.carregar_clientes_juridicos()
            self.janela_editar_cliente.close()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao atualizar cliente: {e}")

    def marcar_alteracao(self):
        self.alteracoes_realizadas = True
    
    def cadastrar_clientes_juridicos(self):
        self.informacoes_obrigatorias_cadastro_clientes()
        try:
            with self.db.connecta() as conexao:
                cursor = conexao.cursor()
                usuario_logado = self.config.obter_usuario_logado()

                # Coletar dados dos campos
                get = lambda campo: self.campos_cliente_juridico[campo].text().strip() \
                    if isinstance(self.campos_cliente_juridico[campo], QLineEdit) \
                    else self.campos_cliente_juridico[campo].currentText()

                nome = get("Nome do Cliente")
                razao_social = get("Razão Social")
                cnpj = get("CNPJ")
                rg = get("RG")
                cpf = get("CPF")
                email = get("Email")
                cnh = get("CNH")
                categoria_cnh = get("Categoria da CNH")
                data_emissao_cnh = get("Data de Emissão da CNH")
                data_vencimento_cnh = get("Data de Vencimento da CNH")
                telefone = get("Telefone")
                cep = get("CEP")
                endereco = get("Endereço")
                numero = get("Número")
                complemento = get("Complemento")
                cidade = get("Cidade")
                bairro = get("Bairro")
                estado = get("Estado")
                categoria = get("Categoria do Cliente")
                status = get("Status do Cliente")


                # Verificar campos únicos individualmente
                cursor.execute('SELECT 1 FROM clientes_juridicos WHERE CNPJ = ?',(cnpj,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com este CNPJ.")
                    return
                cursor.execute('SELECT 1 FROM clientes_juridicos WHERE "Razão Social" = ?',(razao_social,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com esta Razão Social.")
                    return
                cursor.execute('SELECT 1 FROM clientes_juridicos WHERE Telefone = ?',(telefone,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com este Telefone.")
                    return
                cursor.execute('SELECT 1 FROM clientes_juridicos WHERE RG = ?',(rg,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com este RG.")
                    return
                
                cursor.execute('SELECT 1 FROM clientes_juridicos WHERE CPF = ?',(cpf,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com este CPF.")
                    return
                
                cursor.execute('SELECT 1 FROM clientes_juridicos WHERE Email = ?',(email,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com este Email.")
                    return
                
                cursor.execute('SELECT 1 FROM clientes_juridicos WHERE CNH = ?',(cnh,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com esta CNH.")
                    return

                # Validação de todos os campos obrigatórios
                for campo, mensagem in self.campos_obrigatorios_clientes.items():
                    widget = self.campos_cliente_juridico[campo]
                    valor = widget.text().strip() if isinstance(widget, QLineEdit) else widget.currentText()
                    
                    if not valor or (isinstance(widget, QComboBox) and valor == "Selecionar"):
                        QMessageBox.warning(self, "Atenção", mensagem)
                        return
                    

                data_inclusao = datetime.now().strftime("%d/%m/%Y %H:%M")
                
                # Valores padrão para os campos não preenchidos manualmente
                valor_gasto_total = "Não Cadastrado"
                ultima_compra = "Não Cadastrado"
                ultima_atualizacao = "Não Cadastrado"

                if not cnh:
                    cnh = "Não Cadastrado"
                    categoria_cnh = "Não Cadastrado"
                    data_emissao_cnh = "Não Cadastrado"
                    data_vencimento_cnh = "Não Cadastrado"

                if not complemento:
                    complemento = "Não se aplica"
                
                # Inserir no banco
                cursor.execute("""
                    INSERT INTO clientes_juridicos(
                        "Nome do Cliente", "Razão Social", "Data da Inclusão", CNPJ, RG, 
                         CPF,Email, CNH, "Categoria da CNH", "Data de Emissão da CNH", "Data de Vencimento da CNH",  
                        Telefone, CEP, Endereço, Número, Complemento, Cidade, Bairro, Estado, 
                        "Status do Cliente", "Categoria do Cliente", "Última Atualização", "Valor Gasto Total", "Última Compra"
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?,?)
                """, (
                    nome, razao_social, data_inclusao, cnpj, rg, 
                    cpf,email, cnh, categoria_cnh, data_emissao_cnh, data_vencimento_cnh, 
                    telefone, cep, endereco, numero, complemento, cidade, bairro, estado, 
                    status, categoria, ultima_atualizacao, valor_gasto_total,ultima_compra
                ))
                conexao.commit()

                self.main_window.registrar_historico_clientes_juridicos(
                    "Cadastro de Cliente", f"Cliente {nome} cadastrado com sucesso"
                )
                
                QMessageBox.information(self, "Sucesso", "Cliente cadastrado com sucesso!")
                # Redimensiona apenas uma vez após preencher
                self.table_clientes_juridicos.resizeColumnsToContents()
                self.table_clientes_juridicos.resizeRowsToContents()
                self.limpar_campos_clientes()
                self.carregar_clientes_juridicos()
                

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao cadastrar cliente: \n{e}")

        
    def limpar_campos_clientes(self):
        for campo, widget in self.campos_cliente_juridico.items():
            if isinstance(widget,QLineEdit):
                widget.clear()
            elif isinstance(widget, QComboBox):
                widget.setCurrentIndex(0) # Volta para "Selecionar"
                    
    def on_cep_editing_finished_cadastro(self, cep_widget):
        cep_digitado = cep_widget.text()
        dados_cep = self.main_window.buscar_cep(cep_digitado)
        if dados_cep:
            self.preencher_campos_cep_cadastro(dados_cep)

    def preencher_campos_cep_cadastro(self, dados):
        if dados is None:
            return

        # Preencher os campos do seu formulário usando self.campos_cliente_juridico
        self.campos_cliente_juridico["Endereço"].setText(dados.get("logradouro", ""))
        self.campos_cliente_juridico["Bairro"].setText(dados.get("bairro", ""))
        self.campos_cliente_juridico["Cidade"].setText(dados.get("localidade", ""))

        complemento = dados.get("complemento", "")
        if any(char.isdigit() for char in complemento):
            self.campos_cliente_juridico["Número"].setText(complemento)

        estado = dados.get("uf", "")
        # Se tiver um combobox para estado, aqui você ajusta o índice
        if "Estado" in self.campos_cliente_juridico:
            estado_combobox = self.campos_cliente_juridico["Estado"]
            index_estado = estado_combobox.findText(estado)
            if index_estado != -1:
                estado_combobox.setCurrentIndex(index_estado)
                
    def excluir_clientes_juridicos(self):
        try:
            total_linhas = self.table_clientes_juridicos.rowCount()
            if total_linhas == 0:
                QMessageBox.warning(self, "Aviso", "Nenhum cliente encontrado para excluir.")
                return

            clientes_para_excluir = []

            if self.coluna_checkboxes_clientes_adicionada:
                # Modo com checkbox
                for linha in range(total_linhas):
                    checkbox_widget = self.table_clientes_juridicos.cellWidget(linha, 0)
                    if checkbox_widget:
                        checkbox = checkbox_widget.findChild(QCheckBox)
                        if checkbox and checkbox.isChecked():
                            nome_cliente = self.table_clientes_juridicos.item(linha, 1).text()
                            clientes_para_excluir.append((linha, nome_cliente))
            else:
                # Modo sem checkbox (seleção direta)
                linha_selecionadas = self.table_clientes_juridicos.selectionModel().selectedRows()
                if not linha_selecionadas:
                    QMessageBox.information(self, "Aviso", "Nenhum cliente selecionado para exclusão.")
                    return
                
                for index in linha_selecionadas:
                    linha = index.row()
                    nome_cliente = self.table_clientes_juridicos.item(linha,0).text()
                    clientes_para_excluir.append((linha,nome_cliente))

            if not clientes_para_excluir:
                QMessageBox.information(self, "Aviso", "Nenhum cliente selecionado para exclusão.")
                return
            
            # Mensagem personalizada
            if len(clientes_para_excluir) == 1:
                _, nome = clientes_para_excluir[0]
                mensagem = f"Tem certeza que deseja excluir o cliente:\n\n• {nome}?"
            else:
                nomes = "\n• " + "\n• ".join(nome for _, nome in clientes_para_excluir)
                mensagem = f"Tem certeza que deseja excluir os seguintes clientes?\n{nomes}"

            # Criar QMessageBox manualmente para alterar os textos dos botões
            msgbox = QMessageBox(self)
            msgbox.setIcon(QMessageBox.Question)
            msgbox.setWindowTitle("Confirmar Exclusão")
            msgbox.setText(mensagem)
            msgbox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)

            # Renomear os botões
            botao_sim = msgbox.button(QMessageBox.Yes)
            botao_nao = msgbox.button(QMessageBox.No)
            botao_sim.setText("Sim")
            botao_nao.setText("Não")

            resposta = msgbox.exec()

            if resposta != QMessageBox.Yes:
                return
            
            # Executa exclusão no banco e remove as linhas da tabela
            db = DataBase()
            db.connecta()
            cursor = db.connection.cursor()

            for linha, nome_cliente in reversed(clientes_para_excluir):
                cursor.execute("""
                    DELETE FROM clientes_juridicos WHERE "Nome do Cliente" = ?
                """, (nome_cliente,))
                self.table_clientes_juridicos.removeRow(linha)

            db.connection.commit()
            
            nomes_excluidos = ", ".join(nome for _, nome in clientes_para_excluir)
            self.main_window.registrar_historico_clientes_juridicos(
                "Exclusão de Cliente(s)",
                f"Cliente(s) excluído(s): {nomes_excluidos}"
            )
            QMessageBox.information(self.main_window, "Sucesso", "Cliente(s) excluído(s) com sucesso.")
        except Exception as e:
            QMessageBox.critical(self.main_window, "Erro", f"Erro ao excluir clientes:\n{e}")
            
    def marcar_como_clientes(self):
        if self.table_clientes_juridicos.rowCount() == 0:
            QMessageBox.warning(self,"Aviso","Nenhum cliente cadastrado para selecionar.")
             # Desmarca o checkbox header visualmente
            if hasattr(self, "checkbox_header_clientes") and isinstance(self.checkbox_header_clientes, QCheckBox):
                QTimer.singleShot(0, lambda: self.checkbox_header_clientes.setChecked(False))

            return  # Impede que o restante da função execute!

        if self.coluna_checkboxes_clientes_adicionada:
            # Animação de saída (fade out)
            self.animar_coluna_checkbox(mostrar=False)
            QTimer.singleShot(300, lambda: self.remover_coluna_checkboxes())
            return

        # Adiciona a coluna de checkboxes
        self.table_clientes_juridicos.insertColumn(0)
        self.table_clientes_juridicos.setHorizontalHeaderItem(0, QTableWidgetItem(""))
        self.table_clientes_juridicos.setColumnWidth(0, 30)
        self.table_clientes_juridicos.horizontalHeader().setMinimumSectionSize(30)
        self.table_clientes_juridicos.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)

        # Carregar tema
        config = self.temas.carregar_config_arquivo()
        tema = config.get("tema", "claro")


        if tema == "escuro":
            checkbox_style = """
            QCheckBox {
                background-color: transparent;
                border: none;
            }
            QCheckBox::indicator {
                background-color: white;
            }
            QToolTip {
                background-color: #2e2e2e;
                color: white;
                border: 1px solid #555555;
            }
            """
        elif tema == "claro":
            checkbox_style = """
            QCheckBox {
                background-color: transparent;
                border: none;
            }
            QCheckBox::indicator {
                background-color: transparent;
                border: 1px solid gray;
                width: 11px;
                height: 11px;
            }

            QCheckBox::indicator:checked {
                background-color: #0078d7;  /* azul para tema claro */
                border: 1px solid #0078d7;
                image: url(imagens/visto_branco.png);
            }
            QToolTip {
                background-color: white;
                color: black;
                border: 1px solid gray;
            }
            """
        else: # classico
            checkbox_style = """
            QCheckBox {
                background-color: transparent;
                border: none;
            }
            QCheckBox::indicator {
                background-color: transparent;
                border: 1px solid gray;
                border-radius: 2px;
                width: 11px;
                height: 11px;
            }
            QCheckBox::indicator:checked {
                background-color: #0078d7;  /* azul para tema claro */
                border: 1px solid #0078d7;
                image: url(imagens/visto_branco.png);
            }

            QToolTip {
                background-color: white;
                color: black;
                border: 1px solid gray;
            }
            """

        # Checkbox no cabeçalho
        header_clientes_juridicos = self.table_clientes_juridicos.horizontalHeader()
        self.checkbox_header_clientes = QCheckBox(header_clientes_juridicos.viewport())
        self.checkbox_header_clientes.setToolTip("Selecionar todos")
        self.checkbox_header_clientes.setChecked(False)
        self.checkbox_header_clientes.stateChanged.connect(self.selecionar_todos_clientes)
        self.checkbox_header_clientes.setFixedSize(20, 20)
        self.checkbox_header_clientes.setStyleSheet(checkbox_style)
        self.checkbox_header_clientes.show()

        QTimer.singleShot(0, self.atualizar_posicao_checkbox_header_clientes)

        for linha in range(self.table_clientes_juridicos.rowCount()):
            checkbox = QCheckBox()
            checkbox.stateChanged.connect(self.atualizar_selecao_todos_clientes)

            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(checkbox)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)

            self.table_clientes_juridicos.setCellWidget(linha, 0, container)
            self.checkboxes_clientes.append(checkbox)

        
        header_clientes_juridicos.sectionResized.connect(self.atualizar_posicao_checkbox_header_clientes)
        self.table_clientes_juridicos.horizontalScrollBar().valueChanged.connect(self.atualizar_posicao_checkbox_header_clientes)
        header_clientes_juridicos.geometriesChanged.connect(self.atualizar_posicao_checkbox_header_clientes)

        self.coluna_checkboxes_clientes_adicionada = True
        
        # Animação de entrada (fade in)
        self.animar_coluna_checkbox(mostrar=True)
        
    def animar_coluna_checkbox(self, mostrar=True):
        self.animacoes_clientes = []

        for row in range(self.table_clientes_juridicos.rowCount()):
            widget = self.table_clientes_juridicos.cellWidget(row, 0)
            if widget:
                efeito = QGraphicsOpacityEffect(widget)
                widget.setGraphicsEffect(efeito)

                # Cria animação corretamente
                anim = QPropertyAnimation(efeito, b"opacity")
                anim.setDuration(300)
                anim.setStartValue(0.0 if mostrar else 1.0)
                anim.setEndValue(1.0 if mostrar else 0.0)
                anim.start()

                self.animacoes_clientes.append(anim)

                
    def remover_coluna_checkboxes(self):
        self.table_clientes_juridicos.removeColumn(0)
        self.table_clientes_juridicos.verticalHeader().setVisible(True)
        self.coluna_checkboxes_clientes_adicionada = False

        if hasattr(self, "checkbox_header_clientes"):
            self.checkbox_header_clientes.setChecked(False)
            self.checkbox_header_clientes.deleteLater()
            del self.checkbox_header_clientes

        self.checkboxes_clientes.clear()

        
    def selecionar_todos_clientes(self):
        # Evita erro ao clicar quando a coluna já foi removida
        if not getattr(self, "coluna_checkboxes_clientes_adicionada", False):
            return  # Simplesmente sai sem mostrar mensagem

        estado = self.checkbox_header_clientes.checkState() == Qt.Checked
        self.checkboxes_clientes.clear()

        for row in range(self.table_clientes_juridicos.rowCount()):
            widget = self.table_clientes_juridicos.cellWidget(row, 0)
            if widget is not None:
                checkbox = widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.blockSignals(True)
                    checkbox.setChecked(estado)
                    checkbox.blockSignals(False)
                    self.checkboxes_clientes.append(checkbox)


    def atualizar_selecao_todos_clientes(self):
        self.checkbox_header_clientes.blockSignals(True)

        all_checked = all(cb.isChecked() for cb in self.checkboxes_clientes if cb)
        any_checked = any(cb.isChecked() for cb in self.checkboxes_clientes if cb)

        if all_checked:
            self.checkbox_header_clientes.setCheckState(Qt.Checked)
        elif any_checked:
            self.checkbox_header_clientes.setCheckState(Qt.PartiallyChecked)
        else:
            self.checkbox_header_clientes.setCheckState(Qt.Unchecked)

        self.checkbox_header_clientes.blockSignals(False)

    def atualizar_posicao_checkbox_header_clientes(self):
        if hasattr(self, "checkbox_header_clientes") and self.coluna_checkboxes_clientes_adicionada:
            header = self.table_clientes_juridicos.horizontalHeader()
            
            x = header.sectionViewportPosition(0) + (header.sectionSize(0) - self.checkbox_header_clientes.width()) // 2 + 4.3
            y = (header.height() - self.checkbox_header_clientes.height()) // 2
            self.checkbox_header_clientes.move(x, y)
            
    # Limpa a coluna selecionada clicando em qualquer lugar da tabela
    def eventFilter(self, source, event):
        if event.type() == QEvent.MouseButtonPress:
            if source == self.table_clientes_juridicos.viewport():
                index = self.table_clientes_juridicos.indexAt(event.pos())
                if not index.isValid():
                    self.table_clientes_juridicos.clearSelection() # remove a linha selecinada da tabela
                    self.table_clientes_juridicos.clearFocus() # remove o foco da tabela      
        return super().eventFilter(source,event)
    
    
    


    def historico_clientes_juridicos(self):
        self.janela_historico_clientes = QMainWindow()
        self.janela_historico_clientes.resize(800,650)
        self.janela_historico_clientes.setWindowTitle("Histórico de Clientes Jurídico")
        self.janela_historico_clientes.setObjectName("janela_historico_clientes")

        # Centralizar a janela na tela
        screen = QGuiApplication.primaryScreen()
        screen_geometry = screen.availableGeometry()
        window_geometry = self.janela_historico_clientes.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.janela_historico_clientes.move(window_geometry.topLeft())


        # Criação do layout e tabela para exibir o histórico
        central_widget = QWidget()
        layout = QVBoxLayout(central_widget)

        # Carregar tema
        config = self.temas.carregar_config_arquivo()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(60,60,60),
                                                stop:1 rgb(100,100,100));
                    color: white;
                    border-radius: 8px;
                    font-size: 12px;
                    border: 2px solid #666666;
                    padding: 3px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
            """
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            table_view_style = """
            /* QTableView com seleção diferenciada */
            QTableView {
                background-color: #202124;
                color: white;
                gridline-color: #555555;
                selection-background-color: #7a7a7a;
                selection-color: white;
            }
            /* Coluna dos cabeçalhos */
            QHeaderView::section {
                background-color: #202124;
                color: white;
                border: 1px solid #aaaaaa;
                padding: 1px;
            }

            /* QTabWidget headers brancos */
            QTabWidget::pane {
                border: 1px solid #444444;
                background-color: #202124;
            }
            /* Estiliza a barra de rolagem horizontal */
            QTableView QScrollBar:horizontal {
                border: none;
                background-color: none;
                height: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* Estiliza a barra de rolagem vertical */
            QTableView QScrollBar:vertical {
                border: none;
                background-color: none; 
                width: 12px;
                margin: 0px;
                border-radius: 5px;
            }
            

            /* Parte que você arrasta */
            QTableView QScrollBar::handle:vertical {
                background-color: #777777;  /* cinza médio */
                min-height: 22px;
                border-radius: 5px;
            }

            QTableView QScrollBar::handle:horizontal {
                background-color: #777777;
                min-width: 22px;
                border-radius: 5px;
            }

            /* Groove horizontal */
            QTableView QScrollBar::groove:horizontal {
                background-color: transparent;
                border-radius: 5px;
                height: 15px;
                margin: 0px 10px 0px 10px;
            }

            /* Groove vertical */
            QTableView QScrollBar::groove:vertical {
                background-color: transparent;
                border-radius: 5px;
                width: 25px;
                margin: 10px 0px 10px 10px;
            }

            /* Estilo para item selecionado */
            QTableWidget::item:selected {
                background-color: #555555;  /* cinza de seleção */
                color: white;
            }
            QTableCornerButton::section {
                background-color: #202124;  /* mesma cor da tabela */
                border: none;
            }

            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """

        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            scroll_style = """
                QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }

                QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }

                QScrollBar::add-line:vertical,
                QScrollBar::sub-line:vertical,
                QScrollBar::add-page:vertical,
                QScrollBar::sub-page:vertical {
                    background: none;
                    height: 0px;
                }
            """
            table_view_style = """
                QTableView {
                    background-color: white;
                    color: black;
                    gridline-color: #ccc;
                    selection-background-color: #e5f3ff;
                    selection-color: black;
                }

                QHeaderView::section {
                    background-color: #f0f0f0;
                    color: black;
                    border: 1px solid #ccc;
                    padding: 1px;
                }

                QTabWidget::pane {
                    border: 1px solid #ccc;
                    background-color: white;
                }

                QTableView QScrollBar:horizontal,
                QTableView QScrollBar:vertical {
                    border: none;
                    background-color: #f5f5f5;
                    border-radius: 5px;
                }

                QTableView QScrollBar::handle:vertical,
                QTableView QScrollBar::handle:horizontal {
                    background-color: #cccccc;
                    min-height: 22px;
                    border-radius: 5px;
                }

                QTableView QScrollBar::groove:horizontal {
                    background-color: transparent;
                    border-radius: 5px;
                    height: 15px;
                    margin: 0px 10px 0px 10px;
                }

                QTableView QScrollBar::groove:vertical {
                    background-color: transparent;
                    border-radius: 5px;
                    width: 25px;
                    margin: 10px 0px 10px 10px;
                }

                QTableWidget::item:selected {
                    background-color: #cce7ff;
                    color: black;
                }

                QTableCornerButton::section {
                    background-color: #f0f0f0;
                    border: 1px solid #ccc;
                }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"

            button_style = """
                QPushButton {
                    color: rgb(255, 255, 255);
                    border-radius: 8px;
                    font-size: 16px;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgb(50, 150, 250), stop:1 rgb(100, 200, 255)); /* Gradiente de azul claro para azul mais claro */
                    border: 4px solid transparent;
                }

                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgb(100, 180, 255), stop:1 rgb(150, 220, 255)); /* Gradiente de azul mais claro para azul ainda mais claro */
                    color: black;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 3px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            scroll_style = """
                QScrollBar:vertical {
                border: none;
                background-color: rgb(255, 255, 256); /* branco */
                width: 30px;
                margin: 0px 10px 0px 10px;
            }
             QScrollBar::handle:vertical {
                background-color: rgb(180, 180,180);  /* cinza */
                min-height: 30px;
                border-radius: 5px;
            }
            """
            table_view_style = """
                QTableView {
                    background-color: rgb(0,80,121);
                    color: white;
                    gridline-color: black;
                    border: 1px solid white;
                    selection-background-color: #007acc;
                    selection-color: white;
                }

                QHeaderView::section {
                    background-color: white;
                    color: black;
                    border: 1px solid #eeeeee;  /* Borda branco-acinzentada */
                    padding: 1px;
                }

                QTabWidget::pane {
                    border: 1px solid #004466;
                    background-color: #003355;
                }

                /* Scrollbars da QTableView - vertical e horizontal */
                QTableView QScrollBar:vertical,
                QTableView QScrollBar:horizontal {
                    border: none;
                    background-color: rgb(255, 255, 255); /* fundo do track */
                    border-radius: 5px;
                    width: 10px; /* largura da barra vertical */
                    margin: 0px;
                }

                /* Handle dos scrolls (a parte que você arrasta) */
                QTableView QScrollBar::handle:vertical,
                QTableView QScrollBar::handle:horizontal {
                    background-color: rgb(180, 180, 150);  /* cor do handle */
                    min-height: 10px;
                    min-width: 10px;
                    border-radius: 5px;
                }
                /* Groove vertical */
                QTableView QScrollBar::groove:vertical {
                    background-color: rgb(100, 240, 240);  /* faixa visível no vertical */
                    border-radius: 2px;
                    width: 10px;
                    margin: 0px 10px 0px 10px;
                }

                /* Groove horizontal (faixa por onde o handle desliza) */
                QTableView QScrollBar::groove:horizontal {
                    background-color: rgb(100, 240, 240);  /* faixa visível no horizontal */
                    border-radius: 2px;
                    height: 15px;
                    margin: 0px 10px 0px 10px;
                }

                QTableWidget::item:selected {
                    background-color: rgb(0, 120, 215);
                    color: white;
                }

                QTableCornerButton::section {
                    background-color: white;
                }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        # Tabela do histórico
        self.tabela_historico_clientes = QTableWidget()
        self.tabela_historico_clientes.setColumnCount(4)
        self.tabela_historico_clientes.setHorizontalHeaderLabels(["Data/Hora", "Usuário", "Ação", "Descrição"])
        # Ocultar o quadrado do canto
        self.tabela_historico_clientes.setCornerButtonEnabled(False)

        # Botão Atualizar
        botao_atualizar = QPushButton("Atualizar Histórico")
        botao_atualizar.clicked.connect(self.atualizar_historico_clientes_jurididos)
        botao_atualizar.setStyleSheet(button_style)

        # Botão Apagar
        botao_apagar = QPushButton("Apagar Histórico")
        botao_apagar.clicked.connect(self.apagar_historico_cliente_juridicos)
        botao_apagar.setStyleSheet(button_style)

        # Botão Exportar CSV
        botao_exportar_csv = QPushButton("Exportar para CSV")
        botao_exportar_csv.clicked.connect(self.exportar_csv_juridicos)
        botao_exportar_csv.setStyleSheet(button_style)

        
        # Botão Exportar Excel
        botao_exportar_excel = QPushButton("Exportar para Excel")
        botao_exportar_excel.clicked.connect(self.exportar_excel_juridicos)
        botao_exportar_excel.setStyleSheet(button_style)

        # Botão PDF
        botao_exportar_pdf = QPushButton("Exportar PDF")
        botao_exportar_pdf.clicked.connect(self.exportar_pdf_juridicos)
        botao_exportar_pdf.setStyleSheet(button_style)

        botao_pausar_historico = QPushButton("Pausar Histórico")
        botao_pausar_historico.clicked.connect(self.pausar_historico_juridicos)
        botao_pausar_historico.setStyleSheet(button_style)


        botao_filtrar_historico = QPushButton("Filtrar Histórico")
        botao_filtrar_historico.clicked.connect(self.filtrar_historico_clientes_juridicos)
        botao_filtrar_historico.setStyleSheet(button_style)

        botao_ordenar_historico = QPushButton("Ordenar Histórico")
        botao_ordenar_historico.clicked.connect(self.ordenar_historico_clientes_juridicos)
        botao_ordenar_historico.setStyleSheet(button_style)


        # Criar checkbox "Selecionar Individualmente" toda vez que a janela for aberta
        self.checkbox_selecionar = QCheckBox("Selecionar")
        self.checkbox_selecionar.stateChanged.connect(self.selecionar_individual_juridicos)
        self.checkbox_selecionar.setStyleSheet(f"color: {text_cor};")

        # Adicionar outros botões ao layout
        layout.addWidget(botao_atualizar)
        layout.addWidget(botao_apagar)
        layout.addWidget(botao_exportar_csv)
        layout.addWidget(botao_exportar_excel)
        layout.addWidget(botao_exportar_pdf)
        layout.addWidget(botao_pausar_historico)
        layout.addWidget(botao_ordenar_historico)
        layout.addWidget(botao_filtrar_historico)
        layout.addWidget(self.checkbox_selecionar)
        layout.addWidget(self.tabela_historico_clientes)

        botoes = [
            botao_ordenar_historico,
            botao_filtrar_historico,
            botao_pausar_historico,
            botao_exportar_pdf,
            botao_exportar_excel,
            botao_exportar_csv,
            botao_apagar,
            botao_atualizar

        ]

        for btn in botoes:
            btn.setCursor(Qt.PointingHandCursor)

        # Configurar o widget central e exibir a janela
        self.janela_historico_clientes.setCentralWidget(central_widget)
        self.janela_historico_clientes.show()
        self.janela_historico_clientes.setStyleSheet(f"background-color: {bg_cor}; color: {text_cor};")

        # Estilo da tabela
        self.tabela_historico_clientes.setStyleSheet(table_view_style)

         # Redimensionar colunas e linhas
        self.carregar_historico_clientes_juridicos()
        self.tabela_historico_clientes.resizeColumnsToContents()
        self.tabela_historico_clientes.resizeRowsToContents()

    def carregar_historico_clientes_juridicos(self):
            with sqlite3.connect('banco_de_dados.db') as cn:
                cursor = cn.cursor()
                cursor.execute('SELECT * FROM historico_clientes_juridicos ORDER BY "Data e Hora" DESC')
                registros = cursor.fetchall()

            self.tabela_historico_clientes.clearContents()
            self.tabela_historico_clientes.setRowCount(len(registros))

            deslocamento = 1 if self.coluna_checkboxes_clientes_adicionada else 0
            self.checkboxes_clientes = []  # Zerar e recriar lista de checkboxes

            for i, (data, usuario, acao, descricao) in enumerate(registros):
                if self.coluna_checkboxes_clientes_adicionada:
                    checkbox = QCheckBox()
                    checkbox.setStyleSheet("margin-left:9px; margin-right:9px;")
                    self.tabela_historico_clientes.setCellWidget(i,0,checkbox)
                    self.checkboxes_clientes.append(checkbox)
                self.tabela_historico_clientes.setItem(i, 0 + deslocamento, QTableWidgetItem(data))
                self.tabela_historico_clientes.setItem(i, 1 + deslocamento, QTableWidgetItem(usuario))
                self.tabela_historico_clientes.setItem(i, 2 + deslocamento, QTableWidgetItem(acao))
                self.tabela_historico_clientes.setItem(i, 3 + deslocamento, QTableWidgetItem(descricao))

    def atualizar_historico_clientes_jurididos(self):
        QMessageBox.information(self, "Sucesso", "Dados carregados com sucesso!")
        self.carregar_historico_clientes_juridicos()

    def confirmar_historico_cliente_juridicos_apagado(self, mensagem):
        msgbox = QMessageBox(self)
        msgbox.setWindowTitle("Confirmação")
        msgbox.setText(mensagem)

        btn_sim = QPushButton("Sim")
        btn_nao = QPushButton("Não")
        msgbox.addButton(btn_sim, QMessageBox.ButtonRole.YesRole)
        msgbox.addButton(btn_nao, QMessageBox.ButtonRole.NoRole)

        msgbox.setDefaultButton(btn_nao)
        resposta = msgbox.exec()

        return msgbox.clickedButton() == btn_sim
    
    def apagar_historico_cliente_juridicos(self):
        # Caso checkboxes estejam ativados
        if self.coluna_checkboxes_clientes_adicionada and self.checkboxes_clientes:
            linhas_para_remover = []
            datas_para_remover = []

            # Identificar as linhas com checkboxes selecionados
            for row, checkbox in enumerate(self.checkboxes_clientes):
                if checkbox and checkbox.isChecked():
                    linhas_para_remover.append(row)
                    coluna_data_hora = 0 if not self.coluna_checkboxes_clientes_adicionada else 1
                    item_data_widget = self.tabela_historico_clientes.item(row, coluna_data_hora)  # Coluna de Data/Hora
                    if item_data_widget:
                        data_text = item_data_widget.text().strip()
                        # Excluir com base na data e hora
                        if data_text:
                            datas_para_remover.append(data_text)
                        else:
                            print(f"[AVISO] Linha {row} com Data e Hora vazia.")
                    else:
                        print(f"[ERRO] Coluna 'Data e Hora' não encontrada na linha {row}.")

            if not datas_para_remover:
                QMessageBox.warning(self, "Erro", "Nenhum item válido foi selecionado para apagar!")
                return

            # Confirmar exclusão
            mensagem = (
                f"Você tem certeza que deseja apagar os {len(datas_para_remover)} itens selecionados?"
                if len(datas_para_remover) > 1
                else "Você tem certeza que deseja apagar o item selecionado?"
            )

            if not self.confirmar_historico_apagado_clientes(mensagem):
                return

            # Excluir do banco de dados
            with sqlite3.connect('banco_de_dados.db') as cn:
                cursor = cn.cursor()
                try:
                    for data in datas_para_remover:
                        cursor.execute('DELETE FROM historico_clientes_juridicos WHERE "Data e Hora" = ?', (data,))
                    cn.commit()
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao excluir do banco de dados: {e}")
                    return

            # Remover as linhas na interface
            for row in sorted(linhas_para_remover, reverse=True):
                self.tabela_historico_clientes.removeRow(row)

            QMessageBox.information(self, "Sucesso", "Itens removidos com sucesso!")

        # Caso sem checkboxes (seleção manual)
        else:
            linha_selecionada = self.tabela_historico_clientes.currentRow()

            if linha_selecionada < 0:
                QMessageBox.warning(self, "Erro", "Nenhum item foi selecionado para apagar!")
                return

            # Capturar a Data/Hora da célula correspondente (coluna 0)
            coluna_data_hora = 0 if not self.coluna_checkboxes_clientes_adicionada else 1
            item_data_widget = self.tabela_historico_clientes.item(linha_selecionada, coluna_data_hora)  # Coluna de Data/Hora
            if not item_data_widget:
                QMessageBox.warning(self, "Erro", "Não foi possível identificar a Data/Hora do item a ser apagado!")
                return

            item_data_text = item_data_widget.text().strip()

            # Verificar se há um valor válido de data/hora
            if not item_data_text:
                QMessageBox.warning(self, "Erro", "Não foi possível identificar a Data/Hora do item a ser apagado!")
                return


            # Confirmar exclusão
            mensagem = "Você tem certeza que deseja apagar o item selecionado?"

            if not self.confirmar_historico_cliente_juridicos_apagado(mensagem):
                return

            # Excluir do banco de dados
            with sqlite3.connect('banco_de_dados.db') as cn:
                cursor = cn.cursor()
                try:
                    cursor.execute('DELETE FROM historico_clientes_juridicos WHERE "Data e Hora" = ?', (item_data_text,))
                    cn.commit()
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao excluir do banco de dados: {e}")
                    return

            # Remover a linha da interface
            self.tabela_historico_clientes.removeRow(linha_selecionada)

            QMessageBox.information(self, "Sucesso", "Item removido com sucesso!")

    def selecionar_todos_clientes_juridicos(self):
        if not self.coluna_checkboxes_clientes_adicionada:
            QMessageBox.warning(self, "Aviso", "Ative a opção 'Selecionar Individualmente' antes.")
            if hasattr(self, "checkbox_header"):
                self.checkbox_header_juridicos.setChecked(False)
            return

        estado = self.checkbox_header_juridicos.checkState() == Qt.Checked
        self.checkboxes_clientes.clear()  # Reinicia a lista para manter consistência
        
        for row in range(self.tabela_historico_clientes.rowCount()):
            widget = self.tabela_historico_clientes.cellWidget(row, 0)
            if widget is not None:
                checkbox = widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.blockSignals(True)
                    checkbox.setChecked(estado)
                    checkbox.blockSignals(False)



     # Função para adicionar checkboxes selecionar_individual na tabela de histórico
    def selecionar_individual_juridicos(self):
        if self.tabela_historico_clientes.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico para selecionar.")
            self.checkbox_selecionar_individual.setChecked(False)
            return

        # ---- Se a coluna já existe, remove ----
        if self.coluna_checkboxes_clientes_adicionada:
            self.remover_coluna_checkboxes_historico()
            return

        self.tabela_historico_clientes.insertColumn(0)
        self.tabela_historico_clientes.setHorizontalHeaderItem(0, QTableWidgetItem(""))
        self.tabela_historico_clientes.setColumnWidth(0, 30)
        self.tabela_historico_clientes.horizontalHeader().setMinimumSectionSize(30)
        self.tabela_historico_clientes.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)

        # Checkbox do cabeçalho
        header = self.tabela_historico_clientes.horizontalHeader()
        self.checkbox_header_juridicos = QCheckBox(header.viewport())
        self.checkbox_header_juridicos.setToolTip("Selecionar todos")
        self.checkbox_header_juridicos.setStyleSheet("""QCheckBox{background:transparent;}""")
        self.checkbox_header_juridicos.setChecked(False)
        self.checkbox_header_juridicos.stateChanged.connect(self.selecionar_todos_clientes_juridicos)
        self.checkbox_header_juridicos.setFixedSize(20, 20)
        self.checkbox_header_juridicos.show()

        
        self.atualizar_posicao_checkbox_header_juridicos()
        header.sectionResized.connect(self.atualizar_posicao_checkbox_header_juridicos)

        self.checkboxes_clientes.clear()

        QTimer.singleShot(0,self.atualizar_posicao_checkbox_header_juridicos)

        for row in range(self.tabela_historico_clientes.rowCount()):
            checkbox = QCheckBox()
            checkbox.stateChanged.connect(self.atualizar_selecao_todos)

            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(checkbox)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)

            self.tabela_historico_clientes.setCellWidget(row, 0, container)
            self.checkboxes_clientes.append(checkbox)

        self.tabela_historico_clientes.verticalHeader().setVisible(False)
        header.sectionResized.connect(self.atualizar_posicao_checkbox_header_juridicos)
        self.tabela_historico_clientes.horizontalScrollBar().valueChanged.connect(self.atualizar_posicao_checkbox_header_juridicos)
        header.geometriesChanged.connect(self.atualizar_posicao_checkbox_header_juridicos)
        self.coluna_checkboxes_clientes_adicionada = True

    def atualizar_selecao_todos(self):
        self.checkbox_header_juridicos.blockSignals(True)

        all_checked = all(checkbox.isChecked() for checkbox in self.checkboxes_clientes if checkbox)
        any_checked = any(checkbox.isChecked() for checkbox in self.checkboxes_clientes if checkbox)

        if all_checked:
            self.checkbox_header_juridicos.setCheckState(Qt.Checked)
        elif any_checked:
            self.checkbox_header_juridicos.setCheckState(Qt.PartiallyChecked)
        else:
            self.checkbox_header_juridicos.setCheckState(Qt.Unchecked)

        self.checkbox_header_juridicos.blockSignals(False)

    def remover_coluna_checkboxes_historico(self):
        """Remove a coluna de checkboxes de forma segura"""
        if self.checkbox_header_juridicos is not None:
            try:
                # Desconecta sinais
                self.checkbox_header_juridicos.stateChanged.disconnect()
            except (RuntimeError, TypeError):
                pass

            try:
                # Remove do parent e agenda deleção
                self.checkbox_header_juridicos.setParent(None)
                self.checkbox_header_juridicos.deleteLater()
            except RuntimeError:
                pass

            self.checkbox_header_juridicos = None

        # Remove a coluna de fato
        try:
            self.tabela_historico_clientes.removeColumn(0)
        except Exception:
            pass

        self.tabela_historico_clientes.verticalHeader().setVisible(True)
        self.coluna_checkboxes_clientes_adicionada = False
        self.checkboxes_clientes.clear()

    def closeEvent(self, event):
        if getattr(self, "coluna_checkboxes_clientes_adicionada", False):
            self.remover_coluna_checkboxes_historico()
        # Desmarcar o checkbox principal se existir
        if hasattr(self, "checkbox_selecionar") and isinstance(self.checkbox_selecionar, QCheckBox):
            self.checkbox_selecionar.setChecked(False)
        super().closeEvent(event)

        
    def atualizar_posicao_checkbox_header_juridicos(self):
        if (
            getattr(self,"checkbox_header_juridicos",None) is not None
            and self.coluna_checkboxes_clientes_adicionada
        ):
            try:
                header = self.tabela_historico_clientes.horizontalHeader()

                # largura da seção da coluna 0
                section_width = header.sectionSize(0)
                section_pos = header.sectionViewportPosition(0)

                # centralizar horizontalmente
                x = section_pos + (section_width - self.checkbox_header_juridicos.width()) // 2 + 4

                # centralizar verticalmente
                y = (header.height() - self.checkbox_header_juridicos.height()) // 2

                self.checkbox_header_juridicos.move(x,y)
            except RuntimeError:
                # Objeto já foi deletado do Qt
                self.checkbox_header_juridicos
            
    def ordenar_historico_clientes_juridicos(self):
        if getattr(self, "checkbox_header_juridicos",None) and self.checkbox_header_juridicos.isChecked():
            QMessageBox.warning(
                self,
                "Aviso",
                "Desmarque o checkbox antes de ordenar o histórico."
            )
            return
        # Obter a coluna pela qual o usuário deseja ordenar
        coluna = self.obter_coluna_para_ordenar_clientes_juridicos()  # Função fictícia para capturar escolha
        if coluna is None:
            return  # Cancela o processo todo
        
        # Determinar a direção de ordenação (ascendente ou descendente)
        direcao = self.obter_direcao_ordenacao_clientes_juridicos()  # Função fictícia para capturar escolha
        if direcao is None:
            return  # Cancela o processo todo
        
        # Mapeamento de colunas para índices (ajustar conforme sua tabela)
        colunas_para_indices = {
            "Data/Hora": 0,
            "Usuário": 1,
            "Ação": 2,
            "Descrição": 3
        }
        
        # Verificar se a coluna escolhida é válida
        if coluna not in colunas_para_indices:
            QMessageBox.warning(self, "Erro", "Coluna inválida para ordenação!")
            return
        
        # Obter o índice da coluna escolhida
        indice_coluna = colunas_para_indices[coluna]
        
        # Obter os dados atuais da tabela
        dados = []
        for row in range(self.tabela_historico_clientes.rowCount()):
            linha = [
                self.tabela_historico_clientes.item(row, col).text() if self.tabela_historico_clientes.item(row, col) else ""
                for col in range(self.tabela_historico_clientes.columnCount())
            ]
            dados.append(linha)
        
        # Ordenar os dados com base na coluna escolhida e direção
        dados.sort(key=lambda x: x[indice_coluna], reverse=(direcao == "Decrescente"))
        
        # Atualizar a tabela com os dados ordenados
        self.tabela_historico_clientes.setRowCount(0)  # Limpar tabela
        for row_data in dados:
            row = self.tabela_historico_clientes.rowCount()
            self.tabela_historico_clientes.insertRow(row)
            for col, value in enumerate(row_data):
                self.tabela_historico_clientes.setItem(row, col, QTableWidgetItem(value))

    def obter_coluna_para_ordenar_clientes_juridicos(self):
        colunas = ["Data/Hora", "Usuário", "Ação", "Descrição"]
        dialog = ComboDialog("Ordenar por", "Escolha a coluna:", colunas, self)
        if dialog.exec() == QDialog.Accepted:
            return dialog.escolha()
        return None


    def obter_direcao_ordenacao_clientes_juridicos(self):
        direcoes = ["Crescente", "Decrescente"]
        dialog = ComboDialog("Direção da Ordenação", "Escolha a direção:", direcoes, self)
        if dialog.exec() == QDialog.Accepted:
            return dialog.escolha()
        return None
    
    def filtrar_historico_clientes_juridicos(self):
        if getattr(self, "checkbox_selecionar", None) and self.checkbox_selecionar.isChecked():
            QMessageBox.warning(
                self,
                "Aviso",
                "Desmarque o checkbox antes de filtrar o histórico."
            )
            return

        # Criar a janela de filtro
        janela_filtro = QDialog(self)
        janela_filtro.setWindowTitle("Filtrar Histórico")
        layout = QVBoxLayout(janela_filtro)

        # Campo para inserir a data
        campo_data = QLineEdit()
        campo_data.setPlaceholderText("formato DD/MM/AAAA")
        campo_data.textChanged.connect(lambda: self.formatar_data_clientes_juridicos(campo_data))

        # Carregar tema
        config = self.temas.carregar_config_arquivo()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            groupbox_style = """
                QGroupBox {
                    background-color: #2b2b2b;
                    border: 1px solid #555555;
                    border-radius: 5px;
                    margin-top: 10px;
                    padding: 10px;
                    color: white;
                }

                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 3px 0 3px;
                }

                QRadioButton {
                    color: white;
                    background: transparent;
                }

                QGroupBox QLineEdit {
                    color: black;
                    background-color: rgb(240, 240, 240);
                    border: 3px solid rgb(50, 150,250);
                    border-radius: 12px;
                    padding: 3px;
                }
            """
        elif tema == "claro":
            groupbox_style = """
                QGroupBox {
                    background-color: #f0f0f0;
                    border: 1px solid #cccccc;
                    border-radius: 5px;
                    margin-top: 10px;
                    padding: 10px;
                    color: black;
                }

                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 3px 0 3px;
                }

                QRadioButton {
                    color: black;
                    background: transparent;
                }

                QGroupBox QLineEdit {
                    color: black;
                    background-color: rgb(240, 240, 240);
                    border: 3px solid rgb(50, 150,250);
                    border-radius: 12px;
                    padding: 3px;
                }
            """
        else:  # clássico
            groupbox_style = """
                QGroupBox {
                    background-color: #00557a;
                    border: 1px solid #003f5c;
                    border-radius: 5px;
                    margin-top: 10px;
                    padding: 10px;
                    color: white;
                }

                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 3px 0 3px;
                }

                QRadioButton {
                    color: white;
                    background: transparent;
                }

                QGroupBox QLineEdit {
                    color: black;
                    background-color: rgb(240, 240, 240);
                    border: 3px solid rgb(50, 150,250);
                    border-radius: 12px;
                    padding: 3px;
                }
            """

        # Grupo para o campo de data (agora estilizado corretamente)
        grupo_data = QGroupBox("Filtros Disponíveis")
        grupo_data.setStyleSheet(groupbox_style)
        layout_data = QVBoxLayout(grupo_data)
        layout_data.addWidget(campo_data)
        grupo_data.setLayout(layout_data)

        # Grupo de radio buttons (filtrar por hora)
        grupo_hora = QGroupBox("Filtrar por Hora")
        grupo_hora.setStyleSheet(groupbox_style)
        layout_hora = QVBoxLayout(grupo_hora)

        radio_mais_novo = QRadioButton("Mais Recente")
        radio_mais_velho = QRadioButton("Mais Antigo")

        layout_hora.addWidget(radio_mais_novo)
        layout_hora.addWidget(radio_mais_velho)
        grupo_hora.setLayout(layout_hora)

        # Botão para aplicar o filtro
        botao_filtrar = QPushButton("Aplicar Filtro")
        botao_filtrar.clicked.connect(
            lambda: self.aplicar_filtro_clientes_juridicos(
                campo_data.text(),
                radio_mais_novo.isChecked(),
                radio_mais_velho.isChecked()
            )
        )

        # Adicionar widgets ao layout principal
        layout.addWidget(grupo_data)
        layout.addWidget(grupo_hora)
        layout.addWidget(botao_filtrar)

        # Exibir a janela
        janela_filtro.setLayout(layout)
        janela_filtro.exec()


    def formatar_data_clientes_juridicos(self, campo_data):
        # Obter o texto do campo de data
        texto_data = campo_data.text().replace("/", "")  # Remover as barras existentes
        texto_data = ''.join(filter(str.isdigit, texto_data))  # Permite apenas números

        # Verificar se há caracteres alfabéticos (letras)
        if any(char.isalpha() for char in texto_data):
            # Mostrar mensagem de erro caso haja letras
            QMessageBox.warning(self, "Erro", "Somente números são permitidos.")
            campo_data.clear()
            return  # Não aplica a formatação se houver letras

        # Limitar a entrada para no máximo 8 dígitos
        if len(texto_data) > 8:
            texto_data = texto_data[:8]

         # Formatar a data no formato DD/MM/AAAA
        if len(texto_data) >= 8:
            data_formatada = "{}/{}/{}".format(texto_data[:2], texto_data[2:4], texto_data[4:])  # DD/MM/AAAA
        elif len(texto_data) > 6:
            data_formatada = "{}/{}".format(texto_data[:2], texto_data[2:8])  # DD/MM
        else:
            data_formatada = texto_data[:8]  # Apenas o DD

        # Atualizar o texto do campo de data se houver mudança
        if campo_data.text() != data_formatada:
            campo_data.setText(data_formatada)  # Atualiza o texto do campo de data
            campo_data.setCursorPosition(len(data_formatada))  # Move o cursor para o final do texto

    def aplicar_filtro_clientes_juridicos(self, data, filtrar_novo, filtrar_velho):
        with sqlite3.connect('banco_de_dados.db') as cn:
            cursor = cn.cursor()

            query = "SELECT * FROM historico_clientes_juridicos"
            params = []

            # Filtrar pela data, se fornecida
            if data:
                try:
                    # Garantir que a data seja no formato correto (DD/MM/AAAA)
                    data_formatada = datetime.strptime(data, "%d/%m/%Y").strftime("%d/%m/%Y")  # Formato DD/MM/YYYY
                    query += " WHERE SUBSTR([Data e Hora], 1, 10) = ?"
                    params.append(data_formatada)
                except ValueError:
                    QMessageBox.warning(self, "Erro", "Data inválida. Use o formato DD/MM/AAAA.")
                    return

            # Ordenar por hora, se aplicável
            if filtrar_novo:
                query += " ORDER BY [Data e Hora] DESC LIMIT 1"
            elif filtrar_velho:
                query += " ORDER BY [Data e Hora] ASC LIMIT 1"

            # Executar a consulta
            cursor.execute(query, params)
            registros = cursor.fetchall()

        # Atualizar a tabela com os registros filtrados
        self.tabela_historico_clientes.clearContents()
        self.tabela_historico_clientes.setRowCount(len(registros))

        for i, row in enumerate(registros):
            self.tabela_historico_clientes.setItem(i, 0, QTableWidgetItem(row[0]))  # Data/Hora
            self.tabela_historico_clientes.setItem(i, 1, QTableWidgetItem(row[1]))  # Usuário
            self.tabela_historico_clientes.setItem(i, 2, QTableWidgetItem(row[2]))  # Ação
            self.tabela_historico_clientes.setItem(i, 3, QTableWidgetItem(row[3]))  # Descrição

        QMessageBox.information(self, "Filtro Aplicado", f"{len(registros)} registro(s) encontrado(s)!")
    
    def exportar_csv_juridicos(self):
        num_linhas = self.tabela_historico_clientes.rowCount()
        num_colunas = self.tabela_historico_clientes.columnCount()

        # Verificar se a tabela está vazia
        if self.tabela_historico_clientes.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico encontrado para gerar arquivo CSV.")
            return  # Se a tabela estiver vazia, encerra a função sem prosseguir

        nome_arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Arquivo CSV",
            "historico.csv",
            "Arquivos CSV (*.csv)"

        )

        if not nome_arquivo:
            return
        
        #Criar o arquivo CSV
        try:
            with open(nome_arquivo, mode="w",newline="",encoding="utf-8-sig") as arquivo_csv:
                escritor = csv.writer(arquivo_csv, delimiter=";")

                 # Adicionar cabeçalhos ao CSV
                cabecalhos = [self.tabela_historico_clientes.horizontalHeaderItem(col).text() for col in range (num_colunas)]
                escritor.writerow(cabecalhos)

                # Adicionar os dados da tabela ao CSV
                for linha in range(num_linhas):
                    dados_linhas = [
                        self.tabela_historico_clientes.item(linha, col).text() if self.tabela_historico_clientes.item(linha, col) else ""
                        for col in range(num_colunas)

                    ]
                    escritor.writerow(dados_linhas)

                QMessageBox.information(self, "Sucesso", f"Arquivo CSV salvo com sucesso em:\n{nome_arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao salvar o arquivo CSV:\n{str(e)}")

    def exportar_excel_juridicos(self):
        num_linhas = self.tabela_historico_clientes.rowCount()
        num_colunas = self.tabela_historico_clientes.columnCount()

        # Verificar se a tabela está vazia
        if self.tabela_historico_clientes.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico encontrado para gerar arquivo Excel.")
            return  # Se a tabela estiver vazia, encerra a função sem prosseguir
        
        nome_arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Arquivo Excel",
            "historico.xlsx",
            "Arquivos Excel (*.xlsx)"

        )

        if not nome_arquivo:
            return
        
        # Garantir que o arquivo tenha extensão .xlsx
        if not nome_arquivo.endswith(".xlsx"):
            nome_arquivo += ".xlsx"

        # Criar uma lista para armazenar os dados da tabela
        dados = []


        for linha in range(num_linhas):
            linha_dados = []
            for coluna in range(num_colunas):
                item = self.tabela_historico_clientes.item(linha, coluna)
                linha_dados.append(item.text() if item else "") # Adicionar o texto ou vazio se o item for None
            dados.append(linha_dados)

        # Obter os cabeçalhos da tabela        
        cabecalhos = [self.tabela_historico_clientes.horizontalHeaderItem(coluna).text() for coluna in range (num_colunas)]

        try:
            # Criar um DataFrame do pandas com os dados e cabeçalhos
            df = pd.DataFrame(dados, columns=cabecalhos)

            # Exportar para Excel
            df.to_excel(nome_arquivo, index=False,engine="openpyxl")
            QMessageBox.information(self, "Sucesso",f"Arquivo Excel gerado com sucesso em: \n{nome_arquivo}")
        except Exception as e:
            QMessageBox.critical(self, "Erro",f"Erro ao salvar arquivo Excel: {str(e)}")

    def exportar_pdf_juridicos(self):
        num_linhas = self.tabela_historico_clientes.rowCount()
        num_colunas = self.tabela_historico_clientes.columnCount()

        # Verificar se a tabela está vazia
        if self.tabela_historico_clientes.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico encontrado para gerar arquivo PDF.")
            return  # Se a tabela estiver vazia, encerra a função sem prosseguir

        nome_arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Arquivo PDF",
            "historico.pdf",
            "Arquivos PDF (*.pdf)"
        )

        if not nome_arquivo:
            return

        # Garantir que o arquivo tenha extensão .pdf
        if not nome_arquivo.endswith(".pdf"):
            nome_arquivo += ".pdf"

        # Criar uma lista para armazenar os dados da tabela
        dados = []

        # Obter os cabeçalhos da tabela
        cabecalhos = [self.tabela_historico_clientes.horizontalHeaderItem(coluna).text() for coluna in range(num_colunas)]
        dados.append(cabecalhos)  # Adicionar os cabeçalhos como a primeira linha do PDF

        # Adicionar os dados da tabela
        for linha in range(num_linhas):
            linha_dados = []
            for coluna in range(num_colunas):
                item = self.tabela_historico_clientes.item(linha, coluna)
                linha_dados.append(item.text() if item else "")  # Adicionar o texto ou vazio se o item for None
            dados.append(linha_dados)

        try:
            # Criar o PDF
            pdf = SimpleDocTemplate(nome_arquivo, pagesize=landscape(letter))
            estilos = getSampleStyleSheet()

            titulo = Paragraph("<b>Histórico de Clientes Jurídicos</b>", estilos['Title'])

            # Espaço entre título e tabela
            espacamento = Spacer(1, 12)

            tabela = Table(dados)

            # Adicionar estilo à tabela
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Cabeçalho com fundo cinza
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Texto do cabeçalho branco
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centralizar texto
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fonte do cabeçalho
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Espaçamento inferior no cabeçalho
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Fundo das linhas de dados
                ('GRID', (0, 0), (-1, -1), 1, colors.black)  # Bordas da tabela
            ])
            tabela.setStyle(estilo)

            # Construir PDF com título + espaço + tabela
            pdf.build([titulo,espacamento,tabela])
            QMessageBox.information(self, "Sucesso", f"Arquivo PDF gerado com sucesso em: \n{nome_arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar arquivo PDF: {str(e)}")

    def pausar_historico_juridicos(self):
        # Criação da nova janela de histórico como QMainWindow
        self.janela_escolha = QMainWindow(self)
        self.janela_escolha.setWindowTitle("Pausar Histórico")
        self.janela_escolha.resize(255, 150)


        # Botão "Sim"
        botao_sim = QPushButton("Sim")
        botao_sim.clicked.connect(self.historico_ativo_juridicos)

        # Botão "Não"
        botao_nao = QPushButton("Não")
        botao_nao.clicked.connect(self.historico_inativo_juridicos)


        # Criação do layout e tabela para exibir o histórico
        central_widget = QWidget()
        layout = QVBoxLayout(central_widget)

        # Texto centralizado
        label = QLabel("Deseja pausar o histórico?")
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)  # Alinha o texto ao centro

        layout.addWidget(label)  # Adiciona o texto centralizado
        layout.addWidget(botao_sim)
        layout.addWidget(botao_nao)

        self.janela_escolha.setCentralWidget(central_widget)
        self.janela_escolha.show()


    def historico_ativo_juridicos(self):
        # Atualiza o estado do histórico para ativo
        self.main_window.historico_pausado_clientes_juridicos = True  # Atualiza a variável no MainWindow
        QMessageBox.information(self, "Histórico", "O registro do histórico foi pausado.")
        self.janela_escolha.close()

    def historico_inativo_juridicos(self):
        # Atualiza o estado do histórico para inativo (continua registrando)
        self.main_window.historico_pausado_clientes_juridicos = False  # Atualiza a variável no MainWindow
        QMessageBox.information(self, "Histórico", "O registro do histórico continua ativo.")
        self.janela_escolha.close()
        

    def abrir_janela_relatorio_clientes_juridicos(self):
        self.janela_historico_clientes = QMainWindow()
        self.janela_historico_clientes.setWindowTitle("Relatório de Clientes Jurídicos")
        self.janela_historico_clientes.resize(800, 600)
        self.janela_historico_clientes.setObjectName("janela_historico_clientes")

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)

        # Carregar tema
        config = self.temas.carregar_config_arquivo()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(60,60,60),
                                                stop:1 rgb(100,100,100));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid #666666;
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
            """
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """
            checkbox_style = """
                QMessageBox QCheckBox{
                    background: transparent;
                    color: white;
                }
            """
            date_style = """
                /* Estilo geral do QDateEdit */
                QDateEdit {
                    color: white; 
                    background-color: #2b2b2b; 
                    border: 2px solid #ffffff;
                    border-radius: 5px;
                    padding: 2px 5px;
                }

                /* Remove o fundo das setas */
                QDateEdit::up-button, 
                QDateEdit::down-button {
                    background: transparent;
                    border: none;
                }

                /* Cor de fundo do calendário popup */
                QDateEdit QCalendarWidget {
                    background-color: #2b2b2b;
                    border: 1px solid #555555;
                }
                /* Cabeçalho dos dias da semana (dom, seg, ..., sáb) */
                QDateEdit QCalendarWidget QHeaderView::section {
                    background-color: #2b2b2b;
                    color: black;  /* padrão: dias úteis */
                    padding: 5px;
                    font-weight: bold;
                }

                /* Dias normais */
                QDateEdit QCalendarWidget QAbstractItemView:enabled {
                    background-color: #2b2b2b;
                    color: black;
                    selection-background-color: rgb(0, 120, 215); /* Azul no dia selecionado */
                    selection-color: black;
                }
                /* Botões de navegação (setas) do calendário */
                QDateEdit QCalendarWidget QToolButton {
                    background: transparent;   /* tira o fundo azul */
                    color: white;              
                    border: none;              /* sem borda */
                    icon-size: 16px 16px;      /* ajusta o tamanho do ícone */
                    padding: 2px;
                }
                QDateEdit QCalendarWidget QToolButton:hover {
                    background: rgb(220, 220, 220); /* cinza claro no hover */
                    border-radius: 4px;
                }
                /* Popup de meses/anos do calendário (QMenu) */
                QDateEdit QCalendarWidget QMenu {
                    background-color: white;
                    border: 1px solid #ccc;
                    color: black;
                }
                

                QDateEdit QCalendarWidget QMenu::item {
                    background: transparent;
                    color: black;
                    padding: 6px 16px;
                }

                QDateEdit QCalendarWidget QMenu::item:selected {
                    background: rgb(0, 120, 215);
                    color: white;
                }
                QDateEdit QCalendarWidget QToolButton::menu-indicator {
                    image: none;   /* remove o ícone padrão */
                    width: 0px;    /* remove o espaço reservado */
                }
                /* Finais de semana - domingos */
                QDateEdit QCalendarWidget QTableView::item:nth-child(7n+1) {
                    color: red;
                }

                /* Finais de semana - sábados */
                QDateEdit QCalendarWidget QTableView::item:nth-child(7n) {
                    color: red;
                }

                /* Domingo (primeira coluna do cabeçalho) */
                QDateEdit QCalendarWidget QHeaderView::section:first-child {
                    color: red;
                }

                /* Sábado (última coluna do cabeçalho) */
                QDateEdit QCalendarWidget QHeaderView::section:last-child {
                    color: red;
                }
            """
        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """
            checkbox_style = """
                QMessageBox QCheckBox{
                    background: transparent;
                    color: black;
                }
            """
            date_style = """
                /* Estilo geral do QDateEdit */
                QDateEdit {
                    color: black; 
                    background-color: 2b2b2b; 
                    border: 1px solid 3399ff;
                    border-radius: 5px;
                    padding: 2px 5px;
                }

                /* Remove o fundo das setas */
                QDateEdit::up-button, 
                QDateEdit::down-button {
                    background: transparent;
                    border: none;
                }

                /* Cor de fundo do calendário popup */
                QDateEdit QCalendarWidget {
                    background-color: 2b2b2b;
                    border: 1px solid #555555;
                }

                /* Dias normais */
                QDateEdit QCalendarWidget QAbstractItemView:enabled {
                    background-color: #2b2b2b;
                    color: black;
                    selection-background-color: rgb(0, 120, 215); /* Azul no dia selecionado */
                    selection-color: white;
                }
                /* Botões de navegação (setas) do calendário */
                QDateEdit QCalendarWidget QToolButton {
                    background: transparent;   /* tira o fundo azul */
                    color: black;              /* deixa as setas pretas */
                    border: none;              /* sem borda */
                    icon-size: 16px 16px;      /* ajusta o tamanho do ícone */
                    padding: 2px;
                }
                QDateEdit QCalendarWidget QToolButton:hover {
                    background: rgb(220, 220, 220); /* cinza claro no hover */
                    border-radius: 4px;
                }
                /* Popup de meses/anos do calendário (QMenu) */
                QDateEdit QCalendarWidget QMenu {
                    background-color: white;
                    border: 1px solid #ccc;
                    color: black;
                }

                QDateEdit QCalendarWidget QMenu::item {
                    background: transparent;
                    color: black;
                    padding: 6px 16px;
                }

                QDateEdit QCalendarWidget QMenu::item:selected {
                    background: rgb(0, 120, 215);
                    color: white;
                }
                QDateEdit QCalendarWidget QToolButton::menu-indicator {
                    image: none;   /* remove o ícone padrão */
                    width: 0px;    /* remove o espaço reservado */
                }

            """

        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"

            button_style = """
            QPushButton {
                color: rgb(255, 255, 255);
                border-radius: 8px;
                font-size: 16px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgb(50, 150, 250), stop:1 rgb(100, 200, 255)); /* Gradiente de azul claro para azul mais claro */
                border: 4px solid transparent;
            }

            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 rgb(100, 180, 255), stop:1 rgb(150, 220, 255)); /* Gradiente de azul mais claro para azul ainda mais claro */
                color: black;
            }
            QPushButton:pressed {
                background-color: #006bb3;
                border: 2px solid #005c99;
            }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 3px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            scroll_style = """
                QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QScrollBar::handle:vertical {
                    background: #b4b4b4;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """
            checkbox_style = """
                QMessageBox QCheckBox{
                    background: transparent;
                    color: black;
                }
            """
            
            date_style = """
                /* Estilo geral do QDateEdit */
                QDateEdit {
                    color: black; 
                    background-color: white; 
                    border: 2px solid #3399ff;
                    border-radius: 6px;
                    padding: 2px 5px;
                }

                /* Remove o fundo das setas */
                QDateEdit::up-button, 
                QDateEdit::down-button {
                    background: transparent;
                    border: none;
                }

                /* Cor de fundo do calendário popup */
                QDateEdit QCalendarWidget {
                    background-color: white;
                    border: 1px solid #555555;
                }
                /* Dias normais */
                QDateEdit QCalendarWidget QAbstractItemView:enabled {
                    background-color: white;
                    color: black;
                    selection-background-color: rgb(0, 120, 215); /* Azul no dia selecionado */
                    selection-color: black;
                }
                QDateEdit QCalendarWidget QToolButton:hover {
                    background: rgb(120, 120, 120); /* cinza claro no hover */
                    border-radius: 4px;
                }
                /* Popup de meses/anos do calendário (QMenu) */
                QDateEdit QCalendarWidget QMenu {
                    background-color: white;
                    border: 1px solid #ccc;
                    color: black;
                }
                QDateEdit QCalendarWidget QMenu::item {
                    background: transparent;
                    color: black;
                    padding: 6px 16px;
                }
                QDateEdit QCalendarWidget QMenu::item:selected {
                    background: rgb(0, 120, 215);
                    color: white;
                }
                QDateEdit QCalendarWidget QToolButton::menu-indicator {
                    image: none;   /* remove o ícone padrão */
                    width: 0px;    /* remove o espaço reservado */
                }

            """

        central_widget = QWidget()
        layout_principal = QVBoxLayout(central_widget)

        grupo_filtros = QGroupBox("Filtros do Relatório")
        layout_filtros = QFormLayout()

        self.combo_status = QComboBox()
        self.combo_status.addItems(["Todos", "Ativo", "Inativo", "Bloqueado", "Pendente"])
        layout_filtros.addRow("Status do Cliente:", self.combo_status)
        self.combo_status.setStyleSheet(combobox_style)

        self.combo_categoria = QComboBox()
        layout_filtros.addRow("Categoria do Cliente:", self.combo_categoria)
        self.combo_categoria.setStyleSheet(combobox_style)

        self.date_de = QDateEdit()
        self.date_de.setDate(QDate.currentDate().addMonths(-1))
        self.date_de.setCalendarPopup(True)
        self.date_de.setStyleSheet(date_style)

        self.date_ate = QDateEdit()
        self.date_ate.setDate(QDate.currentDate())
        self.date_ate.setCalendarPopup(True)
        self.date_ate.setStyleSheet(date_style)

        layout_filtros.addRow("Última Compra - De:", self.date_de)
        layout_filtros.addRow("Última Compra - Até:", self.date_ate)

        self.combo_origem = QComboBox()
        layout_filtros.addRow("Origem do Cliente:", self.combo_origem)
        self.combo_origem.setStyleSheet(combobox_style)

        grupo_filtros.setLayout(layout_filtros)
        layout_principal.addWidget(grupo_filtros)

        self.carregar_opcoes_combo("Categoria do Cliente", self.combo_categoria)
        self.carregar_opcoes_combo("Origem do Cliente", self.combo_origem)

        grupo_campos = QGroupBox("Informações a Incluir no Relatório")
        layout_campos = QVBoxLayout()

        self.combo_selecionar_todos = QCheckBox("Selecionar Todos")
        self.combo_selecionar_todos.setChecked(True)
        self.combo_selecionar_todos.stateChanged.connect(self.selecionar_todos_checkboxes)
        layout_campos.addWidget(self.combo_selecionar_todos)

        checks = [
            "Nome", "Razão Social", "Data da Inclusão", "CNPJ",
            "Contato", "CEP", "Endereço", "Número", "Complemento", "Cidade", "Bairro", "Estado",
            "Status", "Categoria", "Última Atualização", "Origem",
            "Valor Gasto Total", "Última Compra"
        ]

        self.checkboxes_relatorio = []

        for texto in checks:
            check = QCheckBox(texto)
            check.setChecked(True)
            self.checkboxes_relatorio.append(check)
            layout_campos.addWidget(check)

        grupo_campos.setLayout(layout_campos)
        layout_principal.addWidget(grupo_campos)

        layout_botoes = QHBoxLayout()
        btn_gerar = QPushButton("Gerar Relatório em PDF")
        btn_gerar_excel = QPushButton("Gerar Relatório em Excel")
        btn_cancelar = QPushButton("Cancelar")
        btn_gerar.setStyleSheet(button_style)
        btn_gerar_excel.setStyleSheet(button_style)
        btn_cancelar.setStyleSheet(button_style)

        botoes = [
            btn_cancelar,
            btn_gerar,
            btn_gerar_excel
        ]
        for btn in botoes:
            btn.setCursor(Qt.PointingHandCursor)


        layout_botoes.addStretch()
        layout_botoes.addWidget(btn_gerar)
        layout_botoes.addWidget(btn_gerar_excel)
        layout_botoes.addWidget(btn_cancelar)
        layout_principal.addLayout(layout_botoes)

        btn_cancelar.clicked.connect(self.janela_historico_clientes.close)
        btn_gerar.clicked.connect(self.gerar_pdf_clientes_juridicos)
        btn_gerar_excel.clicked.connect(self.gerar_excel_clientes_juridicos)

        scroll.setWidget(central_widget)
        scroll.setStyleSheet(scroll_style)
        self.janela_historico_clientes.setStyleSheet(f"background-color: {bg_cor}; color: {text_cor};")
        self.janela_historico_clientes.setCentralWidget(scroll)
        self.janela_historico_clientes.show()
        
        configuracoes = Configuracoes_Login(self.main_window)
        if  configuracoes.nao_mostrar_mensagem_arquivo_excel:
            return
        
        # AVISO: Recomendação
        aviso = QMessageBox(self)
        aviso.setWindowTitle("Aviso")
        aviso.setIcon(QMessageBox.Information) # Ícone de aviso
        aviso.setText("Para uma melhor experiência recomendamos a geração do relatório em Excel.")
        aviso.setStandardButtons(QMessageBox.Ok)
        
        
        
        checkbox_nao_mostrar = QCheckBox("Não mostrar essa mensagem novamente")
        checkbox_nao_mostrar.setStyleSheet(checkbox_style)
        aviso.setCheckBox(checkbox_nao_mostrar)
        aviso.exec()

        # Verifica se o usuário marcou a opção para não mostrar novamente
        if checkbox_nao_mostrar.isChecked():
                configuracoes.nao_mostrar_mensagem_arquivo_excel = True # Define que o usuário não quer mais ver este aviso
                configuracoes.salvar(configuracoes.usuario,configuracoes.senha,configuracoes.mantem_conectado)


    def gerar_pdf_clientes_juridicos(self):
        status = self.combo_status.currentText()
        categoria = self.combo_categoria.currentText()
        origem = self.combo_origem.currentText()
        data_de = self.date_de.date().toPython()
        data_ate = self.date_ate.date().toPython()

        campos_selecionados = [
            checkbox.text()
            for checkbox in self.checkboxes_relatorio
            if checkbox.isChecked()
        ]

        if not campos_selecionados:
            QMessageBox.warning(self, "Aviso", "Selecione ao menos um campo para incluir no relatório.")
            return

        mapeamento_colunas = {
            "Nome": "Nome",
            "Razão Social": "Razão Social",
            "Data da Inclusão": "Data da Inclusão",
            "CNPJ": "CNPJ",
            "Contato": "Telefone",
            "CEP": "CEP",
            "Endereço": "Endereço",
            "Número": "Número",
            "Complemento": "Complemento",
            "Cidade": "Cidade",
            "Bairro": "Bairro",
            "Estado": "Estado",
            "Status": "Status",
            "Categoria": "Categoria",
            "Última Atualização": "Última Atualização",
            "Origem": "Origem",
            "Valor Gasto Total": "Valor Gasto Total",
            "Última Compra": "Última Compra"
        }

        colunas_sql = [f'"{mapeamento_colunas[nome]}"' for nome in campos_selecionados]
        sql = f"SELECT {', '.join(colunas_sql)} FROM clientes_juridicos WHERE 1=1 "

        params = []
        if status != "Todos":
            sql += " AND `Status do Cliente` = ?"
            params.append(status)
        if categoria != "Todos":
            sql += " AND `Categoria do Cliente` = ?"
            params.append(categoria)
        if origem != "Todos":
            sql += " AND `Origem do Cliente` = ?"
            params.append(origem)

        sql += """
            AND (
                `Última Compra` = 'Não Cadastrado' OR
                (
                    length(`Última Compra`) = 10 AND
                    substr(`Última Compra`, 3, 1) = '/' AND
                    substr(`Última Compra`, 6, 1) = '/' AND
                    substr(`Última Compra`, 7, 4) || '-' || 
                    substr(`Última Compra`, 4, 2) || '-' || 
                    substr(`Última Compra`, 1, 2)
                    BETWEEN ? AND ?
                )
            )
        """

        params.append(data_de.strftime("%Y-%m-%d"))
        params.append(data_ate.strftime("%Y-%m-%d"))

        try:
            conn = sqlite3.connect("banco_de_dados.db")
            cursor = conn.cursor()
            cursor.execute(sql, params)
            resultados = cursor.fetchall()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao buscar dados: \n{e}")
            return

        if not resultados:
            QMessageBox.information(self, "Aviso", "Nenhum cliente encontrado com os filtros aplicados")
            return

        caminho_pdf, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Relatório",
            "relatorio_clientes_juridicos.pdf",
            "Arquivos PDF (*.pdf)"
        )

        if not caminho_pdf:
            return

        try:
            pdf = FPDF(orientation="L", format="A4")  # Cria um objeto PDF com orientação paisagem ("L") e tamanho A4
            pdf.set_left_margin(5) # Define as margens do PDF
            pdf.set_right_margin(5) # Define as margens do PDF
            pdf.add_page() # Adiciona uma página ao PDF, obrigatória antes de começar a escrever
            pdf.set_font("Arial", size=8) # Define a fonte para o texto:
            # Cria uma célula que ocupa toda a largura (0 = largura total da página),
            # com altura 10, com o texto "Relatório de Clientes Jurídicos",
            # quebra de linha após a célula (ln=True) e alinhamento centralizado
            pdf.cell(0, 10, "Relatório de Clientes Jurídicos", ln=True, align="C")
            pdf.ln(5) # Insere uma linha em branco com altura 5 para espaçamento vertical

            pdf.set_font("Arial", size=5) # tamanho da fonte
            cel_height = 5 # Define a altura das células que serão usadas para o cabeçalho da tabela


            larguras_automaticas = {}
            margem = 4  # margem interna da célula, para deixar texto mais confortável
            largura_minima = 10
            largura_maxima = 80  # máximo para evitar colunas absurdamente largas

            for i, campo in enumerate(campos_selecionados):
                # largura do cabeçalho
                largura_cabecalho = pdf.get_string_width(campo) + margem

                # largura máxima entre todos os textos daquela coluna
                largura_max = largura_cabecalho

                for linha in resultados:
                    item = linha[i]
                    texto = str(item) if item is not None else ""
                    largura_texto = pdf.get_string_width(texto) + margem
                    if largura_texto > largura_max:
                        largura_max = largura_texto

                # Limita largura para mínimo e máximo
                if largura_max < largura_minima:
                    largura_max = largura_minima
                elif largura_max > largura_maxima:
                    largura_max = largura_maxima

                larguras_automaticas[campo] = largura_max

            # Cabeçalho
            for campo in campos_selecionados:
                pdf.cell(larguras_automaticas[campo], cel_height, campo, border=1, align="C")
            pdf.ln()

            # Dados
            for linha in resultados:
                y_inicial = pdf.get_y()
                max_altura = cel_height
                for i, item in enumerate(linha):
                    campo = campos_selecionados[i]
                    largura = larguras_automaticas[campo]
                    texto = str(item) if item is not None else ""
                    x = pdf.get_x()
                    y = pdf.get_y()

                    # Usar multi_cell para texto que possa ter quebras, mas com largura suficiente para minimizar linhas
                    pdf.multi_cell(largura, cel_height, texto, border=1, align="L")
                    altura = pdf.get_y() - y
                    if altura > max_altura:
                        max_altura = altura
                    pdf.set_xy(x + largura, y)
                pdf.set_y(y_inicial + max_altura)
            pdf.output(caminho_pdf)
            QMessageBox.information(self, "Sucesso", "Relatório gerado com sucesso.")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao gerar PDF:\n{e}")

    def gerar_excel_clientes_juridicos(self):
        status = self.combo_status.currentText()
        categoria = self.combo_categoria.currentText()
        origem = self.combo_origem.currentText()
        data_de = self.date_de.date().toPython()
        data_ate = self.date_ate.date().toPython()

        campos_selecionados = [
            checkbox.text()
            for checkbox in self.checkboxes_relatorio
            if checkbox.isChecked()
        ]

        if not campos_selecionados:
            QMessageBox.warning(self, "Aviso", "Selecione ao menos um campo para incluir no relatório.")
            return

        mapeamento_colunas = {
            "Nome": "Nome",
            "Razão Social": "Razão Social",
            "Data da Inclusão": "Data da Inclusão",
            "CNPJ": "CNPJ",
            "Contato": "Telefone",
            "CEP": "CEP",
            "Endereço": "Endereço",
            "Número": "Número",
            "Complemento": "Complemento",
            "Cidade": "Cidade",
            "Bairro": "Bairro",
            "Estado": "Estado",
            "Status": "Status",
            "Categoria": "Categoria",
            "Última Atualização": "Última Atualização",
            "Origem": "Origem",
            "Valor Gasto Total": "Valor Gasto Total",
            "Última Compra": "Última Compra"
        }

        colunas_sql = [f'"{mapeamento_colunas[nome]}"' for nome in campos_selecionados]
        sql = f"SELECT {', '.join(colunas_sql)} FROM clientes_juridicos WHERE 1=1 "

        params = []
        
        if status != "Todos":
            sql += " AND `Status do Cliente` = ?"
            params.append(status)
        if categoria != "Todos":
            sql += " AND `Categoria do Cliente` = ?"
            params.append(categoria)
        if origem != "Todos":
            sql += " AND `Origem do Cliente` = ?"
            params.append(origem)

        sql += """
            AND (
                `Última Compra` = 'Não Cadastrado' OR
                (
                    length(`Última Compra`) = 10 AND
                    substr(`Última Compra`, 3, 1) = '/' AND
                    substr(`Última Compra`, 6, 1) = '/' AND
                    substr(`Última Compra`, 7, 4) || '-' || 
                    substr(`Última Compra`, 4, 2) || '-' || 
                    substr(`Última Compra`, 1, 2)
                    BETWEEN ? AND ?
                )
            )
        """
        params.append(data_de.strftime("%Y-%m-%d"))
        params.append(data_ate.strftime("%Y-%m-%d"))

        try:
            conn = sqlite3.connect("banco_de_dados.db")
            cursor = conn.cursor()
            cursor.execute(sql, params)
            resultados = cursor.fetchall()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao buscar dados: \n{e}")
            return

        if not resultados:
            QMessageBox.information(self, "Aviso", "Nenhum cliente encontrado com os filtros aplicados")
            return

        caminho_excel, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Relatório",
            "relatorio_clientes_juridicos.xlsx",
            "Arquivos Excel (*.xlsx)"
        )

        if not caminho_excel:
            return

        try:
            # Criar um DataFrame do pandas com os dados
            df = pd.DataFrame(resultados,columns=campos_selecionados)
            
            # Salva como Excel usando openpyxl
            sheet_name = "Relatório de Clientes Jurídicos"
            df.to_excel(caminho_excel,index=False,engine='openpyxl',sheet_name=sheet_name)
            
            # Reabre o Excel com openpyxl para aplicar formatações
            wb = load_workbook(caminho_excel)
            ws = wb[sheet_name]
            
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length,len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2  # Ajusta a largura da coluna    
            wb.save(caminho_excel)
            QMessageBox.information(self, "Sucesso", "Relatório Excel gerado com sucesso.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao gerar Excel:\n{e}")


    def selecionar_todos_checkboxes(self,estado):
        for checkbox in self.checkboxes_relatorio:
            checkbox.setChecked(bool(estado))

    def carregar_opcoes_combo(self, nome_coluna, combo_box):
        try:
            conn = sqlite3.connect("banco_de_dados.db")  # substitua com seu caminho
            cursor = conn.cursor()

            # Aspas duplas se o nome da coluna tiver espaços
            cursor.execute(f'''
                SELECT DISTINCT "{nome_coluna}"
                FROM clientes_juridicos
                WHERE "{nome_coluna}" IS NOT NULL AND "{nome_coluna}" != ''
            ''')
            resultados = cursor.fetchall()

            opcoes = sorted(set([row[0] for row in resultados]))
            combo_box.clear()
            combo_box.addItem("Todos")
            combo_box.addItems(opcoes)

        except Exception as e:
            print(f"Erro ao carregar opções de {nome_coluna}:", e)

    def confirmar_historico_apagado_clientes(self, mensagem):
        """
        Exibe uma caixa de diálogo para confirmar a exclusão.
        """
        msgbox = QMessageBox(self)
        msgbox.setWindowTitle("Confirmação")
        msgbox.setText(mensagem)

        btn_sim = QPushButton("Sim")
        btn_nao = QPushButton("Não")
        msgbox.addButton(btn_sim, QMessageBox.ButtonRole.YesRole)
        msgbox.addButton(btn_nao, QMessageBox.ButtonRole.NoRole)

        msgbox.setDefaultButton(btn_nao)
        resposta = msgbox.exec()

        return msgbox.clickedButton() == btn_sim
    
    def configurar_menu_contexto(self):
        self.table_clientes_juridicos.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_clientes_juridicos.customContextMenuRequested.connect(self.abrir_menu_contexto)

    def abrir_menu_contexto(self,pos:QPoint):
        index = self.table_clientes_juridicos.indexAt(pos)
        if not index.isValid():
            return # Clicou fora de uma célula
        
        # Captura os dados da linha
        linha = index.row()
        nome_cliente = self.table_clientes_juridicos.item(linha,0).text()  # Coluna 0 = Nome do Cliente

        menu = QMenu()

        detalhes_action = menu.addAction("Detalhes")
        editar_action = menu.addAction("Editar Cliente")
        excluir_action = menu.addAction("Excluir Cliente")


        action = menu.exec(self.table_clientes_juridicos.viewport().mapToGlobal(pos))

        # Executa conforme a opção clicada
        if action == detalhes_action:
            QMessageBox.information(self, "Aviso", "Essa função ainda não está disponível")
        elif action == editar_action:
            self.editar_cliente_juridico()
        elif action == excluir_action:
            self.excluir_clientes_juridicos()



    

    
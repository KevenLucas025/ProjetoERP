from PySide6.QtWidgets import (QLineEdit, QToolButton,QTableWidgetItem,
                               QMessageBox,QMainWindow,QVBoxLayout,QWidget,QLabel,QCheckBox,
                               QPushButton,QScrollArea,QComboBox,QGridLayout,QHeaderView,QHBoxLayout,
                               QGraphicsOpacityEffect,QTableWidget,QInputDialog,QDialog,
                               QRadioButton,QGroupBox,QFileDialog,QFormLayout,QDateEdit,QMenu,QApplication)
from PySide6.QtGui import QPixmap, QIcon,QColor,QBrush,QGuiApplication
from PySide6.QtCore import Qt,QTimer,QPropertyAnimation,QEvent,QDate,QPoint
from database import DataBase
import sqlite3
import pandas as pd
from configuracoes import Configuracoes_Login
from datetime import datetime
from PySide6.QtGui import QKeySequence, QShortcut
import csv
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.pagesizes import letter,landscape,A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from fpdf import FPDF
import json


class Clientes_Fisicos(QWidget):
    def __init__(self,line_clientes_fisicos: QLineEdit,main_window,btn_adicionar_cliente_fisico,btn_editar_clientes_fisicos,
                 btn_excluir_clientes_fisicos,btn_gerar_relatorio_clientes_fisicos,
                 btn_historico_clientes_fisicos,btn_marcar_como_clientes_fisicos):
        super().__init__()
        self.line_clientes_fisicos = line_clientes_fisicos
        self.db = DataBase("banco_de_dados.db")
        self.config = Configuracoes_Login(self)
        self.coluna_checkboxes_clientes_fisicos_adicionada = False
        self.checkboxes_clientes_fisicos = []

        self.timer_buscar_fisicos = QTimer()
        self.timer_buscar_fisicos.setSingleShot(True)
        self.timer_buscar_fisicos.timeout.connect(self._executar_busca_dinamica_fisicos)
        self.line_clientes_fisicos.textChanged.connect(self._iniciar_timer_busca_fisicos)
   

        self.main_window = main_window
        self.table_clientes_fisicos = self.main_window.table_clientes_fisicos # Referência para a tabela no main window
        self.btn_adicionar_cliente_fisico = btn_adicionar_cliente_fisico
        self.btn_editar_clientes_fisicos = btn_editar_clientes_fisicos
        self.btn_excluir_clientes_fisicos = btn_excluir_clientes_fisicos
        self.btn_gerar_relatorio_clientes_fisicos = btn_gerar_relatorio_clientes_fisicos
        self.btn_historico_clientes_fisicos = btn_historico_clientes_fisicos
        self.btn_marcar_como_clientes_fisicos = btn_marcar_como_clientes_fisicos


        self.btn_excluir_clientes_fisicos.clicked.connect(self.excluir_clientes_fisicos)
        self.btn_marcar_como_clientes_fisicos.clicked.connect(self.marcar_como_clientes_fisicos)
        self.btn_adicionar_cliente_fisico.clicked.connect(self.exibir_janela_cadastro_cliente_fisico)
        self.btn_editar_clientes_fisicos.clicked.connect(self.editar_cliente_fisico)
        self.btn_historico_clientes_fisicos.clicked.connect(self.historico_clientes_fisicos)
        self.btn_gerar_relatorio_clientes_fisicos.clicked.connect(self.abrir_janela_relatorio_clientes_fisicos)
        self.imagem_line_fisico()
        
        
       # ENTER → busca manual
        self.line_clientes_fisicos.returnPressed.connect(
            lambda: self.buscar_cliente_fisico_dinamico(manual=True)
        )

        # Botão lupa → busca manual
        self.botao_lupa_fisicos.clicked.connect(
            lambda: self.buscar_cliente_fisico_dinamico(manual=True)
        )

        # Digitação → busca dinâmica
        self.line_clientes_fisicos.textChanged.connect(
            lambda: self.buscar_cliente_fisico_dinamico(manual=False)
        )

        self.configurar_menu_contexto_fisicos()
        

        
        
        self.installEventFilter(self)
        self.table_clientes_fisicos.viewport().installEventFilter(self)


    def _iniciar_timer_busca_fisicos(self, texto):
        self.ultimo_texto = texto
        self.timer_buscar_fisicos.start(300)  # espera 300ms após digitar

    def _executar_busca_dinamica_fisicos(self):
        self.buscar_cliente_fisico_dinamico(self.ultimo_texto)

    def _buscar_clientes_fisicos(self, texto):
        """Executa a busca no banco e retorna lista de resultados."""
        if not isinstance(texto, str):
            return []

        texto = texto.strip()
        if not texto:
            return self._listar_todos_clientes()

        with sqlite3.connect('banco_de_dados.db') as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT  
                    "Nome do Cliente",
                    "Data da Inclusão",
                    RG,       
                    CPF,
                    Email,
                    CNH,
                    "Categoria da CNH",
                    "Data de Emissão da CNH",
                    "Data de Vencimento da CNH",
                    Telefone,
                    CEP,
                    Endereço,
                    Número,
                    Complemento,
                    Cidade,
                    Bairro,
                    Estado,
                    "Status do Cliente",
                    "Categoria do Cliente",
                    "Última Atualização",
                    "Origem do Cliente",
                    "Valor Gasto Total",
                    "Última Compra"
                FROM clientes_fisicos
                WHERE 
                    LOWER("Nome do Cliente") LIKE LOWER(?) OR
                    CPF LIKE ? OR
                    RG LIKE ? OR
                    Telefone LIKE ?
            """, (f'%{texto}%', f'%{texto}%', f'%{texto}%', f'%{texto}%'))
            return cursor.fetchall()

    def _listar_todos_clientes(self):
        with sqlite3.connect('banco_de_dados.db') as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT  
                    "Nome do Cliente",
                    "Data da Inclusão",
                    RG,       
                    CPF,
                    Email,
                    CNH,
                    "Categoria da CNH",
                    "Data de Emissão da CNH",
                    "Data de Vencimento da CNH",
                    Telefone,
                    CEP,
                    Endereço,
                    Número,
                    Complemento,
                    Cidade,
                    Bairro,
                    Estado,
                    "Status do Cliente",
                    "Categoria do Cliente",
                    "Última Atualização",
                    "Origem do Cliente",
                    "Valor Gasto Total",
                    "Última Compra"
                FROM clientes_fisicos
            """)
            return cursor.fetchall()

    def buscar_cliente_fisico_dinamico(self, texto=None,manual=None):
        # Captura o texto do campo se não foi passado
        if not isinstance(texto, str):
            texto = self.line_clientes_fisicos.text()
            
        texto = texto.strip()
        
        if not texto:
            if manual:
                QMessageBox.warning(self, "Atenção", "Por favor, digite um cliente para pesquisar.")
            else:
                # Busca dinâmica sem texto → mostra todos
                self.preencher_resultado_busca_fisica(self._listar_todos_clientes())
            return

        resultados = self._buscar_clientes_fisicos(texto)

        if not resultados and manual:
            QMessageBox.information(self, "Resultado", "Nenhum cliente encontrado.")

        self.preencher_resultado_busca_fisica(resultados)


    def buscar_cliente_fisico_manual(self):
        texto = self.line_clientes_fisicos.text().strip()
        if not texto:
            self.preencher_resultado_busca_fisica(self._listar_todos_clientes())
            return

        resultados = self._buscar_clientes_fisicos(texto)
        if resultados:
            self.preencher_resultado_busca_fisica(resultados)
        else:
            QMessageBox.information(self, "Resultado", "Nenhum cliente encontrado.")
            self.preencher_resultado_busca_fisica(self._listar_todos_clientes())

    def preencher_resultado_busca_fisica(self, resultados):
        self.table_clientes_fisicos.setRowCount(0)

        for row_data in resultados:
            row_index = self.table_clientes_fisicos.rowCount()
            self.table_clientes_fisicos.insertRow(row_index)
            for col_index, data in enumerate(row_data):
                item = self.formatar_texto_fisico(str(data))
                self.table_clientes_fisicos.setItem(row_index, col_index, item)
        self.table_clientes_fisicos.resizeColumnsToContents()
        self.table_clientes_fisicos.resizeRowsToContents()


    # Função auxiliar para criar um QTableWidgetItem com texto centralizado
    def formatar_texto_fisico(self, text):
        item = QTableWidgetItem(text)
        item.setTextAlignment(Qt.AlignCenter)  # Centraliza o texto
        item.setForeground(QBrush(QColor("white")))
        return item

    
    def carregar_clientes_fisicos(self):
        try:
            with self.db.connecta() as conexao:
                cursor = conexao.cursor()

                cursor.execute("""
                    SELECT 
                        "Nome do Cliente",
                        "Data da Inclusão",
                        RG,      
                        CPF, 
                        Email,
                        CNH,
                        "Categoria da CNH",
                        "Data de Emissão da CNH",
                        "Data de Vencimento da CNH",
                        Telefone, 
                        CEP, 
                        Endereço, 
                        Número,
                        Complemento,
                        Cidade, 
                        Bairro,
                        Estado,
                        "Status do Cliente", 
                        "Categoria do Cliente",
                        "Última Atualização",
                        "Valor Gasto Total",
                        "Última Compra"
                    FROM clientes_fisicos
                    ORDER BY "Data da Inclusão" DESC
                """)

                dados = cursor.fetchall()

                # Limpa a tabela antes de recarregar
                self.table_clientes_fisicos.clearContents()
                self.table_clientes_fisicos.setRowCount(0)

                deslocamento = 1 if self.coluna_checkboxes_clientes_fisicos_adicionada else 0
                self.checkboxes_clientes_fisicos = []

                for linha_idx, linha_dados in enumerate(dados):
                    self.table_clientes_fisicos.insertRow(linha_idx)

                    # Adiciona checkbox se estiver ativado
                    if self.coluna_checkboxes_clientes_fisicos_adicionada:
                        checkbox = QCheckBox()
                        checkbox.setStyleSheet("margin-left: 9px; margin-right: 9px;")
                        self.table_clientes_fisicos.setCellWidget(linha_idx, 0, checkbox)
                        self.checkboxes_clientes_fisicos.append(checkbox)

                    # Preenche os dados nas colunas (com deslocamento)
                    for coluna_idx, dado in enumerate(linha_dados):
                        item = self.formatar_texto_fisico(str(dado))
                        self.table_clientes_fisicos.setItem(linha_idx, coluna_idx + deslocamento, item)

                self.table_clientes_fisicos.resizeColumnsToContents()
                self.table_clientes_fisicos.resizeRowsToContents()

        except Exception as e:
            QMessageBox.critical(self.main_window, "Erro", f"Erro ao carregar clientes:\n{e}")

    def imagem_line_fisico(self):
        # Criar botão da lupa
        self.botao_lupa_fisicos = QPushButton(self.line_clientes_fisicos)
        self.botao_lupa_fisicos.setCursor(Qt.PointingHandCursor)  # Muda o cursor ao passar o mouse
        self.botao_lupa_fisicos.setObjectName("botao_lupa_fisicos")
        
        # Definir tamanho do botão
        altura = self.line_clientes_fisicos.height() - 4  # Ajustar altura conforme a LineEdit
        self.botao_lupa_fisicos.setFixedSize(altura, altura)

        # Posicionar o botão no canto direito da LineEdit
        self.botao_lupa_fisicos.move(self.line_clientes_fisicos.width() - altura + 160, 2)

        # Conectar clique do botão a uma função
        self.botao_lupa_fisicos.clicked.connect(self.buscar_cliente_fisico_dinamico)


    

    def formatar_e_buscar_cep_fisico(self, widget):
        texto_cep = widget.text()
        self.main_window.formatar_cep(texto_cep, widget)

        dados = self.main_window.buscar_cep(texto_cep)
        if dados:
            self.campos_cliente_fisico["Endereço"].setText(dados.get("logradouro", ""))
            self.campos_cliente_fisico["Complemento"].setText(dados.get("complemento", ""))
            self.campos_cliente_fisico["Bairro"].setText(dados.get("bairro", ""))
            self.campos_cliente_fisico["Cidade"].setText(dados.get("localidade", ""))
            
            estado = dados.get("uf", "")
            index_estado = self.campos_cliente_fisico["Estado"].findText(estado)
            if index_estado >= 0:
                self.campos_cliente_fisico["Estado"].setCurrentIndex(index_estado)

    

    def editar_cliente_fisico(self):
        linha_selecionada = self.table_clientes_fisicos.currentRow()
        if linha_selecionada < 0:
            QMessageBox.warning(self, "Aviso", "Nenhum cliente selecionado para edição.")
            return
        
        coluna_offset = 1 if self.coluna_checkboxes_clientes_fisicos_adicionada else 0

        colunas = [
            "Nome do Cliente", "Data da Inclusão", "RG","CPF","Email","CNH","Categoria da CNH","Data de Emissão da CNH",
            "Data de Vencimento da CNH","Telefone","CEP", "Endereço", "Número", "Complemento", "Cidade", "Bairro",
            "Estado", "Status do Cliente","Categoria do Cliente","Última Atualização","Valor Gasto Total", "Última Compra"
        ]
        

        dados_cliente = {}
        for col, nome_coluna in enumerate(colunas):
            item = self.table_clientes_fisicos.item(linha_selecionada, col + coluna_offset)
            dados_cliente[nome_coluna] = item.text() if item else ""    

        self.exibir_edicao_clientes_fisicos(dados_cliente)   

    
    def carregar_config(self):
        with open("config.json", "r", encoding="utf-8") as f:
            return json.load(f)

    def exibir_janela_cadastro_cliente_fisico(self):
        self.campos_cliente_fisico = {}
        self.janela_cadastro_fisico = QMainWindow()


        self.janela_cadastro_fisico.resize(700, 550)
        self.janela_cadastro_fisico.setWindowTitle("Cadastro de Cliente Físico")
        self.janela_cadastro_fisico.setObjectName("janela_cadastro_fisico")

        # Centralizar a janela
        screen = QGuiApplication.primaryScreen()
        screen_geometry = screen.availableGeometry()
        window_geometry = self.janela_cadastro_fisico.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.janela_cadastro_fisico.move(window_geometry.topLeft())
        
        # Carregar tema
        config = self.carregar_config()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(60,60,60),
                                                stop:1 rgb(100,100,100));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid #666666;
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
            """
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """

        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(0,120,180),
                                                stop:1 rgb(0,150,220));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(0,100,160);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #007acc;
                }
                QPushButton:pressed {
                    background-color: #006bb3;
                    border: 2px solid #005c99;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 3px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            scroll_style = """
                QScrollBar:vertical {
                border: none;
                background-color: rgb(255, 255, 256); /* branco */
                width: 30px;
                margin: 0px 10px 0px 10px;
            }
             QScrollBar::handle:vertical {
                background-color: rgb(180, 180,180);  /* cinza */
                min-height: 30px;
                border-radius: 5px;
            }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        # Conteúdo do formulário
        conteudo = QWidget()
        layout = QVBoxLayout(conteudo)
        conteudo.setStyleSheet(f"background-color: {bg_cor}; color: {text_cor};")

        def add_linha(titulo, widget=None):
            label = QLabel(titulo)
            label.setStyleSheet("color: white;")
            layout.addWidget(label)
            if widget is None:
                widget = QLineEdit()
                widget.setStyleSheet(lineedit_style)
            layout.addWidget(widget)
            chave_sem_ponto = titulo.rstrip(":")
            self.campos_cliente_fisico[chave_sem_ponto] = widget
            return label, widget

        # ComboBox Categoria CNH criado antecipadamente para uso no formulário
        combobox_categoria_cnh_fisico = QComboBox()
        combobox_categoria_cnh_fisico.addItems(["Selecionar", "AB", "A", "B", "C", "D", "E", "Nenhuma"])
        combobox_categoria_cnh_fisico.setCurrentIndex(0)
        combobox_categoria_cnh_fisico.setStyleSheet(combobox_style)


        add_linha("Nome do Cliente")
        add_linha("RG")
        self.campos_cliente_fisico["RG"].setPlaceHolderText("Opcional")
        rg_widget = self.campos_cliente_fisico["RG"]
        rg_widget.textChanged.connect(lambda text, w=rg_widget: self.main_window.formatar_rg(text,w))
        add_linha("CPF")
        cpf_widget = self.campos_cliente_fisico["CPF"]
        cpf_widget.textChanged.connect(lambda text,w=cpf_widget: self.main_window.formatar_cpf(text, w))
        add_linha("Email")
        email_widget = self.campos_cliente_fisico["Email"]
        email_widget.textChanged.connect(lambda text: self.main_window.validar_email(text,email_widget))
        add_linha("CNH")
        self.campos_cliente_fisico["CNH"].setPlaceholderText("Opcional")
        label_categoria_cnh, widget_categoria_cnh = add_linha("Categoria da CNH", combobox_categoria_cnh_fisico)

        # Campos Data Emissão e Vencimento CNH, inicialmente escondidos
        label_emissao_cnh, widget_emissao_cnh = add_linha("Data de Emissão da CNH")
        label_vencimento_cnh, widget_vencimento_cnh = add_linha("Data de Vencimento da CNH")

        # Aplicar formatação de data (mesma usada em Data de Nascimento)
        widget_emissao_cnh.textChanged.connect(lambda text: self.main_window.formatar_data_nascimento(text, widget_emissao_cnh))
        widget_vencimento_cnh.textChanged.connect(lambda text: self.main_window.formatar_data_nascimento(text, widget_vencimento_cnh))
        label_emissao_cnh.hide()
        widget_emissao_cnh.hide()
        label_vencimento_cnh.hide()
        widget_vencimento_cnh.hide()

        cnh_widget = self.campos_cliente_fisico["CNH"]
        cnh_widget.textChanged.connect(lambda text: self.main_window.formatar_cnh(text, cnh_widget))

        # Função para mostrar/esconder datas da CNH conforme categoria selecionada
        def on_categoria_cnh_change_fisicos(text):
            if text not in ("Selecionar", "Nenhuma"):
                label_emissao_cnh.show()
                widget_emissao_cnh.show()
                label_vencimento_cnh.show()
                widget_vencimento_cnh.show()
            else:
                label_emissao_cnh.hide()
                widget_emissao_cnh.hide()
                label_vencimento_cnh.hide()
                widget_vencimento_cnh.hide()
                widget_emissao_cnh.clear()
                widget_vencimento_cnh.clear()

        combobox_categoria_cnh_fisico.currentTextChanged.connect(on_categoria_cnh_change_fisicos)


        
        add_linha("Telefone")
        telefone_widget = self.campos_cliente_fisico["Telefone"]
        telefone_widget.textChanged.connect(lambda text: self.main_window.formatar_telefone(text, telefone_widget))        
        add_linha("CEP")
        cep_widget = self.campos_cliente_fisico["CEP"]
        # Formatação do CEP enquanto digita
        cep_widget.textChanged.connect(lambda text: self.main_window.formatar_cep(text,cep_widget))
        # Buscar dados do CEP ao terminar de digitar
        cep_widget.editingFinished.connect(lambda: self.on_cep_editing_finished_cadastro(cep_widget))
        add_linha("Endereço")
        add_linha("Número")
        add_linha("Complemento")
        self.campos_cliente_fisico["Complemento"].setPlaceholderText("Opcional")
        add_linha("Cidade")
        add_linha("Bairro")

        # Aqui adiciona Estado logo após Bairro
        combobox_estado_cliente_fisico = QComboBox()
        combobox_estado_cliente_fisico.addItems([
            "Selecionar","AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT",
            "MS","MG","PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"
        ])
        combobox_estado_cliente_fisico.setCurrentIndex(0) # Seleciona "Selecionar" por padrão
        combobox_estado_cliente_fisico.setStyleSheet(combobox_style)
        add_linha("Estado", combobox_estado_cliente_fisico)
        add_linha("Categoria do Cliente")

        combobox_status_cliente = QComboBox()
        combobox_status_cliente.addItems(["Selecionar","Ativo","Inativo","Pendente","Bloqueado"])
        combobox_status_cliente.setCurrentIndex(0) # Seleciona "Selecionar" por padrão
        combobox_status_cliente.setStyleSheet(combobox_style)
        add_linha("Status do Cliente:", combobox_status_cliente)    

        btn_fazer_cadastro_fisico = QPushButton("Fazer o Cadastro")
        btn_fazer_cadastro_fisico.clicked.connect(self.cadastrar_clientes_fisicos)
        btn_fazer_cadastro_fisico.setStyleSheet(button_style)
        layout.addWidget(btn_fazer_cadastro_fisico)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(conteudo)
        scroll_area.setStyleSheet(scroll_style)

        self.janela_cadastro_fisico.setCentralWidget(scroll_area)
        self.janela_cadastro_fisico.show()

    def exibir_edicao_clientes_fisicos(self, dados_cliente: dict):
        self.dados_originais_cliente_fisico = dados_cliente.copy()
        self.alteracoes_realizadas = False
        # Inicializa o dicionário para armazenar os widgets
        self.campos_cliente_fisico = {}

        self.janela_editar_cliente_fisico = QMainWindow()
        self.janela_editar_cliente_fisico.setWindowTitle("Editar Cliente Físico")
        self.janela_editar_cliente_fisico.resize(683, 600)
        self.janela_editar_cliente_fisico.setStyleSheet("background-color: rgb(0, 80, 121);")

        screen = QGuiApplication.primaryScreen()
        screen_geometry = screen.availableGeometry()
        window_geometry = self.janela_editar_cliente_fisico.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.janela_editar_cliente_fisico.move(window_geometry.topLeft())
        
        
        # Carregar tema
        config = self.carregar_config()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(60,60,60),
                                                stop:1 rgb(100,100,100));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid #666666;
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
            """
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """

        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(0,120,180),
                                                stop:1 rgb(0,150,220));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(0,100,160);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #007acc;
                }
                QPushButton:pressed {
                    background-color: #006bb3;
                    border: 2px solid #005c99;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 3px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            scroll_style = """
                QScrollBar:vertical {
                border: none;
                background-color: rgb(255, 255, 255); /* branco */
                width: 30px;
                margin: 0px 10px 0px 10px;
            }
             QScrollBar::handle:vertical {
                background-color: rgb(180, 180,180);  /* cinza */
                min-height: 30px;
                border-radius: 5px;
            }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        # Widget central e layout
        conteudo = QWidget()
        conteudo.setStyleSheet(f"background-color: {bg_cor}; color: {text_cor};")
        layout = QVBoxLayout(conteudo)

        # Função auxiliar para adicionar campos
        def add_linha(titulo, widget=None):
            label = QLabel(titulo)
            label.setStyleSheet(f"color: {text_cor};")
            layout.addWidget(label)
            if widget is None:
                widget = QLineEdit()
                widget.setStyleSheet(lineedit_style)
                # Preencher valor existente
                chave = titulo.rstrip(":")
                widget.setText(dados_cliente.get(chave, ""))
            layout.addWidget(widget)
            chave_sem_ponto = titulo.rstrip(":")
            self.campos_cliente_fisico[chave_sem_ponto] = widget
            return label, widget

        add_linha("Nome do Cliente")
        add_linha("Data da Inclusão")
        add_linha("RG")
        add_linha("CPF")
        add_linha("Email")
        add_linha("CNH")

        # Categoria CNH
        combobox_categoria_cnh = QComboBox()
        combobox_categoria_cnh.addItems(["Selecionar","AB","A","B","C","D","E","Nenhuma"])
        valor_categoria = dados_cliente.get("Categoria da CNH", "Selecionar")
        index_categoria = combobox_categoria_cnh.findText(valor_categoria)
        combobox_categoria_cnh.setCurrentIndex(index_categoria if index_categoria >=0 else 0)
        combobox_categoria_cnh.setStyleSheet(combobox_style)
        label_categoria, widget_categoria = add_linha("Categoria da CNH", combobox_categoria_cnh)


        # Datas CNH
        label_emissao_cnh, widget_emissao_cnh = add_linha("Data de Emissão da CNH")
        widget_emissao_cnh.setText(dados_cliente.get("Data de Emissão da CNH",""))
        label_vencimento_cnh, widget_vencimento_cnh = add_linha("Data de Vencimento da CNH")
        widget_vencimento_cnh.setText(dados_cliente.get("Data de Vencimento da CNH",""))

        # Mostrar/esconder datas conforme categoria
        def on_categoria_cnh_change(text):
            if text not in ("Selecionar","Nenhuma"):
                label_emissao_cnh.show()
                widget_emissao_cnh.show()
                label_vencimento_cnh.show()
                widget_vencimento_cnh.show()
            else:
                label_emissao_cnh.hide()
                widget_emissao_cnh.hide()
                label_vencimento_cnh.hide()
                widget_vencimento_cnh.hide()
                widget_emissao_cnh.clear()
                widget_vencimento_cnh.clear()


        
        combobox_categoria_cnh.currentTextChanged.connect(on_categoria_cnh_change)
        on_categoria_cnh_change(valor_categoria)

        add_linha("Telefone")
        add_linha("CEP")
        add_linha("Endereço")
        add_linha("Número")
        add_linha("Complemento")
        add_linha("Cidade")
        add_linha("Bairro")

        # Estado
        combobox_estado_cliente = QComboBox()
        combobox_estado_cliente.addItems([
            "Selecionar","AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS","MG",
            "PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"
        ])
        valor_estado = dados_cliente.get("Estado", "Selecionar")
        index_estado = combobox_estado_cliente.findText(valor_estado)
        combobox_estado_cliente.setCurrentIndex(index_estado if index_estado >=0 else 0)
        combobox_estado_cliente.setStyleSheet(combobox_style)
        add_linha("Estado", combobox_estado_cliente)

        # Categoria do cliente
        add_linha("Categoria do Cliente")

         # Status do cliente
        combobox_status_cliente = QComboBox()
        combobox_status_cliente.addItems(["Selecionar","Ativo","Inativo","Pendente","Bloqueado"])
        valor_status = dados_cliente.get("Status do Cliente","Selecionar")
        index_status = combobox_status_cliente.findText(valor_status)
        combobox_status_cliente.setCurrentIndex(index_status if index_status >=0 else 0)
        combobox_status_cliente.setStyleSheet(combobox_style)
        add_linha("Status do Cliente", combobox_status_cliente)

        add_linha("Última Atualização")


        # Valor gasto total
        add_linha("Valor Gasto Total")
        self.campos_cliente_fisico["Valor Gasto Total"].setText(dados_cliente.get("Valor Gasto Total",""))

        # Última Compra
        add_linha("Última Compra")
        self.campos_cliente_fisico["Última Compra"].setText(dados_cliente.get("Última Compra",""))

     
        # Conectar formatações
        self.campos_cliente_fisico["RG"].textChanged.connect(
            lambda text: self.main_window.formatar_rg(text, self.campos_cliente_fisico["RG"])
        )
        self.campos_cliente_fisico["CPF"].textChanged.connect(
            lambda text: self.main_window.formatar_cpf(text, self.campos_cliente_fisico["CPF"])
        )
        self.campos_cliente_fisico["Email"].textChanged.connect(
            lambda text: self.main_window.validar_email(text, self.campos_cliente_fisico["Email"])
        )
        self.campos_cliente_fisico["CNH"].textChanged.connect(
            lambda text: self.main_window.formatar_cnh(text, self.campos_cliente_fisico["CNH"])
        )
        widget_emissao_cnh.textChanged.connect(
            lambda text: self.main_window.formatar_data_nascimento(text, widget_emissao_cnh)
        )
        widget_vencimento_cnh.textChanged.connect(
            lambda text: self.main_window.formatar_data_nascimento(text, widget_vencimento_cnh)
        )
        self.campos_cliente_fisico["Telefone"].textChanged.connect(
            lambda text: self.main_window.formatar_telefone(text, self.campos_cliente_fisico["Telefone"])
        )
        self.campos_cliente_fisico["CEP"].textChanged.connect(
            lambda text: self.main_window.formatar_cep(text, self.campos_cliente_fisico["CEP"])
        )
        self.campos_cliente_fisico["CEP"].editingFinished.connect(
            lambda: self.formatar_e_buscar_cep(self.campos_cliente_fisico["CEP"])
        )
        self.campos_cliente_fisico["Valor Gasto Total"].editingFinished.connect(
            lambda: self.main_window.formatar_moeda(self.campos_cliente_fisico["Valor Gasto Total"])
        )

         # Campos de data conectados corretamente
        campos_data = ["Data de Emissão da CNH", "Data de Vencimento da CNH",
                   "Última Compra", "Data da Inclusão", "Última Atualização"]
        
        for campo in campos_data:
            if campo in self.campos_cliente_fisico:
                widget = self.campos_cliente_fisico[campo]
                widget.textChanged.connect(
                    lambda texto, w=widget: self.main_window.formatar_data_nascimento(texto, w)
                )
                widget.editingFinished.connect(
                    lambda w=widget: self.main_window.validar_data_quando_finalizar(w.text(), w)
                )
        for campo, widget in self.campos_cliente_fisico.items():
            if isinstance(widget, QLineEdit):
                widget.textChanged.connect(self.marcar_alteracao)
            elif isinstance(widget, QComboBox):
                widget.currentTextChanged.connect(self.marcar_alteracao)

        
        
        # Botão atualizar
        botao_atualizar = QPushButton("Atualizar")
        botao_atualizar.setStyleSheet(button_style)
        botao_atualizar.clicked.connect(self.atualizar_dados_clientes_fisicos)
        layout.addWidget(botao_atualizar)


        # Scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(conteudo)
        scroll_area.setStyleSheet(scroll_style)

        self.janela_editar_cliente_fisico.setCentralWidget(scroll_area)
        self.janela_editar_cliente_fisico.show()

    def informacoes_obrigatorias_cadastro_clientes_fisicos(self):
        self.campos_obrigatorios_clientes_fisicos = {
            "Nome do Cliente": "O campo Nome do Cliente é obrigatório.",
            "CPF": "O campo de CPF é obrigatório",
            "Email": "O campo de E-mail é obrigatório",
            "Telefone": "O campo Telefone é obrigatório.",
            "CEP": "O campo CEP é obrigatório.",
            "Endereço": "O campo Endereço é obrigatório.",
            "Número": "O campo Número é obrigatório.",
            "Cidade": "O campo Cidade é obrigatório.",
            "Bairro": "O campo Bairro é obrigatório.",
            "Estado": "O campo de Estado é obrigatório",
            "Categoria do Cliente": "O campo Categoria do Cliente é obrigatório.",
            "Status do Cliente": "Você deve selecionar um status válido para o cliente.",
        }

        # Se CNH foi informada, exige os outros campos relacionados
        cnh_widget = self.campos_cliente_fisico.get("CNH")
        cnh_valor = ""
        if isinstance(cnh_widget, QLineEdit):
            cnh_valor = cnh_widget.text().strip()
        elif isinstance(cnh_widget, QComboBox):
            cnh_valor = cnh_widget.currentText().strip()
                
        # Se CNH tiver valor, obriga mais campos
        if cnh_valor and cnh_valor.lower() != "não cadastrado" and cnh_valor.lower() != "selecione":
            self.campos_obrigatorios_clientes_fisicos.update({
                "Categoria da CNH": "Informe a Categoria da CNH.",
                "Data de Emissão da CNH": "Informe a Data de Emissão da CNH.",
                "Data de Vencimento da CNH": "Informe a Data de Vencimento da CNH."
            })

        # Validação genérica
        for campo, mensagem_erro in self.campos_obrigatorios_clientes_fisicos.items():
            widget = self.campos_cliente_fisico.get(campo)
            if widget is None:
                continue
            
            valor = ""
            if isinstance(widget, QLineEdit):
                valor = widget.text().strip()
            elif isinstance(widget, QComboBox):
                valor = widget.currentText().strip()
                if valor.lower() == "selecione":
                    valor = ""
            else:
                valor = str(widget).strip()

            print(f"Validando campo '{campo}': valor capturado '{valor}'")  # <-- print para debug

        return True
    
    def informacoes_obrigatorias_edicao_clientes_fisicos(self):
        """Valida também campos automáticos na edição de clientes físicos"""
        # Primeiro roda a validação padrão de cadastro
        self.informacoes_obrigatorias_cadastro_clientes_fisicos()

        # Adiciona os extras que só são exigidos na edição
        self.campos_obrigatorios_clientes_fisicos.update({
            "Última Atualização": "O campo Última Atualização é obrigatório.",
            "Data da Inclusão": "O campo Data da Inclusão é obrigatório.",
            "Valor Gasto Total": "O campo Valor Gasto Total é obrigatório.",
            "Última Compra": "O campo Última Compra é obrigatório."
        })

        # Validação genérica novamente
        for campo, mensagem_erro in self.campos_obrigatorios_clientes_fisicos.items():
            widget = self.campos_cliente_fisico.get(campo)
            if widget is None:
                continue

            valor = ""
            if isinstance(widget, QLineEdit):
                valor = widget.text().strip()
            elif isinstance(widget, QComboBox):
                valor = widget.currentText().strip()
                if valor.lower() in ["selecione", "selecionar"]:
                    valor = ""
            else:
                valor = str(widget).strip()

            print(f"[Edição Físico] Validando campo '{campo}': valor '{valor}'")  # debug

        return True

    def cadastrar_clientes_fisicos(self):
        # Atualiza a lista de campos obrigatórios de acordo com o preenchimento
        self.informacoes_obrigatorias_cadastro_clientes_fisicos()
        try:
            with self.db.connecta() as conexao:
                cursor = conexao.cursor()
                usuario_logado = self.config.obter_usuario_logado()

                # Coletar dados dos campos
                get = lambda campo: self.campos_cliente_fisico[campo].text().strip() \
                    if isinstance(self.campos_cliente_fisico[campo], QLineEdit) \
                    else self.campos_cliente_fisico[campo].currentText()

                nome = get("Nome do Cliente")
                rg = get("RG")
                cpf = get("CPF")
                email = get("Email")
                cnh = get("CNH")
                categoria_cnh = get("Categoria da CNH")
                data_emissao_cnh = get("Data de Emissão da CNH")
                data_vencimento_cnh = get("Data de Vencimento da CNH")
                telefone = get("Telefone")
                cep = get("CEP")
                endereco = get("Endereço")
                numero = get("Número")
                complemento = get("Complemento")
                cidade = get("Cidade")
                bairro = get("Bairro")
                estado = get("Estado")
                categoria = get("Categoria do Cliente")
                status = get("Status do Cliente")

                # Verificar campos únicos individualmente
                cursor.execute('SELECT 1 FROM clientes_fisicos WHERE RG = ?',(rg,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com este RG.")
                    return
                cursor.execute("SELECT 1 FROM clientes_fisicos WHERE CPF = ?",(cpf,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com este CPF.")
                    return
                cursor.execute('SELECT 1 FROM clientes_fisicos WHERE Email = ?',(email,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com este Email.")
                    return
                cursor.execute('SELECT 1 FROM clientes_fisicos WHERE RG = ?',(telefone,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com este Telefone.")
                    return

                cursor.execute('SELECT 1 FROM clientes_fisicos WHERE CNH = ?',(cnh,))
                if cursor.fetchone():
                    QMessageBox.information(self,"Duplicidade","Já existe um cliente cadastrado com esta CNH.")
                    return

                # Validação de todos os campos obrigatórios
                for campo, mensagem in self.campos_obrigatorios_clientes_fisicos.items():
                    widget = self.campos_cliente_fisico[campo]
                    valor = widget.text().strip() if isinstance(widget, QLineEdit) else widget.currentText()
                    
                    if not valor or (isinstance(widget, QComboBox) and valor == "Selecionar"):
                        QMessageBox.warning(self, "Atenção", mensagem)
                        return
                    

                data_inclusao = datetime.now().strftime("%d/%m/%Y %H:%M")
                
                # Valores padrão para os campos não preenchidos manualmente
                valor_gasto_total = "Não Cadastrado"
                ultima_compra = "Não Cadastrado"
                ultima_atualizacao = "Não Cadastrado"

                if not cnh:
                    cnh = "Não Cadastrado"
                    categoria_cnh = "Não Cadastrado"
                    data_emissao_cnh = "Não Cadastrado"
                    data_vencimento_cnh = "Não Cadastrado"

                if not complemento:
                    complemento = "Não se aplica"
                
                # Inserir no banco
                cursor.execute("""
                    INSERT INTO clientes_fisicos(
                        "Nome do Cliente", "Data da Inclusão",RG, CPF,Email, CNH, "Categoria da CNH", "Data de Emissão da CNH", 
                        "Data de Vencimento da CNH",  Telefone, CEP, Endereço, Número, Complemento, Cidade, Bairro, Estado, 
                        "Status do Cliente", "Categoria do Cliente", "Última Atualização", "Valor Gasto Total", "Última Compra"
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    nome, data_inclusao, rg, cpf,email, cnh, categoria_cnh, data_emissao_cnh, data_vencimento_cnh, 
                    telefone, cep, endereco, numero, complemento, cidade, bairro, estado, 
                    status, categoria, ultima_atualizacao, valor_gasto_total,ultima_compra
                ))
                conexao.commit()

                self.main_window.registrar_historico_clientes_fisicos(
                    "Cadastro de Cliente", f"Cliente {nome} cadastrado com sucesso"
                )
                
                QMessageBox.information(self, "Sucesso", "Cliente cadastrado com sucesso!")
                # Redimensiona apenas uma vez após preencher
                self.table_clientes_fisicos.resizeColumnsToContents()
                self.table_clientes_fisicos.resizeRowsToContents()
                self.limpar_campos_clientes_fisicos()
                self.carregar_clientes_fisicos()       
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao cadastrar cliente: \n{e}")
            
    def atualizar_dados_clientes_fisicos(self):
        # Atualiza campos obrigatórios de acordo com a CNH
        if not self.informacoes_obrigatorias_edicao_clientes_fisicos():
            return
        
        if not self.alteracoes_realizadas:
            QMessageBox.information(self, "Sem alterações", "Nenhuma modificação foi feita.")
            return
        
        # --- VALIDAÇÃO DOS CAMPOS OBRIGATÓRIOS ---
        for campo, mensagem in self.campos_obrigatorios_clientes_fisicos.items():
            widget = self.campos_cliente_fisico[campo]
            if isinstance(widget, QLineEdit):
                valor = widget.text().strip()
            elif isinstance(widget, QComboBox):
                valor = widget.currentText().strip()
            else:
                valor = str(widget).strip()

            if not valor or (isinstance(widget, QComboBox) and valor.lower() == "selecione"):
                QMessageBox.warning(self, "Atenção", mensagem)
                return


        # Monta dados atualizados
        dados_atualizados = {}
        for campo, widget in self.campos_cliente_fisico.items():
            if isinstance(widget, QComboBox):
                dados_atualizados[campo] = widget.currentText()
            else:
                dados_atualizados[campo] = widget.text()

        # --- VERIFICAÇÃO DE CAMPOS SENSÍVEIS ---
        campos_sensiveis_fisicos = ["Data da Inclusão", "Última Atualização", "Valor Gasto Total", "Última Compra"]
        alterou_campo_sensivel_fisico = any(
            dados_atualizados[c] != self.dados_originais_cliente_fisico[c]
            for c in campos_sensiveis_fisicos
        )
        if alterou_campo_sensivel_fisico:
            try:
                with open("config.json", "r", encoding="utf-8") as f:
                    config = json.load(f)
                    senha_correta = config.get("senha", "")
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Não foi possível carregar a senha do sistema\n {e}")
                return
            tentativas = 0
            while tentativas < 3:
                senha, ok = QInputDialog.getText(
                    self, "Confirmação de Segurança",
                    "Digite a senha do sistema:", QLineEdit.Password
                )
                if not ok:  # Cancelou
                    return
                
                if senha.strip() == senha_correta.strip():
                    break

                tentativas += 1
                if tentativas < 3:
                    QMessageBox.critical(self, "Acesso Negado",
                                        f"Senha incorreta. Tentativas restantes: {3 - tentativas}")
                else:
                    QMessageBox.critical(self, "Acesso Negado",
                                        "Você excedeu o número máximo de tentativas.\nO sistema será encerrado.")
                    QApplication.quit()
                    return

        try:
            cursor = self.db.connection.cursor()

            # CPF ORIGINAL — garante que vamos atualizar o cliente certo
            cpf_original = self.dados_originais_cliente_fisico["CPF"]

            # Corrigir campos vazios antes do UPDATE
            if not dados_atualizados["CNH"]:
                dados_atualizados["CNH"] = "Não Cadastrado"
                dados_atualizados["Categoria da CNH"] = "Não Cadastrado"
                dados_atualizados["Data de Emissão da CNH"] = "Não Cadastrado"
                dados_atualizados["Data de Vencimento da CNH"] = "Não Cadastrado"

            if not dados_atualizados["Complemento"]:
                dados_atualizados["Complemento"] = "Não se aplica"

            query = """
            UPDATE clientes_fisicos SET
                `Nome do Cliente` = ?, `Data da Inclusão` = ?,
                RG = ?, CPF = ?, CNH = ?, `Categoria da CNH` = ?, `Data de Emissão da CNH` = ?, 
                `Data de Vencimento da CNH` = ?, Telefone = ?, CEP = ?, Endereço = ?, 
                Número = ?, Complemento = ?, Cidade = ?, Bairro = ?, Estado = ?, `Status do Cliente` = ?, `Categoria do Cliente` = ?,
                `Última Atualização` = ?, `Valor Gasto Total` = ?, `Última Compra` = ?
            WHERE CPF = ?
            """

            valores = (
                dados_atualizados["Nome do Cliente"], dados_atualizados["Data da Inclusão"],
                dados_atualizados["RG"], dados_atualizados["CPF"], dados_atualizados["CNH"], dados_atualizados["Categoria da CNH"], dados_atualizados["Data de Emissão da CNH"], 
                dados_atualizados["Data de Vencimento da CNH"], dados_atualizados["Telefone"], dados_atualizados["CEP"], dados_atualizados["Endereço"], dados_atualizados["Número"], dados_atualizados["Complemento"],
                dados_atualizados["Cidade"], dados_atualizados["Bairro"], dados_atualizados["Estado"], dados_atualizados["Status do Cliente"], dados_atualizados["Categoria do Cliente"],
                dados_atualizados["Última Atualização"], dados_atualizados["Valor Gasto Total"], dados_atualizados["Última Compra"],
                cpf_original
            )

            cursor.execute(query, valores)
            self.db.connection.commit()

            
            self.main_window.registrar_historico_clientes_fisicos(
                "Edição de Clientes", f"Cliente {dados_atualizados['Nome do Cliente']} editado com sucesso"
            )
            QMessageBox.information(self, "Sucesso", "Cliente atualizado com sucesso.")
            self.carregar_clientes_fisicos()
            self.janela_editar_cliente_fisico.close()
            

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao atualizar cliente: {e}")

    def limpar_campos_clientes_fisicos(self):
        for campo, widget in self.campos_cliente_fisico.items():
            if isinstance(widget,QLineEdit):
                widget.clear()
            elif isinstance(widget, QComboBox):
                widget.setCurrentIndex(0) # Volta para "Selecionar"

    def on_cep_editing_finished_cadastro(self, cep_widget):
        cep_digitado = cep_widget.text()
        dados_cep = self.main_window.buscar_cep(cep_digitado)
        if dados_cep:
            self.preencher_campos_cep_cadastro_fisico(dados_cep)

    def preencher_campos_cep_cadastro_fisico(self, dados):
        if dados is None:
            return

        # Preencher os campos do seu formulário usando self.campos_cliente_fisicos
        self.campos_cliente_fisico["Endereço"].setText(dados.get("logradouro", ""))
        self.campos_cliente_fisico["Bairro"].setText(dados.get("bairro", ""))
        self.campos_cliente_fisico["Cidade"].setText(dados.get("localidade", ""))

        complemento = dados.get("complemento", "")
        if any(char.isdigit() for char in complemento):
            self.campos_cliente_fisico["Número"].setText(complemento)

        estado = dados.get("uf", "")
        # Se tiver um combobox para estado, aqui você ajusta o índice
        if "Estado" in self.campos_cliente_fisico:
            estado_combobox = self.campos_cliente_fisico["Estado"]
            index_estado = estado_combobox.findText(estado)
            if index_estado != -1:
                estado_combobox.setCurrentIndex(index_estado)
                
    def excluir_clientes_fisicos(self):
        try:
            total_linhas = self.table_clientes_fisicos.rowCount()
            if total_linhas == 0:
                QMessageBox.warning(self, "Aviso", "Nenhum cliente encontrado para excluir.")
                return

            clientes_para_excluir = []

            if self.coluna_checkboxes_clientes_fisicos_adicionada:
                # Modo com checkbox
                for linha in range(total_linhas):
                    checkbox_widget = self.table_clientes_fisicos.cellWidget(linha, 0)
                    if checkbox_widget:
                        checkbox = checkbox_widget.findChild(QCheckBox)
                        if checkbox and checkbox.isChecked():
                            nome_cliente = self.table_clientes_fisicos.item(linha, 1).text()
                            clientes_para_excluir.append((linha, nome_cliente))
            else:
                # Modo sem checkbox (seleção direta)
                linha_selecionadas = self.table_clientes_fisicos.selectionModel().selectedRows()
                if not linha_selecionadas:
                    QMessageBox.information(self, "Aviso", "Nenhum cliente selecionado para exclusão.")
                    return
                
                for index in linha_selecionadas:
                    linha = index.row()
                    nome_cliente = self.table_clientes_fisicos.item(linha,0).text()
                    clientes_para_excluir.append((linha,nome_cliente))

            if not clientes_para_excluir:
                QMessageBox.information(self, "Aviso", "Nenhum cliente selecionado para exclusão.")
                return
            
            # Mensagem personalizada
            if len(clientes_para_excluir) == 1:
                _, nome = clientes_para_excluir[0]
                mensagem = f"Tem certeza que deseja excluir o cliente:\n\n• {nome}?"
            else:
                nomes = "\n• " + "\n• ".join(nome for _, nome in clientes_para_excluir)
                mensagem = f"Tem certeza que deseja excluir os seguintes clientes?\n{nomes}"

            # Criar QMessageBox manualmente para alterar os textos dos botões
            msgbox = QMessageBox(self)
            msgbox.setIcon(QMessageBox.Question)
            msgbox.setWindowTitle("Confirmar Exclusão")
            msgbox.setText(mensagem)
            msgbox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)

            # Renomear os botões
            botao_sim = msgbox.button(QMessageBox.Yes)
            botao_nao = msgbox.button(QMessageBox.No)
            botao_sim.setText("Sim")
            botao_nao.setText("Não")

            resposta = msgbox.exec()

            if resposta != QMessageBox.Yes:
                return
            
            # Executa exclusão no banco e remove as linhas da tabela
            db = DataBase()
            db.connecta()
            cursor = db.connection.cursor()

            for linha, nome_cliente in reversed(clientes_para_excluir):
                cursor.execute("""
                    DELETE FROM clientes_fisicos WHERE "Nome do Cliente" = ?
                """, (nome_cliente,))
                self.table_clientes_fisicos.removeRow(linha)

            db.connection.commit()
            
            nomes_excluidos = ", ".join(nome for _, nome in clientes_para_excluir)
            self.main_window.registrar_historico_clientes_fisicos(
                "Exclusão de Cliente(s)",
                f"Cliente(s) excluído(s): {nomes_excluidos}"
            )
            QMessageBox.information(self, "Sucesso", "Cliente(s) excluído(s) com sucesso.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao excluir clientes:\n{e}")

    def marcar_como_clientes_fisicos(self):
        if self.table_clientes_fisicos.rowCount() == 0:
            QMessageBox.warning(self,"Aviso","Nenhum cliente cadastrado para selecionar.")
             # Desmarca o checkbox header visualmente
            if hasattr(self, "checkbox_header_clientes_fisicos_table") and isinstance(self.checkbox_header_clientes_fisicos_table, QCheckBox):
                QTimer.singleShot(0, lambda: self.checkbox_header_clientes_fisicos_table.setChecked(False))

            return  # Impede que o restante da função execute!

        if self.coluna_checkboxes_clientes_fisicos_adicionada:
            # Animação de saída (fade out)
            self.animar_coluna_checkbox_fisicos(mostrar=False)
            QTimer.singleShot(300, lambda: self.remover_coluna_checkboxes_fisicos_table())
            return

        # Adiciona a coluna de checkboxes
        self.table_clientes_fisicos.insertColumn(0)
        self.table_clientes_fisicos.setHorizontalHeaderItem(0, QTableWidgetItem(""))
        self.table_clientes_fisicos.setColumnWidth(0, 30)
        self.table_clientes_fisicos.horizontalHeader().setMinimumSectionSize(30)
        self.table_clientes_fisicos.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)

        # Checkbox no cabeçalho
        self.checkbox_header_clientes_fisicos_table = QCheckBox(self.table_clientes_fisicos)
        self.checkbox_header_clientes_fisicos_table.setToolTip("Selecionar todos")
        self.checkbox_header_clientes_fisicos_table.setChecked(False)
        self.checkbox_header_clientes_fisicos_table.stateChanged.connect(self.selecionar_todos_clientes_fisicos)
        self.checkbox_header_clientes_fisicos_table.setFixedSize(20, 20)
        self.checkbox_header_clientes_fisicos_table.setStyleSheet("background-color: transparent; border: none;")
        self.checkbox_header_clientes_fisicos_table.show()

        header = self.table_clientes_fisicos.horizontalHeader()
        header.sectionResized.connect(self.atualizar_posicao_checkbox_header_clientes_fisicos_table)

        self.checkboxes_clientes_fisicos = []

        for linha in range(self.table_clientes_fisicos.rowCount()):
            checkbox = QCheckBox()
            checkbox.stateChanged.connect(self.atualizar_selecao_todos_clientes)

            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(checkbox)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)

            self.table_clientes_fisicos.setCellWidget(linha, 0, container)
            self.checkboxes_clientes_fisicos.append(checkbox)

        QTimer.singleShot(0, self.atualizar_posicao_checkbox_header_clientes_fisicos_table)

        self.coluna_checkboxes_clientes_fisicos_adicionada = True
        
        # Animação de entrada (fade in)
        self.animar_coluna_checkbox_fisicos(mostrar=True)

    def atualizar_posicao_checkbox_header_clientes_fisicos_table(self):
        if hasattr(self, "checkbox_header_clientes_fisicos_table") and self.coluna_checkboxes_clientes_fisicos_adicionada:
            header = self.table_clientes_fisicos.horizontalHeader()
            
            x = header.sectionViewportPosition(0) + (header.sectionSize(0) - self.checkbox_header_clientes_fisicos_table.width()) // 2 + 20
            y = (header.height() - self.checkbox_header_clientes_fisicos_table.height()) // 2
            self.checkbox_header_clientes_fisicos_table.move(x, y)

    def animar_coluna_checkbox_fisicos(self, mostrar=True):
        self.animacoes_clientes = []

        for row in range(self.table_clientes_fisicos.rowCount()):
            widget = self.table_clientes_fisicos.cellWidget(row, 0)
            if widget:
                efeito = QGraphicsOpacityEffect(widget)
                widget.setGraphicsEffect(efeito)

                # Cria animação corretamente
                anim = QPropertyAnimation(efeito, b"opacity")
                anim.setDuration(300)
                anim.setStartValue(0.0 if mostrar else 1.0)
                anim.setEndValue(1.0 if mostrar else 0.0)
                anim.start()

                self.animacoes_clientes.append(anim)

                
    def remover_coluna_checkboxes_fisicos_table(self):
        self.table_clientes_fisicos.removeColumn(0)
        self.table_clientes_fisicos.verticalHeader().setVisible(True)
        self.coluna_checkboxes_clientes_fisicos_adicionada = False

        if hasattr(self, "checkbox_header_clientes_fisicos_table"):
            self.checkbox_header_clientes_fisicos_table.setChecked(False)
            self.checkbox_header_clientes_fisicos_table.deleteLater()
            del self.checkbox_header_clientes_fisicos_table

        self.checkboxes_clientes_fisicos.clear()

        
    def selecionar_todos_clientes_fisicos(self):
        # Evita erro ao clicar quando a coluna já foi removida
        if not getattr(self, "coluna_checkboxes_clientes_fisicos_adicionada", False):
            return  # Simplesmente sai sem mostrar mensagem

        estado = self.checkbox_header_clientes_fisicos_table.checkState() == Qt.Checked
        self.checkboxes_clientes_fisicos.clear()

        for row in range(self.table_clientes_fisicos.rowCount()):
            widget = self.table_clientes_fisicos.cellWidget(row, 0)
            if widget is not None:
                checkbox = widget.findChild(QCheckBox)
                if checkbox:
                    checkbox.blockSignals(True)
                    checkbox.setChecked(estado)
                    checkbox.blockSignals(False)
                    self.checkboxes_clientes_fisicos.append(checkbox)


    def atualizar_selecao_todos_clientes(self):
        self.checkbox_header_clientes_fisicos_table.blockSignals(True)

        all_checked = all(cb.isChecked() for cb in self.checkboxes_clientes_fisicos if cb)
        any_checked = any(cb.isChecked() for cb in self.checkboxes_clientes_fisicos if cb)

        if all_checked:
            self.checkbox_header_clientes_fisicos_table.setCheckState(Qt.Checked)
        elif any_checked:
            self.checkbox_header_clientes_fisicos_table.setCheckState(Qt.PartiallyChecked)
        else:
            self.checkbox_header_clientes_fisicos_table.setCheckState(Qt.Unchecked)

        self.checkbox_header_clientes_fisicos_table.blockSignals(False)

    
            
    # Limpa a coluna selecionada clicando em qualquer lugar da tabela
    def eventFilter(self, source, event):
        if event.type() == QEvent.MouseButtonPress:
            if source == self.table_clientes_fisicos.viewport():
                index = self.table_clientes_fisicos.indexAt(event.pos())
                if not index.isValid():
                    self.table_clientes_fisicos.clearSelection() # remove a linha selecinada da tabela
                    self.table_clientes_fisicos.clearFocus() # remove o foco da tabela      
        return super().eventFilter(source,event)
    
    
    def marcar_alteracao(self):
        self.alteracoes_realizadas = True


    def historico_clientes_fisicos(self):
        self.janela_historico_clientes_fisicos = QMainWindow()
        self.janela_historico_clientes_fisicos.resize(800,650)
        self.janela_historico_clientes_fisicos.setWindowTitle("Histórico de Clientes Físicos")

        # Centralizar a janela na tela
        screen = QGuiApplication.primaryScreen()
        screen_geometry = screen.availableGeometry()
        window_geometry = self.janela_historico_clientes_fisicos.frameGeometry()
        window_geometry.moveCenter(screen_geometry.center())
        self.janela_historico_clientes_fisicos.move(window_geometry.topLeft())

        # Criação do layout e tabela para exibir o histórico
        central_widget = QWidget()
        layout = QVBoxLayout(central_widget)

        # Carregar tema
        config = self.carregar_config()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(60,60,60),
                                                stop:1 rgb(100,100,100));
                    color: white;
                    border-radius: 8px;
                    font-size: 12px;
                    border: 2px solid #666666;
                    padding: 3px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
            """
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            table_view_style = """
            /* QTableView com seleção diferenciada */
            QTableView {
                background-color: #202124;
                color: white;
                gridline-color: #555555;
                selection-background-color: #7a7a7a;
                selection-color: white;
            }
            /* Coluna dos cabeçalhos */
            QHeaderView::section {
                background-color: #202124;
                color: white;
                border: 1px solid #aaaaaa;
                padding: 1px;
            }

            /* QTabWidget headers brancos */
            QTabWidget::pane {
                border: 1px solid #444444;
                background-color: #202124;
            }
            /* Estiliza a barra de rolagem horizontal */
            QTableView QScrollBar:horizontal {
                border: none;
                background-color: none;
                height: 12px;
                margin: 0px;
                border-radius: 5px;
            }

            /* Estiliza a barra de rolagem vertical */
            QTableView QScrollBar:vertical {
                border: none;
                background-color: none; 
                width: 12px;
                margin: 0px;
                border-radius: 5px;
            }
            

            /* Parte que você arrasta */
            QTableView QScrollBar::handle:vertical {
                background-color: #777777;  /* cinza médio */
                min-height: 22px;
                border-radius: 5px;
            }

            QTableView QScrollBar::handle:horizontal {
                background-color: #777777;
                min-width: 22px;
                border-radius: 5px;
            }

            /* Groove horizontal */
            QTableView QScrollBar::groove:horizontal {
                background-color: transparent;
                border-radius: 5px;
                height: 15px;
                margin: 0px 10px 0px 10px;
            }

            /* Groove vertical */
            QTableView QScrollBar::groove:vertical {
                background-color: transparent;
                border-radius: 5px;
                width: 25px;
                margin: 10px 0px 10px 10px;
            }

            /* Estilo para item selecionado */
            QTableWidget::item:selected {
                background-color: #555555;  /* cinza de seleção */
                color: white;
            }
            QTableCornerButton::section {
                background-color: #202124;  /* mesma cor da tabela */
                border: none;
            }

            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """

        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(0,120,180),
                                                stop:1 rgb(0,150,220));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(0,100,160);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #007acc;
                }
                QPushButton:pressed {
                    background-color: #006bb3;
                    border: 2px solid #005c99;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 3px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            scroll_style = """
                QScrollBar:vertical {
                border: none;
                background-color: rgb(255, 255, 255); /* branco */
                width: 30px;
                margin: 0px 10px 0px 10px;
            }
             QScrollBar::handle:vertical {
                background-color: rgb(180, 180,180);  /* cinza */
                min-height: 30px;
                border-radius: 5px;
            }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        # Tabela do histórico
        self.tabela_historico_clientes_fisicos = QTableWidget()
        self.tabela_historico_clientes_fisicos.setColumnCount(4)
        self.tabela_historico_clientes_fisicos.setHorizontalHeaderLabels(["Data/Hora", "Usuário", "Ação", "Descrição"])

        # Botão Atualizar
        botao_atualizar = QPushButton("Atualizar Histórico")
        botao_atualizar.clicked.connect(self.atualizar_historico_clientes_fisicos)
        botao_atualizar.setStyleSheet(button_style)


        # Botão Apagar
        botao_apagar = QPushButton("Apagar Histórico")
        botao_apagar.clicked.connect(self.apagar_historico_cliente_fisicos)
        botao_apagar.setStyleSheet(button_style)


        # Botão Exportar CSV
        botao_exportar_csv = QPushButton("Exportar para CSV")
        botao_exportar_csv.clicked.connect(self.exportar_csv_fisicos)
        botao_exportar_csv.setStyleSheet(button_style)

        
        # Botão Exportar Excel
        botao_exportar_excel = QPushButton("Exportar para Excel")
        botao_exportar_excel.clicked.connect(self.exportar_excel_fisicos)
        botao_exportar_excel.setStyleSheet(button_style)


        # Botão PDF
        botao_exportar_pdf = QPushButton("Exportar PDF")
        botao_exportar_pdf.clicked.connect(self.exportar_pdf_fisicos)
        botao_exportar_pdf.setStyleSheet(button_style)

        botao_pausar_historico = QPushButton("Pausar Histórico")
        botao_pausar_historico.clicked.connect(self.pausar_historico_fisicos)
        botao_pausar_historico.setStyleSheet(button_style)


        botao_filtrar_historico = QPushButton("Filtrar Histórico")
        botao_filtrar_historico.clicked.connect(self.filtrar_historico_clientes_fisicos)
        botao_filtrar_historico.setStyleSheet(button_style)

        botao_ordenar_historico = QPushButton("Ordenar Histórico")
        botao_ordenar_historico.clicked.connect(self.ordenar_historico_clientes_fisicos)
        botao_ordenar_historico.setStyleSheet(button_style)


        # Criar checkbox "Selecionar Individualmente" toda vez que a janela for aberta
        self.checkbox_selecionar_fisicos = QCheckBox("Selecionar")
        self.checkbox_selecionar_fisicos.stateChanged.connect(self.selecionar_individual_fisicos)
        self.checkbox_selecionar_fisicos.setStyleSheet(f"color: {text_cor};")

        # Adicionar outros botões ao layout
        layout.addWidget(botao_atualizar)
        layout.addWidget(botao_apagar)
        layout.addWidget(botao_exportar_csv)
        layout.addWidget(botao_exportar_excel)
        layout.addWidget(botao_exportar_pdf)
        layout.addWidget(botao_pausar_historico)
        layout.addWidget(botao_ordenar_historico)
        layout.addWidget(botao_filtrar_historico)
        layout.addWidget(self.checkbox_selecionar_fisicos)
        layout.addWidget(self.tabela_historico_clientes_fisicos)
        


        # Configurar o widget central e exibir a janela
        self.janela_historico_clientes_fisicos.setCentralWidget(central_widget)
        self.janela_historico_clientes_fisicos.show()
        self.janela_historico_clientes_fisicos.setStyleSheet(f"background-color: {bg_cor}; color: {text_cor};")

        # Estilo da tabela
        self.tabela_historico_clientes_fisicos.setStyleSheet(table_view_style)

         # Redimensionar colunas e linhas
        self.carregar_historico_clientes_fisicos()
        self.tabela_historico_clientes_fisicos.resizeColumnsToContents()
        self.tabela_historico_clientes_fisicos.resizeRowsToContents()

    def carregar_historico_clientes_fisicos(self):
            with sqlite3.connect('banco_de_dados.db') as cn:
                cursor = cn.cursor()
                cursor.execute('SELECT * FROM historico_clientes_fisicos ORDER BY "Data e Hora" DESC')
                registros = cursor.fetchall()

            self.tabela_historico_clientes_fisicos.clearContents()
            self.tabela_historico_clientes_fisicos.setRowCount(len(registros))

            deslocamento = 1 if self.coluna_checkboxes_clientes_fisicos_adicionada else 0
            self.checkboxes_clientes_fisicos = []  # Zerar e recriar lista de checkboxes

            for i, (data, usuario, acao, descricao) in enumerate(registros):
                if self.coluna_checkboxes_clientes_fisicos_adicionada:
                    checkbox = QCheckBox()
                    checkbox.setStyleSheet("margin-left:9px; margin-right:9px;")
                    self.tabela_historico_clientes_fisicos.setCellWidget(i,0,checkbox)
                    self.checkboxes_clientes_fisicos.append(checkbox)
                self.tabela_historico_clientes_fisicos.setItem(i, 0 + deslocamento, QTableWidgetItem(data))
                self.tabela_historico_clientes_fisicos.setItem(i, 1 + deslocamento, QTableWidgetItem(usuario))
                self.tabela_historico_clientes_fisicos.setItem(i, 2 + deslocamento, QTableWidgetItem(acao))
                self.tabela_historico_clientes_fisicos.setItem(i, 3 + deslocamento, QTableWidgetItem(descricao))

    def atualizar_historico_clientes_fisicos(self):
        QMessageBox.information(self, "Sucesso", "Dados carregados com sucesso!")
        self.carregar_historico_clientes_fisicos()

    def confirmar_historico_cliente_fisicos_apagado(self, mensagem):
        msgbox = QMessageBox(self)
        msgbox.setWindowTitle("Confirmação")
        msgbox.setText(mensagem)

        btn_sim = QPushButton("Sim")
        btn_nao = QPushButton("Não")
        msgbox.addButton(btn_sim, QMessageBox.ButtonRole.YesRole)
        msgbox.addButton(btn_nao, QMessageBox.ButtonRole.NoRole)

        msgbox.setDefaultButton(btn_nao)
        resposta = msgbox.exec()

        return msgbox.clickedButton() == btn_sim
    
    def apagar_historico_cliente_fisicos(self):
        # Caso checkboxes estejam ativados
        if self.coluna_checkboxes_clientes_fisicos_adicionada and self.checkboxes_clientes:
            linhas_para_remover = []
            datas_para_remover = []

            # Identificar as linhas com checkboxes selecionados
            for row, checkbox in enumerate(self.checkboxes_clientes_fisicos):
                if checkbox and checkbox.isChecked():
                    linhas_para_remover.append(row)
                    coluna_data_hora = 0 if not self.coluna_checkboxes_clientes_fisicos_adicionada else 1
                    item_data_widget = self.tabela_historico_clientes_fisicos.item(row, coluna_data_hora)  # Coluna de Data/Hora
                    if item_data_widget:
                        data_text = item_data_widget.text().strip()
                        # Excluir com base na data e hora
                        if data_text:
                            datas_para_remover.append(data_text)
                        else:
                            print(f"[AVISO] Linha {row} com Data e Hora vazia.")
                    else:
                        print(f"[ERRO] Coluna 'Data e Hora' não encontrada na linha {row}.")

            if not datas_para_remover:
                QMessageBox.warning(self, "Erro", "Nenhum item válido foi selecionado para apagar!")
                return

            # Confirmar exclusão
            mensagem = (
                f"Você tem certeza que deseja apagar os {len(datas_para_remover)} itens selecionados?"
                if len(datas_para_remover) > 1
                else "Você tem certeza que deseja apagar o item selecionado?"
            )

            if not self.confirmar_historico_apagado_clientes(mensagem):
                return

            # Excluir do banco de dados
            with sqlite3.connect('banco_de_dados.db') as cn:
                cursor = cn.cursor()
                try:
                    for data in datas_para_remover:
                        cursor.execute('DELETE FROM historico_clientes_fisicos WHERE "Data e Hora" = ?', (data,))
                    cn.commit()
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao excluir do banco de dados: {e}")
                    return

            # Remover as linhas na interface
            for row in sorted(linhas_para_remover, reverse=True):
                self.tabela_historico_clientes_fisicos.removeRow(row)

            QMessageBox.information(self, "Sucesso", "Itens removidos com sucesso!")

        # Caso sem checkboxes (seleção manual)
        else:
            linha_selecionada = self.tabela_historico_clientes_fisicos.currentRow()

            if linha_selecionada < 0:
                QMessageBox.warning(self, "Erro", "Nenhum item foi selecionado para apagar!")
                return

            # Capturar a Data/Hora da célula correspondente (coluna 0)
            coluna_data_hora = 0 if not self.coluna_checkboxes_clientes_fisicos_adicionada else 1
            item_data_widget = self.tabela_historico_clientes_fisicos.item(linha_selecionada, coluna_data_hora)  # Coluna de Data/Hora
            if not item_data_widget:
                QMessageBox.warning(self, "Erro", "Não foi possível identificar a Data/Hora do item a ser apagado!")
                return

            item_data_text = item_data_widget.text().strip()

            # Verificar se há um valor válido de data/hora
            if not item_data_text:
                QMessageBox.warning(self, "Erro", "Não foi possível identificar a Data/Hora do item a ser apagado!")
                return


            # Confirmar exclusão
            mensagem = "Você tem certeza que deseja apagar o item selecionado?"

            if not self.confirmar_historico_cliente_fisicos_apagado(mensagem):
                return

            # Excluir do banco de dados
            with sqlite3.connect('banco_de_dados.db') as cn:
                cursor = cn.cursor()
                try:
                    cursor.execute('DELETE FROM historico_clientes_fisicos WHERE "Data e Hora" = ?', (item_data_text,))
                    cn.commit()
                except Exception as e:
                    QMessageBox.critical(self, "Erro", f"Erro ao excluir do banco de dados: {e}")
                    return

            # Remover a linha da interface
            self.tabela_historico_clientes_fisicos.removeRow(linha_selecionada)

            QMessageBox.information(self, "Sucesso", "Item removido com sucesso!")

     # Função para adicionar checkboxes selecionar_individual na tabela de histórico
    def selecionar_individual_fisicos(self):
        if self.tabela_historico_clientes_fisicos.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico para selecionar.")
            self.checkbox_selecionar_individual.setChecked(False)
            return

        if self.coluna_checkboxes_clientes_fisicos_adicionada:
            self.tabela_historico_clientes_fisicos.removeColumn(0)
            self.tabela_historico_clientes_fisicos.verticalHeader().setVisible(True)
            self.coluna_checkboxes_clientes_fisicos_adicionada = False

            if hasattr(self, "checkbox_header_clientes_fisicos"):
                self.checkbox_header_clientes_fisicos.deleteLater()
                del self.checkbox_header_clientes_fisicos

            self.checkboxes_clientes_fisicos.clear()
            return

        self.tabela_historico_clientes_fisicos.insertColumn(0)
        self.tabela_historico_clientes_fisicos.setHorizontalHeaderItem(0, QTableWidgetItem(""))
        self.tabela_historico_clientes_fisicos.setColumnWidth(0, 30)
        self.tabela_historico_clientes_fisicos.horizontalHeader().setMinimumSectionSize(30)
        self.tabela_historico_clientes_fisicos.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)

        # Checkbox do cabeçalho
        self.checkbox_header_clientes_fisicos = QCheckBox(self.tabela_historico_clientes_fisicos)
        self.checkbox_header_clientes_fisicos.setToolTip("Selecionar todos")
        self.checkbox_header_clientes_fisicos.setChecked(False)
        self.checkbox_header_clientes_fisicos.stateChanged.connect(self.selecionar_todos_clientes_fisicos)
        self.checkbox_header_clientes_fisicos.setFixedSize(20, 20)
        self.checkbox_header_clientes_fisicos.show()

        header = self.tabela_historico_clientes_fisicos.horizontalHeader()
        self.atualizar_posicao_checkbox_header_clientes_fisicos()
        header.sectionResized.connect(self.atualizar_posicao_checkbox_header_clientes_fisicos)

        self.checkboxes_clientes_fisicos.clear()

        QTimer.singleShot(0,self.atualizar_posicao_checkbox_header_clientes_fisicos)

        for row in range(self.tabela_historico_clientes_fisicos.rowCount()):
            checkbox = QCheckBox()
            checkbox.stateChanged.connect(self.atualizar_selecao_todos)

            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(checkbox)
            layout.setAlignment(Qt.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)

            self.tabela_historico_clientes_fisicos.setCellWidget(row, 0, container)
            self.checkboxes_clientes_fisicos.append(checkbox)

        self.tabela_historico_clientes_fisicos.verticalHeader().setVisible(False)
        self.coluna_checkboxes_clientes_fisicos_adicionada = True

    def atualizar_selecao_todos(self):
        self.checkbox_header_clientes_fisicos.blockSignals(True)

        all_checked = all(checkbox.isChecked() for checkbox in self.checkboxes_clientes_fisicos if checkbox)
        any_checked = any(checkbox.isChecked() for checkbox in self.checkboxes_clientes_fisicos if checkbox)

        if all_checked:
            self.checkbox_header_clientes_fisicos.setCheckState(Qt.Checked)
        elif any_checked:
            self.checkbox_header_clientes_fisicos.setCheckState(Qt.PartiallyChecked)
        else:
            self.checkbox_header_clientes_fisicos.setCheckState(Qt.Unchecked)

        self.checkbox_header_clientes_fisicos.blockSignals(False)

        
    def atualizar_posicao_checkbox_header_clientes_fisicos(self):
        if hasattr(self, "checkbox_header_clientes_fisicos") and self.coluna_checkboxes_clientes_fisicos_adicionada:
            header = self.tabela_historico_clientes_fisicos.horizontalHeader()

            x = header.sectionViewportPosition(0) + (header.sectionSize(0) - self.checkbox_header_clientes_fisicos.width()) // 2 + 4
            y = (header.height() - self.checkbox_header_clientes_fisicos.height()) // 2

            self.checkbox_header_clientes_fisicos.move(x, y)
            
    def ordenar_historico_clientes_fisicos(self):
        if hasattr(self, "checkbox_header_clientes_fisicos"):
            QMessageBox.warning(
                self,
                "Aviso",
                "Desmarque o checkbox antes de ordenar o histórico."
            )
            return
        # Obter a coluna pela qual o usuário deseja ordenar
        coluna = self.obter_coluna_para_ordenar_clientes_fisicos()  # Função fictícia para capturar escolha
        if coluna is None:
            return  # Cancela o processo todo
        
        # Determinar a direção de ordenação (ascendente ou descendente)
        direcao = self.obter_direcao_ordenacao_clientes_fisicos()  # Função fictícia para capturar escolha
        if direcao is None:
            return  # Cancela o processo todo
        
        # Mapeamento de colunas para índices (ajustar conforme sua tabela)
        colunas_para_indices = {
            "Data/Hora": 0,
            "Usuário": 1,
            "Ação": 2,
            "Descrição": 3
        }
        
        # Verificar se a coluna escolhida é válida
        if coluna not in colunas_para_indices:
            QMessageBox.warning(self, "Erro", "Coluna inválida para ordenação!")
            return
        
        # Obter o índice da coluna escolhida
        indice_coluna = colunas_para_indices[coluna]
        
        # Obter os dados atuais da tabela
        dados = []
        for row in range(self.tabela_historico_clientes_fisicos.rowCount()):
            linha = [
                self.tabela_historico_clientes_fisicos.item(row, col).text() if self.tabela_historico_clientes_fisicos.item(row, col) else ""
                for col in range(self.tabela_historico_clientes_fisicos.columnCount())
            ]
            dados.append(linha)
        
        # Ordenar os dados com base na coluna escolhida e direção
        dados.sort(key=lambda x: x[indice_coluna], reverse=(direcao == "Decrescente"))
        
        # Atualizar a tabela com os dados ordenados
        self.tabela_historico_clientes_fisicos.setRowCount(0)  # Limpar tabela
        for row_data in dados:
            row = self.tabela_historico_clientes_fisicos.rowCount()
            self.tabela_historico_clientes_fisicos.insertRow(row)
            for col, value in enumerate(row_data):
                self.tabela_historico_clientes_fisicos.setItem(row, col, QTableWidgetItem(value))

    def obter_coluna_para_ordenar_clientes_fisicos(self):
        colunas = ["Data/Hora", "Usuário", "Ação", "Descrição"]
        coluna, ok = QInputDialog.getItem(self, "Ordenar por", "Escolha a coluna:", colunas, 0, False)
        return coluna if ok else None

    def obter_direcao_ordenacao_clientes_fisicos(self):
        direcoes = ["Crescente", "Decrescente"]
        direcao, ok = QInputDialog.getItem(self, "Direção da Ordenação", "Escolha a direção:", direcoes, 0, False)
        return direcao if ok else None

    def filtrar_historico_clientes_fisicos(self):
        if hasattr(self,"checkbox_header_clientes_fisicos"):
            QMessageBox.warning(
                self,
                "Aviso",
                "Desmarque o checkbox antes de filtrar o histórico."
            )
            return
        
        # Criar a janela de filtro
        janela_filtro = QDialog(self)
        janela_filtro.setWindowTitle("Filtrar Histórico")
        layout = QVBoxLayout(janela_filtro)

        # Campo para inserir a data
        campo_data = QLineEdit()
        campo_data.setPlaceholderText("Digite a data no formato DD/MM/AAAA")
        
        # Conectar ao método de formatação, passando o texto
        campo_data.textChanged.connect(lambda: self.formatar_data(campo_data))


        # Campo para selecionar se quer o mais recente ou mais antigo (filtro por hora)
        grupo_hora = QGroupBox("Filtrar por Hora")
        layout_hora = QVBoxLayout(grupo_hora)

        radio_mais_novo = QRadioButton("Mais Recente")
        radio_mais_velho = QRadioButton("Mais Antigo")

        layout_hora.addWidget(radio_mais_novo)
        layout_hora.addWidget(radio_mais_velho)
        grupo_hora.setLayout(layout_hora)

        # Botão para aplicar o filtro
        botao_filtrar = QPushButton("Aplicar Filtro")
        botao_filtrar.clicked.connect(
            lambda: self.aplicar_filtro_clientes_fisicos(
                campo_data.text(),
                radio_mais_novo.isChecked(),
                radio_mais_velho.isChecked()
            )
        )

        # Adicionar widgets ao layout
        layout.addWidget(QLabel("Filtros Disponíveis"))
        layout.addWidget(campo_data)
        layout.addWidget(grupo_hora)
        layout.addWidget(botao_filtrar)

        # Exibir a janela de filtro
        janela_filtro.setLayout(layout)
        janela_filtro.exec()

    def formatar_data_clientes_fisicos(self, campo_data):
        # Obter o texto do campo de data
        texto_data = campo_data.text().replace("/", "")  # Remover as barras existentes
        texto_data = ''.join(filter(str.isdigit, texto_data))  # Permite apenas números

        # Verificar se há caracteres alfabéticos (letras)
        if any(char.isalpha() for char in texto_data):
            # Mostrar mensagem de erro caso haja letras
            QMessageBox.warning(self, "Erro", "Somente números são permitidos.")
            campo_data.clear()
            return  # Não aplica a formatação se houver letras

        # Limitar a entrada para no máximo 8 dígitos
        if len(texto_data) > 8:
            texto_data = texto_data[:8]

         # Formatar a data no formato DD/MM/AAAA
        if len(texto_data) >= 8:
            data_formatada = "{}/{}/{}".format(texto_data[:2], texto_data[2:4], texto_data[4:])  # DD/MM/AAAA
        elif len(texto_data) > 6:
            data_formatada = "{}/{}".format(texto_data[:2], texto_data[2:8])  # DD/MM
        else:
            data_formatada = texto_data[:8]  # Apenas o DD

        # Atualizar o texto do campo de data se houver mudança
        if campo_data.text() != data_formatada:
            campo_data.setText(data_formatada)  # Atualiza o texto do campo de data
            campo_data.setCursorPosition(len(data_formatada))  # Move o cursor para o final do texto

    def aplicar_filtro_clientes_fisicos(self, data, filtrar_novo, filtrar_velho):
        with sqlite3.connect('banco_de_dados.db') as cn:
            cursor = cn.cursor()

            query = "SELECT * FROM historico_clientes_fisicos"
            params = []

            # Filtrar pela data, se fornecida
            if data:
                try:
                    # Garantir que a data seja no formato correto (DD/MM/AAAA)
                    data_formatada = datetime.strptime(data, "%d/%m/%Y").strftime("%d/%m/%Y")  # Formato DD/MM/YYYY
                    query += " WHERE SUBSTR([Data e Hora], 1, 10) = ?"
                    params.append(data_formatada)
                except ValueError:
                    QMessageBox.warning(self, "Erro", "Data inválida. Use o formato DD/MM/AAAA.")
                    return

            # Ordenar por hora, se aplicável
            if filtrar_novo:
                query += " ORDER BY [Data e Hora] DESC LIMIT 1"
            elif filtrar_velho:
                query += " ORDER BY [Data e Hora] ASC LIMIT 1"

            # Executar a consulta
            cursor.execute(query, params)
            registros = cursor.fetchall()

        # Atualizar a tabela com os registros filtrados
        self.tabela_historico_clientes_fisicos.clearContents()
        self.tabela_historico_clientes_fisicos.setRowCount(len(registros))

        for i, row in enumerate(registros):
            self.tabela_historico_clientes_fisicos.setItem(i, 0, QTableWidgetItem(row[0]))  # Data/Hora
            self.tabela_historico_clientes_fisicos.setItem(i, 1, QTableWidgetItem(row[1]))  # Usuário
            self.tabela_historico_clientes_fisicos.setItem(i, 2, QTableWidgetItem(row[2]))  # Ação
            self.tabela_historico_clientes_fisicos.setItem(i, 3, QTableWidgetItem(row[3]))  # Descrição

        QMessageBox.information(self, "Filtro Aplicado", f"{len(registros)} registro(s) encontrado(s)!")
    
    def exportar_csv_fisicos(self):
        num_linhas = self.tabela_historico_clientes_fisicos.rowCount()
        num_colunas = self.tabela_historico_clientes_fisicos.columnCount()

        # Verificar se a tabela está vazia
        if self.tabela_historico_clientes_fisicos.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico encontrado para gerar arquivo CSV.")
            return  # Se a tabela estiver vazia, encerra a função sem prosseguir

        nome_arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Arquivo CSV",
            "historico.csv",
            "Arquivos CSV (*.csv)"

        )

        if not nome_arquivo:
            return
        
        #Criar o arquivo CSV
        try:
            with open(nome_arquivo, mode="w",newline="",encoding="utf-8-sig") as arquivo_csv:
                escritor = csv.writer(arquivo_csv, delimiter=";")

                 # Adicionar cabeçalhos ao CSV
                cabecalhos = [self.tabela_historico_clientes_fisicos.horizontalHeaderItem(col).text() for col in range (num_colunas)]
                escritor.writerow(cabecalhos)

                # Adicionar os dados da tabela ao CSV
                for linha in range(num_linhas):
                    dados_linhas = [
                        self.tabela_historico_clientes_fisicos.item(linha, col).text() if self.tabela_historico_clientes_fisicos.item(linha, col) else ""
                        for col in range(num_colunas)

                    ]
                    escritor.writerow(dados_linhas)

                    QMessageBox.information(self, "Sucesso", f"Arquivo CSV salvo com sucesso em:\n{nome_arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Falha ao salvar o arquivo CSV:\n{str(e)}")

    def exportar_excel_fisicos(self):
        num_linhas = self.tabela_historico_clientes_fisicos.rowCount()
        num_colunas = self.tabela_historico_clientes_fisicos.columnCount()

        # Verificar se a tabela está vazia
        if self.tabela_historico_clientes_fisicos.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico encontrado para gerar arquivo Excel.")
            return  # Se a tabela estiver vazia, encerra a função sem prosseguir
        
        nome_arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Arquivo Excel",
            "historico.xlsx",
            "Arquivos Excel (*.xlsx)"

        )

        if not nome_arquivo:
            return
        
        # Garantir que o arquivo tenha extensão .xlsx
        if not nome_arquivo.endswith(".xlsx"):
            nome_arquivo += ".xlsx"

        # Criar uma lista para armazenar os dados da tabela
        dados = []


        for linha in range(num_linhas):
            linha_dados = []
            for coluna in range(num_colunas):
                item = self.tabela_historico_clientes_fisicos.item(linha, coluna)
                linha_dados.append(item.text() if item else "") # Adicionar o texto ou vazio se o item for None
            dados.append(linha_dados)

        # Obter os cabeçalhos da tabela        
        cabecalhos = [self.tabela_historico_clientes_fisicos.horizontalHeaderItem(coluna).text() for coluna in range (num_colunas)]

        try:
            # Criar um DataFrame do pandas com os dados e cabeçalhos
            df = pd.DataFrame(dados, columns=cabecalhos)

            # Exportar para Excel
            df.to_excel(nome_arquivo, index=False,engine="openpyxl")
            QMessageBox.information(self, "Sucesso",f"Arquivo Excel gerado com sucesso em: \n{nome_arquivo}")
        except Exception as e:
            QMessageBox.critical(self, "Erro",f"Erro ao salvar arquivo Excel: {str(e)}")

    def exportar_pdf_fisicos(self):
        num_linhas = self.tabela_historico_clientes_fisicos.rowCount()
        num_colunas = self.tabela_historico_clientes_fisicos.columnCount()

        # Verificar se a tabela está vazia
        if self.tabela_historico_clientes_fisicos.rowCount() == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum histórico encontrado para gerar arquivo PDF.")
            return  # Se a tabela estiver vazia, encerra a função sem prosseguir

        nome_arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Arquivo PDF",
            "historico.pdf",
            "Arquivos PDF (*.pdf)"
        )

        if not nome_arquivo:
            return

        # Garantir que o arquivo tenha extensão .pdf
        if not nome_arquivo.endswith(".pdf"):
            nome_arquivo += ".pdf"

        # Criar uma lista para armazenar os dados da tabela
        dados = []

        # Obter os cabeçalhos da tabela
        cabecalhos = [self.tabela_historico_clientes_fisicos.horizontalHeaderItem(coluna).text() for coluna in range(num_colunas)]
        dados.append(cabecalhos)  # Adicionar os cabeçalhos como a primeira linha do PDF

        # Adicionar os dados da tabela
        for linha in range(num_linhas):
            linha_dados = []
            for coluna in range(num_colunas):
                item = self.tabela_historico_clientes_fisicos.item(linha, coluna)
                linha_dados.append(item.text() if item else "")  # Adicionar o texto ou vazio se o item for None
            dados.append(linha_dados)

        try:
            # Criar o PDF
            pdf = SimpleDocTemplate(nome_arquivo, pagesize=landscape(letter))
            tabela = Table(dados)

            # Adicionar estilo à tabela
            estilo = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Cabeçalho com fundo cinza
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),  # Texto do cabeçalho branco
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centralizar texto
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Fonte do cabeçalho
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Espaçamento inferior no cabeçalho
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),  # Fundo das linhas de dados
                ('GRID', (0, 0), (-1, -1), 1, colors.black)  # Bordas da tabela
            ])
            tabela.setStyle(estilo)

            # Gerar o PDF
            pdf.build([tabela])
            QMessageBox.information(self, "Sucesso", f"Arquivo PDF gerado com sucesso em: \n{nome_arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar arquivo PDF: {str(e)}")

    def pausar_historico_fisicos(self):
        msg = QMessageBox(self)
        msg.setWindowTitle("Pausar Histórico")
        msg.setText("Deseja pausar o histórico?")
        msg.setIcon(QMessageBox.Question)

        botao_sim = msg.addButton("Sim", QMessageBox.YesRole)
        botao_nao = msg.addButton("Não", QMessageBox.NoRole)

        msg.exec()

        if msg.clickedButton() == botao_sim:
            self.historico_ativo_fisicos()
        elif msg.clickedButton() == botao_nao:
            self.historico_inativo_fisicos()


    def historico_ativo_fisicos(self):
        # Atualiza o estado do histórico para ativo
        self.main_window.historico_pausado_clientes_fisicos = True  # Atualiza a variável no MainWindow
        QMessageBox.information(self, "Histórico", "O registro do histórico foi pausado.")


    def historico_inativo_fisicos(self):
        # Atualiza o estado do histórico para inativo (continua registrando)
        self.main_window.historico_pausado_clientes_fisicos = False  # Atualiza a variável no MainWindow
        QMessageBox.information(self, "Histórico", "O registro do histórico continua ativo.")

        

    def abrir_janela_relatorio_clientes_fisicos(self):
        self.janela_historico_clientes_fisicos = QMainWindow()
        self.janela_historico_clientes_fisicos.setWindowTitle("Relatório de Clientes Físicos")
        self.janela_historico_clientes_fisicos.resize(800, 600)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)

        # Carregar tema
        config = self.carregar_config()
        tema = config.get("tema", "claro")

        # Definições de tema
        if tema == "escuro":
            bg_cor = "#202124"
            text_cor = "white"
            lineedit_bg = "#303030"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(60,60,60),
                                                stop:1 rgb(100,100,100));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid #666666;
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #444444;
                }
                QPushButton:pressed {
                    background-color: #555555;
                    border: 2px solid #888888;
                }
            """
            combobox_style = """
                QComboBox {
                    color: #f0f0f0;
                    border: 2px solid #ffffff;
                    border-radius: 6px;
                    padding: 4px 10px;
                    background-color: #2b2b2b;
                }
                QComboBox QAbstractItemView::item:hover {
                    background-color: #444444;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item:selected {
                    background-color: #696969;
                    color: #f0f0f0;
                }
                QComboBox QAbstractItemView::item {
                    height: 24px;
                }
                QComboBox QScrollBar:vertical {
                    background: #ffffff;
                    width: 12px;
                    border-radius: 6px;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #555555;
                    border-radius: 6px;
                }
            """
            scroll_style = """
            /* Scrollbar vertical */
            QScrollBar:vertical {
                background: #ffffff;   /* fundo do track */
                width: 12px;
                margin: 0px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: #555555;   /* cor do handle */
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: #777777;   /* hover no handle */
            }

            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                background: none;
                height: 0px;
            }

            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: none;
            }
            """
            lineedit_style = f"""
                QLineEdit {{
                    background-color: {lineedit_bg};
                    color: {text_cor};
                    border: 2px solid white;
                    border-radius: 6px;
                    padding: 3px;
                }}
            """

        elif tema == "claro":
            bg_cor = "white"
            text_cor = "black"
            lineedit_bg = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(50,150,250),
                                                stop:1 rgb(100,200,255));
                    color: black;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(50,150,250);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #e5f3ff;
                }
                QPushButton:pressed {
                    background-color: #cce7ff;
                    border: 2px solid #3399ff;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
                
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        else:  # clássico
            bg_cor = "rgb(0,80,121)"
            text_cor = "white"

            button_style = """
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                stop:0 rgb(0,120,180),
                                                stop:1 rgb(0,150,220));
                    color: white;
                    border-radius: 8px;
                    font-size: 16px;
                    border: 2px solid rgb(0,100,160);
                    padding: 6px;
                }
                QPushButton:hover {
                    background-color: #007acc;
                }
                QPushButton:pressed {
                    background-color: #006bb3;
                    border: 2px solid #005c99;
                }
            """
            combobox_style = """
                QComboBox {
                    background-color: white;
                    border: 3px solid rgb(50,150,250);
                    border-radius: 5px;
                    color: black;
                    padding: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: white;
                    color: black;
                    border: 1px solid #ccc;
                    selection-background-color: #e5e5e5;
                    selection-color: black;
                }
                QComboBox QScrollBar:vertical {
                    background: #f5f5f5;
                    width: 12px;
                    border: none;
                }
                QComboBox QScrollBar::handle:vertical {
                    background: #cccccc;
                    min-height: 20px;
                    border-radius: 5px;
                }
            """
            scroll_style = """
                QScrollBar:vertical {
                border: none;
                background-color: rgb(255, 255, 255); /* branco */
                width: 30px;
                margin: 0px 10px 0px 10px;
            }
             QScrollBar::handle:vertical {
                background-color: rgb(180, 180,180);  /* cinza */
                min-height: 30px;
                border-radius: 5px;
            }
            """
            lineedit_style = """
                QLineEdit {
                    background-color: white;
                    color: black;
                    border: 2px solid rgb(50,150,250);
                    border-radius: 6px;
                    padding: 3px;
                }
            """

        central_widget = QWidget()
        layout_principal = QVBoxLayout(central_widget)

        grupo_filtros = QGroupBox("Filtros do Relatório")
        layout_filtros = QFormLayout()

        self.combo_status = QComboBox()
        self.combo_status.addItems(["Todos", "Ativo", "Inativo", "Bloqueado", "Pendente"])
        layout_filtros.addRow("Status do Cliente:", self.combo_status)
        self.combo_status.setStyleSheet(combobox_style)

        self.combo_categoria = QComboBox()
        layout_filtros.addRow("Categoria do Cliente:", self.combo_categoria)
        self.combo_categoria.setStyleSheet(combobox_style)

        self.date_de = QDateEdit()
        self.date_de.setDate(QDate.currentDate().addMonths(-1))
        self.date_de.setCalendarPopup(True)
        

        self.date_ate = QDateEdit()
        self.date_ate.setDate(QDate.currentDate())
        self.date_ate.setCalendarPopup(True)

        layout_filtros.addRow("Última Compra - De:", self.date_de)
        layout_filtros.addRow("Última Compra - Até:", self.date_ate)

        self.combo_origem = QComboBox()
        layout_filtros.addRow("Origem do Cliente:", self.combo_origem)
        self.combo_origem.setStyleSheet(combobox_style)

        grupo_filtros.setLayout(layout_filtros)
        layout_principal.addWidget(grupo_filtros)

        self.carregar_opcoes_combo_fisicos("Categoria do Cliente", self.combo_categoria)
        self.carregar_opcoes_combo_fisicos("Origem do Cliente", self.combo_origem)

        grupo_campos = QGroupBox("Informações a Incluir no Relatório")
        layout_campos = QVBoxLayout()

        self.combo_selecionar_todos = QCheckBox("Selecionar Todos")
        self.combo_selecionar_todos.setChecked(True)
        self.combo_selecionar_todos.stateChanged.connect(self.selecionar_todos_checkboxes_fisicos)
        layout_campos.addWidget(self.combo_selecionar_todos)

        checks = [
            "Nome",  "Data da Inclusão", "CPF","RG",
            "Contato", "CEP", "Endereço", "Número", "Complemento", "Cidade", "Bairro", "Estado",
            "Status", "Categoria", "Última Atualização", "Origem",
            "Valor Gasto Total", "Última Compra"
        ]

        self.checkboxes_relatorio_fisicos = []

        for texto in checks:
            check = QCheckBox(texto)
            check.setChecked(True)
            self.checkboxes_relatorio_fisicos.append(check)
            layout_campos.addWidget(check)

        grupo_campos.setLayout(layout_campos)
        layout_principal.addWidget(grupo_campos)

        layout_botoes = QHBoxLayout()
        btn_gerar = QPushButton("Gerar Relatório em PDF")
        btn_gerar_excel = QPushButton("Gerar Relatório em Excel")
        btn_cancelar = QPushButton("Cancelar")
        btn_gerar.setStyleSheet(button_style)
        btn_gerar_excel.setStyleSheet(button_style)
        btn_cancelar.setStyleSheet(button_style)

        layout_botoes.addStretch()
        layout_botoes.addWidget(btn_gerar)
        layout_botoes.addWidget(btn_gerar_excel)
        layout_botoes.addWidget(btn_cancelar)
        layout_principal.addLayout(layout_botoes)

        btn_cancelar.clicked.connect(self.janela_historico_clientes_fisicos.close)
        btn_gerar.clicked.connect(self.gerar_pdf_clientes_fisicos)
        btn_gerar_excel.clicked.connect(self.gerar_excel_clientes_fisicos)

        scroll.setWidget(central_widget)
        scroll.setStyleSheet(scroll_style)
        self.janela_historico_clientes_fisicos.setCentralWidget(scroll)
        self.janela_historico_clientes_fisicos.setStyleSheet(f"background-color: {bg_cor}; color: {text_cor};")
        self.janela_historico_clientes_fisicos.show()
        
        configuracoes = Configuracoes_Login(self.main_window)
        if  configuracoes.nao_mostrar_mensagem_arquivo_excel_fisicos:
            return
        
        # AVISO: Recomendação
        aviso = QMessageBox(self)
        aviso.setWindowTitle("Aviso")
        aviso.setIcon(QMessageBox.Information) # Ícone de aviso
        aviso.setText("Para uma melhor experiência recomendamos a geração do relatório em Excel.")
        aviso.setStandardButtons(QMessageBox.Ok)
        
        checkbox_nao_mostrar = QCheckBox("Não mostrar essa mensagem novamente")
        aviso.setCheckBox(checkbox_nao_mostrar)
        aviso.exec()

        # Verifica se o usuário marcou a opção para não mostrar novamente
        if checkbox_nao_mostrar.isChecked():
                configuracoes.nao_mostrar_mensagem_arquivo_excel_fisicos = True # Define que o usuário não quer mais ver este aviso
                configuracoes.salvar(configuracoes.usuario,configuracoes.senha,configuracoes.mantem_conectado)


    def gerar_pdf_clientes_fisicos(self):
        status = self.combo_status.currentText()
        categoria = self.combo_categoria.currentText()
        origem = self.combo_origem.currentText()
        data_de = self.date_de.date().toPython()
        data_ate = self.date_ate.date().toPython()

        campos_selecionados = [
            checkbox.text()
            for checkbox in self.checkboxes_relatorio_fisicos
            if checkbox.isChecked()
        ]

        if not campos_selecionados:
            QMessageBox.warning(self, "Aviso", "Selecione ao menos um campo para incluir no relatório.")
            return

        mapeamento_colunas = {
            "Nome": "Nome",
            "Data da Inclusão": "Data da Inclusão",
            "CPF": "CPF",
            "RG":"RG",
            "Contato": "Telefone",
            "CEP": "CEP",
            "Endereço": "Endereço",
            "Número": "Número",
            "Complemento": "Complemento",
            "Cidade": "Cidade",
            "Bairro": "Bairro",
            "Estado": "Estado",
            "Status": "Status",
            "Categoria": "Categoria",
            "Última Atualização": "Última Atualização",
            "Origem": "Origem",
            "Valor Gasto Total": "Valor Gasto Total",
            "Última Compra": "Última Compra"
        }

        colunas_sql = [f'"{mapeamento_colunas[nome]}"' for nome in campos_selecionados]
        sql = f"SELECT {', '.join(colunas_sql)} FROM clientes_fisicos WHERE 1=1 "

        params = []
        if status != "Todos":
            sql += " AND `Status do Cliente` = ?"
            params.append(status)
        if categoria != "Todos":
            sql += " AND `Categoria do Cliente` = ?"
            params.append(categoria)
        if origem != "Todos":
            sql += " AND `Origem do Cliente` = ?"
            params.append(origem)

        sql += """
            AND (
                `Última Compra` = 'Não Cadastrado' OR
                (
                    length(`Última Compra`) = 10 AND
                    substr(`Última Compra`, 3, 1) = '/' AND
                    substr(`Última Compra`, 6, 1) = '/' AND
                    substr(`Última Compra`, 7, 4) || '-' || 
                    substr(`Última Compra`, 4, 2) || '-' || 
                    substr(`Última Compra`, 1, 2)
                    BETWEEN ? AND ?
                )
            )
        """

        params.append(data_de.strftime("%Y-%m-%d"))
        params.append(data_ate.strftime("%Y-%m-%d"))

        try:
            conn = sqlite3.connect("banco_de_dados.db")
            cursor = conn.cursor()
            cursor.execute(sql, params)
            resultados = cursor.fetchall()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao buscar dados: \n{e}")
            return

        if not resultados:
            QMessageBox.information(self, "Aviso", "Nenhum cliente encontrado com os filtros aplicados")
            return

        caminho_pdf, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Relatório",
            "relatorio_clientes_fisicos.pdf",
            "Arquivos PDF (*.pdf)"
        )

        if not caminho_pdf:
            return

        try:
            pdf = FPDF(orientation="L", format="A4")  # Cria um objeto PDF com orientação paisagem ("L") e tamanho A4
            pdf.set_left_margin(5) # Define as margens do PDF
            pdf.set_right_margin(5) # Define as margens do PDF
            pdf.add_page() # Adiciona uma página ao PDF, obrigatória antes de começar a escrever
            pdf.set_font("Arial", size=8) # Define a fonte para o texto:
            # Cria uma célula que ocupa toda a largura (0 = largura total da página),
            # com altura 10, com o texto "Relatório de Clientes Físicos",
            # quebra de linha após a célula (ln=True) e alinhamento centralizado
            pdf.cell(0, 10, "Relatório de Clientes Fisícos", ln=True, align="C")
            pdf.ln(5) # Insere uma linha em branco com altura 5 para espaçamento vertical

            pdf.set_font("Arial", size=5) # tamanho da fonte
            cel_height = 5 # Define a altura das células que serão usadas para o cabeçalho da tabela


            larguras_automaticas = {}
            margem = 4  # margem interna da célula, para deixar texto mais confortável
            largura_minima = 10
            largura_maxima = 80  # máximo para evitar colunas absurdamente largas

            for i, campo in enumerate(campos_selecionados):
                # largura do cabeçalho
                largura_cabecalho = pdf.get_string_width(campo) + margem

                # largura máxima entre todos os textos daquela coluna
                largura_max = largura_cabecalho

                for linha in resultados:
                    item = linha[i]
                    texto = str(item) if item is not None else ""
                    largura_texto = pdf.get_string_width(texto) + margem
                    if largura_texto > largura_max:
                        largura_max = largura_texto

                # Limita largura para mínimo e máximo
                if largura_max < largura_minima:
                    largura_max = largura_minima
                elif largura_max > largura_maxima:
                    largura_max = largura_maxima

                larguras_automaticas[campo] = largura_max

            # Cabeçalho
            for campo in campos_selecionados:
                pdf.cell(larguras_automaticas[campo], cel_height, campo, border=1, align="C")
            pdf.ln()

            # Dados
            for linha in resultados:
                y_inicial = pdf.get_y()
                max_altura = cel_height
                for i, item in enumerate(linha):
                    campo = campos_selecionados[i]
                    largura = larguras_automaticas[campo]
                    texto = str(item) if item is not None else ""
                    x = pdf.get_x()
                    y = pdf.get_y()

                    # Usar multi_cell para texto que possa ter quebras, mas com largura suficiente para minimizar linhas
                    pdf.multi_cell(largura, cel_height, texto, border=1, align="L")
                    altura = pdf.get_y() - y
                    if altura > max_altura:
                        max_altura = altura
                    pdf.set_xy(x + largura, y)
                pdf.set_y(y_inicial + max_altura)
            pdf.output(caminho_pdf)
            QMessageBox.information(self, "Sucesso", "Relatório gerado com sucesso.")

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao gerar PDF:\n{e}")

    def gerar_excel_clientes_fisicos(self):
        status = self.combo_status.currentText()
        categoria = self.combo_categoria.currentText()
        origem = self.combo_origem.currentText()
        data_de = self.date_de.date().toPython()
        data_ate = self.date_ate.date().toPython()

        campos_selecionados = [
            checkbox.text()
            for checkbox in self.checkboxes_relatorio_fisicos
            if checkbox.isChecked()
        ]

        if not campos_selecionados:
            QMessageBox.warning(self, "Aviso", "Selecione ao menos um campo para incluir no relatório.")
            return

        mapeamento_colunas = {
            "Nome": "Nome",
            "Data da Inclusão": "Data da Inclusão",
            "CPF": "CPF",
            "RG":"RG", 
            "Contato": "Telefone",
            "CEP": "CEP",
            "Endereço": "Endereço",
            "Número": "Número",
            "Complemento": "Complemento",
            "Cidade": "Cidade",
            "Bairro": "Bairro",
            "Estado": "Estado",
            "Status": "Status",
            "Categoria": "Categoria",
            "Última Atualização": "Última Atualização",
            "Origem": "Origem",
            "Valor Gasto Total": "Valor Gasto Total",
            "Última Compra": "Última Compra"
        }

        colunas_sql = [f'"{mapeamento_colunas[nome]}"' for nome in campos_selecionados]
        sql = f"SELECT {', '.join(colunas_sql)} FROM clientes_fisicos WHERE 1=1 "

        params = []
        
        if status != "Todos":
            sql += " AND `Status do Cliente` = ?"
            params.append(status)
        if categoria != "Todos":
            sql += " AND `Categoria do Cliente` = ?"
            params.append(categoria)
        if origem != "Todos":
            sql += " AND `Origem do Cliente` = ?"
            params.append(origem)

        sql += """
            AND (
                `Última Compra` = 'Não Cadastrado' OR
                (
                    length(`Última Compra`) = 10 AND
                    substr(`Última Compra`, 3, 1) = '/' AND
                    substr(`Última Compra`, 6, 1) = '/' AND
                    substr(`Última Compra`, 7, 4) || '-' || 
                    substr(`Última Compra`, 4, 2) || '-' || 
                    substr(`Última Compra`, 1, 2)
                    BETWEEN ? AND ?
                )
            )
        """
        params.append(data_de.strftime("%Y-%m-%d"))
        params.append(data_ate.strftime("%Y-%m-%d"))

        try:
            conn = sqlite3.connect("banco_de_dados.db")
            cursor = conn.cursor()
            cursor.execute(sql, params)
            resultados = cursor.fetchall()
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao buscar dados: \n{e}")
            return

        if not resultados:
            QMessageBox.information(self, "Aviso", "Nenhum cliente encontrado com os filtros aplicados")
            return

        caminho_excel, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Relatório",
            "relatorio_clientes_fisicos.xlsx",
            "Arquivos Excel (*.xlsx)"
        )

        if not caminho_excel:
            return

        try:
            # Criar um DataFrame do pandas com os dados
            df = pd.DataFrame(resultados,columns=campos_selecionados)
            
            # Salva como Excel usando openpyxl
            sheet_name = "Relatório de Clientes Físicos"
            df.to_excel(caminho_excel,index=False,engine='openpyxl',sheet_name=sheet_name)
            
            # Reabre o Excel com openpyxl para aplicar formatações
            wb = load_workbook(caminho_excel)
            ws = wb[sheet_name]
            
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length,len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2  # Ajusta a largura da coluna    
            wb.save(caminho_excel)
            QMessageBox.information(self, "Sucesso", "Relatório Excel gerado com sucesso.")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao gerar Excel:\n{e}")


    def selecionar_todos_checkboxes_fisicos(self,estado):
        for checkbox in self.checkboxes_relatorio_fisicos:
            checkbox.setChecked(bool(estado))

    def carregar_opcoes_combo_fisicos(self, nome_coluna, combo_box):
        try:
            conn = sqlite3.connect("banco_de_dados.db")  # substitua com seu caminho
            cursor = conn.cursor()

            # Aspas duplas se o nome da coluna tiver espaços
            cursor.execute(f'''
                SELECT DISTINCT "{nome_coluna}"
                FROM clientes_fisicos
                WHERE "{nome_coluna}" IS NOT NULL AND "{nome_coluna}" != ''
            ''')
            resultados = cursor.fetchall()

            opcoes = sorted(set([row[0] for row in resultados]))
            combo_box.clear()
            combo_box.addItem("Todos")
            combo_box.addItems(opcoes)

        except Exception as e:
            print(f"Erro ao carregar opções de {nome_coluna}:", e)

    def confirmar_historico_apagado_clientes_fisicos(self, mensagem):
        """
        Exibe uma caixa de diálogo para confirmar a exclusão.
        """
        msgbox = QMessageBox(self)
        msgbox.setWindowTitle("Confirmação")
        msgbox.setText(mensagem)

        btn_sim = QPushButton("Sim")
        btn_nao = QPushButton("Não")
        msgbox.addButton(btn_sim, QMessageBox.ButtonRole.YesRole)
        msgbox.addButton(btn_nao, QMessageBox.ButtonRole.NoRole)

        msgbox.setDefaultButton(btn_nao)
        resposta = msgbox.exec()

        return msgbox.clickedButton() == btn_sim
    
    def configurar_menu_contexto_fisicos(self):
        self.table_clientes_fisicos.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_clientes_fisicos.customContextMenuRequested.connect(self.abrir_menu_contexto_fisicos)
    
    def abrir_menu_contexto_fisicos(self,pos:QPoint):
        index = self.table_clientes_fisicos.indexAt(pos)
        if not index.isValid():
            return # Clicou fora de uma célula
        
        # Captura os dados da linha
        linha = index.row()
        nome_cliente = self.table_clientes_fisicos.item(linha,0).text()  # Coluna 0 = Nome do Cliente

        menu = QMenu()

        detalhes_action = menu.addAction("Detalhes")
        editar_action = menu.addAction("Editar Cliente")
        excluir_action = menu.addAction("Excluir Cliente")


        action = menu.exec(self.table_clientes_fisicos.viewport().mapToGlobal(pos))

        # Executa conforme a opção clicada
        if action == detalhes_action:
            QMessageBox.information(self, "Aviso", "Essa função ainda não está disponível")
        elif action == editar_action:
            self.editar_cliente_fisico()
        elif action == excluir_action:
            self.excluir_clientes_fisicos()